from __future__ import annotations

import tkinter as tk
from tkinter import messagebox, filedialog, Menu, END
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from pipeline import DataPipeline
import sys
import traceback
import pandas as pd
import sqlite3
from core.logging_utils import log_exception
from core.platform import configure_macos_scaling
from core.window import center_window, center_window_smart, maximize_window
from core.icons import load_logo_photo, safe_iconphoto
from core.security import hash_pass, LOGIN_USER, LOGIN_PASS
from core import *  # şimdilik hızlı çözüm; sonra tek tek isimle temizleriz

class App(ttk.Window):
    def __init__(self):
        super().__init__(themename="flatly")
        configure_macos_scaling(self)
        # TEK PENCERE LOGIN: En stabil yaklaşım (Toplevel/grab/topmost yok)
        self.title(f"Giriş - Leta  |  {APP_VERSION}")
        center_window(self, 520, 420)
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.report_callback_exception = self._tk_report_callback_exception
        self.bind_class("Toplevel", "<Map>", self._auto_adjust_mapped_toplevel, add="+")
        # Logo (varsa) - GC olmaması için referans tutuyoruz
        self._logo_small = load_logo_photo(28, 28)
        self._logo_big = load_logo_photo(80, 80)
        self._logo_icon = load_logo_photo(64, 64)
        self._toolbar_logo_lbl = None
        self._login_logo_lbl = None
        safe_iconphoto(self, self._logo_icon)

        # giriş yapan kullanıcı
        self.kullanici = None  # (id, kullanici_adi, ad_soyad, yetki, terapist_adi)
        self.kullanici_yetki = "kurum_muduru"
        self.kullanici_terapist = None

        self._vcmd_money = (self.register(self._validate_money), "%P")
        self._build_login_ui()
        # Windows odak sorunlarına karşı: kısa süre üste getir
        try:
            self.lift()
            self.focus_force()
            self.attributes("-topmost", True)
            self.after(250, lambda: self.attributes("-topmost", False))
        except Exception:
            pass

    

    # NOT: class gövdesinde "self.report_callback_exception = ..." gibi satır OLMAYACAK.
    # bunu __init__ içinde zaten yapıyoruz.

    


    
    def _on_close(self):
        self.destroy()

    def _fit_window_to_screen(self, win):
        """Pencereyi ekran sınırlarına sığdır ve ortala (tüm popup'lar görünür kalsın)."""
        try:
            win.update_idletasks()
            sw = max(1, int(win.winfo_screenwidth() or 1))
            sh = max(1, int(win.winfo_screenheight() or 1))

            ww = max(320, int(max(win.winfo_width() or 0, win.winfo_reqwidth() or 0, 320)))
            wh = max(220, int(max(win.winfo_height() or 0, win.winfo_reqheight() or 0, 220)))

            max_w = max(320, sw - 60)
            max_h = max(220, sh - 100)
            ww = min(ww, max_w)
            wh = min(wh, max_h)

            x = max(0, (sw - ww) // 2)
            y = max(0, (sh - wh) // 2)
            win.geometry(f"{ww}x{wh}+{x}+{y}")
            try:
                win.minsize(min(ww, 420), min(wh, 260))
            except Exception:
                pass
        except Exception:
            pass

    def _auto_adjust_mapped_toplevel(self, event=None):
        """Açılan popup'ları otomatik ekrana sığdırır."""
        try:
            w = event.widget if event is not None else None
            if not isinstance(w, tk.Toplevel):
                return
            if w is self:
                return
            if getattr(w, "_auto_screen_fitted", False):
                return
            try:
                w.resizable(True, True)
            except Exception:
                pass
            self._fit_window_to_screen(w)
            w._auto_screen_fitted = True
        except Exception:
            pass
    def popup_personel_avans(self, **kwargs):
        """Personel Avans/Ekstra Ödeme Penceresi"""
        win = ttk.Toplevel(self)
        win.title("Personel Avans / Ödeme Ver")
        center_window(win, 400, 350)
        
        ttk.Label(win, text="Personel Seç:", font=("Segoe UI", 10, "bold")).pack(pady=5)
        # Terapist listesini settings'den çekiyoruz
        c_personel = ttk.Combobox(win, values=DEFAULT_THERAPISTS, width=30, state="readonly")
        conn_avans = None
        try:
            conn_avans = self.veritabani_baglan()
            cur = conn_avans.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1")
            lst = [r[0] for r in cur.fetchall()]
            c_personel["values"] = lst
        except Exception:
            pass
        finally:
            try:
                if conn_avans is not None:
                    conn_avans.close()
            except Exception:
                pass
        c_personel.pack(pady=5)

        ttk.Label(win, text="Tutar (TL):", font=("Segoe UI", 10, "bold")).pack(pady=5)
        e_tutar = ttk.Entry(win)
        e_tutar.pack(pady=5)

        ttk.Label(win, text="İşlem Türü:", font=("Segoe UI", 10)).pack(pady=5)
        c_tur = ttk.Combobox(win, values=["Avans", "Maaş Ödemesi", "Prim", "Yol/Yemek"], state="readonly")
        c_tur.current(0) # Varsayılan: Avans
        c_tur.pack(pady=5)

        ttk.Label(win, text="Kısa Açıklama:", font=("Segoe UI", 10)).pack(pady=5)
        e_not = ttk.Entry(win)
        e_not.pack(pady=5)

        def kaydet():
            personel = c_personel.get()
            try:
                tutar = float(e_tutar.get())
            except Exception:
                messagebox.showerror("Hata", "Geçerli bir tutar girin.")
                return

            if not personel or tutar <= 0:
                messagebox.showwarning("Eksik", "Personel seçin ve tutar girin.")
                return

            conn_kaydet = None
            try:
                conn_kaydet = self.veritabani_baglan()
                pipeline = DataPipeline(conn_kaydet, self.kullanici[0] if self.kullanici else None)
                pipeline.personel_harici_islem(personel, tutar, islem_turu=c_tur.get(), aciklama=e_not.get())
                messagebox.showinfo("Başarılı", f"{personel} adına {tutar} TL {c_tur.get()} çıkışı yapıldı.\nKasa defterine işlendi.")
                win.destroy()
            except Exception as e:
                messagebox.showerror("Hata", str(e))
            finally:
                try:
                    if conn_kaydet is not None:
                        conn_kaydet.close()
                except Exception:
                    pass

        ttk.Button(win, text="KASADAN ÖDE", bootstyle="danger", command=kaydet).pack(pady=15)
    def _kasa_secili_sil(self, parent):
        """Kasa defterinden seçili satırı güvenli şekilde sil."""
        tree = getattr(parent, "_tree_kasa", None) or getattr(self, "_tree_kasa", None)
        if not tree:
            messagebox.showerror("Hata", "Kasa tablosu bulunamadı.")
            return

        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Silmek için bir satır seç.")
            return

        values = tree.item(sel[0], "values") or []
        if not values:
            messagebox.showerror("Hata", "Seçili satır okunamadı.")
            return

        try:
            hareket_id = int(values[0])
        except Exception:
            messagebox.showerror("Hata", f"Geçersiz ID: {values[0]!r}")
            return

        if not messagebox.askyesno("Onay", f"Kasa hareketi silinsin mi?\nID: {hareket_id}"):
            return

        conn = None
        try:
            conn = self.veritabani_baglan()
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            ok = bool(pipeline.kasa_hareketi_sil(hareket_id))
            conn.close()

            if not ok:
                messagebox.showerror("Hata", "Silme başarısız. Kayıt bulunamadı veya silinemedi.")
                return

            try:
                if hasattr(self, "_kasa_rapor_yukle"):
                    self._kasa_rapor_yukle(parent)
                elif hasattr(self, "kasa_rapor_yukle"):
                    self.kasa_rapor_yukle(parent)
            except Exception:
                pass

            self._refresh_open_reports()
            messagebox.showinfo("OK", "Kasa hareketi silindi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Silme başarısız:\n{e}")
            log_exception("_kasa_secili_sil", e)
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass


    def _kasa_popup_odeme_ekle(self, parent):
     tree = getattr(parent, "_tree_kasa", None)
     ent_tarih = getattr(parent, "_ent_tarih", None)

     if not tree:
        messagebox.showerror("Hata", "Kasa tablosu hazır değil.")
        return

     sel = tree.selection()
     if not sel:
        messagebox.showwarning("Uyarı", "Ödeme eklemek için bir satır seç.")
        return

     values = tree.item(sel[0], "values") or []
     if len(values) < 7:
        messagebox.showerror("Hata", "Seçili satır formatı beklenenden farklı.")
        return

     kayit_txt = str(values[6] or "").strip()
     record_id = None
     if kayit_txt.startswith("Kayıt #"):
        try:
            record_id = int(kayit_txt.replace("Kayıt #", "").strip())
        except Exception:
            record_id = None

     if not record_id:
        messagebox.showwarning(
            "Uyarı",
            "Bu kasa hareketi bir 'Kayıt #...' ile bağlı değil.\n"
            "Ödeme ekleme için Seans/Record bağlı satır seçmelisin.",
        )
        return

     default_tarih = str(values[1] or "").strip()
     if not default_tarih:
        default_tarih = (ent_tarih.get().strip() if ent_tarih and hasattr(ent_tarih, "get") else "")
     if not default_tarih:
        default_tarih = datetime.datetime.now().strftime("%Y-%m-%d")

     win = ttk.Toplevel(self)
     win.title("Ödeme Ekle")
     win.geometry("420x220")
     win.transient(self)
     win.grab_set()

     frm = ttk.Frame(win, padding=12)
     frm.pack(fill="both", expand=True)

     ttk.Label(frm, text=f"İlgili Kayıt: #{record_id}", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 10))

     row1 = ttk.Frame(frm)
     row1.pack(fill="x", pady=5)
     ttk.Label(row1, text="Tarih:", width=10).pack(side="left")
     e_tarih = ttk.Entry(row1)
     e_tarih.pack(side="left", fill="x", expand=True)
     e_tarih.insert(0, default_tarih)

     row2 = ttk.Frame(frm)
     row2.pack(fill="x", pady=5)
     ttk.Label(row2, text="Tutar:", width=10).pack(side="left")
     e_tutar = ttk.Entry(row2, validate="key", validatecommand=self._vcmd_money)
     e_tutar.pack(side="left", fill="x", expand=True)
     e_tutar.insert(0, "")

     row3 = ttk.Frame(frm)
     row3.pack(fill="x", pady=5)
     ttk.Label(row3, text="Ödeme:", width=10).pack(side="left")
     cb = ttk.Combobox(row3, values=["Nakit", "Kart", "Havale", "EFT", "Diğer"], state="readonly")
     cb.pack(side="left", fill="x", expand=True)
     cb.set("Nakit")

     btns = ttk.Frame(frm)
     btns.pack(fill="x", pady=(15, 0))

    def _ok():
        tarih = (e_tarih.get() or "").strip()
        try:
            tutar = float((e_tutar.get() or "0").replace(",", "."))
        except Exception:
            tutar = 0.0
        odeme_sekli = (cb.get() or "").strip()

        if not tarih:
            messagebox.showerror("Hata", "Tarih boş olamaz.", parent=win)
            return
        if tutar <= 0:
            messagebox.showerror("Hata", "Tutar > 0 olmalı.", parent=win)
            return

        ok = False
        try:
            ok = bool(self.pipeline.odeme_ekle(record_id, tutar, tarih, odeme_sekli, "Kasa Defteri Üzerinden"))
        except Exception:
            ok = False

        if not ok:
            messagebox.showerror("Hata", "Ödeme eklenemedi. (Log'a bak)", parent=win)
            return

        try:
            self._kasa_rapor_yukle(parent)
        except Exception:
            pass

        win.destroy()
        messagebox.showinfo("OK", "Ödeme eklendi ve kasa güncellendi.")

        ttk.Button(btns, text="Kaydet", bootstyle="success", command=self.safe_call(_ok, "kasa_odeme_ok")).pack(side="left", padx=5)
        ttk.Button(btns, text="İptal", bootstyle="secondary", command=win.destroy).pack(side="left", padx=5)

    def girise_don(self):
        """Giriş ekranına dön (logout) - bilgisiz kullanıcı için basit."""
        try:
            # kullanıcıyı temizle
            self.kullanici = None
            self.kullanici_yetki = "kurum_muduru"
            self.kullanici_terapist = None

            # tüm mevcut UI'ı kaldır
            for w in list(self.winfo_children()):
                try:
                    w.destroy()
                except Exception:
                    pass
            try:
                self.config(menu="")
            except Exception:
                pass

            # giriş ekranını tekrar kur
            self.title(f"Giriş - Leta  |  {APP_VERSION}")
            center_window_smart(self, 520, 420, max_ratio=0.85)
            self.resizable(False, False)
            self._build_login_ui()
        except Exception:
            # worst-case: uygulamayı kapat
            try:
                self.destroy()
            except Exception:
                pass

    def sistemi_sifirla(self):
        """Kurum müdürü: yedek al + DB'yi direkt sil + temiz başlat + Danışan-Hoca bazlı fiyatlandırma kurulumu."""
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Yetki", "Bu işlem sadece Kurum Müdürü tarafından yapılabilir.")
            return

        msg = (
            "Bu işlem veritabanını SIFIRLAR.\n\n"
            "• Tüm kayıtlar silinir (DB dosyası).\n"
            "• Önce otomatik yedek alınır.\n"
            "• Sıfırlama sonrası danışan-hoca bazlı fiyatlandırma kurulumu yapılacak.\n\n"
            "Devam etmek istiyor musunuz?"
        )
        if not messagebox.askyesno("Sistemi Sıfırla", msg):
            return
        if not messagebox.askyesno("Son Onay", "Emin misiniz? Bu işlem geri alınamaz."):
            return

        try:
            # 1) Yedek al
            backup_path = backup_now(prefix="reset_before_delete")
            
            # 2) Tüm bağlantıları kapat
            try:
                # Mevcut bağlantıları kapat
                if hasattr(self, 'conn') and self.conn:
                    try:
                        self.conn.close()
                    except Exception:
                        pass
            except Exception:
                pass
            
            # 3) DB dosyalarını direkt sil (CMD açmadan)
            db = db_path()
            files_to_delete = [db, db + "-wal", db + "-shm"]
            deleted_count = 0
            
            for f in files_to_delete:
                try:
                    if os.path.exists(f):
                        # WAL checkpoint yap (eğer DB dosyasıysa)
                        if f == db:
                            try:
                                conn_temp = sqlite3.connect(db)
                                conn_temp.execute("PRAGMA wal_checkpoint(FULL);")
                                conn_temp.close()
                            except Exception:
                                pass
                        os.remove(f)
                        deleted_count += 1
                except Exception as e:
                    log_exception("sistemi_sifirla_delete", e)
            
            # 4) Kullanıcı bilgilendir
            if backup_path:
                messagebox.showinfo(
                    "Başarılı",
                    f"✅ Sistem sıfırlandı!\n\n"
                    f"Yedek alındı:\n{backup_path}\n\n"
                    f"Silinen dosyalar: {deleted_count}\n\n"
                    f"Uygulama yeniden başlatılıyor..."
                )
            else:
                messagebox.showinfo(
                    "Başarılı",
                    f"✅ Sistem sıfırlandı!\n\n"
                    f"Silinen dosyalar: {deleted_count}\n\n"
                    f"Uygulama yeniden başlatılıyor..."
                )
            
            # 5) Uygulamayı yeniden başlat (basit yöntem: destroy + init_db çağrılacak)
            try:
                # Giriş ekranına dön
                self.girise_don()
                # Veritabanını yeniden oluştur
                init_db()
                
                # 6) Danışan-Hoca bazlı fiyatlandırma kurulum penceresi aç
                messagebox.showinfo("Bilgi", "Temiz veritabanı oluşturuldu. Şimdi danışan-hoca bazlı fiyatlandırma kurulumu yapılacak.")
                self._danisan_hoca_fiyatlandirma_kurulum()
            except Exception as e:
                log_exception("sistemi_sifirla_restart", e)
                messagebox.showwarning("Uyarı", f"Veritabanı silindi ama yeniden başlatma sırasında hata oluştu:\n{e}\n\nLütfen uygulamayı manuel olarak yeniden başlatın.")
                
        except Exception as e:
            messagebox.showerror("Hata", f"Sistem sıfırlama hatası:\n{e}")
            log_exception("sistemi_sifirla", e)

    def _build_login_ui(self):
        self.login_frame = ttk.Frame(self, padding=18)
        self.login_frame.pack(fill=BOTH, expand=True)

        header = ttk.Frame(self.login_frame)
        header.pack(fill=X, pady=(6, 10))
        if getattr(self, "_logo_big", None):
            self._login_logo_lbl = ttk.Label(header, image=self._logo_big)
            self._login_logo_lbl.pack(side=LEFT, padx=(0, 12))
        title_box = ttk.Frame(header)
        title_box.pack(side=LEFT, fill=X, expand=True)
        ttk.Label(title_box, text="Leta Aile ve Çocuk", font=("Segoe UI", 18, "bold"), bootstyle="primary").pack(anchor=W)
        ttk.Label(title_box, text="Seans ve Ücret Takip", font=("Segoe UI", 11), foreground="gray").pack(anchor=W, pady=(2, 0))

        ttk.Label(self.login_frame, text="Kullanıcı Adı:").pack(anchor=W)
        self.login_user = ttk.Entry(self.login_frame)
        self.login_user.pack(fill=X, pady=(4, 10))

        ttk.Label(self.login_frame, text="Şifre:").pack(anchor=W)
        # Bazı makinelerde tema/maskeleme yüzünden "yazamıyorum" sanılabiliyor.
        # Stabilite için burada maskelemeyi kapatıyoruz; istersen sonra tekrar "*" yaparız.
        self.login_pass = ttk.Entry(self.login_frame, show="*")  # Şifre gizli
        self.login_pass.pack(fill=X, pady=(4, 6))

        self.lbl_pw = ttk.Label(self.login_frame, text="Şifre uzunluğu: 0", foreground="gray")
        self.lbl_pw.pack(anchor=W, pady=(0, 10))

        self.btn_login = ttk.Button(self.login_frame, text="GİRİŞ YAP", bootstyle="success", command=self._do_login)
        self.btn_login.pack(fill=X)
        self.btn_register = ttk.Button(self.login_frame, text="KAYIT OL", bootstyle="info-outline", command=self._open_register)
        self.btn_register.pack(fill=X, pady=(8, 0))
        self.btn_first = ttk.Button(self.login_frame, text="İLK KURULUM (Kurum Müdürü Oluştur)", bootstyle="warning", command=self._first_setup)
        self.btn_first.pack(fill=X, pady=(10, 0))
        ttk.Label(self.login_frame, text=f"Sürüm: {APP_VERSION}  •  Build: {APP_BUILD}", font=("Segoe UI", 8), foreground="gray").pack(pady=(10, 0))
        self.lbl_hint = ttk.Label(self.login_frame, text="", font=("Segoe UI", 8), foreground="gray")
        self.lbl_hint.pack(pady=(2, 0))

        self.login_user.bind("<Return>", lambda e: self.login_pass.focus_set())
        self.login_pass.bind("<Return>", lambda e: self._do_login())
        self.login_pass.bind("<KeyRelease>", lambda e: self.lbl_pw.config(text=f"Şifre uzunluğu: {len(self.login_pass.get())}"))
        self.login_user.bind("<Button-1>", lambda e: self.login_user.focus_set())
        self.login_pass.bind("<Button-1>", lambda e: self.login_pass.focus_set())
        self.after(150, self.login_pass.focus_set)
        self.after(10, self._refresh_first_run_state)

    def _open_register(self):
        RegisterDialog(self, first_setup=False)

    def _first_setup(self):
        RegisterDialog(self, first_setup=True)

    def _refresh_first_run_state(self):
        """İlk kurulumda (DB boş) kurum müdürü oluşturulmadan giriş kapalı olsun."""
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM users WHERE is_active=1")
            n = int((cur.fetchone() or [0])[0] or 0)
            conn.close()
        except Exception:
            n = 0

        if n <= 0:
            # henüz hiç kullanıcı yok
            try:
                self.btn_login.configure(state="disabled")
            except Exception:
                pass
            try:
                self.btn_register.pack_forget()
            except Exception:
                pass
            try:
                self.btn_first.configure(state="normal")
            except Exception:
                pass
            try:
                self.lbl_hint.config(text="İlk kullanım: önce kurum müdürü hesabı oluşturun.")
            except Exception:
                pass
        else:
            try:
                self.btn_login.configure(state="normal")
            except Exception:
                pass
            # kayıt ol butonu görünür kalsın
            try:
                if not self.btn_register.winfo_ismapped():
                    self.btn_register.pack(fill=X, pady=(8, 0))
            except Exception:
                pass
            try:
                self.btn_first.configure(state="disabled")
            except Exception:
                pass
            try:
                self.lbl_hint.config(text="Kayıt olmak için 'KAYIT OL' kullanabilirsiniz.")
            except Exception:
                pass

    def _do_login(self):
        u = self.login_user.get().strip()
        p = self.login_pass.get()
        # DB üzerinden doğrula
        ok = False
        user_row = None
        try:
            conn = self.veritabani_baglan()
            self.pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
            cur = conn.cursor()
            cur.execute(
                "SELECT id, username, full_name, COALESCE(access_role, role, 'egitim_gorevlisi'), therapist_name, password_hash FROM users WHERE is_active=1 AND username=?",
                (u,),
            )
            row = cur.fetchone()
            if row and (row[5] or "") == hash_pass(p):
                ok = True
                user_row = (row[0], row[1], row[2] or row[1], "kurum_muduru", row[4])
                try:
                    cur.execute(
                        "UPDATE users SET last_login=? WHERE id=?",
                        (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), row[0]),
                    )
                    conn.commit()
                except Exception:
                    pass
            conn.close()
        except Exception:
            # fallback (eski sürüm)
            ok = (LOGIN_USER != "" and u == LOGIN_USER and p == LOGIN_PASS)

        if not ok:
            messagebox.showerror("Hata", "Kullanıcı adı veya şifre yanlış!")
            return

        self.kullanici = user_row
        self.kullanici_yetki = "kurum_muduru"
        self.kullanici_terapist = (user_row[4] if user_row else None)

        # Login başarılı -> ana UI (hata olursa kullanıcıya göster + logla)
        try:
            self.login_frame.destroy()
            self.title(f"{APP_TITLE}  |  {APP_VERSION}")
            self.geometry("1400x800")
            self.resizable(True, True)
            try:
                self.state("zoomed")
            except Exception:
                pass

            self._build_toolbar()
            # Menu bar kaldırıldı - kullanıcı sadece butonlarla erişsin
            # self._build_menu()
            self._build_tabs()
            self.terapistleri_yukle()
            self._apply_user_restrictions()
            self.kayitlari_listele()
        except Exception as e:
            log_exception("UI_BUILD_ERROR", e)
            messagebox.showerror(
                "Kritik Hata",
                "Girişten sonra arayüz oluşturulamadı.\n\n"
                f"Hata logu: {error_log_path()}\n\n"
                "Lütfen bu log dosyasını bana gönderin.",
            )
            # En azından kullanıcı boş ekran görmesin:
            try:
                fallback = ttk.Frame(self, padding=20)
                fallback.pack(fill=BOTH, expand=True)
                ttk.Label(
                    fallback,
                    text="Arayüz yüklenemedi.\nLütfen yöneticinize haber verin.",
                    font=("Segoe UI", 14, "bold"),
                    bootstyle="danger",
                ).pack(pady=20)
                ttk.Label(fallback, text=f"Log: {error_log_path()}").pack()
            except Exception:
                pass

    def _build_toolbar(self):
        # Üst bar (taskbar alanı): sadece temel aksiyonlar (Girişe dön / Kapat / Kılavuz)
        bar = ttk.Frame(self, padding=10, bootstyle="dark")
        bar.pack(fill=X)
        self.toolbar = bar

        user_txt = "Kullanıcı"
        role_txt = ""
        try:
            if self.kullanici and len(self.kullanici) >= 3:
                user_txt = self.kullanici[2] or self.kullanici[1]
        except Exception:
            pass

        if getattr(self, "_logo_small", None):
            self._toolbar_logo_lbl = ttk.Label(bar, image=self._logo_small, bootstyle="inverse-dark")
            self._toolbar_logo_lbl.pack(side=LEFT, padx=(2, 10))
        ttk.Label(
            bar,
            text=f"{APP_TITLE} • {APP_VERSION}",
            font=("Segoe UI", 11, "bold"),
            bootstyle="inverse-dark",
        ).pack(side=LEFT, padx=(4, 14))

        ttk.Label(
            bar,
            text=f"{user_txt}",
            font=("Segoe UI", 10),
            bootstyle="inverse-dark",
        ).pack(side=LEFT, padx=(0, 14))

        # Sağ tarafta: en basit 3 buton
        ttk.Button(bar, text="KILAVUZ", bootstyle="secondary", command=self.kullanim_kilavuzu_ac).pack(side=RIGHT, padx=6)
        ttk.Button(bar, text="KAPAT", bootstyle="danger", command=self._on_close).pack(side=RIGHT, padx=6)
        ttk.Button(bar, text="GİRİŞ EKRANI", bootstyle="warning", command=self.girise_don).pack(side=RIGHT, padx=6)

    def _brand_window(self, win) -> None:
        """Logo varsa pencerede (alt-tab/taskbar) ikon olarak göster."""
        try:
            safe_iconphoto(win, self._logo_icon)
        except Exception:
            pass

    def _style_table_strong(self):
        """Tabloların satır/sütun ayrımını daha belirgin yap - KENARLIKLI."""
        try:
            s = ttk.Style()
            # Ana tablo stili - belirgin kenarlıklar
            s.configure(
                "Strong.Treeview",
                rowheight=30,
                borderwidth=2,
                relief="solid",
                font=("Segoe UI", 10),
                background="#FFFFFF",
                foreground="#000000",
            )
            # Başlık stili - kalın kenarlıklar
            s.configure(
                "Strong.Treeview.Heading",
                font=("Segoe UI", 10, "bold"),
                relief="solid",
                borderwidth=2,
                background="#E9ECEF",
                foreground="#000000",
            )
            # Hücre kenarlıkları için item stili
            s.map(
                "Strong.Treeview",
                background=[("selected", "#007BFF"), ("!selected", "#FFFFFF")],
            )
        except Exception:
            pass

    def _apply_stripes(self, tree):
        """Alternating row colors for readability."""
        try:
            tree.tag_configure("odd", background="#FFFFFF")
            tree.tag_configure("even", background="#F3F6FB")
        except Exception:
            pass

    def _update_sync_badge(self):
        """SEANS TAKİP ekranındaki 'Belirsiz: X' sayacını güncelle (best-effort)."""
        try:
            cnt = len(getattr(self, "_last_sync_ambiguous", []) or [])
        except Exception:
            cnt = 0
        try:
            if getattr(self, "_sync_badge_lbl", None) is not None:
                self._sync_badge_lbl.configure(text=f"Belirsiz: {cnt}")
        except Exception:
            pass

    def _default_saat(self) -> str:
        """Kullanıcı seçmezse: şu anki saat (dakika yuvarlayıp) HH:MM üret."""
        try:
            now = datetime.datetime.now()
            # dakika 0/30'a yuvarla
            m = 0 if now.minute < 15 else (30 if now.minute < 45 else 0)
            h = now.hour if now.minute < 45 else (now.hour + 1) % 24
            return f"{h:02d}:{m:02d}"
        except Exception:
            return "09:00"

    def _sync_from_record_to_seans(self, cur, record_id: int, tarih: str, saat: str, danisan: str, terapist: str, notlar: str):
        """records kaydı varsa seans_takvimi'ne bağla/oluştur."""
        try:
            cur.execute("SELECT seans_id FROM records WHERE id=?", (record_id,))
            seans_id = (cur.fetchone() or [None])[0]
        except Exception:
            seans_id = None

        # seans_id varsa: record_id yaz ve çık
        if seans_id:
            try:
                cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=? AND (record_id IS NULL OR record_id='')", (record_id, seans_id))
            except Exception:
                pass
            return

        # record_id ile daha önce seans var mı?
        try:
            cur.execute("SELECT id FROM seans_takvimi WHERE record_id=? ORDER BY id DESC LIMIT 1", (record_id,))
            row = cur.fetchone()
            if row and row[0]:
                sid = int(row[0])
                cur.execute("UPDATE records SET seans_id=? WHERE id=? AND (seans_id IS NULL OR seans_id='')", (sid, record_id))
                return
        except Exception:
            pass

        # Aynı tarih+saat+danisan+terapist varsa onu bağla, yoksa yeni oluştur
        sid = None
        try:
            cur.execute(
                """
                SELECT id FROM seans_takvimi
                WHERE tarih=? AND saat=? AND danisan_adi=? AND terapist=?
                ORDER BY id DESC LIMIT 1
                """,
                (tarih, saat, danisan, terapist),
            )
            row = cur.fetchone()
            if row and row[0]:
                sid = int(row[0])
                cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=? AND (record_id IS NULL OR record_id='')", (record_id, sid))
        except Exception:
            sid = None

        if not sid:
            try:
                cur.execute(
                    """
                    INSERT INTO seans_takvimi (tarih, saat, danisan_adi, terapist, oda, durum, notlar, olusturma_tarihi, olusturan_kullanici_id, record_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        tarih,
                        saat,
                        danisan,
                        terapist,
                        "",
                        "planlandi",
                        notlar or "",
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        (self.kullanici[0] if self.kullanici else None),
                        record_id,
                    ),
                )
                sid = int(cur.lastrowid or 0) or None
            except Exception:
                sid = None

        if sid:
            try:
                cur.execute("UPDATE records SET seans_id=? WHERE id=? AND (seans_id IS NULL OR seans_id='')", (sid, record_id))
            except Exception:
                pass

    def _sync_from_seans_to_record(self, cur, seans_id: int, tarih: str, saat: str, danisan: str, terapist: str, notlar: str):
        """seans_takvimi kaydı varsa records'a bağla/oluştur."""
        try:
            cur.execute("SELECT record_id FROM seans_takvimi WHERE id=?", (seans_id,))
            record_id = (cur.fetchone() or [None])[0]
        except Exception:
            record_id = None

        if record_id:
            try:
                cur.execute("UPDATE records SET seans_id=? WHERE id=? AND (seans_id IS NULL OR seans_id='')", (seans_id, record_id))
            except Exception:
                pass
            return

        # Aynı tarih+saat+danisan+terapist varsa bağla
        rid = None
        try:
            cur.execute(
                """
                SELECT id FROM records
                WHERE tarih=? AND COALESCE(saat,'')=? AND danisan_adi=? AND terapist=?
                ORDER BY id DESC LIMIT 1
                """,
                (tarih, saat, danisan, terapist),
            )
            row = cur.fetchone()
            if row and row[0]:
                rid = int(row[0])
        except Exception:
            rid = None

        if not rid:
            try:
                cur.execute(
                    """
                    INSERT INTO records (tarih, saat, danisan_adi, terapist, hizmet_bedeli, alinan_ucret, kalan_borc, seans_id, notlar, olusturma_tarihi)
                    VALUES (?,?,?,?,0,0,0,?,?,?)
                    """,
                    (
                        tarih,
                        saat,
                        danisan,
                        terapist,
                        seans_id,
                        notlar or "",
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                rid = int(cur.lastrowid or 0) or None
            except Exception:
                rid = None

        if rid:
            try:
                cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=? AND (record_id IS NULL OR record_id='')", (rid, seans_id))
            except Exception:
                pass

    def belirsizleri_duzelt_pencere(self):
        """Teknik olmayan kullanıcı için basit 'sihirbaz' ekranı."""
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Yetki", "Bu ekran sadece Kurum Müdürü tarafından kullanılabilir.")
            return

        items = []
        try:
            items = list(getattr(self, "_last_sync_ambiguous", []) or [])
        except Exception:
            items = []

        if not items:
            messagebox.showinfo("Bilgi", "Şu anda düzeltilmesi gereken belirsiz kayıt yok.")
            return

        win = ttk.Toplevel(self)
        win.title("Belirsiz Kayıtları Düzelt")
        win.transient(self)
        center_window_smart(win, 1100, 750, min_w=1000, min_h=700)
        maximize_window(win)
        self._brand_window(win)

        wrapper = ttk.Frame(win, padding=12)
        wrapper.pack(fill=BOTH, expand=True)

        head = ttk.Frame(wrapper)
        head.pack(fill=X)
        ttk.Label(head, text="BELİRSİZ KAYITLARI DÜZELT", font=("Segoe UI", 14, "bold"), bootstyle="warning").pack(side=LEFT, anchor=W)
        prog_lbl = ttk.Label(head, text=f"Kalan: {len(items)}", font=("Segoe UI", 11, "bold"))
        prog_lbl.pack(side=RIGHT)
        ttk.Label(
            wrapper,
            text=(
                "Uygulama bazen aynı gün aynı danışan için birden fazla kayıt bulduğu için otomatik eşleştirme yapamaz.\n"
                "Aşağıdan doğru seçeneği seçip 'EŞLEŞTİR' diyerek işlemi tamamlayabilirsin."
            ),
            foreground="gray",
            wraplength=940,
        ).pack(anchor=W, pady=(6, 12))

        # Sol: iş listesi
        body = ttk.Frame(wrapper)
        body.pack(fill=BOTH, expand=True)

        left = ttk.Labelframe(body, text="Düzeltilmesi Gerekenler", padding=10)
        left.pack(side=LEFT, fill=Y, padx=(0, 10))

        lst = tk.Listbox(left, width=44, height=22)
        lst.pack(fill=Y, expand=False)

        # Sağ: detay + seçenekler
        right = ttk.Labelframe(body, text="Detay ve Seçim", padding=10)
        right.pack(side=LEFT, fill=BOTH, expand=True)

        info = ttk.Label(right, text="", font=("Segoe UI", 11, "bold"), wraplength=580)
        info.pack(anchor=W, pady=(0, 10))

        choice_var = tk.StringVar(value="")
        choices_box = ttk.Frame(right)
        choices_box.pack(fill=BOTH, expand=True)

        def _label_for_item(it: dict) -> str:
            if it.get("type") == "record_missing_time":
                return f"Seans Takip kaydı: {it.get('tarih','')} | {it.get('danisan','')} | {it.get('terapist','')}"
            if it.get("type") == "seans_multiple_records":
                return f"Takvim seansı: {it.get('tarih','')} {it.get('saat','')} | {it.get('danisan','')} | {it.get('terapist','')}"
            return "Belirsiz kayıt"

        for it in items:
            lst.insert(END, _label_for_item(it))

        def _clear_choices():
            for w in choices_box.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

        def _render(idx: int):
            _clear_choices()
            if idx < 0 or idx >= len(items):
                info.configure(text="Bitti. Kalan belirsiz kayıt yok.")
                try:
                    prog_lbl.configure(text="Kalan: 0")
                except Exception:
                    pass
                return
            it = items[idx]
            t = it.get("type")
            choice_var.set("")
            try:
                prog_lbl.configure(text=f"Kalan: {len(items)}")
            except Exception:
                pass

            if t == "record_missing_time":
                info.configure(
                    text=(
                        "Bu Seans Takip kaydında saat yok.\n"
                        "Takvimde aynı gün aynı danışan için birden fazla seans var.\n"
                        "Lütfen doğru saati seç:"
                    )
                )
                for c in it.get("candidates", []):
                    sid = c.get("seans_id")
                    saat = c.get("saat") or "(saat yok)"
                    ttk.Radiobutton(
                        choices_box,
                        text=f"{saat}  (Takvim ID: {sid})",
                        value=str(sid),
                        variable=choice_var,
                        bootstyle="warning",
                    ).pack(anchor=W, pady=4)

            elif t == "seans_multiple_records":
                info.configure(
                    text=(
                        "Bu Takvim seansı için Seans Takip'te birden fazla aday kayıt bulundu.\n"
                        "Lütfen doğru kaydı seç:"
                    )
                )
                for c in it.get("candidates", []):
                    rid = c.get("record_id")
                    linked = " (zaten bağlı)" if c.get("has_link") else ""
                    ttk.Radiobutton(
                        choices_box,
                        text=f"Seans Takip ID: {rid}{linked}",
                        value=str(rid),
                        variable=choice_var,
                        bootstyle="warning",
                    ).pack(anchor=W, pady=4)
            else:
                info.configure(text="Bu belirsiz tip tanınmadı.")

        def _selected_index():
            try:
                sel = lst.curselection()
                if not sel:
                    return 0
                return int(sel[0])
            except Exception:
                return 0

        def _apply_choice():
            idx = _selected_index()
            if idx < 0 or idx >= len(items):
                return
            it = items[idx]
            t = it.get("type")
            val = (choice_var.get() or "").strip()
            if not val:
                messagebox.showwarning("Uyarı", "Lütfen bir seçenek seçiniz.")
                return

            conn = None
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()

                if t == "record_missing_time":
                    rid = int(it["record_id"])
                    sid = int(val)
                    cur.execute("SELECT tarih, COALESCE(saat,''), danisan_adi, terapist, COALESCE(notlar,'') FROM seans_takvimi WHERE id=?", (sid,))
                    row = cur.fetchone()
                    if not row:
                        messagebox.showerror("Hata", "Seçtiğiniz takvim seansı bulunamadı.")
                        return
                    tarih, saat, danisan, terapist, notlar = row[0], row[1], row[2], row[3], row[4]
                    # bağla + saat doldur
                    cur.execute("UPDATE records SET seans_id=?, saat=? WHERE id=?", (sid, saat, rid))
                    cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=?", (rid, sid))

                elif t == "seans_multiple_records":
                    sid = int(it["seans_id"])
                    rid = int(val)
                    # bağla
                    cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=?", (rid, sid))
                    cur.execute("UPDATE records SET seans_id=? WHERE id=?", (sid, rid))

                conn.commit()
            except Exception as e:
                messagebox.showerror("Hata", f"Eşleştirme yapılamadı:\n{e}")
                return
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

            # listeden çıkar, sıradakine geç
            try:
                lst.delete(idx)
            except Exception:
                pass
            try:
                items.pop(idx)
            except Exception:
                pass
            try:
                self._last_sync_ambiguous = items
            except Exception:
                pass
            try:
                self.kayitlari_listele()
                self._refresh_borc_tables()
            except Exception:
                pass
            self._update_sync_badge()
            if items:
                try:
                    lst.selection_clear(0, END)
                    lst.selection_set(min(idx, len(items) - 1))
                except Exception:
                    pass
                _render(_selected_index())
            else:
                win.destroy()
                messagebox.showinfo("Bitti", "Tüm belirsiz kayıtlar düzeltildi.")

        def _skip():
            idx = _selected_index()
            if idx < 0 or idx >= len(items):
                return
            # sadece listeden kaldır; DB'ye dokunma
            try:
                lst.delete(idx)
            except Exception:
                pass
            try:
                items.pop(idx)
            except Exception:
                pass
            try:
                self._last_sync_ambiguous = items
            except Exception:
                pass
            self._update_sync_badge()
            if items:
                try:
                    lst.selection_clear(0, END)
                    lst.selection_set(min(idx, len(items) - 1))
                except Exception:
                    pass
                _render(_selected_index())
            else:
                win.destroy()
                messagebox.showinfo("Bitti", "Belirsiz kayıt kalmadı.")

        btns = ttk.Frame(right)
        btns.pack(fill=X, pady=(10, 0))
        ttk.Button(btns, text="EŞLEŞTİR (SEÇİLİ OLANI)", bootstyle="success", command=_apply_choice).pack(side=LEFT, fill=X, expand=True, padx=6)
        ttk.Button(btns, text="ŞİMDİLİK ATLA", bootstyle="secondary", command=_skip).pack(side=LEFT, fill=X, expand=True, padx=6)
        ttk.Button(btns, text="KAPAT", bootstyle="danger", command=win.destroy).pack(side=LEFT, fill=X, expand=True, padx=6)

        def _on_select(_evt=None):
            _render(_selected_index())

        lst.bind("<<ListboxSelect>>", _on_select)
        try:
            lst.selection_set(0)
        except Exception:
            pass
        _render(0)
        self._update_sync_badge()

    def senkronize_takvim_seanslar(self):
        """Toplu senkronizasyon: records <-> seans_takvimi.
        - Net eşleşme varsa bağlar (tarih+saat+danışan+terapist)
        - Saat eksikse tekil eşleşme varsa bağlar ve records.saat'i doldurur
        - Bulunamazsa eksik tarafta kayıt oluşturur
        """
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Yetki", "Toplu senkronizasyon sadece Kurum Müdürü tarafından yapılabilir.")
            return
        if not messagebox.askyesno(
            "Onay",
            "Takvim ve Seans Takip kayıtları senkronize edilecek.\n\n"
            "Bu işlem:\n"
            "- Eşleşenleri bağlar\n"
            "- Eksik tarafta kayıt oluşturabilir\n\n"
            "Devam edilsin mi?",
        ):
            return

        stats = {
            "linked_records": 0,
            "created_seans": 0,
            "linked_seans": 0,
            "created_records": 0,
            "skipped_ambiguous": 0,
        }
        ambiguous: list[dict] = []

        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()

            # 1) records tarafı: seans_id yoksa bağla/oluştur
            cur.execute(
                """
                SELECT id, tarih, COALESCE(saat,''), danisan_adi, terapist, COALESCE(notlar,'')
                FROM records
                WHERE seans_id IS NULL OR seans_id=''
                ORDER BY id ASC
                """
            )
            rec_rows = cur.fetchall() or []

            for rid, tarih, saat, danisan, terapist, notlar in rec_rows:
                rid = int(rid)
                tarih = (tarih or "").strip()
                saat = (saat or "").strip()
                danisan = (danisan or "").strip().upper()
                terapist = (terapist or "").strip()
                notlar = (notlar or "").strip()

                # 1a) önce tam eşleşme
                sid = None
                if saat:
                    cur.execute(
                        """
                        SELECT id, COALESCE(record_id,NULL) FROM seans_takvimi
                        WHERE tarih=? AND saat=? AND danisan_adi=? AND terapist=?
                        ORDER BY id DESC LIMIT 1
                        """,
                        (tarih, saat, danisan, terapist),
                    )
                    row = cur.fetchone()
                    if row and row[0]:
                        sid = int(row[0])
                else:
                    # 1b) saat yoksa: tekil eşleşme varsa bağla + saat doldur
                    cur.execute(
                        """
                        SELECT id, saat, COALESCE(record_id,NULL) FROM seans_takvimi
                        WHERE tarih=? AND danisan_adi=? AND terapist=?
                        ORDER BY id DESC
                        """,
                        (tarih, danisan, terapist),
                    )
                    cands = cur.fetchall() or []
                    # sadece tek aday varsa güvenli
                    if len(cands) == 1:
                        sid = int(cands[0][0])
                        saat = (cands[0][1] or "").strip()
                        if saat:
                            cur.execute("UPDATE records SET saat=? WHERE id=? AND (saat IS NULL OR saat='')", (saat, rid))
                    elif len(cands) > 1:
                        stats["skipped_ambiguous"] += 1
                        try:
                            ambiguous.append(
                                {
                                    "type": "record_missing_time",
                                    "record_id": rid,
                                    "tarih": tarih,
                                    "danisan": danisan,
                                    "terapist": terapist,
                                    "candidates": [{"seans_id": int(x[0]), "saat": (x[1] or "").strip()} for x in cands],
                                }
                            )
                        except Exception:
                            pass

                if sid:
                    # bağla
                    cur.execute("UPDATE records SET seans_id=? WHERE id=? AND (seans_id IS NULL OR seans_id='')", (sid, rid))
                    cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=? AND (record_id IS NULL OR record_id='')", (rid, sid))
                    stats["linked_records"] += 1
                    continue

                # 1c) yoksa seans oluştur
                use_saat = saat or self._default_saat()
                cur.execute(
                    """
                    INSERT INTO seans_takvimi (tarih, saat, danisan_adi, terapist, oda, durum, notlar, olusturma_tarihi, olusturan_kullanici_id, record_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        tarih,
                        use_saat,
                        danisan,
                        terapist,
                        "",
                        "planlandi",
                        notlar,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        (self.kullanici[0] if self.kullanici else None),
                        rid,
                    ),
                )
                sid = int(cur.lastrowid or 0)
                cur.execute("UPDATE records SET seans_id=?, saat=? WHERE id=?", (sid, use_saat, rid))
                stats["created_seans"] += 1

            # 2) seans_takvimi tarafı: record_id yoksa bağla/oluştur
            cur.execute(
                """
                SELECT id, tarih, saat, danisan_adi, terapist, COALESCE(notlar,'')
                FROM seans_takvimi
                WHERE record_id IS NULL OR record_id=''
                ORDER BY id ASC
                """
            )
            seans_rows = cur.fetchall() or []

            for sid, tarih, saat, danisan, terapist, notlar in seans_rows:
                sid = int(sid)
                tarih = (tarih or "").strip()
                saat = (saat or "").strip()
                danisan = (danisan or "").strip().upper()
                terapist = (terapist or "").strip()
                notlar = (notlar or "").strip()

                cur.execute(
                    """
                    SELECT id, COALESCE(seans_id,NULL) FROM records
                    WHERE tarih=? AND COALESCE(saat,'')=? AND danisan_adi=? AND terapist=?
                    ORDER BY id DESC
                    """,
                    (tarih, saat, danisan, terapist),
                )
                cands = cur.fetchall() or []
                # Eğer birden fazla aday varsa yanlış bağlamamak için kullanıcıya soralım
                if len(cands) > 1:
                    stats["skipped_ambiguous"] += 1
                    try:
                        ambiguous.append(
                            {
                                "type": "seans_multiple_records",
                                "seans_id": sid,
                                "tarih": tarih,
                                "saat": saat,
                                "danisan": danisan,
                                "terapist": terapist,
                                "candidates": [{"record_id": int(x[0]), "has_link": bool(x[1])} for x in cands],
                            }
                        )
                    except Exception:
                        pass
                    continue

                rid = None
                if cands:
                    rid = int(cands[0][0])

                if rid:
                    cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=? AND (record_id IS NULL OR record_id='')", (rid, sid))
                    cur.execute("UPDATE records SET seans_id=? WHERE id=? AND (seans_id IS NULL OR seans_id='')", (sid, rid))
                    stats["linked_seans"] += 1
                    continue

                # yoksa record oluştur
                cur.execute(
                    """
                    INSERT INTO records (tarih, saat, danisan_adi, terapist, hizmet_bedeli, alinan_ucret, kalan_borc, notlar, olusturma_tarihi, seans_id)
                    VALUES (?,?,?,?,0,0,0,?,?,?)
                    """,
                    (
                        tarih,
                        saat,
                        danisan,
                        terapist,
                        notlar,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        sid,
                    ),
                )
                rid = int(cur.lastrowid or 0)
                cur.execute("UPDATE seans_takvimi SET record_id=? WHERE id=?", (rid, sid))
                stats["created_records"] += 1

            # 3) Haftalık programa eksik seansları ekle (seans_takvimi'ndeki her kayıt haftalık programda görünsün)
            GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            cur.execute(
                "SELECT id, tarih, saat, danisan_adi, terapist, COALESCE(notlar,''), COALESCE(oda,'') FROM seans_takvimi"
            )
            for row in cur.fetchall() or []:
                try:
                    sid, tarih_s, saat_s, danisan_s, terapist_s, notlar_s, oda_s = row
                    if not tarih_s or not terapist_s:
                        continue
                    dt = datetime.datetime.strptime(tarih_s.strip()[:10], "%Y-%m-%d")
                    weekday = dt.weekday()
                    monday = dt - datetime.timedelta(days=weekday)
                    hafta_bas = monday.strftime("%Y-%m-%d")
                    gun = GUNLER[weekday]
                    saat_norm = (saat_s or "").strip() or "09:00"
                    if len(saat_norm) <= 2 and saat_norm.isdigit():
                        saat_norm = f"{int(saat_norm):02d}:00"
                    olusturma = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    cur.execute(
                        """
                        INSERT OR REPLACE INTO haftalik_seans_programi
                        (personel_adi, hafta_baslangic_tarihi, gun, saat, ogrenci_adi, oda_adi, notlar, olusturma_tarihi, guncelleme_tarihi, olusturan_kullanici_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (terapist_s.strip(), hafta_bas, gun, saat_norm, (danisan_s or "").strip(), (oda_s or "").strip(), (notlar_s or "").strip(), olusturma, olusturma, (self.kullanici[0] if self.kullanici else None)),
                    )
                except Exception:
                    pass

            conn.commit()
            conn.close()
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass
            messagebox.showerror("Hata", f"Senkronizasyon hatası:\n{e}")
            return

        # UI tazele: Seans listesi + borç tabloları + Haftalık program (Windows/macOS)
        try:
            self.kayitlari_listele()
            self._refresh_borc_tables()
        except Exception:
            pass
        try:
            for child in self.tab_haftalik.winfo_children():
                if getattr(child, "_program_table", None):
                    self._haftalik_program_yukle(child)
                    break
        except Exception:
            pass

        # belirsiz listesi sakla (butondan tekrar açılabilsin)
        try:
            self._last_sync_ambiguous = ambiguous
        except Exception:
            pass
        self._update_sync_badge()

        messagebox.showinfo(
            "Senkronizasyon Tamam",
            "Özet:\n"
            f"- Bağlanan kayıtlar (records→seans): {stats['linked_records']}\n"
            f"- Oluşturulan seans (records→seans): {stats['created_seans']}\n"
            f"- Bağlanan seans (seans→records): {stats['linked_seans']}\n"
            f"- Oluşturulan kayıt (seans→records): {stats['created_records']}\n"
            f"- Belirsiz olduğu için atlanan: {stats['skipped_ambiguous']}\n\n"
            "Not: Belirsiz olanları istersen 'Belirsizleri Düzelt' ekranından tek tek seçerek tamamlayabilirsin.",
        )

        if ambiguous:
            if messagebox.askyesno(
                "Belirsiz Kayıtlar Var",
                f"{len(ambiguous)} adet belirsiz kayıt bulundu.\n\n"
                "Şimdi düzeltme ekranı açılsın mı?\n"
                "(İstersen sonra da 'Belirsizleri Düzelt' butonundan açabilirsin.)",
            ):
                self.belirsizleri_duzelt_pencere()

    def _range_summary(self, bas: str, bit: str, terapist: str | None = None) -> dict:
        """Kasa + seans özetini döndür."""
        out = {
            "seans_sayisi": 0,
            "bedel_toplam": 0.0,
            "alinan_toplam": 0.0,
            "kalan_toplam": 0.0,
            "kasa_giren": 0.0,
            "kasa_cikan": 0.0,
        }
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()

            if terapist:
                cur.execute(
                    """
                    SELECT COUNT(*),
                           COALESCE(SUM(hizmet_bedeli),0),
                           COALESCE(SUM(alinan_ucret),0),
                           COALESCE(SUM(kalan_borc),0)
                    FROM records
                    WHERE tarih>=? AND tarih<=? AND terapist=?
                    """,
                    (bas, bit, terapist),
                )
            else:
                cur.execute(
                    """
                    SELECT COUNT(*),
                           COALESCE(SUM(hizmet_bedeli),0),
                           COALESCE(SUM(alinan_ucret),0),
                           COALESCE(SUM(kalan_borc),0)
                    FROM records
                    WHERE tarih>=? AND tarih<=?
                    """,
                    (bas, bit),
                )
            c, b, a, k = cur.fetchone() or (0, 0, 0, 0)
            out["seans_sayisi"] = int(c or 0)
            out["bedel_toplam"] = float(b or 0)
            out["alinan_toplam"] = float(a or 0)
            out["kalan_toplam"] = float(k or 0)

            if terapist:
                cur.execute(
                    """
                    SELECT
                        COALESCE(SUM(CASE WHEN kh.tip='giren' THEN kh.tutar ELSE 0 END),0),
                        COALESCE(SUM(CASE WHEN kh.tip='cikan' THEN kh.tutar ELSE 0 END),0)
                    FROM kasa_hareketleri kh
                    LEFT JOIN records r ON r.id = kh.record_id
                    WHERE kh.tarih>=? AND kh.tarih<=? AND r.terapist=?
                    """,
                    (bas, bit, terapist),
                )
            else:
                cur.execute(
                    """
                    SELECT
                        COALESCE(SUM(CASE WHEN tip='giren' THEN tutar ELSE 0 END),0),
                        COALESCE(SUM(CASE WHEN tip='cikan' THEN tutar ELSE 0 END),0)
                    FROM kasa_hareketleri
                    WHERE tarih>=? AND tarih<=?
                    """,
                    (bas, bit),
                )
            g, ckn = cur.fetchone() or (0, 0)
            out["kasa_giren"] = float(g or 0)
            out["kasa_cikan"] = float(ckn or 0)
            conn.close()
        except Exception:
            pass
        return out

    def gunluk_rapor_pencere(self):
        # SEANS TAKİP ekranındaki tarih alanını baz al
        try:
            s = (self.tarih_var.get() or "").strip()
        except Exception:
            s = ""
        gun = self._tarih_db_from(s)
        self._rapor_pencere(gun, gun, title="Günlük Rapor")

    def haftalik_rapor_pencere(self):
        # SEANS TAKİP tarihine göre haftayı seç
        try:
            s = (self.tarih_var.get() or "").strip()
        except Exception:
            s = ""
        d = datetime.datetime.strptime(self._tarih_db_from(s), "%Y-%m-%d")
        bas = d - datetime.timedelta(days=d.weekday())
        bit = bas + datetime.timedelta(days=6)
        self._rapor_pencere(bas.strftime("%Y-%m-%d"), bit.strftime("%Y-%m-%d"), title="Haftalık Rapor")
    
    def senkronizasyon_kontrol_pencere(self):
        """Genel senkronizasyon kontrolü - Tüm tablolar arası tutarlılık"""
        win = ttk.Toplevel(self)
        win.title("Senkronizasyon Kontrolü")
        win.transient(self)
        center_window_smart(win, 1000, 700, min_w=900, min_h=650)
        maximize_window(win)
        self._brand_window(win)
        
        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="🔍 GENEL SENKRONİZASYON KONTROLÜ", font=("Segoe UI", 14, "bold"), bootstyle="info").pack(side=LEFT)
        
        # Kontrol butonu
        def _kontrol_et():
            conn = None
            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                result = pipeline.validate_sync()
                
                # Sonuçları göster
                for iid in tree.get_children():
                    tree.delete(iid)
                
                # İstatistikler
                stats = result["stats"]
                tree.insert("", END, values=("📊 İSTATİSTİKLER", "", ""), tags=("header",))
                tree.insert("", END, values=("Seans Takvimi", str(stats["seans_takvimi_count"]), ""))
                tree.insert("", END, values=("Records", str(stats["records_count"]), ""))
                tree.insert("", END, values=("Danışanlar", str(stats["danisanlar_count"]), ""))
                tree.insert("", END, values=("Odalar", str(stats["odalar_count"]), ""))
                tree.insert("", END, values=("Kasa Hareketleri", str(stats["kasa_hareketleri_count"]), ""))
                tree.insert("", END, values=("Ödeme Hareketleri", str(stats["odeme_hareketleri_count"]), ""))
                
                # Hatalar
                if result["errors"]:
                    tree.insert("", END, values=("", "", ""))
                    tree.insert("", END, values=("❌ HATALAR", "", ""), tags=("error",))
                    for err in result["errors"]:
                        tree.insert("", END, values=("", err, ""), tags=("error",))
                
                # Uyarılar
                missing_danisanlar_list.clear()
                if result["warnings"]:
                    tree.insert("", END, values=("", "", ""))
                    tree.insert("", END, values=("⚠️ UYARILAR", "", ""), tags=("warning",))
                    for warn in result["warnings"]:
                        tree.insert("", END, values=("", warn, ""), tags=("warning",))
                
                # Eksik danışanları al (validate_sync'den direkt)
                if "missing_danisanlar" in result:
                    missing_danisanlar_list.extend(result["missing_danisanlar"])
                
                # Eksik danışanlar varsa butonu aktif et
                if missing_danisanlar_list:
                    btn_ekle.config(state="normal")
                else:
                    btn_ekle.config(state="disabled")
                
                # Durum
                if result["ok"]:
                    durum_lbl.config(text="✅ Tüm senkronizasyonlar OK!", bootstyle="success")
                else:
                    durum_lbl.config(text=f"❌ {len(result['errors'])} hata bulundu!", bootstyle="danger")
                    
            except Exception as e:
                messagebox.showerror("Hata", f"Senkronizasyon kontrolü hatası:\n{e}")
                log_exception("senkronizasyon_kontrol", e)
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass
        
        # Eksik danışanları otomatik ekle
        missing_danisanlar_list = []
        
        def _eksik_danisanlari_ekle():
            if not missing_danisanlar_list:
                messagebox.showinfo("Bilgi", "Eklenecek danışan bulunamadı.")
                return
            conn = None
            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                
                eklenen = 0
                for danisan_adi in missing_danisanlar_list:
                    pipeline._ensure_danisan_exists(danisan_adi)
                    eklenen += 1
                
                conn.commit()
                
                messagebox.showinfo("Başarılı", f"{eklenen} danışan otomatik olarak eklendi!\n\nLütfen kontrolü tekrar çalıştırın.")
                _kontrol_et()
                
            except Exception as e:
                messagebox.showerror("Hata", f"Danışan ekleme hatası:\n{e}")
                log_exception("eksik_danisan_ekle", e)
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass
        
        btn_frame = ttk.Frame(win, padding=10)
        btn_frame.pack(fill=X)
        ttk.Button(btn_frame, text="🔍 Kontrol Et", bootstyle="info", command=_kontrol_et).pack(side=LEFT, padx=6)
        btn_ekle = ttk.Button(btn_frame, text="➕ Eksik Danışanları Ekle", bootstyle="success", command=_eksik_danisanlari_ekle, state="disabled")
        btn_ekle.pack(side=LEFT, padx=6)
        durum_lbl = ttk.Label(btn_frame, text="Kontrol edilmeyi bekliyor...", font=("Segoe UI", 10, "bold"))
        durum_lbl.pack(side=LEFT, padx=20)
        
        # Sonuçlar
        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("Kategori", "Detay", "Durum")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=20)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=300)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        tree.tag_configure("header", background="#e9ecef", font=("Segoe UI", 10, "bold"))
        tree.tag_configure("error", foreground="#dc3545")
        tree.tag_configure("warning", foreground="#ffc107")
        
        # İlk kontrolü otomatik yap
        _kontrol_et()

    def toplam_rapor_pencere(self):
        self._rapor_pencere("0001-01-01", "9999-12-31", title="Toplam Rapor (Genel)")

    def _rapor_pencere(self, bas: str, bit: str, title: str):
        win = ttk.Toplevel(self)
        win.title(title)
        win.transient(self)
        center_window_smart(win, 1250, 760)
        maximize_window(win)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text=title.upper(), font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(side=LEFT)
        ttk.Label(top, text=f"Tarih: {bas}  →  {bit}", foreground="gray").pack(side=LEFT, padx=10)

        s = self._range_summary(bas, bit, None if self.kullanici_yetki == "kurum_muduru" else (self.kullanici_terapist or None))
        net = float(s["kasa_giren"]) - float(s["kasa_cikan"])

        o = ttk.Labelframe(win, text="📊 DETAYLI ÖZET BİLGİLERİ", padding=12, bootstyle="info")
        o.pack(fill=X, padx=10, pady=(0, 8))
        ttk.Label(o, text=f"📅 Seans: {s['seans_sayisi']}   💰 Ücret: {format_money(s['bedel_toplam'])}   ✅ Alınan: {format_money(s['alinan_toplam'])}   ⚠️ Borç: {format_money(s['kalan_toplam'])}", font=("Segoe UI", 10, "bold")).pack(anchor=W)
        ttk.Label(o, text=f"📥 Kasa Giren: {format_money(s['kasa_giren'])}   📤 Kasa Çıkan: {format_money(s['kasa_cikan'])}   💵 Net: {format_money(net)}", font=("Segoe UI", 10)).pack(anchor=W, pady=(4, 0))

        mid = ttk.Frame(win, padding=10)
        mid.pack(fill=BOTH, expand=True)

        nb = ttk.Notebook(mid, bootstyle="info")
        nb.pack(fill=BOTH, expand=True)
        tab_records = ttk.Frame(nb, padding=6)
        tab_kasa = ttk.Frame(nb, padding=6)
        nb.add(tab_records, text="Seans / Records")
        nb.add(tab_kasa, text="Kasa Defteri")

        cols_r = ("SeansID", "Tarih", "Saat", "Danışan", "Terapist", "Seans Ücreti", "Alınan Ödeme", "Kalan Borç", "Not")
        tree_r = ttk.Treeview(tab_records, columns=cols_r, show="headings", style="Strong.Treeview", selectmode="extended")
        for c in cols_r:
            tree_r.heading(c, text=c)
        tree_r.column("SeansID", width=0, stretch=False)
        tree_r.column("Tarih", width=95, anchor="center")
        tree_r.column("Saat", width=65, anchor="center")
        tree_r.column("Danışan", width=190)
        tree_r.column("Terapist", width=130)
        tree_r.column("Seans Ücreti", width=115, anchor="e")
        tree_r.column("Alınan Ödeme", width=115, anchor="e")
        tree_r.column("Kalan Borç", width=115, anchor="e")
        tree_r.column("Not", width=280)
        tree_r.pack(side=LEFT, fill=BOTH, expand=True)
        sb_r = ttk.Scrollbar(tab_records, orient=VERTICAL, command=tree_r.yview)
        tree_r.configure(yscroll=sb_r.set)
        sb_r.pack(side=RIGHT, fill=Y)

        cols_k = ("KasaID", "Tarih", "Tip", "Açıklama", "Tutar", "Ödeme Şekli", "Kategori")
        tree_k = ttk.Treeview(tab_kasa, columns=cols_k, show="headings", style="Strong.Treeview", selectmode="extended")
        for c in cols_k:
            tree_k.heading(c, text=c)
        tree_k.column("KasaID", width=0, stretch=False)
        tree_k.column("Tarih", width=95, anchor="center")
        tree_k.column("Tip", width=70, anchor="center")
        tree_k.column("Açıklama", width=360)
        tree_k.column("Tutar", width=130, anchor="e")
        tree_k.column("Ödeme Şekli", width=140, anchor="center")
        tree_k.column("Kategori", width=120, anchor="center")
        tree_k.pack(side=LEFT, fill=BOTH, expand=True)
        sb_k = ttk.Scrollbar(tab_kasa, orient=VERTICAL, command=tree_k.yview)
        tree_k.configure(yscroll=sb_k.set)
        sb_k.pack(side=RIGHT, fill=Y)

        def _load_records():
            for iid in tree_r.get_children():
                tree_r.delete(iid)
            conn = None
            try:
                conn = self.veritabani_baglan()
                if self.kullanici_yetki == "kurum_muduru" or not self.kullanici_terapist:
                    df = pd.read_sql_query(
                        """
                        SELECT
                            st.id AS seans_id,
                            st.tarih,
                            COALESCE(st.saat, '') AS saat,
                            COALESCE(r.danisan_adi, st.danisan_adi, '') AS danisan_adi,
                            COALESCE(r.terapist, st.terapist, '') AS terapist,
                            COALESCE(r.hizmet_bedeli, 0) AS seans_ucreti,
                            COALESCE(r.alinan_ucret, 0) AS alinan_odeme,
                            COALESCE(r.kalan_borc, 0) AS kalan_borc,
                            COALESCE(r.notlar, st.notlar, '') AS notlar
                        FROM seans_takvimi st
                        LEFT JOIN records r ON r.id = st.record_id OR r.seans_id = st.id
                        WHERE st.tarih>=? AND st.tarih<=?
                        ORDER BY st.tarih, st.saat, st.id
                        """,
                        conn, params=(bas, bit)
                    )
                else:
                    df = pd.read_sql_query(
                        """
                        SELECT
                            st.id AS seans_id,
                            st.tarih,
                            COALESCE(st.saat, '') AS saat,
                            COALESCE(r.danisan_adi, st.danisan_adi, '') AS danisan_adi,
                            COALESCE(r.terapist, st.terapist, '') AS terapist,
                            COALESCE(r.hizmet_bedeli, 0) AS seans_ucreti,
                            COALESCE(r.alinan_ucret, 0) AS alinan_odeme,
                            COALESCE(r.kalan_borc, 0) AS kalan_borc,
                            COALESCE(r.notlar, st.notlar, '') AS notlar
                        FROM seans_takvimi st
                        LEFT JOIN records r ON r.id = st.record_id OR r.seans_id = st.id
                        WHERE st.tarih>=? AND st.tarih<=? AND COALESCE(r.terapist, st.terapist, '')=?
                        ORDER BY st.tarih, st.saat, st.id
                        """,
                        conn, params=(bas, bit, self.kullanici_terapist)
                    )
            except Exception as e:
                log_exception("_rapor_pencere_records", e)
                df = pd.DataFrame(columns=["seans_id", "tarih", "saat", "danisan_adi", "terapist", "seans_ucreti", "alinan_odeme", "kalan_borc", "notlar"])
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

            for _, r in df.iterrows():
                tree_r.insert("", END, values=(
                    int(r.get("seans_id", 0) or 0), r.get("tarih", ""), r.get("saat", ""), r.get("danisan_adi", ""),
                    r.get("terapist", ""), format_money(r.get("seans_ucreti", 0) or 0), format_money(r.get("alinan_odeme", 0) or 0),
                    format_money(r.get("kalan_borc", 0) or 0), r.get("notlar", "") or ""
                ))
            return df

        def _load_kasa():
            for iid in tree_k.get_children():
                tree_k.delete(iid)
            conn = None
            try:
                conn = self.veritabani_baglan()
                params = [bas, bit]
                sql = """
                    SELECT id, tarih, tip, aciklama, tutar, COALESCE(odeme_sekli,''), COALESCE(gider_kategorisi,'')
                    FROM kasa_hareketleri
                    WHERE tarih>=? AND tarih<=?
                """
                if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                    sql += " AND (record_id IN (SELECT id FROM records WHERE terapist=?))"
                    params.append(self.kullanici_terapist)
                sql += " ORDER BY tarih, id"
                df = pd.read_sql_query(sql, conn, params=tuple(params))
            except Exception as e:
                log_exception("_rapor_pencere_kasa", e)
                df = pd.DataFrame(columns=["id", "tarih", "tip", "aciklama", "tutar", "odeme_sekli", "gider_kategorisi"])
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

            for _, r in df.iterrows():
                tree_k.insert("", END, values=(
                    int(r.get("id", 0) or 0), r.get("tarih", ""), r.get("tip", ""), r.get("aciklama", ""),
                    format_money(r.get("tutar", 0) or 0), r.get("odeme_sekli", "") or "", r.get("gider_kategorisi", "") or ""
                ))
            return df

        def _secim_modu(adet: int, tur_adi: str) -> str | None:
            if adet <= 0:
                return None
            if adet == 1:
                return "tek"
            yanit = messagebox.askyesnocancel("Silme Modu", f"{adet} adet {tur_adi} seçili.\n\nEvet: Sadece ilk seçiliyi sil\nHayır: Tüm seçilileri sil\nİptal: Vazgeç")
            if yanit is None:
                return None
            return "tek" if yanit else "toplu"

        def _sil_rapor_secili():
            sec_tab = nb.select()
            conn = None
            try:
                conn = self.veritabani_baglan()
                pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
                if sec_tab == str(tab_records):
                    sec = list(tree_r.selection())
                    mod = _secim_modu(len(sec), "seans kaydı")
                    if not mod:
                        return
                    hedef = sec[:1] if mod == "tek" else sec
                    silinen = 0
                    for iid in hedef:
                        vals = tree_r.item(iid).get("values") or []
                        seans_id = int(vals[0] or 0) if vals else 0
                        if seans_id <= 0:
                            continue
                        if pipeline.kayit_sil(seans_id):
                            silinen += 1
                    if silinen <= 0:
                        messagebox.showwarning("Uyarı", "Silinebilecek uygun seans kaydı bulunamadı.")
                        return
                    messagebox.showinfo("Tamam", f"{silinen} seans kaydı silindi.")
                    self.kayitlari_listele()
                    self._refresh_borc_tables()
                    _load_records(); _load_kasa(); self._refresh_open_reports()
                else:
                    sec = list(tree_k.selection())
                    mod = _secim_modu(len(sec), "kasa hareketi")
                    if not mod:
                        return
                    hedef = sec[:1] if mod == "tek" else sec
                    silinen = 0
                    for iid in hedef:
                        vals = tree_k.item(iid).get("values") or []
                        kid = int(vals[0] or 0) if vals else 0
                        if kid <= 0:
                            continue
                        if pipeline.kasa_hareketi_sil(kid):
                            silinen += 1
                    if silinen <= 0:
                        messagebox.showwarning("Uyarı", "Silinebilecek kasa kaydı bulunamadı.")
                        return
                    messagebox.showinfo("Tamam", f"{silinen} kasa hareketi silindi.")
                    _load_kasa()
                    self.kayitlari_listele()
                    self._refresh_borc_tables()
                    self._refresh_open_reports()
            except Exception as e:
                messagebox.showerror("Hata", f"Silme hatası:\n{e}")
                log_exception("_rapor_pencere_sil", e)
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        df_cache = {"records": _load_records(), "kasa": _load_kasa()}

        def _yenile():
            df_cache["records"] = _load_records()
            df_cache["kasa"] = _load_kasa()

        self._register_report_refresher(win, _yenile)

        def _excel():
            sec_tab = nb.select()
            df = df_cache.get("records") if sec_tab == str(tab_records) else df_cache.get("kasa")
            if df is None:
                return
            path = filedialog.asksaveasfilename(
                title="Excel Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"rapor_{bas}_{bit}.xlsx",
            )
            if not path:
                return
            try:
                df.to_excel(path, index=False, engine="openpyxl")
                messagebox.showinfo("Başarılı", "Excel'e aktarıldı.")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel aktarma hatası:\n{e}")

        btns = ttk.Frame(win, padding=10)
        btns.pack(fill=X)
        ttk.Button(btns, text="Yenile", bootstyle="secondary", command=_yenile).pack(side=LEFT, padx=6)
        ttk.Button(btns, text="Seçiliyi/Seçilileri Sil", bootstyle="danger", command=_sil_rapor_secili).pack(side=LEFT, padx=6)
        ttk.Button(btns, text="Excel'e Aktar", bootstyle="primary", command=_excel).pack(side=RIGHT, padx=6)

    def _reload_logos(self) -> None:
        """Logo dosyası değiştiyse PhotoImage'ları yeniden yükle ve UI'da güncelle."""
        self._logo_small = load_logo_photo(28, 28)
        self._logo_big = load_logo_photo(80, 80)
        self._logo_icon = load_logo_photo(64, 64)
        safe_iconphoto(self, self._logo_icon)

        try:
            if self._toolbar_logo_lbl is not None:
                if self._logo_small is not None:
                    self._toolbar_logo_lbl.configure(image=self._logo_small)
                    self._toolbar_logo_lbl.image = self._logo_small
                else:
                    self._toolbar_logo_lbl.destroy()
                    self._toolbar_logo_lbl = None
        except Exception:
            pass

        try:
            if self._login_logo_lbl is not None:
                if self._logo_big is not None:
                    self._login_logo_lbl.configure(image=self._logo_big)
                    self._login_logo_lbl.image = self._logo_big
                else:
                    self._login_logo_lbl.destroy()
                    self._login_logo_lbl = None
        except Exception:
            pass

    def logo_yukle_degistir(self):
        """Kurum müdürü için: logo.png seçip AppData'ya kopyalar."""
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Yetki", "Logo değiştirme sadece Kurum Müdürü tarafından yapılabilir.")
            return
        path = filedialog.askopenfilename(
            title="Logo Seç (PNG)",
            filetypes=[("PNG", "*.png")],
        )
        if not path:
            return
        try:
            os.makedirs(data_dir(), exist_ok=True)
            target = os.path.join(data_dir(), "logo.png")
            shutil.copy2(path, target)
        except Exception as e:
            messagebox.showerror("Hata", f"Logo kopyalanamadı:\n{e}")
            return

        self._reload_logos()
        messagebox.showinfo("Başarılı", "Logo kaydedildi.\nGerekirse programı kapatıp tekrar açın.")

    def _apply_user_restrictions(self):
        # Eğitim görevlisi ise kendi adına kilitle
        if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
            try:
                if hasattr(self, "cmb_terapist") and self.cmb_terapist:
                    self.cmb_terapist.set(self.kullanici_terapist)
                    self.cmb_terapist.configure(state="disabled")
            except Exception:
                pass

    def _build_menu(self):
        # Not: Windows'ta ttkbootstrap Menu bazen görünürlük sorunu çıkarabiliyor.
        # Bu yüzden ana menüde tkinter.Menu kullanıyoruz.
        menubar = Menu(self)
        self.config(menu=menubar)
        # Rol bazlı modüller (leta_pro uyumlu)
        if self.kullanici_yetki == "kurum_muduru":
            seans_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Seans Takvimi", menu=seans_menu)
            seans_menu.add_command(label="Günlük Takvim", command=self.seans_takvimi_goster)
            seans_menu.add_command(label="Haftalık Takvim", command=self.haftalik_takvim_goster)
            seans_menu.add_separator()
            seans_menu.add_command(label="Seans Listesi (Düzenle/Sil)", command=self.seans_listesi)

            sekreter_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Sekreterlik", menu=sekreter_menu)
            sekreter_menu.add_command(label="Danışan Yönetimi", command=self.danisan_yonetimi)
            sekreter_menu.add_command(label="Randevu Yönetimi", command=self.randevu_yonetimi)
            sekreter_menu.add_command(label="Görev Takibi", command=self.gorev_takibi)
            sekreter_menu.add_command(label="Oda Yönetimi", command=self.odalar_yonetimi)

            muhasebe_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Muhasebe", menu=muhasebe_menu)
            muhasebe_menu.add_command(label="Ücret Takibi", command=self.ucret_takibi_goster)
            muhasebe_menu.add_command(label="Haftalık Ders/Ücret Takip", command=self.haftalik_ders_ucret_takip)
            muhasebe_menu.add_command(label="Kasa Defteri (Günlük)", command=self.kasa_defteri_goster)
            muhasebe_menu.add_command(label="Gelir-Gider Raporu", command=self.gelir_gider_raporu)
            muhasebe_menu.add_command(label="Ödeme İşlemleri", command=self.odeme_islemleri)

            kullanici_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Kullanıcı Yönetimi", menu=kullanici_menu)
            kullanici_menu.add_command(label="Kullanıcıları Listele", command=self.kullanicilari_listele)
            kullanici_menu.add_command(label="Kullanıcı Sil/Pasif Et", command=self.kullanici_sil)

        elif self.kullanici_yetki in ["egitim_gorevlisi", "normal"]:
            seans_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Seans Takvimi", menu=seans_menu)
            seans_menu.add_command(label="Kendi Seanslarım", command=self.kendi_seanslarim)
            seans_menu.add_command(label="Haftalık Takvim", command=self.haftalik_takvim_goster)

            muhasebe_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Ücret Takibi", menu=muhasebe_menu)
            muhasebe_menu.add_command(label="Kendi Ücretlerim", command=self.kendi_ucretlerim)
            muhasebe_menu.add_command(label="Haftalık Ders/Ücret Takip", command=self.haftalik_ders_ucret_takip)

        dosya_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Dosya İşlemleri", menu=dosya_menu)
        dosya_menu.add_command(label="Excel'e Aktar", command=self.excel_aktar)
        dosya_menu.add_command(label="Yedek Klasörünü Aç", command=self.yedek_klasoru_ac)
        if self.kullanici_yetki == "kurum_muduru":
            dosya_menu.add_separator()
            dosya_menu.add_command(label="Sistemi Sıfırla (DB Sil)", command=self.sistemi_sifirla)
        dosya_menu.add_separator()
        dosya_menu.add_command(label="Giriş Ekranına Dön", command=self.girise_don)
        dosya_menu.add_command(label="Çıkış", command=self._on_close)

        yardim_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Yardım", menu=yardim_menu)
        yardim_menu.add_command(label="İlk 5 Adım (Hızlı Yardım)", command=self.ilk_5_adim_goster)
        yardim_menu.add_command(label="Kullanım Kılavuzu", command=self.kullanim_kilavuzu_ac)
        yardim_menu.add_command(label="Logo Yükle/Değiştir (Kurum Müdürü)", command=self.logo_yukle_degistir)
        yardim_menu.add_command(label="Hakkında", command=self.hakkinda_goster)

    def _build_tabs(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=BOTH, expand=True, padx=10, pady=10)

        self.tab_records = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_records, text="SEANS TAKİP")

        self._build_records_tab()

        # ANA SAYFA: Menü görünür olmasa bile butonlarla erişim (özellikle sahada kullanım için)
        self.tab_modules = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_modules, text="ANA SAYFA")
        self._build_modules_tab()

        # ÜCRET TAKİPİ - Çocuk ve Personel ücret takipleri (tek ekranda 2 sayfa)
        self.tab_ucret_takibi = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_ucret_takibi, text="ÜCRET TAKİPİ")
        self._build_ucret_takibi_tab()

        # Çocuk günlük takip modülü kaldırıldı (talep doğrultusunda).

        # KASA DEFTERİ - Günlük/haftalık/aylık raporlar
        self.tab_kasa = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_kasa, text="KASA DEFTERİ")
        self._build_kasa_defteri_tab()


        # HAFTALIK SEANS TAKİP - Personel bazlı dinamik program
        self.tab_haftalik = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_haftalik, text="HAFTALIK PROGRAM")
        self._build_haftalik_seans_tab()

        # ÖĞRENCİ BİLGİLERİ - Aile ve kimlik bilgileri
        self.tab_ogrenci_bilgileri = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_ogrenci_bilgileri, text="ÖĞRENCİ BİLGİLERİ")
        self._build_ogrenci_bilgileri_tab()

       


        # SİSTEM ŞİFRELERİ - KALDIRILDI (Kullanıcı isteği)

        # AYARLAR sadece kurum müdürü
        if self.kullanici_yetki == "kurum_muduru":
            self.tab_settings = ttk.Frame(self.nb, padding=10)
            self.nb.add(self.tab_settings, text="AYARLAR")
            self._build_settings_tab()

        # ✅ OTOMATIK GÜNCELLEME: Ana sayfa sekmesine geçildiğinde otomatik yenile
        def auto_refresh_on_tab_change(event=None):
            """Ana sayfa sekmesine geçildiğinde otomatik yenile"""
            try:
                selected_tab = self.nb.select()
                if selected_tab == str(self.tab_modules):
                    # Ana sayfa sekmesindeyiz, yenile
                    self._build_modules_tab()
            except Exception:
                pass
        
        # Notebook tab değişikliğinde otomatik yenile
        self.nb.bind("<<NotebookTabChanged>>", auto_refresh_on_tab_change)
        
        # Klasik akış: girişten sonra ilk görünen ANA SAYFA olsun
        try:
            self.nb.select(self.tab_modules)
        except Exception:
            pass

    def _build_modules_tab(self):
        """Modern card-based ana sayfa tasarımı"""
        # ✅ OTOMATIK YENİLEME: Mevcut içeriği temizle
        for widget in self.tab_modules.winfo_children():
            widget.destroy()
        
        wrapper = ttk.Frame(self.tab_modules, padding=20)
        wrapper.pack(fill=BOTH, expand=True)

        ttk.Label(
            wrapper,
            text="Açıklama: Bu ana sayfa kurumun temel modüllerine hızlı erişim sağlar.",
            font=("Segoe UI", 9),
            foreground="gray",
        ).pack(anchor=W, pady=(0, 6))

        # Başlık
        head = ttk.Frame(wrapper)
        head.pack(fill=X, pady=(0, 20))
        if getattr(self, "_logo_small", None):
            ttk.Label(head, image=self._logo_small).pack(side=LEFT, padx=(0, 15))
        ttk.Label(head, text="ANA SAYFA", font=("Segoe UI", 20, "bold"), bootstyle="primary").pack(side=LEFT)
        
        # ✅ OTOMATIK YENİLEME: Ana sayfa yenileme butonu
        def ana_sayfa_yenile():
            """Ana sayfayı tamamen yeniden yükle"""
            try:
                # Ana sayfa tab'ını yeniden oluştur
                self._build_modules_tab()
            except Exception as e:
                messagebox.showerror("Hata", f"Ana sayfa yenilenemedi:\n{e}")
                log_exception("ana_sayfa_yenile", e)
        
        ttk.Button(head, text="🔄 Yenile", bootstyle="info", command=ana_sayfa_yenile).pack(side=RIGHT, padx=10)
        ttk.Label(wrapper, text="Açıklama: Yenile butonu paneldeki canlı özetleri ve borç listelerini günceller.",
                  font=("Segoe UI", 9), foreground="gray").pack(anchor="w", pady=(0, 10))
        
        # ✅ PERSONEL CÜZDANI: Terapistlerin kendi bakiyelerini görebileceği bilgi etiketi
        if self.kullanici_terapist and self.kullanici_yetki != "kurum_muduru":
            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                cuzdan = pipeline.get_personel_cuzdan(self.kullanici_terapist)
                conn.close()
                
                cuzdan_frame = ttk.Labelframe(head, text="💼 Personel Cüzdanı", padding=8, bootstyle="info")
                cuzdan_frame.pack(side=RIGHT, padx=(20, 0))
                
                ttk.Label(cuzdan_frame, text=f"Beklemede: {format_money(cuzdan['beklemede_toplam'])}", 
                         font=("Segoe UI", 10, "bold"), bootstyle="warning").pack(anchor=W)
                ttk.Label(cuzdan_frame, text=f"Ödenen: {format_money(cuzdan['odendi_toplam'])}", 
                         font=("Segoe UI", 9), bootstyle="success").pack(anchor=W)
                ttk.Label(cuzdan_frame, text=f"Toplam: {format_money(cuzdan['toplam_hak_edis'])}", 
                         font=("Segoe UI", 9, "bold"), bootstyle="primary").pack(anchor=W, pady=(5, 0))
            except Exception as e:
                log_exception("personel_cuzdan_ui", e)
        
        # ✅ ENTERPRISE DASHBOARD: Pipeline'dan beslenen dinamik özet paneli
        dashboard_frame = ttk.Labelframe(wrapper, text="📊 Executive Summary", padding=15, bootstyle="info")
        dashboard_frame.pack(fill=X, pady=(0, 25))
        
        # Pipeline'dan dashboard verilerini al
        try:
            conn = self.veritabani_baglan()
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            dashboard_data = pipeline.get_dashboard_data()
            conn.close()
        except Exception as e:
            log_exception("dashboard_data", e)
            dashboard_data = {
                "operasyonel": {"bugun_beklenen_seans": 0, "bugun_tamamlanan_seans": 0, "bugun_toplam_seans": 0},
                "finansal": {"bugun_kasa_giren": 0.0, "beklenen_toplam_alacak": 0.0, "toplam_borc": 0.0},
                "kritik": [],
                "borclular": []
            }
        
        # Üst satır: Operasyonel Metrikler
        operasyonel_frame = ttk.Frame(dashboard_frame)
        operasyonel_frame.pack(fill=X, pady=(0, 15))
        
        ttk.Label(operasyonel_frame, text="📈 Operasyonel", font=("Segoe UI", 12, "bold"), bootstyle="primary").pack(anchor=W, pady=(0, 8))
        
        ops_row = ttk.Frame(operasyonel_frame)
        ops_row.pack(fill=X)
        
        def create_dashboard_card(parent, title, value, icon, color, subtitle=""):
            card = ttk.Labelframe(parent, text="", padding=8, bootstyle="secondary")
            card.pack(side=LEFT, fill=BOTH, expand=True, padx=5)
            
            ttk.Label(card, text=icon, font=("Segoe UI", 14)).pack()
            ttk.Label(card, text=str(value), font=("Segoe UI", 12, "bold"), bootstyle=color).pack(pady=(2, 0))
            ttk.Label(card, text=title, font=("Segoe UI", 8), foreground="gray").pack()
            if subtitle:
                ttk.Label(card, text=subtitle, font=("Segoe UI", 7), foreground="darkgray").pack(pady=(1, 0))
            return card
        
        ops = dashboard_data["operasyonel"]
        create_dashboard_card(ops_row, "Bugün Beklenen", ops["bugun_beklenen_seans"], "⏰", "warning")
        create_dashboard_card(ops_row, "Bugün Tamamlanan", ops["bugun_tamamlanan_seans"], "✅", "success")
        create_dashboard_card(ops_row, "Toplam Seans", ops["bugun_toplam_seans"], "📅", "info")
        
        # Orta satır: Finansal Metrikler
        finansal_frame = ttk.Frame(dashboard_frame)
        finansal_frame.pack(fill=X, pady=(0, 15))
        
        ttk.Label(finansal_frame, text="💰 Finansal", font=("Segoe UI", 12, "bold"), bootstyle="primary").pack(anchor=W, pady=(0, 8))
        
        fin_row = ttk.Frame(finansal_frame)
        fin_row.pack(fill=X)
        
        fin = dashboard_data["finansal"]
        create_dashboard_card(fin_row, "Bugün Kasa Giren", format_money(fin["bugun_kasa_giren"]), "💵", "success", "Nakit")
        create_dashboard_card(fin_row, "Beklenen Alacak", format_money(fin["beklenen_toplam_alacak"]), "📊", "warning", "Toplam")
        create_dashboard_card(fin_row, "Toplam Borç", format_money(fin["toplam_borc"]), "⚠️", "danger", "Kalan")
        
        # Alt satır: Kritik Liste (Kırmızı Liste) - GELİŞTİRİLMİŞ
        if dashboard_data.get("kritik"):
            kritik_frame = ttk.Frame(dashboard_frame)
            kritik_frame.pack(fill=X, pady=(0, 15))
            
            ttk.Label(kritik_frame, text="🚨 Kritik: Ödemesi Geciken Danışanlar (1+ Hafta)", 
                     font=("Segoe UI", 12, "bold"), bootstyle="danger").pack(anchor=W, pady=(0, 8))
            
            kritik_tree_frame = ttk.Frame(kritik_frame)
            kritik_tree_frame.pack(fill=X)
            
            kritik_tree = ttk.Treeview(kritik_tree_frame, columns=("Danışan", "Kalan Borç", "Gecikme"), 
                                      show="headings", height=min(6, len(dashboard_data["kritik"]) + 1))
            kritik_tree.heading("Danışan", text="Danışan Adı")
            kritik_tree.heading("Kalan Borç", text="Kalan Borç")
            kritik_tree.heading("Gecikme", text="Gecikme (Gün)")
            kritik_tree.column("Danışan", width=250)
            kritik_tree.column("Kalan Borç", width=150, anchor="e")
            kritik_tree.column("Gecikme", width=120, anchor="center")
            
            for item in dashboard_data["kritik"]:
                gecikme_gunu = item["gecikme_gunu"]
                oncelik = item.get("oncelik", "orta")
                gecikme_text = f"{gecikme_gunu} gün"
                
                # Renk kodlaması
                tag_name = "kritik_kirmizi" if oncelik == "kritik" else ("kritik_turuncu" if oncelik == "yuksek" else "kritik_sari")
                
                if oncelik == "kritik":
                    gecikme_text += " 🔴"
                elif oncelik == "yuksek":
                    gecikme_text += " 🟠"
                else:
                    gecikme_text += " 🟡"
                
                kritik_tree.insert("", END, values=(
                    item["danisan_adi"],
                    format_money(item["kalan_borc"]),
                    gecikme_text
                ), tags=(tag_name,))
            
            kritik_tree.tag_configure("kritik_kirmizi", foreground="#d32f2f", background="#ffebee")
            kritik_tree.tag_configure("kritik_turuncu", foreground="#f57c00", background="#fff3e0")
            kritik_tree.tag_configure("kritik_sari", foreground="#fbc02d", background="#fffde7")
            kritik_tree.pack(side=LEFT, fill=BOTH, expand=True)
            
            kritik_sb = ttk.Scrollbar(kritik_tree_frame, orient=VERTICAL, command=kritik_tree.yview)
            kritik_tree.configure(yscroll=kritik_sb.set)
            kritik_sb.pack(side=RIGHT, fill=Y)
        else:
            # Kritik liste boşsa bilgi göster
            ttk.Label(dashboard_frame, text="✅ Tüm danışanların ödemeleri güncel!", 
                     font=("Segoe UI", 10), bootstyle="success").pack(pady=10)
        
        # ✅ BORÇ ALARMI: Borcu olan danışanları göster
        borclular = dashboard_data.get("borclular") or []
        if borclular:
            borc_frame = ttk.Frame(dashboard_frame)
            borc_frame.pack(fill=X)

            ttk.Label(borc_frame, text="💳 Borcu Olan Danışanlar", 
                     font=("Segoe UI", 12, "bold"), bootstyle="danger").pack(anchor=W, pady=(0, 8))

            borc_tree_frame = ttk.Frame(borc_frame)
            borc_tree_frame.pack(fill=X)

            borc_tree = ttk.Treeview(borc_tree_frame, columns=("Danışan", "Kalan Borç", "Kayıt"), 
                                     show="headings", height=min(8, len(borclular) + 1))
            borc_tree.heading("Danışan", text="Danışan Adı")
            borc_tree.heading("Kalan Borç", text="Kalan Borç")
            borc_tree.heading("Kayıt", text="Açık Kayıt")
            borc_tree.column("Danışan", width=260)
            borc_tree.column("Kalan Borç", width=160, anchor="e")
            borc_tree.column("Kayıt", width=120, anchor="center")

            for item in borclular:
                borc_tree.insert("", END, values=(
                    item.get("danisan_adi", ""),
                    format_money(item.get("kalan_borc", 0)),
                    str(item.get("acik_kayit", 0)),
                ))

            borc_tree.pack(side=LEFT, fill=BOTH, expand=True)
            borc_sb = ttk.Scrollbar(borc_tree_frame, orient=VERTICAL, command=borc_tree.yview)
            borc_tree.configure(yscroll=borc_sb.set)
            borc_sb.pack(side=RIGHT, fill=Y)
    

    def _validate_money(self, newval: str) -> bool:
        try:
            if newval is None:
                return True
            s = str(newval).strip()
            if s == "":
                return True
            s = s.replace(",", ".")
            float(s)
            return True
        except Exception:
            return False

    def veritabani_baglan(self) -> sqlite3.Connection:
        return connect_db()

    def _table_has_column(self, conn: sqlite3.Connection, table: str, column: str) -> bool:
        try:
            cur = conn.cursor()
            cur.execute(f"PRAGMA table_info({table})")
            cols = [str(r[1]) for r in (cur.fetchall() or []) if len(r) > 1]
            return column in cols
        except Exception:
            return False

    def _refresh_borc_tables(self) -> None:
        """Balance/records değişince danışan listesindeki borç sütununu yenile.
        Hem hemen çağrılır (güncel veri görünsün) hem after(0, ...) ile tekrar (pencere/platform gecikmeleri için)."""
        def _do_refresh():
            try:
                wrapper = getattr(self, "_ogrenci_bilgileri_wrapper", None)
                if wrapper and hasattr(wrapper, "_tree_danisanlar") and hasattr(wrapper, "_ent_ara_danisan"):
                    self._tum_danisanlari_listele(wrapper)
            except Exception:
                pass
        _do_refresh()
        try:
            self.after(0, _do_refresh)
        except Exception:
            pass

    def _register_report_refresher(self, win, callback) -> None:
        refs = getattr(self, "_open_report_refreshers", {})
        refs[str(win)] = callback
        self._open_report_refreshers = refs
        try:
            win.bind("<Destroy>", lambda e, k=str(win): self._open_report_refreshers.pop(k, None), add="+")
        except Exception:
            pass

    def _refresh_open_reports(self) -> None:
        refs = dict(getattr(self, "_open_report_refreshers", {}))
        for key, cb in refs.items():
            try:
                cb()
            except Exception:
                self._open_report_refreshers.pop(key, None)

    def _reload_open_toplevels(self) -> None:
        """Açık olan Toplevel pencerelerinde _reload varsa çağır (danışan/liste güncellemeleri)."""
        try:
            for child in self.winfo_children():
                if isinstance(child, ttk.Toplevel) and hasattr(child, "_reload"):
                    try:
                        child._reload()
                    except Exception:
                        pass
        except Exception:
            pass

    # --- TAB 1: SEANS TAKİP ---
    def _build_records_tab(self):
        top = ttk.Labelframe(self.tab_records, text="Yeni Seans Kaydı (Tarih, Danışan Adı, Terapist, Alınacak Ücret, Alınan Ücret, Kalan Borç)", padding=16, bootstyle="primary")
        top.pack(fill=X, pady=(0, 12))
        # Farklı ekran genişliklerinde taşmayı azaltmak için kolonlar esnek
        for col in (1, 3, 5):
            try:
                top.grid_columnconfigure(col, weight=1)
            except Exception:
                pass

        # Tabloya göre: Tarih, Danışan Adı, Terapist (saat otomatik atanır; formda yok)
        ttk.Label(top, text="Tarih:", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=8, pady=8, sticky=W)
        self.tarih_var = ttk.StringVar(value=datetime.datetime.now().strftime("%d.%m.%Y"))
        ttk.Entry(top, textvariable=self.tarih_var, width=16, font=("Segoe UI", 10)).grid(
            row=0, column=1, padx=8, pady=8, sticky=W
        )
        fin_frame = ttk.Labelframe(self.tab_records, text="💰 Hızlı Finansal İşlemler (Eski Borç & Toplu Ödeme)", padding=10, bootstyle="warning")
        fin_frame.pack(fill=X, pady=10, padx=5)

        ttk.Button(fin_frame, text="📉 Eski Borç Yükle (Devir)", bootstyle="danger", command=self.safe_call(self.popup_eski_borc, "popup_eski_borc")).pack(side=LEFT, padx=10)
        ttk.Button(fin_frame, text="🗑️ Eski Borç Sil", bootstyle="secondary", command=self.safe_call(self.popup_eski_borc_sil, "popup_eski_borc_sil")).pack(side=LEFT, padx=10)
        ttk.Button(fin_frame, text="💳 Toplu Ödeme Al (Bakiyeden Düş)", bootstyle="success", command=self.popup_toplu_odeme).pack(side=LEFT, padx=10)
        ttk.Button(fin_frame, text="🧩 Toplu Ödeme Yönetimi", bootstyle="info", command=self.safe_call(self.popup_toplu_odeme_yonet, "popup_toplu_odeme_yonet")).pack(side=LEFT, padx=10)
        ttk.Label(top, text="Danışan Adı:", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, padx=8, pady=8, sticky=W)
        danisan_frame = ttk.Frame(top)
        danisan_frame.grid(row=0, column=3, padx=8, pady=8, sticky=W)
        
        self.cmb_danisan = ttk.Combobox(danisan_frame, width=28, state="normal")
        self._yenile_danisan_listesi()
        self.cmb_danisan.pack(side=LEFT)
        
        def _yeni_danisan_ekle():
            self._yeni_danisan_ekle_ve_guncelle(self.cmb_danisan, top)
        
        btn_yeni_danisan = ttk.Button(danisan_frame, text="+", bootstyle="success-outline", width=4,
                                      command=_yeni_danisan_ekle)
        btn_yeni_danisan.pack(side=LEFT, padx=(4, 0))

        ttk.Label(top, text="Terapist:", font=("Segoe UI", 10, "bold")).grid(row=0, column=4, padx=8, pady=8, sticky=W)
        self.cmb_terapist = ttk.Combobox(top, state="readonly", width=24)
        self.cmb_terapist.grid(row=0, column=5, padx=8, pady=8, sticky=W)
        
        # ✅ ENTERPRISE: AKILLI VARSAYILANLAR (Zero-Effort UI)
        def _akilli_varsayilanlar_ata(*args):
            """Enterprise Smart Defaults: Otomatik fiyat, oda ve çakışma kontrolü"""
            try:
                danisan_adi = (self.cmb_danisan.get() or "").strip().upper()
                terapist_adi = (self.cmb_terapist.get() or "").strip()
                
                if danisan_adi and terapist_adi:
                    conn = self.veritabani_baglan()
                    kullanici_id = self.kullanici[0] if self.kullanici else None
                    pipeline = DataPipeline(conn, kullanici_id)
                    
                    # Enterprise Smart Defaults: Tek metod çağrısı ile tüm bilgileri al (varsayılan saat 09:00)
                    try:
                        tarih = self._tarih_db()
                        saat = "09:00"
                        smart_defaults = pipeline.get_smart_defaults(danisan_adi, terapist_adi, tarih, saat)
                        
                        # 1) Otomatik fiyat (oda seçimi kaldırıldı)
                        if float(smart_defaults.get("price", 0) or 0) > 0:
                            self._otomatik_bedel = float(smart_defaults["price"] or 0)
                            if hasattr(self, "lbl_bedel"):
                                self.lbl_bedel.config(text=format_money(self._otomatik_bedel))
                            if hasattr(self, "ent_alinan"):
                                mevcut_alinan = (self.ent_alinan.get() or "").strip()
                                if (not mevcut_alinan) or parse_money(mevcut_alinan) == 0:
                                    self.ent_alinan.delete(0, END)
                                    self.ent_alinan.insert(0, format_money(self._otomatik_bedel))
                    except Exception as e:
                        log_exception("smart_defaults_kayit_ekle", e)
                    
                    conn.close()
            except Exception as e:
                log_exception("_akilli_varsayilanlar_ata", e)
        
        # Event binding - Danışan veya terapist değiştiğinde akıllı varsayılanları ata
        self.cmb_danisan.bind("<<ComboboxSelected>>", _akilli_varsayilanlar_ata)
        self.cmb_danisan.bind("<KeyRelease>", lambda e: self.after(300, _akilli_varsayilanlar_ata))
        self.cmb_terapist.bind("<<ComboboxSelected>>", _akilli_varsayilanlar_ata)
        
        # Otomatik bedel için gizli değişken
        self._otomatik_bedel = 0.0

        # Tabloya göre: Alınacak Ücret (hizmet bedeli), Alınan Ücret (alınan), NOTLAR
        # Hizmet bedeli kullanıcıdan alınmaz; fiyatlandırmadan otomatik gelir
        ttk.Label(top, text="Hizmet Bedeli (otomatik):", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, padx=8, pady=8, sticky=W)
        self.lbl_bedel = ttk.Label(top, text="0.00 ₺", font=("Segoe UI", 10))
        self.lbl_bedel.grid(row=1, column=1, padx=8, pady=8, sticky=W)

        ttk.Label(top, text="Alınan Ücret (₺):", font=("Segoe UI", 10, "bold")).grid(row=1, column=2, padx=8, pady=8, sticky=W)
        self.ent_alinan = ttk.Entry(top, width=18, validate="key", validatecommand=self._vcmd_money, font=("Segoe UI", 10))
        self.ent_alinan.insert(0, "0")
        self.ent_alinan.grid(row=1, column=3, padx=8, pady=8, sticky=W)

        ttk.Label(top, text="NOTLAR:", font=("Segoe UI", 10, "bold")).grid(row=1, column=4, padx=8, pady=8, sticky=W)
        self.ent_not = ttk.Entry(top, width=36, font=("Segoe UI", 10))
        self.ent_not.grid(row=1, column=5, columnspan=2, padx=8, pady=8, sticky=W+E)
        
        info_label = ttk.Label(top, text="ℹ️ Alınan ücret otomatik dolar (istersen değiştirebilirsin). Seans kaydı haftalık programdan bağımsızdır.", 
                              font=("Segoe UI", 9), foreground="gray")
        info_label.grid(row=2, column=0, columnspan=7, padx=8, pady=(0, 8), sticky=W)

        ttk.Button(top, text="KAYDET", bootstyle="success", command=self.kayit_ekle, width=14).grid(
            row=0, column=6, rowspan=1, padx=8, pady=8, sticky="ew"
        )

        # RAPORLAR (günlük / haftalık / toplam) - kasa için gereken özetler burada
        rep = ttk.Labelframe(self.tab_records, text="Raporlar", padding=8, bootstyle="secondary")
        rep.pack(fill=X, pady=(0, 10))
        rep.grid_columnconfigure(0, weight=1)
        rep.grid_columnconfigure(1, weight=1)
        ttk.Button(rep, text="Günlük Rapor", width=18, bootstyle="primary", command=self.gunluk_rapor_pencere).grid(row=0, column=0, padx=4, pady=4, sticky="ew")
        ttk.Button(rep, text="Haftalık Rapor", width=18, bootstyle="primary", command=self.haftalik_rapor_pencere).grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        ttk.Button(rep, text="Toplam Rapor", width=18, bootstyle="secondary", command=self.toplam_rapor_pencere).grid(row=1, column=0, padx=4, pady=(2,4), sticky="ew")
        ttk.Button(rep, text="📘 Kılavuz", width=18, bootstyle="info", command=self.kullanim_kilavuzu_ac).grid(row=1, column=1, padx=4, pady=(2,4), sticky="ew")
        ttk.Label(
            self.tab_records,
            text="İpucu: Seans kaydı ve ödeme işlemleri otomatik işlenir; ek senkronizasyon ayarı gerekmez.",
            foreground="gray",
        ).pack(anchor=W, pady=(0, 8), padx=4)

        mid = ttk.Frame(self.tab_records)
        mid.pack(fill=X, pady=(0, 8))
        ttk.Label(mid, text="İsim ile ara...", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(4, 6))
        self.ent_ara = ttk.Entry(mid, width=40)
        self.ent_ara.pack(side=LEFT, padx=6)
        self.ent_ara.bind("<KeyRelease>", lambda e: self.kayitlari_listele())

        self.lbl_ozet = ttk.Label(
            mid, text="TOPLAM İÇERİDEKİ ALACAK: 0.00 TL", font=("Segoe UI", 12, "bold"), bootstyle="danger"
        )
        self.lbl_ozet.pack(side=RIGHT, padx=6)
        
        # ✅ Notlar gösterimi için label (tıklanabilir - başlangıçta gizli)
        self.lbl_notlar = ttk.Label(
            self.tab_records, text="", font=("Segoe UI", 10), wraplength=800, justify="left",
            foreground="#0066cc", background="#f0f8ff", relief="solid", borderwidth=1, padding=10
        )

        # Seçim/action butonlarını tablonun üstüne al: küçük ekranlarda görünür kalsın
        sec_toolbar = ttk.Frame(self.tab_records)
        sec_toolbar.pack(fill=X, pady=(0, 4), padx=4)
        ttk.Label(sec_toolbar, text="Seç:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 8))
        ttk.Button(sec_toolbar, text="Hepsini Seç", bootstyle="secondary-outline", command=self._seans_hepsini_sec).pack(side=LEFT, padx=4)
        ttk.Button(sec_toolbar, text="Seçimi Kaldır", bootstyle="secondary-outline", command=self._seans_secimi_kaldir).pack(side=LEFT, padx=4)
        ttk.Button(sec_toolbar, text="Seçilileri Sil", bootstyle="danger-outline", command=self.seclileri_sil).pack(side=LEFT, padx=4)

        action_toolbar = ttk.Frame(self.tab_records)
        action_toolbar.pack(fill=X, pady=(0, 8), padx=4)
        ttk.Label(action_toolbar, text="Seçili Kayıt İşlemleri:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 10))

        self.btn_odeme_ekle = ttk.Button(action_toolbar, text="💰 Ödeme Ekle", bootstyle="success", command=self.odeme_ekle, state="disabled")
        self.btn_odeme_ekle.pack(side=LEFT, padx=5)

        self.btn_kayit_sil = ttk.Button(action_toolbar, text="🗑️ Kaydı Sil", bootstyle="danger", command=self.kayit_sil, state="disabled")
        self.btn_kayit_sil.pack(side=LEFT, padx=5)

        table = ttk.Frame(self.tab_records)
        table.pack(fill=BOTH, expand=True)

        # Seans tablosu: Tarih, Danışan Adı, Terapist, Alınacak Ücret, Alınan Ücret, Kalan Borç
        cols = ("ID", "Tarih", "Danışan Adı", "Terapist", "Alınacak Ücret", "Alınan Ücret", "Kalan Borç")
        self._style_table_strong()
        self.tree = ttk.Treeview(table, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview", selectmode="extended")
        for c in cols:
            self.tree.heading(c, text=c)
        self.tree.column("ID", width=0, stretch=False)  # ID gizli (sadece iç kullanım için)
        self.tree.column("Tarih", width=110, anchor="center")
        self.tree.column("Danışan Adı", width=250)
        self.tree.column("Terapist", width=150)
        self.tree.column("Alınacak Ücret", width=120, anchor="e")
        self.tree.column("Alınan Ücret", width=120, anchor="e")
        self.tree.column("Kalan Borç", width=120, anchor="e")
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)

        sb = ttk.Scrollbar(table, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)

        self.tree.tag_configure("borclu", background="#f8d7da", foreground="#721c24")
        self.tree.tag_configure("tamam", background="#d4edda", foreground="#155724")
        
        # Notlar kolonu kaldırıldı; tıklama özelliği yok
        
        # Seçim değiştiğinde butonları aktif/pasif yap
        def _on_selection_change(event):
            sel = self.tree.selection()
            if sel:
                self.btn_odeme_ekle.config(state="normal")
                self.btn_kayit_sil.config(state="normal")
            else:
                self.btn_odeme_ekle.config(state="disabled")
                self.btn_kayit_sil.config(state="disabled")
        
        self.tree.bind("<<TreeviewSelect>>", _on_selection_change)
        

    def _selected_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        vals = self.tree.item(sel[0]).get("values") or []
        if not vals:
            return None
        try:
            return int(vals[0])
        except Exception:
            return None

    def _selected_ids(self):
        """Seçili tüm satırların ID (seans_id) listesini döndürür."""
        ids = []
        for iid in self.tree.selection():
            vals = self.tree.item(iid).get("values") or []
            if vals:
                try:
                    ids.append(int(vals[0]))
                except (ValueError, TypeError):
                    pass
        return ids

    def _secim_modu_sor(self, adet: int, tur_adi: str = "kayıt") -> str | None:
        if adet <= 0:
            return None
        if adet == 1:
            return "tek"
        yanit = messagebox.askyesnocancel("Silme Modu", f"{adet} adet {tur_adi} seçili.\n\nEvet: Sadece ilk seçiliyi sil\nHayır: Tüm seçilileri sil\nİptal: Vazgeç")
        if yanit is None:
            return None
        return "tek" if yanit else "toplu"

    def _seans_hepsini_sec(self):
        """Seans Takip listesindeki tüm satırları seç."""
        self.tree.selection_set(self.tree.get_children())

    def _seans_secimi_kaldir(self):
        """Seans Takip listesindeki seçimi kaldır."""
        self.tree.selection_remove(*self.tree.selection())

    def seclileri_sil(self):
        """Seçili kayıtlar için tekli/toplu silme modu sunar."""
        ids = self._selected_ids()
        if not ids:
            messagebox.showwarning("Uyarı", "Lütfen silmek istediğiniz kayıtları seçin!")
            return
        mod = self._secim_modu_sor(len(ids), "seans kaydı")
        if not mod:
            return
        hedef_ids = ids[:1] if mod == "tek" else ids
        if not messagebox.askyesno(
            "Onay",
            f"{len(hedef_ids)} adet kayıt silinecek.\n\nİlgili tüm veriler (seans takvimi, records, ödemeler, kasa kayıtları) silinecektir!\n\nDevam etmek istiyor musunuz?"
        ):
            return

        conn = None
        try:
            conn = self.veritabani_baglan()
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            silinen = 0
            for seans_id in hedef_ids:
                try:
                    if pipeline.kayit_sil(seans_id=seans_id):
                        silinen += 1
                except Exception as e:
                    log_exception("seclileri_sil_one", e)
            if silinen > 0:
                self.kayitlari_listele()
                self._refresh_borc_tables()
                self._refresh_open_reports()
            messagebox.showinfo("Tamamlandı", f"{silinen} kayıt silindi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Silme hatası:\n{e}")
            log_exception("seclileri_sil", e)
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass

    def _tarih_db(self) -> str:
        try:
            s = (self.tarih_var.get() or "").strip()
        except Exception:
            s = ""
        try:
            dt = datetime.datetime.strptime(s, "%d.%m.%Y")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return datetime.datetime.now().strftime("%Y-%m-%d")

    def _tarih_db_from(self, s: str) -> str:
        s = (s or "").strip()
        if not s:
            return datetime.datetime.now().strftime("%Y-%m-%d")
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        return datetime.datetime.now().strftime("%Y-%m-%d")

    def _yenile_danisan_listesi(self):
        """Danışan listesini veritabanından yükle ve combobox'a ekle"""
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            danisan_listesi = [row[0] for row in cur.fetchall()]
            conn.close()
            if hasattr(self, 'cmb_danisan'):
                self.cmb_danisan["values"] = danisan_listesi
            return danisan_listesi
        except Exception as e:
            log_exception("_yenile_danisan_listesi", e)
            if hasattr(self, 'cmb_danisan'):
                self.cmb_danisan["values"] = []
            return []
    
    def _yeni_danisan_ekle_ve_guncelle(self, combobox, parent):
        """Yeni danışan ekle - tabloya göre: AD SOYAD, DOĞUM TARİHİ, VELİ ADI, VELİ TELEFON, ADRES (telefon/email opsiyonel - çoğu danışanda yok)"""
        try:
            win = ttk.Toplevel(self)
            win.title("Yeni Danışan Ekle")
            win.resizable(True, False)
            center_window(win, 520, 520)
            win.transient(self)
            win.grab_set()
            self._brand_window(win)
            
            wrapper = ttk.Frame(win, padding=20)
            wrapper.pack(fill=BOTH, expand=True)
            
            ttk.Label(wrapper, text="Yeni Danışan Ekle (tablo: AD SOYAD, DOĞUM TARİHİ, VELİ ADI, VELİ TELEFON, ADRES – iletişim veli bilgilerinden)", font=("Segoe UI", 11, "bold"), bootstyle="primary").pack(pady=(0, 12))
            
            def _field(label, init=""):
                ttk.Label(wrapper, text=label, font=("Segoe UI", 10)).pack(anchor=W, pady=(8, 0))
                e = ttk.Entry(wrapper, width=45, font=("Segoe UI", 10))
                e.insert(0, init or "")
                e.pack(fill=X, pady=2)
                return e
            
            ent_ad = _field("AD SOYAD *:", "")
            ent_ad.focus_set()
            ent_dogum = _field("DOĞUM TARİHİ (GG.AA.YYYY):", "")
            ent_veli = _field("VELİ ADI:", "")
            ent_veli_tel = _field("VELİ TELEFON:", "")
            ttk.Label(wrapper, text="ADRES:", font=("Segoe UI", 10)).pack(anchor=W, pady=(8, 0))
            txt_adres = ttk.Text(wrapper, height=2, width=45, font=("Segoe UI", 10))
            txt_adres.pack(fill=X, pady=2)
            
            def _kaydet():
                ad_soyad = (ent_ad.get() or "").strip()
                if not ad_soyad:
                    messagebox.showwarning("Uyarı", "AD SOYAD zorunludur!")
                    return
                
                try:
                    conn = self.veritabani_baglan()
                    cur = conn.cursor()
                    cur.execute("SELECT id FROM danisanlar WHERE UPPER(ad_soyad) = UPPER(?) AND aktif=1", (ad_soyad,))
                    if cur.fetchone():
                        messagebox.showinfo("Bilgi", f"{ad_soyad} zaten kayıtlı!")
                        conn.close()
                        win.destroy()
                        self._yenile_danisan_listesi()
                        if combobox:
                            combobox.set(ad_soyad)
                        return
                    
                    dogum = (ent_dogum.get() or "").strip()
                    veli_adi = (ent_veli.get() or "").strip()
                    veli_tel = (ent_veli_tel.get() or "").strip()
                    adres = (txt_adres.get("1.0", END) or "").strip()
                    olusturma = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    cur.execute(
                        """INSERT INTO danisanlar (ad_soyad, telefon, email, dogum_tarihi, veli_adi, veli_telefon, adres, aktif, olusturma_tarihi)
                           VALUES (?, '', '', ?, ?, ?, ?, 1, ?)""",
                        (ad_soyad, dogum, veli_adi, veli_tel, adres, olusturma)
                    )
                    conn.commit()
                    conn.close()
                    
                    messagebox.showinfo("Başarılı", f"{ad_soyad} eklendi!")
                    win.destroy()
                    self._yenile_danisan_listesi()
                    if combobox:
                        combobox.set(ad_soyad)
                    
                except Exception as e:
                    messagebox.showerror("Hata", f"Danışan eklenemedi:\n{e}")
                    log_exception("_yeni_danisan_ekle_ve_guncelle", e)
            
            ttk.Button(wrapper, text="Kaydet", bootstyle="success", command=_kaydet, width=20).pack(pady=16)
            ent_ad.bind("<Return>", lambda e: ent_dogum.focus_set())
            ent_dogum.bind("<Return>", lambda e: _kaydet())
            
        except Exception as e:
            messagebox.showerror("Hata", f"Pencere açılamadı:\n{e}")
            log_exception("_yeni_danisan_ekle_ve_guncelle", e)

    def hizli_seans_kaydi_ekle(self):
        """Yeni kullanıcılar için en basit kayıt ekranı (ANA SAYFA)."""
        win = ttk.Toplevel(self)
        win.title("Seans Kaydı Ekle")
        win.resizable(False, False)
        center_window(win, 700, 500)
        win.transient(self)
        win.grab_set()
        self._brand_window(win)

        ttk.Label(win, text="SEANS KAYDI EKLE (tablo: TARİH, DANIŞANIN ADI, TERAPİSTİN ADI, ALINAN ÜCRET, GÜNCEL, NOTLAR)", font=("Segoe UI", 12, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(win, padding=14)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text="TARİH:").grid(row=0, column=0, sticky=W, padx=6, pady=6)
        tarih_var = ttk.StringVar(value=datetime.datetime.now().strftime("%d.%m.%Y"))
        ent_tarih = ttk.Entry(frm, textvariable=tarih_var, width=16)
        ent_tarih.grid(row=0, column=1, sticky=W, padx=6, pady=6)

        ttk.Label(frm, text="DANIŞANIN ADI:").grid(row=1, column=0, sticky=W, padx=6, pady=6)
        danisan_frame = ttk.Frame(frm)
        danisan_frame.grid(row=1, column=1, sticky=W, padx=6, pady=6)
        
        cb_dan = ttk.Combobox(danisan_frame, width=30, state="normal")
        # Danışan listesini yükle
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            danisan_listesi = [row[0] for row in cur.fetchall()]
            conn.close()
            cb_dan["values"] = danisan_listesi
        except Exception:
            cb_dan["values"] = []
        cb_dan.pack(side=LEFT)
        cb_dan.focus_set()
        
        def _yeni_danisan_ekle():
            self._yeni_danisan_ekle_ve_guncelle(cb_dan, frm)
        
        btn_yeni_dan = ttk.Button(danisan_frame, text="+", bootstyle="success-outline", width=4, command=_yeni_danisan_ekle)
        btn_yeni_dan.pack(side=LEFT, padx=(4, 0))

        ttk.Label(frm, text="TERAPİSTİN ADI:", font=("Segoe UI", 10)).grid(row=2, column=0, sticky=W, padx=8, pady=8)
        cb_ter = ttk.Combobox(frm, state="readonly", width=30)
        cb_ter.grid(row=2, column=1, sticky=W, padx=8, pady=8)

        # Terapist listesi (settings -> fallback)
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            tnames = [r[0] for r in cur.fetchall()]
            conn.close()
        except Exception:
            tnames = DEFAULT_THERAPISTS[:]

        if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
            cb_ter["values"] = [self.kullanici_terapist]
            cb_ter.set(self.kullanici_terapist)
            cb_ter.configure(state="disabled")
        else:
            cb_ter["values"] = tnames
            if tnames:
                cb_ter.current(0)

        # ✅ ENTERPRISE: AKILLI VARSAYILANLAR (Zero-Effort UI)
        def _akilli_varsayilanlar_ata_hizli(*args):
            """Yeni seans kaydı için danışan+hoca fiyatını otomatik doldurur."""
            try:
                danisan_adi = (cb_dan.get() or "").strip()
                terapist_adi = (cb_ter.get() or "").strip()
                if not danisan_adi or not terapist_adi:
                    return

                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                try:
                    tarih = self._tarih_db_from(tarih_var.get())
                    smart_defaults = pipeline.get_smart_defaults(danisan_adi, terapist_adi, tarih, "09:00")
                    bedel = float(smart_defaults.get("price", 0) or 0)
                    ent_bedel.configure(state="normal")
                    ent_bedel.delete(0, END)
                    ent_bedel.insert(0, format_money(bedel))
                    ent_bedel.configure(state="readonly")
                except Exception as e:
                    log_exception("smart_defaults_hizli", e)
                finally:
                    conn.close()
            except Exception as e:
                log_exception("_akilli_varsayilanlar_ata_hizli", e)
        
        # Event binding - Akıllı varsayılanlar
        cb_dan.bind("<<ComboboxSelected>>", _akilli_varsayilanlar_ata_hizli)
        cb_dan.bind("<KeyRelease>", lambda e: win.after(300, _akilli_varsayilanlar_ata_hizli))
        cb_ter.bind("<<ComboboxSelected>>", _akilli_varsayilanlar_ata_hizli)

        ttk.Label(frm, text="ALINAN ÜCRET (₺):").grid(row=3, column=0, sticky=W, padx=6, pady=6)
        ent_bedel = ttk.Entry(frm, validate="key", validatecommand=self._vcmd_money, width=18, state="readonly")
        ent_bedel.grid(row=3, column=1, sticky=W, padx=6, pady=6)

        ttk.Label(frm, text="GÜNCEL (Alınan - ₺):").grid(row=4, column=0, sticky=W, padx=6, pady=6)
        ent_alinan = ttk.Entry(frm, validate="key", validatecommand=self._vcmd_money, width=18)
        ent_alinan.insert(0, "0")
        ent_alinan.grid(row=4, column=1, sticky=W, padx=6, pady=6)

        ttk.Label(frm, text="NOTLAR (opsiyonel):").grid(row=5, column=0, sticky=W, padx=6, pady=6)
        ent_not = ttk.Entry(frm, width=34)
        ent_not.grid(row=5, column=1, sticky=W, padx=6, pady=6)

        def _save():
            danisan = (cb_dan.get() or "").strip()
            terapist = (cb_ter.get() or "").strip()
            if not danisan:
                messagebox.showwarning("Uyarı", "Lütfen danışan adını giriniz!")
                return
            if not terapist:
                messagebox.showwarning("Uyarı", "Lütfen terapist seçiniz!")
                return
            try:
                bedel = parse_money(ent_bedel.get())
                alinan = parse_money(ent_alinan.get())
            except Exception:
                messagebox.showerror("Hata", "Lütfen sayı giriniz! (Alınan ücret / Güncel)")
                return

            kalan = max(0.0, bedel - alinan)
            notlar_manuel = (ent_not.get() or "").strip()
            saat = "09:00"
            
            # Otomatik açıklama oluştur
            tarih_formatted = tarih_var.get() or datetime.datetime.now().strftime("%d.%m.%Y")
            otomatik_aciklama = f"{danisan} danışanın {terapist} hocadan {tarih_formatted} tarihli seansından {format_money(bedel)} ücret bedelinin {format_money(alinan)} miktarı alınmıştır. {format_money(kalan)} miktar borcu kalmıştır."
            
            # Manuel not varsa ekle
            if notlar_manuel:
                notlar = f"{otomatik_aciklama}\n\nNot: {notlar_manuel}"
            else:
                notlar = otomatik_aciklama

            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                
                # Oda seçimi kaldırıldı; kayıt oda olmadan yapılır
                tarih_db = self._tarih_db_from(tarih_var.get())
                pipeline = DataPipeline(conn, kullanici_id)
                seans_id = pipeline.seans_kayit(
                    tarih=tarih_db,
                    saat=saat,
                    danisan_adi=danisan,
                    terapist=terapist,
                    hizmet_bedeli=bedel,
                    alinan_ucret=alinan,
                    notlar=notlar,
                    oda=oda,
                    check_oda_cakisma=False
                )
                
                conn.close()
                
                if not seans_id:
                    messagebox.showerror("Hata", "Seans kaydı oluşturulamadı!")
                    return
                    
            except Exception as e:
                messagebox.showerror("Hata", f"Kayıt ekleme hatası:\n{e}")
                log_exception("hizli_seans_kaydi_ekle_pipeline", e)
                return

            # peş peşe kayıt için formu temizle
            cb_dan.set("")
            ent_bedel.configure(state="normal")
            ent_bedel.delete(0, END)
            ent_bedel.configure(state="readonly")
            ent_alinan.delete(0, END)
            ent_alinan.insert(0, "0")
            ent_not.delete(0, END)
            cb_dan.focus_set()
            
            # Danışan listesini yenile (yeni eklenen danışanlar görünsün)
            try:
                conn2 = self.veritabani_baglan()
                cur2 = conn2.cursor()
                cur2.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
                danisan_listesi = [row[0] for row in cur2.fetchall()]
                conn2.close()
                cb_dan["values"] = danisan_listesi
            except Exception:
                pass
            
            # Ana sayfadaki combobox'ı da güncelle
            if hasattr(self, 'cmb_danisan'):
                try:
                    self.cmb_danisan["values"] = danisan_listesi
                except Exception:
                    pass
            try:
                if hasattr(self, "kayitlari_listele"):
                    self.kayitlari_listele()
                    self._refresh_borc_tables()
                self._reload_open_toplevels()
            except Exception:
                pass

        btns = ttk.Frame(frm)
        btns.grid(row=7, column=0, columnspan=2, sticky=EW, padx=6, pady=(14, 0))
        ttk.Button(btns, text="KAYDET", bootstyle="success", command=_save).pack(side=LEFT, fill=X, expand=True)
        ttk.Button(btns, text="KAPAT", bootstyle="secondary", command=win.destroy).pack(side=LEFT, padx=8, fill=X, expand=True)

        cb_dan.bind("<Return>", lambda e: ent_bedel.focus_set())
        ent_bedel.bind("<Return>", lambda e: ent_alinan.focus_set())
        ent_alinan.bind("<Return>", lambda e: ent_not.focus_set())
        ent_not.bind("<Return>", lambda e: _save())

    def _haftalik_programda_seans_bul(self, danisan: str, terapist: str, tarih_db: str):
        """Haftalık programdan ilgili gün/personel/danışan seans slotunu döndürür (saat, oda)."""
        try:
            dt = datetime.datetime.strptime(tarih_db, "%Y-%m-%d")
            hafta_bas = (dt - datetime.timedelta(days=dt.weekday())).strftime("%Y-%m-%d")
            gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            gun = gunler[dt.weekday()]
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT saat, COALESCE(oda_adi,'')
                FROM haftalik_seans_programi
                WHERE personel_adi=? AND hafta_baslangic_tarihi=? AND gun=? AND UPPER(TRIM(COALESCE(ogrenci_adi,'')))=UPPER(TRIM(?))
                ORDER BY saat
                LIMIT 1
                """,
                (terapist, hafta_bas, gun, danisan),
            )
            row = cur.fetchone()
            conn.close()
            if row:
                return str(row[0] or "09:00"), str(row[1] or "")
            return None
        except Exception:
            return None

    def kayit_ekle(self):
        """
        PIPELINE ENTEGRASYONU: Seans kaydı ekleme (SEANS TAKİP ANA KAYNAK)
        → seans_takvimi (ANA) → records → kasa_hareketleri → odeme_hareketleri → oda_doluluk
        
        ✅ SADELEŞTİRİLMİŞ: Sadece danışan, terapist, alınan ücret ve not manuel girilir.
        Tarih seçilir; saat otomatik atanır. Seans kaydı haftalık programdan bağımsızdır.
        """
        danisan = (self.cmb_danisan.get() or "").strip()
        terapist = (self.cmb_terapist.get() or "").strip()
        
        if not danisan:
            messagebox.showwarning("Uyarı", "Lütfen danışan adını giriniz!")
            return
        if not terapist:
            messagebox.showwarning("Uyarı", "Lütfen terapist seçiniz!")
            return

        # Seans kaydı haftalık programdan bağımsızdır
        tarih = self._tarih_db()
        saat = self._default_saat()
        oda = ""

        # ✅ Hizmet bedeli kullanıcıdan alınmaz; Ücret Takibi fiyatlandırmasından otomatik gelir
        bedel = 0.0
        try:
            conn_check = self.veritabani_baglan()
            kullanici_id_check = self.kullanici[0] if self.kullanici else None
            pipeline_check = DataPipeline(conn_check, kullanici_id_check)
            bedel = float(pipeline_check.get_price_for_danisan_terapist(danisan, terapist) or 0.0)
            conn_check.close()
        except Exception:
            bedel = 0.0

        # UI'da göster
        try:
            if hasattr(self, "lbl_bedel") and self.lbl_bedel:
                self.lbl_bedel.config(text=format_money(bedel))
        except Exception:
            pass

        if bedel <= 0:
            messagebox.showwarning(
                "Fiyat bulunamadı",
                f"{danisan} için {terapist} hocada fiyatlandırma bulunamadı.\n\n"
                f"Lütfen ÜCRET TAKİBİ ekranından fiyatlandırma güncelleyin.",
            )
            return
        
        # Alınan Ücret: Manuel girilir
        try:
            alinan = parse_money(self.ent_alinan.get())
        except Exception:
            messagebox.showerror("Hata", "Lütfen geçerli bir 'Alınan Ücret' giriniz!")
            return

        # Not: Manuel girilir
        notlar_manuel = (self.ent_not.get() or "").strip()
        
        

        try:
            conn = self.veritabani_baglan()
            kullanici_id = self.kullanici[0] if self.kullanici else None
            
            # Otomatik açıklama oluştur
            tarih_formatted = self.tarih_var.get() or datetime.datetime.now().strftime("%d.%m.%Y")
            kalan_borc = max(0.0, bedel - alinan)
            otomatik_aciklama = f"{danisan} danışanın {terapist} hocadan {tarih_formatted} tarihli seansından {format_money(bedel)} ücret bedelinin {format_money(alinan)} miktarı alınmıştır. {format_money(kalan_borc)} miktar borcu kalmıştır."
            if notlar_manuel:
                notlar = f"{otomatik_aciklama}\n\nNot: {notlar_manuel}"
            else:
                notlar = otomatik_aciklama
            
            pipeline = DataPipeline(conn, kullanici_id)
            seans_id = pipeline.seans_kayit(
                tarih=tarih,
                saat=saat,
                danisan_adi=danisan,
                terapist=terapist,
                hizmet_bedeli=bedel,
                alinan_ucret=alinan,
                notlar=notlar,
                oda=oda,
                check_oda_cakisma=False
            )
            
            # Pipeline log'u konsola yaz (debugging için)
            if seans_id:
                print(f"\n{'='*60}")
                print(f"✅ SEANS KAYIT BAŞARILI (SEANS TAKİP ANA KAYNAK) | seans_id={seans_id}")
                print(f"{'='*60}")
                print(pipeline.get_log())
                print(f"{'='*60}\n")
            
            conn.close()
            
            # ✅ OTOMASYON 5: Seans eklendiğinde otomatik senkronizasyon
            # Tüm ilgili tabloları ve listeleri otomatik güncelle
            try:
                # 1) Kayıt listesini yenile
                self.kayitlari_listele()
                
                # 2) Danışan bakiyesini güncelle (zaten pipeline içinde yapılıyor ama emin olmak için)
                try:
                    conn_sync = self.veritabani_baglan()
                    cur_sync = conn_sync.cursor()
                    cur_sync.execute(
                        "UPDATE danisanlar SET balance = (SELECT COALESCE(SUM(kalan_borc), 0) FROM records WHERE danisan_adi=?) WHERE ad_soyad=?",
                        (danisan, danisan)
                    )
                    conn_sync.commit()
                    conn_sync.close()
                except Exception:
                    pass

                # 2b) Öğrenci Bilgileri tabındaki borç tablosunu yenile
                self._refresh_borc_tables()
                
                # 3) Fiyat politikasını güncelle (gelecek seanslar için otomatik atama)
                if bedel > 0:
                    try:
                        conn_price = self.veritabani_baglan()
                        cur_price = conn_price.cursor()
                        # Danışan ID'sini bul
                        cur_price.execute("SELECT id FROM danisanlar WHERE ad_soyad=?", (danisan,))
                        d_row = cur_price.fetchone()
                        if d_row:
                            danisan_id = d_row[0]
                            # pricing_policy tablosuna ekle/güncelle
                            try:
                                cur_price.execute(
                                    "INSERT OR REPLACE INTO pricing_policy (student_id, teacher_name, price, created_at) VALUES (?, ?, ?, ?)",
                                    (danisan_id, terapist, bedel, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                )
                                conn_price.commit()
                            except Exception as e:
                                # Tablo yoksa veya hata varsa sessizce geç
                                log_exception("pricing_policy_update", e)
                        conn_price.close()
                    except Exception:
                        pass
            except Exception:
                pass
            
            # Başarı mesajı (kısa ve öz)
            if seans_id:
                messagebox.showinfo(
                    "✅ Başarılı", 
                    f"Seans kaydı oluşturuldu!\n\n"
                    f"• Seans ID: #{seans_id}\n"
                    f"• Kasa: {'Eklendi' if alinan > 0 else 'İlk ödeme yok'}\n\n"
                    f"Tüm tablolar otomatik senkronize edildi!"
                )
        
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt ekleme hatası:\n{e}")
            log_exception("kayit_ekle_pipeline", e)
            return

        # Formu temizle (sadeleştirilmiş alanlar)
        self.cmb_danisan.set("")
        self.cmb_terapist.set("")
        self.ent_alinan.delete(0, END)
        self.ent_alinan.insert(0, "0")
        self.ent_not.delete(0, END)
        self._otomatik_bedel = 0.0  # Otomatik bedeli sıfırla
        self.cmb_danisan.focus_set()
        
        # Danışan listesini yenile (yeni eklenen danışanlar görünsün)
        conn = None
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            danisan_listesi = [row[0] for row in cur.fetchall()]
            self.cmb_danisan["values"] = danisan_listesi
        except Exception:
            pass
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass
        self._reload_open_toplevels()
        self.kayitlari_listele()

    def kayitlari_listele(self):
        """
        Seans Takip listesi - SEANS_TAKVIMI ANA KAYNAK
        seans_takvimi tablosundan okuyup records ile JOIN yaparak tüm bilgileri gösterir.
        """
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        q = (self.ent_ara.get() or "").strip().upper()
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            where = []
            params = []
            
            # Role göre filtre (eğitim görevlisi sadece kendi kayıtlarını görür)
            if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                where.append("st.terapist = ?")
                params.append(self.kullanici_terapist)
            if q:
                where.append("st.danisan_adi LIKE ?")
                params.append(f"%{q}%")

            # ✅ SEANS_TAKVIMI ANA KAYNAK - records ile JOIN
            sql = """
                SELECT 
                    st.id AS seans_id,
                    st.tarih,
                    COALESCE(st.saat, '') AS saat,
                    st.danisan_adi,
                    st.terapist,
                    COALESCE(st.hizmet_bedeli, r.hizmet_bedeli, 0) AS hizmet_bedeli,
                    COALESCE(r.alinan_ucret, 0) AS alinan_ucret,
                    COALESCE(r.kalan_borc, 0) AS kalan_borc,
                    COALESCE(st.notlar, r.notlar, '') AS notlar,
                    COALESCE(st.seans_alindi, 0) AS seans_alindi,
                    COALESCE(st.ucret_alindi, 0) AS ucret_alindi,
                    r.id AS record_id
                FROM seans_takvimi st
                LEFT JOIN records r ON st.record_id = r.id OR st.id = r.seans_id
            """
            if where:
                sql += " WHERE " + " AND ".join(where)
            sql += " ORDER BY st.tarih DESC, st.saat DESC, st.id DESC"
            
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Listeleme hatası:\n{e}")
            log_exception("kayitlari_listele", e)
            return

        toplam = 0.0
        for r in rows:
            # r[0]=seans_id, r[1]=tarih, r[2]=saat, r[3]=danisan, r[4]=terapist,
            # r[5]=hizmet_bedeli, r[6]=alinan_ucret, r[7]=kalan_borc, r[8]=notlar, r[9]=seans_alindi, r[10]=ucret_alindi, r[11]=record_id
            borc = float(r[7] or 0)
            toplam += borc
            tag = "borclu" if borc > 0 else "tamam"
            
            # tarih gösterimi
            try:
                dt = datetime.datetime.strptime(r[1], "%Y-%m-%d")
                tarih_g = dt.strftime("%d.%m.%Y")
            except Exception:
                tarih_g = r[1]
            
            # Tree'ye ekle: Tarih, Danışan Adı, Terapist, Alınacak Ücret, Alınan Ücret, Kalan Borç
            self.tree.insert(
                "",
                END,
                values=(
                    r[0],  # seans_id (ANA KAYNAK)
                    tarih_g,
                    r[3],  # danisan_adi
                    r[4],  # terapist
                    format_money(r[5]),  # Alınacak Ücret (hizmet_bedeli)
                    format_money(r[6]),  # Alınan Ücret (alinan_ucret)
                    format_money(r[7]),  # Kalan Borç
                ),
                tags=(tag,),
            )

        self.lbl_ozet.config(text=f"TOPLAM İÇERİDEKİ ALACAK: {toplam:,.2f} TL")

    def odeme_ekle(self):
        seans_id = self._selected_id()  # Artık seans_id (ANA KAYNAK)
        if not seans_id:
            messagebox.showwarning("Uyarı", "Lütfen bir kayıt seçiniz!")
            return
        
        # seans_id'den record_id'yi bul
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT record_id FROM seans_takvimi WHERE id=?", (seans_id,))
            row = cur.fetchone()
            record_id = row[0] if row and row[0] else None
            conn.close()
            
            if not record_id:
                messagebox.showerror("Hata", "Bu seans kaydına bağlı bir record bulunamadı!")
                return
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt bulunamadı:\n{e}")
            return

        win = ttk.Toplevel(self)
        win.title("Ödeme Ekle")
        win.resizable(False, False)
        center_window(win, 440, 280)
        win.transient(self)
        win.grab_set()

        frm = ttk.Frame(win, padding=14)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text="Tarih (YYYY-AA-GG):").grid(row=0, column=0, sticky=W, padx=6, pady=(6, 2))
        ent_tarih = ttk.Entry(frm, width=16)
        ent_tarih.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_tarih.grid(row=0, column=1, sticky=W, padx=6, pady=(6, 2))

        ttk.Label(frm, text="Ödeme Şekli:").grid(row=1, column=0, sticky=W, padx=6, pady=2)
        cb_sekil = ttk.Combobox(frm, state="readonly", values=["Nakit", "Havale/EFT", "Kart", "Diğer"], width=14)
        cb_sekil.current(0)
        cb_sekil.grid(row=1, column=1, sticky=W, padx=6, pady=2)

        ttk.Label(frm, text="Eklenen ödeme (₺):").grid(row=2, column=0, sticky=W, padx=6, pady=(10, 2))
        ent = ttk.Entry(frm, validate="key", validatecommand=self._vcmd_money, width=18)
        ent.grid(row=2, column=1, sticky=W, padx=6, pady=(10, 2))
        ent.focus_set()

        ttk.Label(frm, text="Açıklama (opsiyonel):").grid(row=3, column=0, sticky=W, padx=6, pady=(10, 2))
        ent_aciklama = ttk.Entry(frm, width=32)
        ent_aciklama.grid(row=3, column=1, sticky=W, padx=6, pady=(10, 2))

        def _save():
            """
            PIPELINE ENTEGRASYONU: Ödeme ekleme
            → odeme_hareketleri, records (borç güncelle), kasa_hareketleri, seans_takvimi (ücret_alindi)
            """
            try:
                ek = parse_money(ent.get())
            except Exception:
                messagebox.showerror("Hata", "Lütfen geçerli bir sayı giriniz!")
                return
            if ek <= 0:
                messagebox.showwarning("Uyarı", "Ödeme 0'dan büyük olmalıdır!")
                return

            tahsil_tarih = (ent_tarih.get() or "").strip() or datetime.datetime.now().strftime("%Y-%m-%d")
            odeme_sekli = (cb_sekil.get() or "").strip()
            aciklama = (ent_aciklama.get() or "").strip()

            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                
                # ✅ PIPELINE KULLAN (Tüm tablolar otomatik güncellenecek)
                pipeline = DataPipeline(conn, kullanici_id)
                basarili = pipeline.odeme_ekle(
                    record_id=record_id,
                    tutar=ek,
                    tarih=tahsil_tarih,
                    odeme_sekli=odeme_sekli,
                    aciklama=aciklama,
                )
                
                if basarili:
                    # Pipeline log'u konsola yaz (debugging için)
                    print(f"\n{'='*60}")
                    print(f"💰 ÖDEME EKLEME BAŞARILI | seans_id={seans_id} | record_id={record_id} | +{ek} TL")
                    print(f"{'='*60}")
                    print(pipeline.get_log())
                    print(f"{'='*60}\n")
                    
                    # Kalan borcu kontrol et (Pipeline zaten güncelledi, sadece gösterim için)
                    cur = conn.cursor()
                    cur.execute("SELECT kalan_borc FROM records WHERE id=?", (record_id,))
                    kalan = float((cur.fetchone() or [0])[0] or 0)
                    conn.close()
                    
                    # ✅ OTOMASYON: Ödeme eklendiğinde otomatik senkronizasyon (Windows/macOS)
                    self.kayitlari_listele()
                    self._refresh_borc_tables()
                    
                    if kalan <= 0:
                        messagebox.showinfo(
                            "✅ Başarılı!", 
                            f"Ödeme kaydedildi!\n\n"
                            f"• Eklenen: {ek:,.2f} TL\n"
                            f"• Borç tamamen ödendi!\n\n"
                            f"Tüm tablolar otomatik senkronize edildi!"
                        )
                    else:
                        messagebox.showinfo(
                            "✅ Başarılı!", 
                            f"Ödeme kaydedildi!\n\n"
                            f"• Eklenen: {ek:,.2f} TL\n"
                            f"• Kalan Borç: {kalan:,.2f} TL\n\n"
                            f"Tüm tablolar otomatik senkronize edildi!"
                        )
                else:
                    messagebox.showerror("Hata", "Ödeme eklenirken bir hata oluştu!")
                    return
                    
            except Exception as e:
                messagebox.showerror("Hata", f"Ödeme ekleme hatası:\n{e}")
                log_exception("odeme_ekle_pipeline", e)
                return
            
            win.destroy()
            self.kayitlari_listele()
            self._refresh_borc_tables()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).grid(row=4, column=0, columnspan=2, sticky=EW, padx=6, pady=(16, 0))
        ent.bind("<Return>", lambda e: _save())

    def kayit_sil(self):
        """Seans kayıtlarında tekli/toplu silme modu (pipeline üzerinden atomik)."""
        ids = self._selected_ids()
        if not ids:
            messagebox.showwarning("Uyarı", "Lütfen bir kayıt seçiniz!")
            return
        mod = self._secim_modu_sor(len(ids), "seans kaydı")
        if not mod:
            return
        hedef_ids = ids[:1] if mod == "tek" else ids
        if not messagebox.askyesno("Onay", f"{len(hedef_ids)} kayıt silinsin mi?\n\nİlgili tüm veriler (seans takvimi, records, ödemeler, kasa kayıtları) silinecektir!"):
            return

        conn = None
        try:
            conn = self.veritabani_baglan()
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            silinen = 0
            for seans_id in hedef_ids:
                if pipeline.kayit_sil(seans_id=seans_id):
                    silinen += 1
            if silinen <= 0:
                messagebox.showerror("Hata", "Kayıt silinirken bir hata oluştu!")
                return
            messagebox.showinfo("✅ Başarılı!", f"{silinen} kayıt silindi.")
            self.kayitlari_listele()
            self._refresh_borc_tables()
            self._refresh_open_reports()
        except Exception as e:
            messagebox.showerror("Hata", f"Silme hatası:\n{e}")
            log_exception("kayit_sil_pipeline", e)
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass


    # --- TAB 2: AYARLAR ---
    def _build_ucret_takibi_tab(self):
        """Ücret Takibi Tab - Çocuk ve Personel ücret takipleri (tek ekranda 2 sayfa)"""
        wrapper = ttk.Frame(self.tab_ucret_takibi, padding=10)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(
            wrapper,
            text="Açıklama: Bu modül çocuk ve personel ücret/hakediş kayıtlarını yönetir.",
            font=("Segoe UI", 9),
            foreground="gray",
        ).pack(anchor=W, pady=(0, 6))

        # Başlık
        head = ttk.Frame(wrapper)
        head.pack(fill=X, pady=(0, 10))
        if getattr(self, "_logo_small", None):
            ttk.Label(head, image=self._logo_small).pack(side=LEFT, padx=(0, 10))
        ttk.Label(head, text="ÜCRET TAKİPİ", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(side=LEFT)
        
        # Notebook widget - 2 sayfa
        nb_ucret = ttk.Notebook(wrapper)
        nb_ucret.pack(fill=BOTH, expand=True)
        
        # Sayfa 1: Çocuk Ücret Takibi
        page_cocuk = ttk.Frame(nb_ucret, padding=10)
        nb_ucret.add(page_cocuk, text="👶 Çocuk Ücret Takibi")
        self._build_cocuk_ucret_takibi_page(page_cocuk)
        
        # Sayfa 2: Personel Ücret Takibi
        page_personel = ttk.Frame(nb_ucret, padding=10)
        nb_ucret.add(page_personel, text="👨‍🏫 Personel Ücret Takibi")
        self._build_personel_ucret_takibi_page(page_personel)
    
    def _build_cocuk_ucret_takibi_page(self, parent):
        """Çocuk Ücret Takibi Sayfası"""
        # Üst toolbar
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=X, pady=(0, 10))
        
        ttk.Label(toolbar, text="Çocuk Ücret Takibi", font=("Segoe UI", 12, "bold")).pack(side=LEFT)
        
        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=10)
        
        ttk.Button(toolbar, text="📊 Excel'den Ücret Listesi Yükle", bootstyle="info",
                   command=self.excel_ucret_listesi_yukle).pack(side=LEFT, padx=5)
        ttk.Button(toolbar, text="🔄 Yenile", bootstyle="secondary",
                   command=lambda: self._cocuk_ucret_listele(parent)).pack(side=LEFT, padx=5)
        
        # Arama
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill=X, pady=(0, 10))
        ttk.Label(search_frame, text="Ara:").pack(side=LEFT, padx=(0, 5))
        ent_ara_cocuk = ttk.Entry(search_frame, width=30)
        ent_ara_cocuk.pack(side=LEFT, padx=5)
        ent_ara_cocuk.bind("<KeyRelease>", lambda e: self._cocuk_ucret_listele(parent))
        ttk.Button(search_frame, text="Ara", bootstyle="primary",
                   command=lambda: self._cocuk_ucret_listele(parent)).pack(side=LEFT, padx=5)
        
        # Treeview
        frame_tree = ttk.Frame(parent)
        frame_tree.pack(fill=BOTH, expand=True)
        
        cols = ("ID", "Çocuk Adı", "Personel", "Tarih", "Seans Ücreti", "Alınan Ödeme", "Kalan Borç", "Durum")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            if c == "ID":
                tree.column(c, width=50)
            elif c == "Çocuk Adı":
                tree.column(c, width=200)
            elif c == "Personel":
                tree.column(c, width=150)
            elif c == "Tarih":
                tree.column(c, width=100)
            elif c in ("Seans Ücreti", "Alınan Ödeme", "Kalan Borç"):
                tree.column(c, width=120, anchor="e")
            else:
                tree.column(c, width=100)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frame_tree, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        # Tag'ler
        tree.tag_configure("borclu", background="#f8d7da", foreground="#721c24")
        tree.tag_configure("tamam", background="#d4edda", foreground="#155724")
        tree.tag_configure("even", background="#f8f9fa")
        tree.tag_configure("odd", background="#ffffff")
        
        action_row = ttk.Frame(parent)
        action_row.pack(fill=X, pady=(8, 4))
        ttk.Label(action_row, text="Seçili satır işlemleri:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 8))
        ttk.Button(action_row, text="💰 Fiyatlandırma Güncelle", bootstyle="warning-outline", command=lambda: self._fiyatlandirma_guncelle(parent, tree)).pack(side=LEFT, padx=4)
        ttk.Button(action_row, text="📊 Detaylı Rapor", bootstyle="info-outline", command=lambda: self._cocuk_detayli_rapor(parent, tree)).pack(side=LEFT, padx=4)
        
        # Özet bilgiler
        summary_frame = ttk.Labelframe(parent, text="Özet", padding=10, bootstyle="secondary")
        summary_frame.pack(fill=X, pady=(10, 0))
        
        summary_labels = ttk.Frame(summary_frame)
        summary_labels.pack(fill=X)
        
        # Treeview'i parent'a kaydet
        parent._tree_cocuk = tree
        parent._ent_ara_cocuk = ent_ara_cocuk
        parent._summary_labels_cocuk = summary_labels
        
        # İlk yükleme
        self._cocuk_ucret_listele(parent)
    
    def _build_personel_ucret_takibi_page(self, parent):
        """Personel Ücret Takibi Sayfası"""
        # Üst toolbar
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=X, pady=(0, 10))
        
        ttk.Label(toolbar, text="Personel Ücret Takibi", font=("Segoe UI", 12, "bold")).pack(side=LEFT)
        # Bu satırı diğer butonların yanına yapıştır
        ttk.Button(toolbar, text="💸 Avans/Ödeme Ver", bootstyle="danger", command=self.popup_personel_avans).pack(side=LEFT, padx=5)
        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=10)
        
        ttk.Button(toolbar, text="💰 Ödeme Yap", bootstyle="warning",
                   command=self.personel_ucret_odeme_yap).pack(side=LEFT, padx=5)
        ttk.Button(toolbar, text="🔄 Yenile", bootstyle="secondary",
                   command=lambda: (self._personel_ucret_listele(parent), self._refresh_kasa_views())).pack(side=LEFT, padx=5)
        
        # Filtreler
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(filter_frame, text="Personel:").pack(side=LEFT, padx=(0, 5))
        cmb_personel = ttk.Combobox(filter_frame, state="readonly", width=25)
        cmb_personel.pack(side=LEFT, padx=5)
        
        # Personel listesi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            personel_listesi = [""] + [row[0] for row in cur.fetchall()]
            conn.close()
            cmb_personel["values"] = personel_listesi
            cmb_personel.current(0)
        except Exception:
            cmb_personel["values"] = [""]
        
        ttk.Label(filter_frame, text="Ödeme Durumu:").pack(side=LEFT, padx=(20, 5))
        cmb_durum = ttk.Combobox(filter_frame, state="readonly", width=20)
        cmb_durum["values"] = ["", "Beklemede", "Ödendi", "İptal"]
        cmb_durum.current(0)
        cmb_durum.pack(side=LEFT, padx=5)
        
        ttk.Button(filter_frame, text="Filtrele", bootstyle="primary",
                   command=lambda: self._personel_ucret_listele(parent)).pack(side=LEFT, padx=10)
        
        # Treeview
        frame_tree = ttk.Frame(parent)
        frame_tree.pack(fill=BOTH, expand=True)
        
        cols = ("ID", "Personel", "Tarih", "Seans Ücreti", "Personel Ücreti", "Oran", "Ödeme Durumu", "Ödeme Tarihi")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            if c == "ID":
                tree.column(c, width=50)
            elif c == "Personel":
                tree.column(c, width=150)
            elif c == "Tarih":
                tree.column(c, width=100)
            elif c in ("Seans Ücreti", "Personel Ücreti"):
                tree.column(c, width=130, anchor="e")
            elif c == "Oran":
                tree.column(c, width=80, anchor="e")
            else:
                tree.column(c, width=120)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frame_tree, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        # Tag'ler
        tree.tag_configure("beklemede", background="#fff3cd", foreground="#856404")
        tree.tag_configure("odendi", background="#d4edda", foreground="#155724")
        tree.tag_configure("iptal", background="#f8d7da", foreground="#721c24")
        tree.tag_configure("even", background="#f8f9fa")
        tree.tag_configure("odd", background="#ffffff")
        
        # Özet bilgiler
        summary_frame = ttk.Labelframe(parent, text="Özet", padding=10, bootstyle="secondary")
        summary_frame.pack(fill=X, pady=(10, 0))
        
        summary_labels = ttk.Frame(summary_frame)
        summary_labels.pack(fill=X)
        
        parent._tree_personel = tree
        parent._cmb_personel = cmb_personel
        parent._cmb_durum = cmb_durum
        parent._summary_labels = summary_labels
        
        # İlk yükleme
        self._personel_ucret_listele(parent)
    
    def _cocuk_ucret_listele(self, parent):
        """
        SADELEŞTİRİLMİŞ VERSİYON:
        Bu ekranda SADECE 'Fiyat Listesi' (Excel'den yüklenenler) görünür.
        Günlük seans kayıtları buraya gelmez (onlar Seans Takip'te kalır).
        """
        tree = parent._tree_cocuk
        ent_ara = parent._ent_ara_cocuk
        summary_labels = parent._summary_labels_cocuk
        
        for iid in tree.get_children():
            tree.delete(iid)
        
        # Özet etiketlerini temizle
        for widget in summary_labels.winfo_children():
            widget.destroy()
        
        q = (ent_ara.get() or "").strip().upper()
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            where_clauses = ["opf.aktif = 1"]
            params = []
            
            if q:
                where_clauses.append("(UPPER(d.ad_soyad) LIKE ? OR UPPER(opf.personel_adi) LIKE ?)")
                params.extend([f"%{q}%", f"%{q}%"])
            
            # Role göre filtre
            if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                where_clauses.append("opf.personel_adi = ?")
                params.append(self.kullanici_terapist)
            
            where_sql = " AND ".join(where_clauses)
            
            # SADECE FİYAT LİSTESİNİ ÇEKİYORUZ (Seans tablosunu karıştırmıyoruz)
            sql = f"""
                SELECT 
                    opf.id,
                    d.ad_soyad,
                    opf.personel_adi,
                    opf.baslangic_tarihi,
                    opf.seans_ucreti,
                    0, -- Alınan (Burada gösterilmiyor)
                    0, -- Kalan (Burada gösterilmiyor)
                    'Tanımlı Ücret' -- Durum
                FROM ogrenci_personel_fiyatlandirma opf
                LEFT JOIN danisanlar d ON opf.ogrenci_id = d.id
                WHERE {where_sql}
                ORDER BY d.ad_soyad, opf.personel_adi
            """
            
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
            conn.close()
            
            total_fee_sum = 0
            
            for idx, row in enumerate(rows):
                opf_id, ad, per, tarih, ucret, _, _, durum = row
                
                total_fee_sum += float(ucret or 0)
                tag = "even" if idx % 2 == 0 else "odd"
                
                # Tabloya Ekle
                # Not: Alınan/Kalan sütunları bu modda boş gelecektir çünkü burası sadece Tarife Listesi
                tree.insert("", END, values=(
                    opf_id,
                    ad,
                    per,
                    tarih, # Başlangıç tarihi
                    format_money(ucret),
                    "-", # Alınan yok
                    "-", # Kalan yok
                    durum
                ), tags=(tag,))
            
            # Özet Bilgi (Sadece Toplam Tanımlı Ücret)
            ttk.Label(summary_labels, text=f"📋 Toplam Kayıt: {len(rows)}", font=("Segoe UI", 10)).pack(side=LEFT, padx=15)
            # ttk.Label(summary_labels, text=f"💰 Ortalama Tarife: {format_money(total_fee_sum/len(rows) if len(rows)>0 else 0)}", font=("Segoe UI", 10, "bold"), bootstyle="info").pack(side=LEFT, padx=15)
        
        except Exception as e:
            messagebox.showerror("Hata", f"Liste yüklenemedi:\n{e}")
            log_exception("_cocuk_ucret_listele", e)
    
    def _fiyatlandirma_guncelle(self, parent, tree):
        """Öğrenci-personel bazlı fiyatlandırma güncelle"""
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir kayıt seçin.")
            return
        
        values = tree.item(sel[0])["values"]
        seans_id = values[0]
        cocuk_adi = values[1]
        personel_adi = values[2]
        
        # Öğrenci ID'sini bul
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT id FROM danisanlar WHERE UPPER(ad_soyad) = UPPER(?) AND aktif = 1 LIMIT 1", (cocuk_adi,))
            row = cur.fetchone()
            if not row:
                messagebox.showerror("Hata", "Öğrenci bulunamadı.")
                conn.close()
                return
            ogrenci_id = row[0]
            
            # Mevcut fiyatlandırmayı kontrol et
            cur.execute(
                """
                SELECT seans_ucreti, zam_orani, baslangic_tarihi, bitis_tarihi
                FROM ogrenci_personel_fiyatlandirma
                WHERE ogrenci_id = ? AND personel_adi = ? AND aktif = 1
                ORDER BY baslangic_tarihi DESC
                LIMIT 1
                """,
                (ogrenci_id, personel_adi)
            )
            mevcut_fiyat = cur.fetchone()
            conn.close()
            
            # Fiyatlandırma güncelleme penceresi
            win = ttk.Toplevel(self)
            win.title("Fiyatlandırma Güncelle")
            win.geometry("500x400")
            center_window(win, 500, 400)
            win.transient(self)
            self._brand_window(win)
            
            wrapper = ttk.Frame(win, padding=20)
            wrapper.pack(fill=BOTH, expand=True)
            
            ttk.Label(wrapper, text="Fiyatlandırma Güncelle", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
            
            ttk.Label(wrapper, text=f"Öğrenci: {cocuk_adi}", font=("Segoe UI", 11)).pack(pady=5)
            ttk.Label(wrapper, text=f"Personel: {personel_adi}", font=("Segoe UI", 11)).pack(pady=5)
            
            ttk.Separator(wrapper).pack(fill=X, pady=20)
            
            ttk.Label(wrapper, text="Seans Ücreti (TL):").pack(anchor=W, pady=(10, 5))
            ent_ucret = ttk.Entry(wrapper, width=30)
            if mevcut_fiyat and mevcut_fiyat[0]:
                ent_ucret.insert(0, str(mevcut_fiyat[0]))
            ent_ucret.pack(fill=X, pady=5)
            
            ttk.Label(wrapper, text="Yıllık Zam Oranı (%):").pack(anchor=W, pady=(10, 5))
            ent_zam = ttk.Entry(wrapper, width=30)
            if mevcut_fiyat and mevcut_fiyat[1]:
                ent_zam.insert(0, str(mevcut_fiyat[1]))
            else:
                ent_zam.insert(0, "0")
            ent_zam.pack(fill=X, pady=5)
            
            ttk.Label(wrapper, text="Başlangıç Tarihi (YYYY-MM-DD):").pack(anchor=W, pady=(10, 5))
            ent_baslangic = ttk.Entry(wrapper, width=30)
            ent_baslangic.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))
            ent_baslangic.pack(fill=X, pady=5)
            
            ttk.Label(wrapper, text="Bitiş Tarihi (YYYY-MM-DD, boş bırakılabilir):").pack(anchor=W, pady=(10, 5))
            ent_bitis = ttk.Entry(wrapper, width=30)
            if mevcut_fiyat and mevcut_fiyat[3]:
                ent_bitis.insert(0, str(mevcut_fiyat[3]))
            ent_bitis.pack(fill=X, pady=5)
            
            def kaydet():
                try:
                    ucret = float(ent_ucret.get() or 0)
                    zam_orani = float(ent_zam.get() or 0)
                    baslangic = ent_baslangic.get().strip()
                    bitis = ent_bitis.get().strip() or None
                    
                    if ucret <= 0:
                        messagebox.showwarning("Uyarı", "Seans ücreti 0'dan büyük olmalıdır.")
                        return
                    
                    if not baslangic:
                        messagebox.showwarning("Uyarı", "Başlangıç tarihi gerekli.")
                        return
                    
                    conn = self.veritabani_baglan()
                    cur = conn.cursor()
                    
                    # Eski kaydı pasif yap
                    if mevcut_fiyat:
                        cur.execute(
                            "UPDATE ogrenci_personel_fiyatlandirma SET aktif = 0 WHERE ogrenci_id = ? AND personel_adi = ? AND aktif = 1",
                            (ogrenci_id, personel_adi)
                        )
                    
                    # Yeni kayıt ekle
                    cur.execute(
                        """
                        INSERT INTO ogrenci_personel_fiyatlandirma
                        (ogrenci_id, personel_adi, seans_ucreti, baslangic_tarihi, bitis_tarihi, zam_orani, aktif, olusturma_tarihi, guncelleme_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?)
                        """,
                        (
                            ogrenci_id,
                            personel_adi,
                            ucret,
                            baslangic,
                            bitis,
                            zam_orani,
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        )
                    )
                    
                    # ✅ DÜZELTME: pricing_policy tablosuna da yansıt (DataPipeline mantığına uygun)
                    # Öğrenci adını bul
                    cur.execute("SELECT ad_soyad FROM danisanlar WHERE id=?", (ogrenci_id,))
                    ogrenci_adi_row = cur.fetchone()
                    ogrenci_adi = ogrenci_adi_row[0] if ogrenci_adi_row else cocuk_adi
                    
                    # pricing_policy tablosuna INSERT OR REPLACE (gelecek seanslar için)
                    cur.execute(
                        """
                        INSERT OR REPLACE INTO pricing_policy 
                        (student_id, teacher_name, price, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            ogrenci_id,
                            personel_adi,
                            ucret,
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        )
                    )
                    
                    # ✅ DÜZELTME: Mevcut gelecek seansların hizmet_bedeli'ni güncelle (sadece planlanmış olanlar)
                    # Bugünden sonraki ve henüz gerçekleşmemiş seanslar için
                    cur.execute(
                        """
                        UPDATE seans_takvimi 
                        SET hizmet_bedeli = ?
                        WHERE danisan_adi = ? AND terapist = ? 
                        AND tarih >= ? 
                        AND (durum = 'planlandi' OR durum IS NULL OR durum = '')
                        """,
                        (ucret, ogrenci_adi, personel_adi, baslangic)
                    )
                    
                    conn.commit()
                    conn.close()
                    
                    messagebox.showinfo("Başarılı", f"Fiyatlandırma güncellendi!\n\nÖğrenci: {cocuk_adi}\nPersonel: {personel_adi}\nÜcret: {format_money(ucret)}\n\nGelecek seansların ücreti otomatik güncellendi.")
                    win.destroy()
                    # ✅ OTOMATIK YENİLENME: Tabloları otomatik yenile (Windows/macOS)
                    self._cocuk_ucret_listele(parent)
                    if hasattr(self, 'kayitlari_listele'):
                        self.kayitlari_listele()
                    self._refresh_borc_tables()
                
                except Exception as e:
                    messagebox.showerror("Hata", f"Fiyatlandırma güncellenemedi:\n{e}")
                    log_exception("_fiyatlandirma_guncelle", e)
            
            ttk.Button(wrapper, text="💾 Kaydet", bootstyle="success", command=kaydet).pack(pady=20)
        
        except Exception as e:
            messagebox.showerror("Hata", f"Fiyatlandırma güncellenemedi:\n{e}")
            log_exception("_fiyatlandirma_guncelle", e)
    
    def _cocuk_detayli_rapor(self, parent, tree):
        """Çocuk detaylı rapor penceresi"""
        sel = tree.selection()
        if not sel:
            return
        
        values = tree.item(sel[0])["values"]
        cocuk_adi = values[1]
        personel_adi = values[2]
        
        win = ttk.Toplevel(self)
        win.title(f"Detaylı Rapor - {cocuk_adi}")
        win.geometry("800x600")
        center_window(win, 800, 600)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text=f"Detaylı Rapor: {cocuk_adi}", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        # Rapor içeriği buraya eklenecek
        # ✅ DÜZELTME: Detaylı rapor özelliği - Basit bir bilgi mesajı
        ttk.Label(wrapper, text="Detaylı rapor özelliği için lütfen 'Raporlar' sekmesini kullanın.", font=("Segoe UI", 10)).pack(pady=20)
    
    def _personel_ucret_listele(self, parent):
        """Personel ücret takibi listesini yükle"""
        tree = parent._tree_personel
        cmb_personel = parent._cmb_personel
        cmb_durum = parent._cmb_durum
        summary_labels = parent._summary_labels
        
        for iid in tree.get_children():
            tree.delete(iid)
        
        # Özet etiketlerini temizle
        for widget in summary_labels.winfo_children():
            widget.destroy()
        
        personel_filtre = cmb_personel.get() or ""
        durum_filtre = cmb_durum.get() or ""
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            has_put_seans = self._table_has_column(conn, "personel_ucret_takibi", "seans_id")
            where = []
            params = []

            if has_put_seans:
                # seans silinmişse bağlı personel ücret kaydını listede göstermeyelim
                where.append("(put.seans_id IS NULL OR st.id IS NOT NULL)")

            if personel_filtre:
                where.append("put.personel_adi = ?")
                params.append(personel_filtre)

            if durum_filtre:
                where.append("put.odeme_durumu = ?")
                params.append(durum_filtre.lower())

            sql = """
                SELECT 
                    put.id,
                    put.personel_adi,
                    put.tarih,
                    put.seans_ucreti,
                    put.personel_ucreti,
                    put.ucret_orani,
                    put.odeme_durumu,
                    put.odeme_tarihi
                FROM personel_ucret_takibi put
            """
            if has_put_seans:
                sql += " LEFT JOIN seans_takvimi st ON st.id = put.seans_id"
            if where:
                sql += " WHERE " + " AND ".join(where)
            sql += " ORDER BY put.tarih DESC, put.id DESC"
            
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
            
            # Özet hesapla
            toplam_beklemede = 0.0
            toplam_odendi = 0.0
            toplam_seans = len(rows)
            
            for idx, row in enumerate(rows):
                ucret_id, personel, tarih, seans_ucreti, personel_ucreti, ucret_orani, odeme_durumu, odeme_tarihi = row
                
                if odeme_durumu == "beklemede":
                    toplam_beklemede += float(personel_ucreti or 0)
                elif odeme_durumu == "odendi":
                    toplam_odendi += float(personel_ucreti or 0)
                
                durum_text = odeme_durumu.capitalize() if odeme_durumu else "Beklemede"
                oran_text = f"%{ucret_orani:.0f}" if ucret_orani > 0 else "Sabit"
                
                tag = odeme_durumu if odeme_durumu else "beklemede"
                if idx % 2 == 0:
                    tag = "even"
                else:
                    tag = "odd"
                
                tree.insert("", END, values=(
                    ucret_id,
                    personel,
                    tarih,
                    format_money(seans_ucreti),
                    format_money(personel_ucreti),
                    oran_text,
                    durum_text,
                    odeme_tarihi or "-"
                ), tags=(tag,))
            
            conn.close()
            
            # Özet göster
            ttk.Label(summary_labels, text=f"📊 Toplam Seans: {toplam_seans}", font=("Segoe UI", 10)).pack(side=LEFT, padx=10)
            ttk.Label(summary_labels, text=f"⏳ Beklemede: {format_money(toplam_beklemede)}", 
                     font=("Segoe UI", 10), bootstyle="warning").pack(side=LEFT, padx=10)
            ttk.Label(summary_labels, text=f"✅ Ödendi: {format_money(toplam_odendi)}", 
                     font=("Segoe UI", 10), bootstyle="success").pack(side=LEFT, padx=10)
        
        except Exception as e:
            messagebox.showerror("Hata", f"Personel ücret listesi yüklenemedi:\n{e}")
            log_exception("_personel_ucret_listele", e)
    
    def excel_ucret_listesi_yukle(self):
        """
        Kullanıcıdan 2026 Seans Ücretleri Excel dosyasını ister 
        ve _import_2026_fees_from_excel fonksiyonunu çağırır.
        """
        path = filedialog.askopenfilename(
            title="2026 Ücret Listesi Excel Dosyasını Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
        )
        
        if not path:
            return

        # Kullanıcıya onay sor
        if not messagebox.askyesno("Onay", f"Seçilen dosya: {os.path.basename(path)}\n\nBu dosyadaki fiyatlandırmalar sisteme aktarılacak.\n('Yeni Değerlendirme' kayıtları atlanacaktır.)\n\nDevam edilsin mi?"):
            return

        # Hata veren karmaşık yapıyı kaldırıp direkt doğru fonksiyonu çağırıyoruz
        try:
            # V4 Import fonksiyonunu çağır (path değişkeni ile)
            self._import_2026_fees_from_excel(path)
            
        except Exception as e:
            messagebox.showerror("Hata", f"İşlem başlatılamadı:\n{e}")
            log_exception("excel_ucret_listesi_yukle", e)

            
    def _import_2026_fees_from_excel(self, excel_path: str):
        """
        GELİŞMİŞ IMPORT V6 (Final):
        1) Sadece sistemde kayıtlı olan danışanları işler (YENİ DANIŞAN OLUŞTURMAZ).
        2) Fiyatları 'Fiyat Listesi' olarak kaydeder.
        3) Gereksiz verileri filtreler.
        """
        import pandas as pd
        import numpy as np
        
        conn = None
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            xls = pd.ExcelFile(excel_path, engine='openpyxl')
            
            toplam_eklenen = 0
            atlanan_danisan = 0
            bulunan_hocalar = set()
            
            # Filtre Listesi
            BLACKLIST = [
                "yeni değerlendirme", "yenı degerlendırme",
                "dakika", "dakıka", "jimnastik", "cimnastik",
                "aas temiz", "toplam", "genel toplam"
            ]
            
            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                    
                    # Başlıkları Bul
                    baslik_koordinatlari = []
                    for r in range(min(30, len(df))):
                        for c in range(len(df.columns)):
                            val = str(df.iloc[r, c]).strip()
                            if "Danışan" in val or "Öğrenci" in val:
                                baslik_koordinatlari.append((r, c))
                    
                    if not baslik_koordinatlari:
                        print(f"Bilgi: '{sheet_name}' sayfasında başlık bulunamadı.")
                        continue

                    for r_baslik, c_danisan in baslik_koordinatlari:
                        # Hoca İsmini Bul
                        hoca_adi = None
                        if r_baslik > 0:
                            val = str(df.iloc[r_baslik-1, c_danisan]).strip()
                            if val and val.lower() != 'nan': hoca_adi = val
                        
                        if not hoca_adi:
                            hoca_adi = sheet_name.replace(".csv", "").replace(".xlsx", "").strip()

                        hoca_adi = hoca_adi.replace("Ödemesi", "").replace("Listesi", "").strip()
                        bulunan_hocalar.add(hoca_adi)

                        # Tutar Sütununu Bul
                        c_tutar = c_danisan + 1
                        for check_col in range(c_danisan + 1, min(c_danisan + 6, len(df.columns))):
                            header_val = str(df.iloc[r_baslik, check_col]).strip()
                            if "Tutar" in header_val or "Ücret" in header_val:
                                c_tutar = check_col
                                break

                        # Satırları Oku
                        for r in range(r_baslik + 1, len(df)):
                            try:
                                danisan = str(df.iloc[r, c_danisan]).strip()
                                
                                # Filtreler
                                if not danisan or danisan.lower() in ['nan', 'none', 'toplam']: continue
                                if "Danışan" in danisan or "Hoca" in danisan: continue
                                if any(x in danisan.lower() for x in BLACKLIST): continue

                                # Tutar Okuma
                                tutar_raw = df.iloc[r, c_tutar]
                                try:
                                    t_str = str(tutar_raw).replace('₺', '').replace('TL', '').strip()
                                    if "," in t_str and "." in t_str: t_str = t_str.replace('.', '').replace(',', '.')
                                    elif "," in t_str: t_str = t_str.replace(',', '.')
                                    tutar = float(t_str)
                                except Exception:
                                    continue

                                if tutar > 0:
                                    # 1. KONTROL: Danışan sistemde var mı? (Türkçe karakter uyumlu)
                                    # Mac ve Windows arasındaki İ/I sorununu çözer
                                    danisan_clean = danisan.strip()
                                    cur.execute("""
                                        SELECT id FROM danisanlar 
                                        WHERE UPPER(REPLACE(REPLACE(ad_soyad, 'İ', 'I'), 'ı', 'I')) = UPPER(REPLACE(REPLACE(?, 'İ', 'I'), 'ı', 'I')) 
                                        AND aktif = 1 LIMIT 1
                                    """, (danisan_clean,))
                                    
                                    d_row = cur.fetchone()

                                    # Eğer bulunamazsa otomatik ekleme mantığı buraya gelecek...
                                    if not d_row:
                                        try:
                                            cur.execute(
                                                "INSERT INTO danisanlar (ad_soyad, aktif, olusturma_tarihi) VALUES (?, 1, ?)",
                                                (danisan_clean, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                            )
                                            danisan_id = cur.lastrowid
                                            toplam_eklenen += 1 # Sayaç (isteğe bağlı)
                                        except Exception:
                                            atlanan_danisan += 1
                                            continue
                                    else:
                                        danisan_id = d_row[0]
                                    
                                    # 2. Fiyat Politikasını Kaydet (Otomatik dolum için)
                                    cur.execute("""
                                        INSERT OR REPLACE INTO pricing_policy 
                                        (student_id, teacher_name, price, created_at, updated_at) 
                                        VALUES (?, ?, ?, ?, ?)
                                    """, (
                                        danisan_id, hoca_adi, tutar, 
                                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                    ))
                                    
                                    # 3. Listede Gözükmesi İçin Kayıt
                                    # ÇAKIŞANLARI OVERWRITE ET: aynı öğrenci/personel için tek aktif kayıt bırak.
                                    now_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                    cur.execute(
                                        "SELECT id FROM ogrenci_personel_fiyatlandirma WHERE ogrenci_id=? AND personel_adi=? ORDER BY id DESC",
                                        (danisan_id, hoca_adi),
                                    )
                                    mevcutlar = [int(r[0]) for r in (cur.fetchall() or []) if r and r[0] is not None]
                                    if mevcutlar:
                                        korunacak = mevcutlar[0]
                                        cur.execute(
                                            """
                                            UPDATE ogrenci_personel_fiyatlandirma
                                            SET seans_ucreti=?, baslangic_tarihi=?, aktif=1, zam_orani=0
                                            WHERE id=?
                                            """,
                                            (tutar, "2026-01-01", korunacak),
                                        )
                                        if len(mevcutlar) > 1:
                                            cur.execute(
                                                f"UPDATE ogrenci_personel_fiyatlandirma SET aktif=0 WHERE id IN ({','.join('?' for _ in mevcutlar[1:])})",
                                                tuple(mevcutlar[1:]),
                                            )
                                    else:
                                        cur.execute("""
                                            INSERT INTO ogrenci_personel_fiyatlandirma
                                            (ogrenci_id, personel_adi, seans_ucreti, baslangic_tarihi, aktif, zam_orani, olusturma_tarihi)
                                            VALUES (?, ?, ?, ?, 1, 0, ?)
                                        """, (
                                            danisan_id, hoca_adi, tutar, "2026-01-01", now_ts
                                        ))

                                    toplam_eklenen += 1

                            except Exception: continue

                except Exception: continue

            conn.commit()
            
            msg = f"İşlem Tamamlandı!\n\nToplam Eşleşen ve Yüklenen: {toplam_eklenen}\n"
            msg += f"Sistemde Bulunamadığı İçin Atlanan: {atlanan_danisan}\n"
            msg += "(Önce Danışan Template'ini yüklediğinizden emin olun)"
            
            messagebox.showinfo("Rapor", msg)
            
            # Ekranı Yenile
            try:
                if hasattr(self, 'tab_ucret_takibi'):
                     for child in self.tab_ucret_takibi.winfo_children():
                        for sub in child.winfo_children():
                             if isinstance(sub, ttk.Notebook):
                                 try:
                                     self._cocuk_ucret_listele(sub.nametowidget(sub.tabs()[0]))
                                 except Exception:
                                     pass
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Hata", f"İçe aktarma sırasında hata:\n{e}")
            log_exception("_import_2026_fees_from_excel", e)
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass
    
    def _normalize_personel_adi(self, personel: str) -> str:
        """Personel adını normalize et"""
        if not personel:
            return ""
        
        personel = str(personel).strip()
        
        # Personel adı eşleştirmeleri
        personel_map = {
            "pervin hoca": "Pervin Hoca",
            "pervin": "Pervin Hoca",
            "pervin hoca (değ.)": "Pervin Hoca",
            "arif hoca": "Arif Hoca",
            "arif": "Arif Hoca",
            "çağlar hoca": "Çağlar Hoca",
            "çağlar": "Çağlar Hoca",
            "çağlar hoca (değ.)": "Çağlar Hoca",
            "sena hoca": "Sena Hoca",
            "sena": "Sena Hoca",
            "aybüke hoca": "Aybüke Hoca",
            "aybüke": "Aybüke Hoca",
            "aybüke hoca (değ.)": "Aybüke Hoca",
            "elif hoca": "Elif Hoca",
            "elif": "Elif Hoca",
        }
        
        personel_lower = personel.lower()
        if personel_lower in personel_map:
            return personel_map[personel_lower]
        
        # İlk harfi büyük yap
        return personel.title()
    
    def personel_ucret_talep_formu(self):
        """Personel ücret talep formu penceresi"""
        win = ttk.Toplevel(self)
        win.title("Personel Ücret Talep Formu")
        win.transient(self)
        self._brand_window(win)
        center_window_smart(win, 700, 600, min_w=650, min_h=550)
        maximize_window(win)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text="Personel Ücret Talep Formu", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        # Personel seçimi
        frm_personel = ttk.Frame(wrapper)
        frm_personel.pack(fill=X, pady=10)
        ttk.Label(frm_personel, text="Personel:", width=15).pack(side=LEFT)
        cmb_personel = ttk.Combobox(frm_personel, state="readonly", width=30)
        cmb_personel.pack(side=LEFT, padx=5)
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            personel_listesi = [row[0] for row in cur.fetchall()]
            conn.close()
            cmb_personel["values"] = personel_listesi
            if personel_listesi:
                cmb_personel.current(0)
        except Exception:
            cmb_personel["values"] = []
        
        # Tarih aralığı
        frm_tarih = ttk.Frame(wrapper)
        frm_tarih.pack(fill=X, pady=10)
        ttk.Label(frm_tarih, text="Başlangıç Tarihi:", width=15).pack(side=LEFT)
        ent_baslangic = ttk.Entry(frm_tarih, width=30)
        ent_baslangic.pack(side=LEFT, padx=5)
        ent_baslangic.insert(0, datetime.datetime.now().strftime("%Y-%m-01"))  # Ayın ilk günü
        
        frm_tarih2 = ttk.Frame(wrapper)
        frm_tarih2.pack(fill=X, pady=10)
        ttk.Label(frm_tarih2, text="Bitiş Tarihi:", width=15).pack(side=LEFT)
        ent_bitis = ttk.Entry(frm_tarih2, width=30)
        ent_bitis.pack(side=LEFT, padx=5)
        ent_bitis.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))  # Bugün
        
        # Özet bilgiler
        frm_ozet = ttk.Labelframe(wrapper, text="Özet Bilgiler", padding=10, bootstyle="secondary")
        frm_ozet.pack(fill=X, pady=10)
        
        lbl_ozet = ttk.Label(frm_ozet, text="Personel ve tarih seçtikten sonra özet bilgiler burada görünecek.", 
                            font=("Segoe UI", 9), foreground="gray")
        lbl_ozet.pack()
        
        def hesapla_ozet():
            personel = cmb_personel.get()
            baslangic = ent_baslangic.get()
            bitis = ent_bitis.get()
            
            if not personel or not baslangic or not bitis:
                lbl_ozet.config(text="Lütfen tüm alanları doldurun.")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                
                cur.execute(
                    """
                    SELECT 
                        COUNT(*) AS seans_sayisi,
                        SUM(personel_ucreti) AS toplam_ucret
                    FROM personel_ucret_takibi
                    WHERE personel_adi = ? AND tarih >= ? AND tarih <= ? AND odeme_durumu = 'beklemede'
                    """,
                    (personel, baslangic, bitis)
                )
                row = cur.fetchone()
                seans_sayisi = row[0] or 0
                toplam_ucret = float(row[1] or 0)
                
                conn.close()
                
                lbl_ozet.config(
                    text=f"📊 Seans Sayısı: {seans_sayisi}\n"
                         f"💰 Toplam Ücret: {format_money(toplam_ucret)}",
                    foreground="black"
                )
            except Exception as e:
                lbl_ozet.config(text=f"Hata: {e}", foreground="red")
        
        ttk.Button(wrapper, text="📊 Özet Hesapla", bootstyle="info", command=hesapla_ozet).pack(pady=10)
        
        def kaydet():
            personel = cmb_personel.get()
            baslangic = ent_baslangic.get()
            bitis = ent_bitis.get()
            
            if not personel or not baslangic or not bitis:
                messagebox.showerror("Hata", "Lütfen tüm alanları doldurun.")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                
                # Özet hesapla
                cur.execute(
                    """
                    SELECT 
                        COUNT(*) AS seans_sayisi,
                        SUM(personel_ucreti) AS toplam_ucret
                    FROM personel_ucret_takibi
                    WHERE personel_adi = ? AND tarih >= ? AND tarih <= ? AND odeme_durumu = 'beklemede'
                    """,
                    (personel, baslangic, bitis)
                )
                row = cur.fetchone()
                seans_sayisi = row[0] or 0
                toplam_ucret = float(row[1] or 0)
                
                if seans_sayisi == 0:
                    messagebox.showwarning("Uyarı", "Seçilen tarih aralığında beklemede ücret bulunamadı.")
                    conn.close()
                    return
                
                # Talep kaydı oluştur
                cur.execute(
                    """
                    INSERT INTO personel_ucret_talepleri
                    (personel_adi, talep_tarihi, baslangic_tarihi, bitis_tarihi, toplam_seans_sayisi, toplam_ucret, durum, olusturma_tarihi, olusturan_kullanici_id)
                    VALUES (?, ?, ?, ?, ?, ?, 'beklemede', ?, ?)
                        """,
                        (
                        personel,
                        datetime.datetime.now().strftime("%Y-%m-%d"),
                        baslangic,
                        bitis,
                        seans_sayisi,
                        toplam_ucret,
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        (self.kullanici[0] if self.kullanici else None)
                    )
                )
                
                conn.commit()
                conn.close()
                
                messagebox.showinfo(
                    "Başarılı",
                    f"Personel ücret talebi oluşturuldu!\n\n"
                    f"• Personel: {personel}\n"
                    f"• Tarih Aralığı: {baslangic} - {bitis}\n"
                    f"• Seans Sayısı: {seans_sayisi}\n"
                    f"• Toplam Ücret: {format_money(toplam_ucret)}"
                )
                
                win.destroy()
                
                # Ücret takibi sayfasını yenile
                try:
                    if hasattr(self, 'tab_ucret_takibi'):
                        for child in self.tab_ucret_takibi.winfo_children():
                            if isinstance(child, ttk.Frame):
                                for subchild in child.winfo_children():
                                    if isinstance(subchild, ttk.Notebook):
                                        for page_id in subchild.tabs():
                                            page = subchild.nametowidget(page_id)
                                            if hasattr(page, '_tree_personel'):
                                                self._personel_ucret_listele(page)
                except Exception:
                    pass

            except Exception as e:
                messagebox.showerror("Hata", f"Talep kaydı oluşturulamadı:\n{e}")
                log_exception("personel_ucret_talep_formu", e)
        
        ttk.Button(wrapper, text="✅ Talep Oluştur", bootstyle="success", command=kaydet).pack(pady=10)
    
    def _refresh_personel_ucret_views(self):
        """Açık Personel Ücret Takibi görünümlerini yenile."""
        try:
            if not hasattr(self, "tab_ucret_takibi"):
                return
            for child in self.tab_ucret_takibi.winfo_children():
                if getattr(child, "_tree_personel", None):
                    self._personel_ucret_listele(child)
                for sub in child.winfo_children():
                    if isinstance(sub, ttk.Notebook):
                        for page_id in sub.tabs():
                            page = sub.nametowidget(page_id)
                            if hasattr(page, "_tree_personel"):
                                self._personel_ucret_listele(page)
        except Exception:
            pass

    def _refresh_kasa_views(self):
        """Açık Kasa Defteri görünümlerini yenile."""
        try:
            if not hasattr(self, "tab_kasa"):
                return
            for child in self.tab_kasa.winfo_children():
                if getattr(child, "_tree_kasa", None):
                    self._kasa_rapor_yukle(child)
        except Exception:
            pass

    def personel_ucret_odeme_yap(self):
        """Personel ücret ödeme yapma penceresi"""
        win = ttk.Toplevel(self)
        win.title("Personel Ücret Ödeme")
        win.geometry("500x400")
        center_window(win, 500, 400)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text="Personel Ücret Ödeme", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        # Personel seçimi
        frm_personel = ttk.Frame(wrapper)
        frm_personel.pack(fill=X, pady=10)
        ttk.Label(frm_personel, text="Personel:", width=15).pack(side=LEFT)
        cmb_personel = ttk.Combobox(frm_personel, state="readonly", width=30)
        cmb_personel.pack(side=LEFT, padx=5)
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            personel_listesi = [row[0] for row in cur.fetchall()]
            conn.close()
            cmb_personel["values"] = personel_listesi
            if personel_listesi:
                cmb_personel.current(0)
        except Exception:
            cmb_personel["values"] = []
        
        # Beklemede ücretler listesi
        frm_liste = ttk.Labelframe(wrapper, text="Beklemede Ücretler", padding=10, bootstyle="secondary")
        frm_liste.pack(fill=BOTH, expand=True, pady=10)
        
        cols = ("ID", "Tarih", "Seans Ücreti", "Personel Ücreti")
        tree = ttk.Treeview(frm_liste, columns=cols, show="headings", height=8)
        for c in cols:
            tree.heading(c, text=c)
            if c == "ID":
                tree.column(c, width=50)
            elif c == "Tarih":
                tree.column(c, width=100)
            else:
                tree.column(c, width=120, anchor="e")
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frm_liste, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        def yukle_liste():
            for iid in tree.get_children():
                tree.delete(iid)
            
            personel = cmb_personel.get()
            if not personel:
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                has_put_seans = self._table_has_column(conn, "personel_ucret_takibi", "seans_id")
                if has_put_seans:
                    cur.execute(
                        """
                        SELECT put.id, put.tarih, put.seans_ucreti, put.personel_ucreti
                        FROM personel_ucret_takibi put
                        LEFT JOIN seans_takvimi st ON st.id = put.seans_id
                        WHERE put.personel_adi = ?
                          AND put.odeme_durumu = 'beklemede'
                          AND (put.seans_id IS NULL OR st.id IS NOT NULL)
                        ORDER BY put.tarih DESC
                        """,
                        (personel,)
                    )
                else:
                    cur.execute(
                        """
                        SELECT id, tarih, seans_ucreti, personel_ucreti
                        FROM personel_ucret_takibi
                        WHERE personel_adi = ? AND odeme_durumu = 'beklemede'
                        ORDER BY tarih DESC
                        """,
                        (personel,)
                    )
                rows = cur.fetchall()
                conn.close()
                
                for row in rows:
                    rid, tarih, seans_ucreti, personel_ucreti = row
                    tree.insert("", END, values=(int(rid), tarih, format_money(seans_ucreti), format_money(personel_ucreti)))
            except Exception as e:
                messagebox.showerror("Hata", f"Liste yüklenemedi:\n{e}")
        
        cmb_personel.bind("<<ComboboxSelected>>", lambda e: yukle_liste())
        yukle_liste()
        
        def odeme_yap():
            secili = tree.selection()
            if not secili:
                messagebox.showwarning("Uyarı", "Lütfen ödeme yapılacak kaydı seçin.")
                return

            vals = tree.item(secili[0]).get("values") or []
            if not vals:
                messagebox.showerror("Hata", "Seçili satır okunamadı.")
                return
            try:
                ucret_id = int(vals[0])
            except Exception:
                messagebox.showerror("Hata", f"Geçersiz ücret ID: {vals[0]!r}")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()

                # Ücret bilgisini al (gösterim için)
                cur.execute(
                    "SELECT personel_adi, personel_ucreti FROM personel_ucret_takibi WHERE id = ?",
                    (ucret_id,)
                )
                row = cur.fetchone()
                if not row:
                    messagebox.showerror("Hata", "Ücret kaydı bulunamadı.")
                    conn.close()
                    return

                personel_adi, tutar = row

                # Tek transaction kaynağı: pipeline (put + kasa birlikte)
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                ok = pipeline.personel_ucret_odeme_kasa_entegrasyonu(
                    personel_adi=personel_adi,
                    tutar=tutar,
                    ucret_takibi_id=ucret_id,
                )
                conn.close()

                if not ok:
                    messagebox.showerror("Hata", "Ödeme işlemi kaydedilemedi.")
                    return

                messagebox.showinfo("Başarılı", f"Personel ücret ödemesi yapıldı!\n\n{personel_adi}: {format_money(tutar)}")

                yukle_liste()
                self._refresh_personel_ucret_views()
                self._refresh_kasa_views()
            except Exception as e:
                messagebox.showerror("Hata", f"Ödeme yapılamadı:\n{e}")
                log_exception("personel_ucret_odeme_yap", e)
        
        ttk.Button(wrapper, text="💰 Ödeme Yap", bootstyle="success", command=odeme_yap).pack(pady=10)
    
    def _build_cocuk_gunluk_tab(self):
        """Çocuk Günlük Takip Tab - Oda ve personel takibi"""
        wrapper = ttk.Frame(self.tab_cocuk_gunluk, padding=10)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(
            wrapper,
            text="Açıklama: Çocuk günlük operasyon kayıtlarını tarih/oda/personel bazında listeler.",
            font=("Segoe UI", 9),
            foreground="gray",
        ).pack(anchor=W, pady=(0, 6))

        # Başlık
        head = ttk.Frame(wrapper)
        head.pack(fill=X, pady=(0, 10))
        if getattr(self, "_logo_small", None):
            ttk.Label(head, image=self._logo_small).pack(side=LEFT, padx=(0, 10))
        ttk.Label(head, text="ÇOCUK GÜNLÜK TAKİP", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(side=LEFT)
        
        # Filtreler
        filter_frame = ttk.Frame(wrapper)
        filter_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(filter_frame, text="Tarih:").pack(side=LEFT, padx=5)
        ent_tarih = ttk.Entry(filter_frame, width=15)
        ent_tarih.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_tarih.pack(side=LEFT, padx=5)
        
        ttk.Label(filter_frame, text="Çocuk:").pack(side=LEFT, padx=(20, 5))
        cmb_cocuk = ttk.Combobox(filter_frame, state="readonly", width=25)
        cmb_cocuk.pack(side=LEFT, padx=5)
        
        # Çocuk listesi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            cocuk_listesi = [""] + [row[0] for row in cur.fetchall()]
            conn.close()
            cmb_cocuk["values"] = cocuk_listesi
            cmb_cocuk.current(0)
        except Exception:
            cmb_cocuk["values"] = [""]
        
        ttk.Label(filter_frame, text="Oda:").pack(side=LEFT, padx=(20, 5))
        cmb_oda = ttk.Combobox(filter_frame, state="readonly", width=20)
        cmb_oda.pack(side=LEFT, padx=5)
        
        # Oda listesi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT oda_adi FROM odalar WHERE aktif=1 ORDER BY oda_adi")
            oda_listesi = [""] + [row[0] for row in cur.fetchall()]
            conn.close()
            cmb_oda["values"] = oda_listesi
            cmb_oda.current(0)
        except Exception:
            cmb_oda["values"] = [""]
        
        ttk.Button(filter_frame, text="🔄 Filtrele", bootstyle="primary",
                   command=lambda: self._cocuk_gunluk_listele(wrapper)).pack(side=LEFT, padx=10)
        
        # Treeview
        frame_tree = ttk.Frame(wrapper)
        frame_tree.pack(fill=BOTH, expand=True)
        
        # ✅ ANLAŞILIR SÜTUN İSİMLERİ: Teknik ID'ler gizli, sadece anlamlı bilgiler gösterilir
        cols = ("ID", "Çocuk Adı", "Tarih", "Çalışılan Oda", "Çalışan Personel", "Seans Bilgisi", "Notlar")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            if c == "ID":
                tree.column(c, width=0, stretch=False)  # Gizli (sadece iç kullanım)
            elif c == "Çocuk Adı":
                tree.column(c, width=200)
            elif c == "Tarih":
                tree.column(c, width=100)
            elif c == "Çalışılan Oda":
                tree.column(c, width=150)
            elif c == "Çalışan Personel":
                tree.column(c, width=150)
            elif c == "Seans Bilgisi":
                tree.column(c, width=120)  # Seans ID yerine "Seans #123" formatı gösterilecek
            else:
                tree.column(c, width=300)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frame_tree, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        # Tag'ler
        tree.tag_configure("even", background="#f8f9fa")
        tree.tag_configure("odd", background="#ffffff")
        
        # Widget'ları kaydet
        wrapper._tree_gunluk = tree
        wrapper._ent_tarih = ent_tarih
        wrapper._cmb_cocuk = cmb_cocuk
        wrapper._cmb_oda = cmb_oda
        
        # İlk yükleme
        self._cocuk_gunluk_listele(wrapper)
    
    def _cocuk_gunluk_listele(self, parent):
        """Çocuk günlük takip listesini yükle"""
        tree = parent._tree_gunluk
        ent_tarih = parent._ent_tarih
        cmb_cocuk = parent._cmb_cocuk
        cmb_oda = parent._cmb_oda
        
        for iid in tree.get_children():
            tree.delete(iid)
        
        tarih_filtre = ent_tarih.get() or ""
        cocuk_filtre = cmb_cocuk.get() or ""
        oda_filtre = ""  # oda filtresi devre dışı
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            where = []
            params = []
            
            if tarih_filtre:
                where.append("cgt.tarih = ?")
                params.append(tarih_filtre)
            
            if cocuk_filtre:
                where.append("d.ad_soyad = ?")
                params.append(cocuk_filtre)
            
            # Role göre filtre
            if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                where.append("cgt.personel_adi = ?")
                params.append(self.kullanici_terapist)
            
            sql = """
                SELECT 
                    cgt.id,
                    d.ad_soyad,
                    cgt.tarih,
                    cgt.oda_adi,
                    cgt.personel_adi,
                    cgt.seans_id,
                    cgt.notlar
                FROM cocuk_gunluk_takip cgt
                LEFT JOIN danisanlar d ON cgt.cocuk_id = d.id
            """
            if where:
                sql += " WHERE " + " AND ".join(where)
            sql += " ORDER BY cgt.tarih DESC, cgt.id DESC"
            
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
            conn.close()
            
            for idx, row in enumerate(rows):
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=row, tags=(tag,))
        
        except Exception as e:
            messagebox.showerror("Hata", f"Çocuk günlük takip listesi yüklenemedi:\n{e}")
            log_exception("_cocuk_gunluk_listele", e)
    
    def _build_kasa_defteri_tab(self):
        """Kasa Defteri Tab - Günlük/haftalık/aylık raporlar"""
        wrapper = ttk.Frame(self.tab_kasa, padding=10)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(
            wrapper,
            text="Açıklama: Bu modül kasa giriş/çıkış hareketlerini dönemsel raporlar.",
            font=("Segoe UI", 9),
            foreground="gray",
        ).pack(anchor=W, pady=(0, 6))

        # Başlık
        head = ttk.Frame(wrapper)
        head.pack(fill=X, pady=(0, 10))
        if getattr(self, "_logo_small", None):
            ttk.Label(head, image=self._logo_small).pack(side=LEFT, padx=(0, 10))
        ttk.Label(head, text="KASA DEFTERİ", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(side=LEFT)
        
        # Rapor tipi seçimi (günlük / haftalık / aylık / bütün veriler)
        report_frame = ttk.Labelframe(wrapper, text="Rapor Tipi", padding=10, bootstyle="secondary")
        report_frame.pack(fill=X, pady=(0, 10))
        
        report_type = ttk.StringVar(value="gunluk")
        rf = ttk.Frame(report_frame)
        rf.pack(fill=X)
        ttk.Radiobutton(rf, text="📅 Günlük", variable=report_type, value="gunluk", command=lambda: self._kasa_rapor_yukle(wrapper)).grid(row=0, column=0, padx=6, pady=2, sticky=W)
        ttk.Radiobutton(rf, text="📆 Haftalık", variable=report_type, value="haftalik", command=lambda: self._kasa_rapor_yukle(wrapper)).grid(row=0, column=1, padx=6, pady=2, sticky=W)
        ttk.Radiobutton(rf, text="📊 Aylık", variable=report_type, value="aylik", command=lambda: self._kasa_rapor_yukle(wrapper)).grid(row=1, column=0, padx=6, pady=2, sticky=W)
        ttk.Radiobutton(rf, text="📋 Bütün veriler", variable=report_type, value="tumu", command=lambda: self._kasa_rapor_yukle(wrapper)).grid(row=1, column=1, padx=6, pady=2, sticky=W)

        # Tarih seçimi (kompakt 2 satır)
        date_frame = ttk.Frame(wrapper)
        date_frame.pack(fill=X, pady=(0, 10))

        top_btn = ttk.Frame(date_frame)
        top_btn.pack(fill=X, pady=(0, 4))
        ttk.Label(top_btn, text="Tarih:").pack(side=LEFT, padx=4)
        ent_tarih = ttk.Entry(top_btn, width=14)
        ent_tarih.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_tarih.pack(side=LEFT, padx=4)
        ttk.Button(top_btn, text="🔄 Göster", bootstyle="primary", width=12, command=lambda: self._kasa_rapor_yukle(wrapper)).pack(side=LEFT, padx=4)
        ttk.Button(top_btn, text="↻ Yenile", bootstyle="secondary", width=10, command=lambda: self._kasa_rapor_yukle(wrapper)).pack(side=LEFT, padx=4)

        bot_btn = ttk.Frame(date_frame)
        bot_btn.pack(fill=X)
        ttk.Button(bot_btn, text="📊 Excel", bootstyle="success", width=12, command=lambda: self._kasa_rapor_hazirla_excel(wrapper)).pack(side=LEFT, padx=4)
        ttk.Button(bot_btn, text="🗑️ Seçiliyi Sil", bootstyle="danger", width=14, command=self.safe_call(lambda: self._kasa_secili_sil(parent=wrapper), "kasa_secili_sil")).pack(side=LEFT, padx=4)

        
        # Özet bilgiler
        summary_frame = ttk.Labelframe(wrapper, text="Özet", padding=10, bootstyle="secondary")
        summary_frame.pack(fill=X, pady=(0, 10))
        
        summary_labels = ttk.Frame(summary_frame)
        summary_labels.pack(fill=X)
        
        # Treeview
        frame_tree = ttk.Frame(wrapper)
        frame_tree.pack(fill=BOTH, expand=True)
        
        # ✅ ANLAŞILIR SÜTUN İSİMLERİ
        cols = ("ID", "Tarih", "Hareket Tipi", "Açıklama", "Tutar", "Ödeme Şekli", "İlgili Kayıt", "İlgili Seans")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            if c == "ID":
                tree.column(c, width=50)
            elif c == "Tarih":
                tree.column(c, width=100)
            elif c == "Tip":
                tree.column(c, width=80)
            elif c == "Açıklama":
                tree.column(c, width=300)
            elif c == "Tutar":
                tree.column(c, width=120, anchor="e")
            elif c == "Ödeme Şekli":
                tree.column(c, width=120)
            else:
                tree.column(c, width=80)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frame_tree, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        # Tag'ler
        tree.tag_configure("giren", background="#d4edda", foreground="#155724")
        tree.tag_configure("cikan", background="#f8d7da", foreground="#721c24")
        tree.tag_configure("even", background="#f8f9fa")
        tree.tag_configure("odd", background="#ffffff")
        
        # Widget'ları kaydet
        wrapper._tree_kasa = tree
        wrapper._ent_tarih = ent_tarih
        wrapper._report_type = report_type
        wrapper._summary_labels = summary_labels
        
        # İlk yükleme
        self._kasa_rapor_yukle(wrapper)
    
    def _kasa_rapor_yukle(self, parent):
        """Kasa defteri raporunu yükle"""
        if not getattr(parent, "_tree_kasa", None) and hasattr(self, "tab_kasa"):
            for child in self.tab_kasa.winfo_children():
                if getattr(child, "_tree_kasa", None):
                    parent = child
                    break
        tree = getattr(parent, "_tree_kasa", None)
        ent_tarih = getattr(parent, "_ent_tarih", None)
        report_type = getattr(parent, "_report_type", None)
        summary_labels = getattr(parent, "_summary_labels", None)
        if not tree or not report_type or not summary_labels:
            messagebox.showerror("Hata", "Kasa raporu yüklenemedi (ekran henüz hazır değil). Lütfen KASA DEFTERİ sekmesine tıklayıp tekrar deneyin.")
            return
        if not ent_tarih:
            ent_tarih = type("_", (), {"get": lambda: datetime.datetime.now().strftime("%Y-%m-%d")})()
        
        for iid in tree.get_children():
            tree.delete(iid)
        
        # Özet etiketlerini temizle
        for widget in summary_labels.winfo_children():
            widget.destroy()
        
        tarih_str = (ent_tarih.get() or "").strip() or datetime.datetime.now().strftime("%Y-%m-%d")
        tarih = None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                tarih = datetime.datetime.strptime(tarih_str, fmt)
                break
            except Exception:
                continue
        if not tarih:
            messagebox.showerror("Hata", "Geçersiz tarih! Örnek: YYYY-MM-DD veya GG.AA.YYYY")
            return

        # Tarih aralığını hesapla (Bütün veriler = tüm kayıtlar)
        if report_type.get() == "tumu":
            baslangic = "2000-01-01"
            bitis = "2099-12-31"
        elif report_type.get() == "gunluk":
            baslangic = bitis = tarih.strftime("%Y-%m-%d")
        elif report_type.get() == "haftalik":
            # Haftanın başlangıcı (Pazartesi)
            baslangic = (tarih - datetime.timedelta(days=tarih.weekday())).strftime("%Y-%m-%d")
            bitis = (tarih + datetime.timedelta(days=6-tarih.weekday())).strftime("%Y-%m-%d")
        else:  # aylik
            baslangic = tarih.replace(day=1).strftime("%Y-%m-%d")
            # Ayın son günü
            if tarih.month == 12:
                bitis = tarih.replace(year=tarih.year+1, month=1, day=1) - datetime.timedelta(days=1)
            else:
                bitis = tarih.replace(month=tarih.month+1, day=1) - datetime.timedelta(days=1)
            bitis = bitis.strftime("%Y-%m-%d")
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # Kasa hareketlerini çek
            cur.execute(
                """
                SELECT 
                    id, tarih, tip, aciklama, tutar, odeme_sekli, record_id, seans_id
                FROM kasa_hareketleri
                WHERE tarih >= ? AND tarih <= ?
                ORDER BY tarih DESC, id DESC
                """,
                (baslangic, bitis)
            )
            rows = cur.fetchall()
            
            # Özet hesapla
            toplam_giren = 0.0
            toplam_cikan = 0.0
            
            for idx, row in enumerate(rows):
                kasa_id, tarih, tip, aciklama, tutar, odeme_sekli, record_id, seans_id = row
                tip_norm = (tip or "").strip().lower().replace("ı", "i").replace("ç", "c")
                if tip_norm == "giren":
                    toplam_giren += float(tutar or 0)
                elif tip_norm == "cikan":
                    toplam_cikan += float(tutar or 0)
                tag = tip_norm if tip_norm in ("giren", "cikan") else ("even" if idx % 2 == 0 else "odd")
                
                # Kayıt ve Seans ID'lerini anlaşılır formata çevir
                kayit_bilgisi = f"Kayıt #{record_id}" if record_id and record_id != "-" else "-"
                seans_bilgisi = f"Seans #{seans_id}" if seans_id and seans_id != "-" else "-"
                
                display_tip = (tip or "").strip()
                if display_tip.lower() in ("cikan", "çıkan"):
                    display_tip = "Çıkan"
                elif display_tip.lower() == "giren":
                    display_tip = "Giren"
                else:
                    display_tip = display_tip.capitalize() if display_tip else "-"
                tree.insert("", END, values=(
                    kasa_id,
                    tarih,
                    display_tip,
                    aciklama,
                    format_money(tutar),
                    odeme_sekli or "-",
                    kayit_bilgisi,
                    seans_bilgisi
                ), tags=(tag,))
            
            net = toplam_giren - toplam_cikan
            
            conn.close()
            
            # Özet göster
            ttk.Label(summary_labels, text=f"📥 Toplam Giren: {format_money(toplam_giren)}", 
                     font=("Segoe UI", 11, "bold"), bootstyle="success").pack(side=LEFT, padx=15)
            ttk.Label(summary_labels, text=f"📤 Toplam Çıkan: {format_money(toplam_cikan)}", 
                     font=("Segoe UI", 11, "bold"), bootstyle="danger").pack(side=LEFT, padx=15)
            ttk.Label(summary_labels, text=f"💰 Net Kasa: {format_money(net)}", 
                     font=("Segoe UI", 12, "bold"), 
                     bootstyle="success" if net >= 0 else "danger").pack(side=LEFT, padx=15)
            ttk.Label(summary_labels, text=f"📊 Tarih Aralığı: {baslangic} - {bitis}", 
                     font=("Segoe UI", 10), foreground="gray").pack(side=LEFT, padx=15)
        
        except Exception as e:
            messagebox.showerror("Hata", f"Kasa raporu yüklenemedi:\n{e}")
            log_exception("_kasa_rapor_yukle", e)
    
    def _kasa_rapor_hazirla_excel(self, parent):
        """Kasa defteri raporunu Excel olarak hazırla ve kaydet"""
        if not getattr(parent, "_tree_kasa", None) and hasattr(self, "tab_kasa"):
            for child in self.tab_kasa.winfo_children():
                if getattr(child, "_tree_kasa", None):
                    parent = child
                    break
        ent_tarih = getattr(parent, "_ent_tarih", None)
        report_type = getattr(parent, "_report_type", None)
        if not report_type:
            messagebox.showerror("Hata", "Kasa raporu hazırlanamadı (ekran henüz hazır değil).")
            return
        tarih_str = (ent_tarih.get() if ent_tarih else "").strip() or datetime.datetime.now().strftime("%Y-%m-%d")
        tarih = None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                tarih = datetime.datetime.strptime(tarih_str, fmt)
                break
            except Exception:
                continue
        if not tarih:
            messagebox.showerror("Hata", "Geçersiz tarih! Örnek: YYYY-MM-DD veya GG.AA.YYYY")
            return
        
        # Tarih aralığını hesapla (Bütün veriler = tüm kayıtlar)
        if report_type.get() == "tumu":
            baslangic = "2000-01-01"
            bitis = "2099-12-31"
            rapor_adi = "Tum_veriler"
        elif report_type.get() == "gunluk":
            baslangic = bitis = tarih.strftime("%Y-%m-%d")
            rapor_adi = f"Gunluk_{baslangic}"
        elif report_type.get() == "haftalik":
            baslangic = (tarih - datetime.timedelta(days=tarih.weekday())).strftime("%Y-%m-%d")
            bitis = (tarih + datetime.timedelta(days=6-tarih.weekday())).strftime("%Y-%m-%d")
            rapor_adi = f"Haftalik_{baslangic}_to_{bitis}"
        else:  # aylik
            baslangic = tarih.replace(day=1).strftime("%Y-%m-%d")
            if tarih.month == 12:
                bitis = tarih.replace(year=tarih.year+1, month=1, day=1) - datetime.timedelta(days=1)
            else:
                bitis = tarih.replace(month=tarih.month+1, day=1) - datetime.timedelta(days=1)
            bitis = bitis.strftime("%Y-%m-%d")
            rapor_adi = f"Aylik_{tarih.year}_{tarih.month:02d}"
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # Kasa hareketlerini çek
            cur.execute(
                """
                SELECT 
                    id, tarih, tip, aciklama, tutar, odeme_sekli, record_id, seans_id
                FROM kasa_hareketleri
                WHERE tarih >= ? AND tarih <= ?
                ORDER BY tarih DESC, id DESC
                """,
                (baslangic, bitis)
            )
            rows = cur.fetchall()
            
            if not rows:
                messagebox.showwarning("Uyarı", "Seçilen rapor kapsamında kasa hareketi bulunamadı!")
                conn.close()
                return
            
            # Özet hesapla
            toplam_giren = 0.0
            toplam_cikan = 0.0
            
            # Excel için veri hazırla
            excel_data = []
            excel_data.append(["ID", "Tarih", "Tip", "Açıklama", "Tutar (₺)", "Ödeme Şekli", "Kayıt ID", "Seans ID"])
            
            for row in rows:
                kasa_id, tarih, tip, aciklama, tutar, odeme_sekli, record_id, seans_id = row
                tutar_val = float(tutar or 0)
                
                if tip == "giren":
                    toplam_giren += tutar_val
                elif tip == "cikan":
                    toplam_cikan += tutar_val
                
                excel_data.append([
                    kasa_id,
                    tarih,
                    tip.capitalize(),
                    aciklama or "",
                    tutar_val,
                    odeme_sekli or "-",
                    record_id or "-",
                    seans_id or "-"
                ])
            
            # Özet satırları ekle
            excel_data.append([])
            excel_data.append(["ÖZET", "", "", "", "", "", "", ""])
            excel_data.append(["Toplam Giren", "", "", "", toplam_giren, "", "", ""])
            excel_data.append(["Toplam Çıkan", "", "", "", toplam_cikan, "", "", ""])
            excel_data.append(["Net Kasa", "", "", "", toplam_giren - toplam_cikan, "", "", ""])
            excel_data.append([])
            excel_data.append(["Tarih Aralığı", f"{baslangic} - {bitis}", "", "", "", "", "", ""])
            
            # DataFrame oluştur
            df = pd.DataFrame(excel_data)
            
            # Excel dosyasını kaydet
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"Kasa_Defteri_{rapor_adi}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if not path:
                conn.close()
                return
            
            # Excel'e yaz
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Kasa Defteri', index=False, header=False)
                
                # Formatlama için worksheet al
                worksheet = writer.sheets['Kasa Defteri']
                
                # Başlık satırını kalın yap
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Özet satırlarını kalın yap
                for row_idx in range(len(excel_data) - 5, len(excel_data)):
                    if row_idx > 0:
                        for cell in worksheet[row_idx]:
                            if cell.value and str(cell.value).strip():
                                cell.font = Font(bold=True)
                
                # Sütun genişliklerini ayarla
                worksheet.column_dimensions['A'].width = 8
                worksheet.column_dimensions['B'].width = 12
                worksheet.column_dimensions['C'].width = 10
                worksheet.column_dimensions['D'].width = 40
                worksheet.column_dimensions['E'].width = 15
                worksheet.column_dimensions['F'].width = 15
                worksheet.column_dimensions['G'].width = 10
                worksheet.column_dimensions['H'].width = 10
                
                # Tutar sütununu sayı formatına çevir
                for row_idx in range(2, len(excel_data) - 5):
                    cell = worksheet.cell(row=row_idx, column=5)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
            
            conn.close()
            messagebox.showinfo("Başarılı", f"Kasa defteri raporu Excel olarak hazırlandı:\n{path}")
        
        except Exception as e:
            messagebox.showerror("Hata", f"Excel raporu oluşturulamadı:\n{e}")
            log_exception("_kasa_rapor_hazirla_excel", e)
    
    def _build_bep_tab(self):
        """Bu modül kaldırıldı."""
        messagebox.showinfo("Bilgi", "Bu modül bu sürümde kaldırılmıştır.")

    def _kasa_manual_entry_popup(self, parent_wrapper=None):
        """Kasa defterine manuel gelir/gider ekleme penceresi."""
        win = ttk.Toplevel(self)
        win.title("Kasa - Gelir/Gider Ekle")
        win.transient(self)
        try:
            center_window_smart(win, 520, 320)
        except Exception:
            pass
        self._brand_window(win)

        frm = ttk.Frame(win, padding=12)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text="KASA DEFTERİ - MANUEL KAYIT", font=("Segoe UI", 12, "bold"), bootstyle="primary").pack(anchor=W, pady=(0, 10))

        row = ttk.Frame(frm)
        row.pack(fill=X, pady=4)
        ttk.Label(row, text="Tarih (YYYY-MM-DD):", width=18).pack(side=LEFT)
        ent_tarih = ttk.Entry(row, width=18)
        ent_tarih.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_tarih.pack(side=LEFT)

        row2 = ttk.Frame(frm)
        row2.pack(fill=X, pady=4)
        ttk.Label(row2, text="Tip:", width=18).pack(side=LEFT)
        tip_var = ttk.StringVar(value="giren")
        ttk.Radiobutton(row2, text="Gelir", variable=tip_var, value="giren").pack(side=LEFT, padx=(0, 10))
        ttk.Radiobutton(row2, text="Gider", variable=tip_var, value="cikan").pack(side=LEFT)

        row3 = ttk.Frame(frm)
        row3.pack(fill=X, pady=4)
        ttk.Label(row3, text="Tutar (₺):", width=18).pack(side=LEFT)
        ent_tutar = ttk.Entry(row3, width=18, validate="key", validatecommand=self._vcmd_money)
        ent_tutar.pack(side=LEFT)

        row4 = ttk.Frame(frm)
        row4.pack(fill=X, pady=4)
        ttk.Label(row4, text="Ödeme Şekli:", width=18).pack(side=LEFT)
        cmb_odeme = ttk.Combobox(row4, state="readonly", width=16, values=["Nakit", "Havale/EFT", "Kredi Kartı", "Diğer"])
        cmb_odeme.set("Nakit")
        cmb_odeme.pack(side=LEFT)

        row5 = ttk.Frame(frm)
        row5.pack(fill=X, pady=4)
        ttk.Label(row5, text="Açıklama:", width=18).pack(side=LEFT)
        ent_aciklama = ttk.Entry(row5)
        ent_aciklama.pack(side=LEFT, fill=X, expand=True)

        btns = ttk.Frame(frm)
        btns.pack(fill=X, pady=(12, 0))

        def _save():
            tarih = (ent_tarih.get() or "").strip()
            tip = tip_var.get()
            try:
                tutar = float(parse_money(ent_tutar.get()))
            except Exception:
                tutar = 0.0
            if not tarih:
                messagebox.showerror("Hata", "Tarih boş olamaz.")
                return
            if tutar <= 0:
                messagebox.showerror("Hata", "Lütfen geçerli bir tutar girin.")
                return

            odeme = (cmb_odeme.get() or "Nakit").strip()
            aciklama = (ent_aciklama.get() or "").strip()

            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                ok = pipeline.add_manual_kasa_entry(tarih=tarih, tip=tip, tutar=tutar, odeme_sekli=odeme, aciklama=aciklama)
                conn.close()
                if ok:
                    try:
                        self._kasa_rapor_yukle(parent_wrapper)
                    except Exception:
                        pass
                    messagebox.showinfo("Başarılı", "Kasa kaydı eklendi.")
                    win.destroy()
                else:
                    messagebox.showerror("Hata", "Kasa kaydı eklenemedi.")
            except Exception as e:
                messagebox.showerror("Hata", f"Kasa kaydı eklenemedi:\n{e}")
                log_exception("_kasa_manual_entry_popup", e)

        ttk.Button(btns, text="Kaydet", bootstyle="success", command=_save).pack(side=RIGHT, padx=5)
        ttk.Button(btns, text="İptal", bootstyle="secondary", command=win.destroy).pack(side=RIGHT, padx=5)

    def _bep_olustur_guncelle(self, parent):
        """BEP oluştur veya güncelle"""
        cmb_cocuk = parent._cmb_cocuk
        cmb_yil = parent._cmb_yil
        bep_table = parent._bep_table
        
        cocuk_text = cmb_cocuk.get()
        if not cocuk_text:
            messagebox.showwarning("Uyarı", "Lütfen bir çocuk seçin.")
            return
        
        # Çocuk ID'sini al
        try:
            cocuk_id = int(cocuk_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            messagebox.showerror("Hata", "Geçersiz çocuk seçimi.")
            return
        
        yil = int(cmb_yil.get())
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # BEP programını kontrol et
            cur.execute(
                "SELECT id FROM bep_programlari WHERE cocuk_id = ? AND program_yili = ?",
                (cocuk_id, yil)
            )
            bep_row = cur.fetchone()
            
            if bep_row:
                bep_id = bep_row[0]
                # Güncelle
                cur.execute(
                    "UPDATE bep_programlari SET guncelleme_tarihi = ? WHERE id = ?",
                    (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), bep_id)
                )
            else:
                # Yeni oluştur
                cur.execute(
                    """
                    INSERT INTO bep_programlari
                    (cocuk_id, program_yili, olusturma_tarihi, olusturan_kullanici_id)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        cocuk_id,
                        yil,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        (self.kullanici[0] if self.kullanici else None)
                    )
                )
                bep_id = cur.lastrowid
            
            # Hedef becerileri kaydet
            for beceri, ay_widgets in bep_table.items():
                for ay_idx, (var, cmb) in enumerate(ay_widgets):
                    durum = var.get()
                    if durum:
                        # Durum mapping
                        durum_map = {
                            "Planlandı": "planlandi",
                            "Devam Ediyor": "devam_ediyor",
                            "Tamamlandı": "tamamlandi"
                        }
                        durum_db = durum_map.get(durum, "planlandi")
                        
                        # Mevcut kaydı kontrol et
                        cur.execute(
                            """
                            SELECT id FROM bep_hedef_beceriler
                            WHERE bep_id = ? AND hedef_beceri = ? AND ay = ?
                            """,
                            (bep_id, beceri, ay_idx + 1)
                        )
                        hedef_row = cur.fetchone()
                        
                        if hedef_row:
                            # Güncelle
                            cur.execute(
                                """
                                UPDATE bep_hedef_beceriler
                                SET durum = ?, olusturma_tarihi = ?
                                WHERE id = ?
                                """,
                                (durum_db, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), hedef_row[0])
                            )
                        else:
                            # Yeni ekle
                            cur.execute(
                                """
                                INSERT INTO bep_hedef_beceriler
                                (bep_id, hedef_beceri, ay, durum, olusturma_tarihi)
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                (
                                    bep_id,
                                    beceri,
                                    ay_idx + 1,
                                    durum_db,
                                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                )
                            )
            
            # Commit ve close döngü dışında olmalı
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Başarılı", f"BEP programı kaydedildi!\n\nÇocuk ID: {cocuk_id}\nYıl: {yil}")
        
        except Exception as e:
            messagebox.showerror("Hata", f"BEP kaydedilemedi:\n{e}")
            log_exception("_bep_olustur_guncelle", e)
    
    def _bep_yukle(self, parent):
        """BEP programını yükle"""
        cmb_cocuk = parent._cmb_cocuk
        cmb_yil = parent._cmb_yil
        bep_table = parent._bep_table
        
        cocuk_text = cmb_cocuk.get()
        if not cocuk_text:
                return

        # Çocuk ID'sini al
        try:
            cocuk_id = int(cocuk_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            return
        
        yil = int(cmb_yil.get())
        
        # Tüm combobox'ları temizle
        for beceri, ay_widgets in bep_table.items():
            for var, cmb in ay_widgets:
                var.set("")
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # BEP programını bul
            cur.execute(
                "SELECT id FROM bep_programlari WHERE cocuk_id = ? AND program_yili = ?",
                (cocuk_id, yil)
            )
            bep_row = cur.fetchone()
            
            if not bep_row:
                return
            
            bep_id = bep_row[0]
            
            # Hedef becerileri yükle
            cur.execute(
                """
                SELECT hedef_beceri, ay, durum
                FROM bep_hedef_beceriler
                WHERE bep_id = ?
                """,
                (bep_id,)
            )
            rows = cur.fetchall()
            
            # Durum mapping (ters)
            durum_map = {
                "planlandi": "Planlandı",
                "devam_ediyor": "Devam Ediyor",
                "tamamlandi": "Tamamlandı"
            }
            
            for hedef_beceri, ay, durum in rows:
                if hedef_beceri in bep_table:
                    ay_widgets = bep_table[hedef_beceri]
                    if 1 <= ay <= 12:
                        var, cmb = ay_widgets[ay - 1]
                        var.set(durum_map.get(durum, "Planlandı"))
            
            conn.close()
        
        except Exception as e:
            log_exception("_bep_yukle", e)
    
    def _build_haftalik_seans_tab(self):
        """Haftalık Seans Takip Tab - Personel bazlı dinamik program"""
        wrapper = ttk.Frame(self.tab_haftalik, padding=10)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(
            wrapper,
            text="Açıklama: Haftalık seans planını personel bazında kaydedip dışa aktarır.",
            font=("Segoe UI", 9),
            foreground="gray",
        ).pack(anchor=W, pady=(0, 6))

        # Başlık
        head = ttk.Frame(wrapper)
        head.pack(fill=X, pady=(0, 10))
        if getattr(self, "_logo_small", None):
            ttk.Label(head, image=self._logo_small).pack(side=LEFT, padx=(0, 10))
        ttk.Label(head, text="HAFTALIK SEANS PROGRAMI", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(side=LEFT)
        
        # Personel seçimi ve hafta seçimi
        filter_frame = ttk.Frame(wrapper)
        filter_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(filter_frame, text="Personel:").pack(side=LEFT, padx=5)
        cmb_personel = ttk.Combobox(filter_frame, state="readonly", width=25)
        cmb_personel.pack(side=LEFT, padx=5)
        
        # Personel listesi: settings + haftalik_seans_programi'ndeki personel (içe aktarılanlar da görünsün)
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            personel_listesi = list({row[0] for row in cur.fetchall() if row[0]})
            cur.execute("SELECT DISTINCT personel_adi FROM haftalik_seans_programi WHERE personel_adi IS NOT NULL AND TRIM(personel_adi) != '' ORDER BY personel_adi")
            for row in cur.fetchall():
                if row[0] and row[0] not in personel_listesi:
                    personel_listesi.append(row[0])
            personel_listesi.sort()
            conn.close()
            cmb_personel["values"] = personel_listesi
            if personel_listesi:
                cmb_personel.current(0)
        except Exception:
            cmb_personel["values"] = []
        
        ttk.Label(filter_frame, text="Hafta:").pack(side=LEFT, padx=(20, 5))
        ent_hafta = ttk.Entry(filter_frame, width=15)
        # Bu haftanın pazartesi gününü hesapla
        today = datetime.datetime.now()
        monday = today - datetime.timedelta(days=today.weekday())
        ent_hafta.insert(0, monday.strftime("%Y-%m-%d"))
        ent_hafta.pack(side=LEFT, padx=5)
        
        ttk.Button(filter_frame, text="💾 Kaydet", bootstyle="success",
                   command=lambda: self._haftalik_program_kaydet(wrapper)).pack(side=LEFT, padx=10)
        ttk.Button(filter_frame, text="📤 Excel'e Aktar", bootstyle="info",
                   command=lambda: self._haftalik_program_excel_aktar(wrapper)).pack(side=LEFT, padx=5)
        
        # Personel seçimi kaydetme/dışa aktarma filtrelerinde kullanılır
        
        # Haftalık program tablosu
        program_frame = ttk.Labelframe(wrapper, text="Haftalık Program", padding=10, bootstyle="secondary")
        program_frame.pack(fill=BOTH, expand=True)
        
        # Scrollable canvas
        canvas_frame = ttk.Frame(program_frame)
        canvas_frame.pack(fill=BOTH, expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg="white")
        scrollbar = ttk.Scrollbar(canvas_frame, orient=VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # Tablo başlıkları
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill=X, pady=(0, 5))
        
        ttk.Label(header_frame, text="Saat", font=("Segoe UI", 10, "bold"), width=12, anchor="w").pack(side=LEFT, padx=2)
        GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        GUN_KISA = ["Paz", "Sal", "Çar", "Per", "Cum", "Cmt", "Pzr"]  # Cumartesi=Cmt, Pazar=Pzr (tekrar önleme)
        for i, gun in enumerate(GUNLER):
            ttk.Label(header_frame, text=GUN_KISA[i], font=("Segoe UI", 9, "bold"), width=15, anchor="center").pack(side=LEFT, padx=1)
        
        # Saatler (09:00 - 18:00, yarım saatler dahil: 16:30 vb. görünsün)
        saatler = [f"{h:02d}:{m:02d}" for h in range(9, 19) for m in (0, 30)]
        program_table = {}
        
        for saat in saatler:
            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill=X, pady=2)
            
            ttk.Label(row_frame, text=saat, font=("Segoe UI", 9), width=12, anchor="w").pack(side=LEFT, padx=2)
            
            gun_widgets = []
            for gun in GUNLER:
                var = ttk.StringVar(value="")
                ent = ttk.Entry(row_frame, textvariable=var, width=15)
                ent.pack(side=LEFT, padx=1)
                gun_widgets.append((var, ent))
            
            program_table[saat] = gun_widgets
        
        # Widget'ları kaydet
        wrapper._cmb_personel = cmb_personel
        wrapper._ent_hafta = ent_hafta
        wrapper._program_table = program_table
        wrapper._canvas = canvas
        wrapper._scrollable_frame = scrollable_frame
        
        # İlk yükleme yapılmaz; kullanıcı kaydetme ve dışa aktarma akışını kullanır.
    
    def _hafta_sec(self, ent_hafta):
        """Hafta seçimi için takvim penceresi"""
        try:
            from tkcalendar import DateEntry
        except ImportError:
            messagebox.showwarning("Uyarı", "tkcalendar modülü bulunamadı. Lütfen 'pip install tkcalendar' komutu ile yükleyin.")
            return
        
        win = ttk.Toplevel(self)
        win.title("Hafta Seç")
        win.geometry("300x200")
        center_window(win, 300, 200)
        
        ttk.Label(win, text="Haftanın Pazartesi gününü seçin:", font=("Segoe UI", 10)).pack(pady=20)
        
        cal = DateEntry(win, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        cal.pack(pady=10)
        
        def sec():
            selected_date = cal.get_date()
            # Pazartesi gününe çevir
            weekday = selected_date.weekday()
            monday = selected_date - datetime.timedelta(days=weekday)
            ent_hafta.delete(0, END)
            ent_hafta.insert(0, monday.strftime("%Y-%m-%d"))
            win.destroy()
        
        ttk.Button(win, text="Seç", bootstyle="primary", command=sec).pack(pady=10)
    
    def _haftalik_program_excel_aktar(self, parent):
        """Seçili personel ve hafta için haftalık programı Excel'e aktarır."""
        cmb_personel = parent._cmb_personel
        ent_hafta = parent._ent_hafta

        personel = (cmb_personel.get() or "").strip()
        hafta = self._normalize_hafta_tarihi((ent_hafta.get() or "").strip())
        if not personel or not hafta:
            messagebox.showwarning("Uyarı", "Lütfen personel ve hafta girin.")
            return

        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT gun, saat, COALESCE(ogrenci_adi,''), COALESCE(notlar,'')
                FROM haftalik_seans_programi
                WHERE personel_adi=? AND hafta_baslangic_tarihi=?
                ORDER BY CASE gun
                    WHEN 'Pazartesi' THEN 1 WHEN 'Salı' THEN 2 WHEN 'Çarşamba' THEN 3
                    WHEN 'Perşembe' THEN 4 WHEN 'Cuma' THEN 5 WHEN 'Cumartesi' THEN 6 WHEN 'Pazar' THEN 7
                    ELSE 99 END,
                    saat
                """,
                (personel, hafta),
            )
            rows = cur.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Program okunamadı\n{e}")
            return

        path = filedialog.asksaveasfilename(
            title="Haftalık Programı Excel'e Aktar",
            defaultextension=".xlsx",
            initialfile=f"haftalik_program_{personel.replace(' ', '_')}_{hafta}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        try:
            df = pd.DataFrame(rows, columns=["Gün", "Saat", "Öğrenci", "Notlar"])
            df.insert(0, "Personel", personel)
            df.insert(1, "Hafta", hafta)
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Haftalık Program")
            messagebox.showinfo("Başarılı", "Haftalık program Excel'e aktarıldı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel aktarımı başarısız\n{e}")
            log_exception("_haftalik_program_excel_aktar", e)

    def _normalize_hafta_tarihi(self, raw: str) -> str:
        """Hafta başlangıç tarihini YYYY-MM-DD formatına çevirir. Veritabanı hep bu formatta tutulur."""
        if not raw or not isinstance(raw, str):
            return ""
        s = raw.strip()
        if not s:
            return ""
        # Zaten YYYY-MM-DD ise
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            try:
                datetime.datetime.strptime(s[:10], "%Y-%m-%d")
                return s[:10]
            except ValueError:
                pass
        # DD.MM.YYYY veya DD/MM/YYYY
        for sep in (".", "/", "-"):
            if sep in s and len(s) >= 10:
                parts = s.replace("/", sep).replace(".", sep).split(sep)
                if len(parts) >= 3:
                    try:
                        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                        if y < 100:
                            y += 2000
                        dt = datetime.datetime(y, m, d)
                        return dt.strftime("%Y-%m-%d")
                    except (ValueError, TypeError):
                        pass
        return s
    
    def _normalize_saat(self, raw: str) -> str:
        """Saat değerini HH:MM formatına getirir (10, 10.00, 16:30 vb.)."""
        if not raw or not isinstance(raw, str):
            return ""
        s = raw.strip().replace(",", ".")
        if not s:
            return ""
        # Zaten HH:MM veya H:MM
        if ":" in s:
            parts = s.split(":")
            try:
                h = int(parts[0].strip())
                m = int(parts[1].strip()) if len(parts) > 1 else 0
                if 0 <= h <= 23 and 0 <= m <= 59:
                    return f"{h:02d}:{m:02d}"
            except (ValueError, TypeError):
                pass
        # Sadece saat (10, 10.0, 16.5 -> 16:30)
        try:
            v = float(s)
            h = int(v)
            m = int(round((v - h) * 60))
            if m >= 60:
                h += 1
                m = 0
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        except (ValueError, TypeError):
            pass
        return s
    
    def _haftalik_program_yukle(self, parent):
        """Haftalık programı veritabanından yükle"""
        cmb_personel = parent._cmb_personel
        ent_hafta = parent._ent_hafta
        program_table = parent._program_table
        
        personel = (cmb_personel.get() or "").strip()
        hafta_baslangic = self._normalize_hafta_tarihi(ent_hafta.get() or "")
        if hafta_baslangic and hafta_baslangic != (ent_hafta.get() or "").strip():
            ent_hafta.delete(0, END)
            ent_hafta.insert(0, hafta_baslangic)
        
        if not personel or not hafta_baslangic:
            return
        
        # Tüm entry'leri temizle
        for saat, gun_widgets in program_table.items():
            for var, ent in gun_widgets:
                var.set("")
        
        conn = None
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            
            cur.execute(
                """
                SELECT gun, saat, ogrenci_adi, oda_adi, notlar
                FROM haftalik_seans_programi
                WHERE TRIM(COALESCE(personel_adi,'')) = ? AND TRIM(COALESCE(hafta_baslangic_tarihi,'')) = ?
                ORDER BY 
                    CASE gun
                        WHEN 'Pazartesi' THEN 1
                        WHEN 'Salı' THEN 2
                        WHEN 'Çarşamba' THEN 3
                        WHEN 'Perşembe' THEN 4
                        WHEN 'Cuma' THEN 5
                        WHEN 'Cumartesi' THEN 6
                        WHEN 'Pazar' THEN 7
                    END,
                    saat
                """,
                (personel, hafta_baslangic)
            )
            rows = cur.fetchall()
            for gun, saat, ogrenci_adi, oda_adi, notlar in rows:
                if saat in program_table:
                    gun_idx = GUNLER.index(gun) if gun in GUNLER else -1
                    if gun_idx >= 0:
                        var, ent = program_table[saat][gun_idx]
                        # Oda gösterimi kaldırıldı; sadece öğrenci ve notlar
                        text = ogrenci_adi or ""
                        if notlar:
                            text += f" [{notlar}]" if text else notlar
                        var.set(text)
        except Exception as e:
            log_exception("_haftalik_program_yukle", e)
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
    
    def _haftalik_program_kaydet(self, parent):
        """Haftalık programı veritabanına kaydet"""
        cmb_personel = parent._cmb_personel
        ent_hafta = parent._ent_hafta
        program_table = parent._program_table
        
        personel = (cmb_personel.get() or "").strip()
        hafta_raw = (ent_hafta.get() or "").strip()
        hafta_baslangic = self._normalize_hafta_tarihi(hafta_raw)
        
        if not personel or not hafta_baslangic:
            messagebox.showwarning("Uyarı", "Lütfen personel ve hafta seçin. Hafta formatı: YYYY-MM-DD veya GG.AA.YYYY")
            return
        
        if hafta_baslangic != hafta_raw:
            ent_hafta.delete(0, END)
            ent_hafta.insert(0, hafta_baslangic)
        
        conn = None
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # Mevcut kayıtları sil (normalize edilmiş tarih ile eşleşir)
            cur.execute(
                "DELETE FROM haftalik_seans_programi WHERE personel_adi = ? AND hafta_baslangic_tarihi = ?",
                (personel, hafta_baslangic)
            )
            
            GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            olusturma_tarihi = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for saat, gun_widgets in program_table.items():
                for gun_idx, (var, ent) in enumerate(gun_widgets):
                    text = var.get().strip()
                    if text:
                        # Oda kaldırıldı: format "Öğrenci [Notlar]" veya sadece "Öğrenci"
                        ogrenci_adi = text
                        notlar = ""
                        if "[" in text and "]" in text:
                            idx = text.index("[")
                            ogrenci_adi = text[:idx].strip()
                            notlar = text[idx+1:].rstrip("]").strip()
                        cur.execute(
                            """
                            INSERT INTO haftalik_seans_programi
                            (personel_adi, hafta_baslangic_tarihi, gun, saat, ogrenci_adi, oda_adi, notlar, olusturma_tarihi, olusturan_kullanici_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                personel,
                                hafta_baslangic,
                                GUNLER[gun_idx],
                                saat,
                                ogrenci_adi,
                                "",  # oda_adi kaldırıldı
                                notlar,
                                olusturma_tarihi,
                                (self.kullanici[0] if self.kullanici else None)
                            )
                        )
            
            conn.commit()
            messagebox.showinfo("Başarılı", f"Haftalık program kaydedildi!\n\nPersonel: {personel}\nHafta: {hafta_baslangic}")
        except sqlite3.IntegrityError as e:
            messagebox.showerror("Hata", f"Program kaydedilemedi (çakışma):\n{e}\n\nAynı personel/hafta/gün/saat zaten var. Önce 'Yükle' ile mevcut programı açıp tekrar deneyin.")
            log_exception("_haftalik_program_kaydet", e)
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
        except Exception as e:
            messagebox.showerror("Hata", f"Program kaydedilemedi:\n{e}")
            log_exception("_haftalik_program_kaydet", e)
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
    
    def _build_ogrenci_bilgileri_tab(self):
        """Öğrenci Bilgileri Tab - Tüm danışanların bilgileri"""
        wrapper = ttk.Frame(self.tab_ogrenci_bilgileri, padding=10)
        wrapper.pack(fill=BOTH, expand=True)
        # Seans ekleme sonrası borç tablosunu yenilemek için sakla
        self._ogrenci_bilgileri_wrapper = wrapper
        
        ttk.Label(
            wrapper,
            text="Açıklama: Danışan kayıtları, iletişim bilgileri ve durum yönetimi burada tutulur.",
            font=("Segoe UI", 9),
            foreground="gray",
        ).pack(anchor=W, pady=(0, 6))

        # Başlık
        head = ttk.Frame(wrapper)
        head.pack(fill=X, pady=(0, 10))
        if getattr(self, "_logo_small", None):
            ttk.Label(head, image=self._logo_small).pack(side=LEFT, padx=(0, 10))
        ttk.Label(head, text="TÜM DANIŞANLAR", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(side=LEFT)
        
        # Toolbar
        toolbar = ttk.Frame(wrapper)
        toolbar.pack(fill=X, pady=(0, 10))
        
        ttk.Label(toolbar, text="Ara:").pack(side=LEFT, padx=5)
        ent_ara_danisan = ttk.Entry(toolbar, width=30)
        ent_ara_danisan.pack(side=LEFT, padx=5)
        ent_ara_danisan.bind("<KeyRelease>", lambda e: self._tum_danisanlari_listele(wrapper))
        
        ttk.Button(toolbar, text="🔄 Yenile", bootstyle="secondary",
                   command=lambda: self._tum_danisanlari_listele(wrapper)).pack(side=LEFT, padx=5)
        ttk.Button(toolbar, text="➕ Yeni Danışan", bootstyle="success",
                   command=lambda: self._yeni_danisan_ekle_ve_guncelle(None, wrapper)).pack(side=LEFT, padx=5)
        
        # Bakiye, records (seans/ödeme) tablosundan hesaplanır; çocuk takip formları ayrı işlevdir
        info_bakiye = ttk.Label(wrapper, text="💡 Bakiye: Seans ve ödeme kayıtlarından otomatik hesaplanır. Çocuk Takip formları ayrı bir modüldür (form kayıtları).", 
            font=("Segoe UI", 9), foreground="gray")
        info_bakiye.pack(anchor=W, pady=(0, 4))
        # ✅ TÜM DANIŞANLAR TABLOSU
        frame_tree = ttk.Frame(wrapper)
        frame_tree.pack(fill=BOTH, expand=True)
        # Tabloya göre: AD SOYAD, DOĞUM TARİHİ, VELİ ADI, VELİ TELEFON, ADRES (danışan telefon/email yok; iletişim veli bilgilerinden)
        cols = ("ID", "AD SOYAD", "DOĞUM TARİHİ", "VELİ ADI", "VELİ TELEFON", "ADRES", "Borç (₺)", "Durum")
        tree_danisanlar = ttk.Treeview(frame_tree, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree_danisanlar.heading(c, text=c)
            if c == "ID":
                tree_danisanlar.column(c, width=50)
            elif c == "AD SOYAD":
                tree_danisanlar.column(c, width=200)
            elif c == "DOĞUM TARİHİ":
                tree_danisanlar.column(c, width=110)
            elif c == "VELİ ADI":
                tree_danisanlar.column(c, width=150)
            elif c == "VELİ TELEFON":
                tree_danisanlar.column(c, width=120)
            elif c == "ADRES":
                tree_danisanlar.column(c, width=180)
            elif c == "Borç (₺)":
                tree_danisanlar.column(c, width=100, anchor="e")
            else:
                tree_danisanlar.column(c, width=80)
        
        tree_danisanlar.pack(side=LEFT, fill=BOTH, expand=True)
        sb_danisanlar = ttk.Scrollbar(frame_tree, orient=VERTICAL, command=tree_danisanlar.yview)
        tree_danisanlar.configure(yscroll=sb_danisanlar.set)
        sb_danisanlar.pack(side=RIGHT, fill=Y)
        
        danisan_actions = ttk.Frame(wrapper)
        danisan_actions.pack(fill=X, pady=(6, 4))
        ttk.Label(danisan_actions, text="Seçili öğrenci işlemleri:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 8))
        ttk.Button(danisan_actions, text="✏️ Düzenle", bootstyle="secondary-outline", command=lambda: self._danisan_duzenle_from_tree(tree_danisanlar)).pack(side=LEFT, padx=4)
        ttk.Button(danisan_actions, text="📊 Detaylı Bilgi", bootstyle="info-outline", command=lambda: self._danisan_detayli_bilgi(tree_danisanlar)).pack(side=LEFT, padx=4)
        ttk.Button(danisan_actions, text="🔁 Aktif/Pasif", bootstyle="warning-outline", command=lambda: self._danisan_aktif_pasif_from_tree(tree_danisanlar)).pack(side=LEFT, padx=4)
        ttk.Button(danisan_actions, text="🗑️ Danışanı Sil (Kalıcı)", bootstyle="danger-outline", command=lambda: self._danisan_sil_from_tree(tree_danisanlar)).pack(side=LEFT, padx=4)
        
        wrapper._tree_danisanlar = tree_danisanlar
        self.tab_ogrenci_bilgileri.danisan_tree = tree_danisanlar
        wrapper._ent_ara_danisan = ent_ara_danisan
        
        # İlk yükleme
        self._tum_danisanlari_listele(wrapper)
        
        # Kimlik bilgileri bölümü kaldırıldı (talep doğrultusunda)
        
    
    def _build_ogrenci_aile_page(self, parent):
        """Öğrenci aile bilgileri sayfası"""
        # Öğrenci seçimi
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(filter_frame, text="Öğrenci:").pack(side=LEFT, padx=5)
        cmb_ogrenci = ttk.Combobox(filter_frame, state="readonly", width=30)
        cmb_ogrenci.pack(side=LEFT, padx=5)
        
        # Öğrenci listesi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT id, ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            ogrenci_listesi = [(row[0], row[1]) for row in cur.fetchall()]
            conn.close()
            cmb_ogrenci["values"] = [f"{c[1]} (ID: {c[0]})" for c in ogrenci_listesi]
            if ogrenci_listesi:
                cmb_ogrenci.current(0)
        except Exception:
            cmb_ogrenci["values"] = []
        
        ttk.Button(filter_frame, text="➕ Yeni Veli Ekle", bootstyle="success",
                   command=lambda: self._yeni_veli_ekle(parent)).pack(side=LEFT, padx=10)
        ttk.Button(filter_frame, text="🔄 Yükle", bootstyle="primary",
                   command=lambda: self._ogrenci_aile_listele(parent)).pack(side=LEFT, padx=5)
        
        # Veli listesi
        frame_tree = ttk.Frame(parent)
        frame_tree.pack(fill=BOTH, expand=True)
        
        cols = ("ID", "Veli Adı", "Yakınlık", "Telefon", "Email", "Adres", "Notlar")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            if c == "ID":
                tree.column(c, width=50)
            elif c == "Veli Adı":
                tree.column(c, width=200)
            elif c == "Yakınlık":
                tree.column(c, width=120)
            elif c == "Telefon":
                tree.column(c, width=120)
            elif c == "Email":
                tree.column(c, width=200)
            else:
                tree.column(c, width=200)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(frame_tree, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        
        veli_actions = ttk.Frame(parent)
        veli_actions.pack(fill=X, pady=(8, 4))
        ttk.Label(veli_actions, text="Seçili veli işlemleri:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 8))
        ttk.Button(veli_actions, text="✏️ Düzenle", bootstyle="secondary-outline", command=lambda: self._veli_duzenle(parent, tree)).pack(side=LEFT, padx=4)
        ttk.Button(veli_actions, text="🗑️ Sil", bootstyle="danger-outline", command=lambda: self._veli_sil(parent, tree)).pack(side=LEFT, padx=4)
        
        parent._tree_aile = tree
        parent._cmb_ogrenci = cmb_ogrenci
        
        # İlk yükleme
        self._ogrenci_aile_listele(parent)
    
    def _build_ogrenci_kimlik_page(self, parent):
        """Öğrenci kimlik bilgileri sayfası"""
        # Öğrenci seçimi
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(filter_frame, text="Öğrenci:").pack(side=LEFT, padx=5)
        cmb_ogrenci = ttk.Combobox(filter_frame, state="readonly", width=30)
        cmb_ogrenci.pack(side=LEFT, padx=5)
        
        # Öğrenci listesi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT id, ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            ogrenci_listesi = [(row[0], row[1]) for row in cur.fetchall()]
            conn.close()
            cmb_ogrenci["values"] = [f"{c[1]} (ID: {c[0]})" for c in ogrenci_listesi]
            if ogrenci_listesi:
                cmb_ogrenci.current(0)
        except Exception:
            cmb_ogrenci["values"] = []
        
        ttk.Button(filter_frame, text="✏️ Düzenle", bootstyle="primary",
                   command=lambda: self._kimlik_duzenle(parent)).pack(side=LEFT, padx=10)
        ttk.Button(filter_frame, text="🔄 Yükle", bootstyle="secondary",
                   command=lambda: self._ogrenci_kimlik_yukle(parent)).pack(side=LEFT, padx=5)
        
        # Kimlik bilgileri formu
        form_frame = ttk.Labelframe(parent, text="Kimlik Bilgileri", padding=20, bootstyle="secondary")
        form_frame.pack(fill=BOTH, expand=True, pady=10)
        
        ttk.Label(form_frame, text="TC Kimlik No:").grid(row=0, column=0, sticky=W, padx=5, pady=5)
        ent_tc = ttk.Entry(form_frame, width=30)
        ent_tc.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Doğum Tarihi:").grid(row=1, column=0, sticky=W, padx=5, pady=5)
        ent_dogum = ttk.Entry(form_frame, width=30)
        ent_dogum.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Doğum Yeri:").grid(row=2, column=0, sticky=W, padx=5, pady=5)
        ent_dogum_yeri = ttk.Entry(form_frame, width=30)
        ent_dogum_yeri.grid(row=2, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Notlar:").grid(row=3, column=0, sticky=W, padx=5, pady=5)
        text_notlar = tk.Text(form_frame, width=30, height=5)
        text_notlar.grid(row=3, column=1, padx=5, pady=5)
        
        ttk.Button(form_frame, text="💾 Kaydet", bootstyle="success",
                   command=lambda: self._kimlik_kaydet(parent)).grid(row=4, column=0, columnspan=2, pady=10)
        
        parent._cmb_ogrenci_kimlik = cmb_ogrenci
        parent._ent_tc = ent_tc
        parent._ent_dogum = ent_dogum
        parent._ent_dogum_yeri = ent_dogum_yeri
        parent._text_notlar = text_notlar
        
        # İlk yükleme
        self._ogrenci_kimlik_yukle(parent)
    
    # Sistem Şifreleri Tab - KALDIRILDI (Kullanıcı isteği)
    
    # ==================== ÖĞRENCİ BİLGİLERİ HELPER METODLARI ====================
    
    def _ogrenci_aile_listele(self, parent):
        """Öğrenci aile bilgilerini listele"""
        tree = parent._tree_aile
        cmb_ogrenci = parent._cmb_ogrenci
        
        for iid in tree.get_children():
            tree.delete(iid)
        
        ogrenci_text = cmb_ogrenci.get()
        if not ogrenci_text:
            return
        
        try:
            ogrenci_id = int(ogrenci_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            return
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT id, veli_adi, veli_yakinlik_derecesi, telefon, email, adres, notlar
                FROM ogrenci_aile_bilgileri
                WHERE ogrenci_id = ?
                ORDER BY id
                """,
                (ogrenci_id,)
            )
            rows = cur.fetchall()
            conn.close()
            
            for idx, row in enumerate(rows):
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=row, tags=(tag,))
        
        except Exception as e:
            log_exception("_ogrenci_aile_listele", e)
    
    def _yeni_veli_ekle(self, parent):
        """Yeni veli ekle penceresi"""
        cmb_ogrenci = parent._cmb_ogrenci
        ogrenci_text = cmb_ogrenci.get()
        
        if not ogrenci_text:
            messagebox.showwarning("Uyarı", "Lütfen bir öğrenci seçin.")
            return
        
        try:
            ogrenci_id = int(ogrenci_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            messagebox.showerror("Hata", "Geçersiz öğrenci seçimi.")
            return
        
        win = ttk.Toplevel(self)
        win.title("Yeni Veli Ekle")
        win.transient(self)
        self._brand_window(win)
        center_window_smart(win, 700, 600, min_w=650, min_h=550)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text="Yeni Veli Ekle", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        ttk.Label(wrapper, text="Veli Adı:").pack(anchor=W)
        ent_veli_adi = ttk.Entry(wrapper, width=40)
        ent_veli_adi.pack(fill=X, pady=5)
        
        ttk.Label(wrapper, text="Yakınlık Derecesi:").pack(anchor=W, pady=(10, 0))
        cmb_yakinlik = ttk.Combobox(wrapper, state="readonly", width=37,
                                   values=["Anne", "Baba", "Büyükanne", "Büyükbaba", "Vasi", "Diğer"])
        cmb_yakinlik.pack(fill=X, pady=5)
        cmb_yakinlik.current(0)
        
        ttk.Label(wrapper, text="Telefon:").pack(anchor=W, pady=(10, 0))
        ent_telefon = ttk.Entry(wrapper, width=40)
        ent_telefon.pack(fill=X, pady=5)
        
        ttk.Label(wrapper, text="Email:").pack(anchor=W, pady=(10, 0))
        ent_email = ttk.Entry(wrapper, width=40)
        ent_email.pack(fill=X, pady=5)
        
        ttk.Label(wrapper, text="Adres:").pack(anchor=W, pady=(10, 0))
        text_adres = tk.Text(wrapper, width=40, height=3)
        text_adres.pack(fill=X, pady=5)
        
        def kaydet():
            veli_adi = ent_veli_adi.get().strip()
            yakinlik = cmb_yakinlik.get()
            
            if not veli_adi or not yakinlik:
                messagebox.showwarning("Uyarı", "Lütfen veli adı ve yakınlık derecesini girin.")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    INSERT INTO ogrenci_aile_bilgileri
                    (ogrenci_id, veli_adi, veli_yakinlik_derecesi, telefon, email, adres, notlar, olusturma_tarihi, guncelleme_tarihi)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ogrenci_id,
                        veli_adi,
                        yakinlik,
                        ent_telefon.get().strip(),
                        ent_email.get().strip(),
                        text_adres.get("1.0", END).strip(),
                        "",
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                )
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Başarılı", "Veli eklendi!")
                win.destroy()
                self._ogrenci_aile_listele(parent)
                
            except Exception as e:
                messagebox.showerror("Hata", f"Veli eklenemedi:\n{e}")
                log_exception("_yeni_veli_ekle", e)
        
        ttk.Button(wrapper, text="💾 Kaydet", bootstyle="success", command=kaydet).pack(pady=20)
    
    def _veli_duzenle(self, parent, tree):
        """Veli düzenle"""
        sel = tree.selection()
        if not sel:
            return
        
        veli_id = tree.item(sel[0])["values"][0]
        # ✅ DÜZELTME: Gerçek veli düzenleme penceresi (basit bir bilgi penceresi)
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT oa.veli_adi, oa.veli_yakinlik_derecesi, oa.telefon, oa.email, oa.adres, oa.notlar,
                       d.ad_soyad as ogrenci_adi
                FROM ogrenci_aile_bilgileri oa
                LEFT JOIN danisanlar d ON oa.ogrenci_id = d.id
                WHERE oa.id = ?
                """,
                (veli_id,)
            )
            row = cur.fetchone()
            conn.close()
            
            if row:
                win = ttk.Toplevel(self)
                win.title(f"Veli Bilgileri - {row[6] or 'Bilinmiyor'}")
                center_window_smart(win, 600, 500, max_ratio=0.9)
                win.transient(self)
                win.grab_set()
                self._brand_window(win)
                
                wrapper = ttk.Frame(win, padding=20)
                wrapper.pack(fill=BOTH, expand=True)
                
                ttk.Label(wrapper, text=f"Veli Bilgileri: {row[0] or '-'}", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
                
                info_frame = ttk.Frame(wrapper)
                info_frame.pack(fill=BOTH, expand=True)
                
                def info_row(label, value):
                    f = ttk.Frame(info_frame)
                    f.pack(fill=X, pady=5)
                    ttk.Label(f, text=f"{label}:", font=("Segoe UI", 10, "bold"), width=25).pack(side=LEFT, anchor=W)
                    ttk.Label(f, text=str(value or "-"), font=("Segoe UI", 10)).pack(side=LEFT, anchor=W, padx=10)
                
                info_row("Öğrenci", row[6] or "-")
                info_row("Veli Adı", row[0])
                info_row("Yakınlık Derecesi", row[1])
                info_row("Telefon", row[2])
                info_row("E-posta", row[3])
                info_row("Adres", row[4] or "-")
                
                if row[5]:
                    ttk.Separator(info_frame).pack(fill=X, pady=10)
                    ttk.Label(info_frame, text="Notlar:", font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(5, 0))
                    txt_not = ttk.Text(info_frame, height=4, wrap=WORD)
                    txt_not.insert("1.0", row[5])
                    txt_not.config(state="disabled")
                    txt_not.pack(fill=X, pady=5)
                
                ttk.Button(wrapper, text="Kapat", bootstyle="secondary", command=win.destroy).pack(pady=20)
            else:
                messagebox.showerror("Hata", "Veli bilgisi bulunamadı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Veli bilgileri yüklenemedi:\n{e}")
            log_exception("_veli_duzenle", e)
    
    def _veli_sil(self, parent, tree):
        """Veli sil"""
        sel = tree.selection()
        if not sel:
            return
        
        veli_id = tree.item(sel[0])["values"][0]
        
        if not messagebox.askyesno("Onay", "Veli kaydını silmek istediğinize emin misiniz?"):
            return
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("DELETE FROM ogrenci_aile_bilgileri WHERE id = ?", (veli_id,))
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Başarılı", "Veli kaydı silindi.")
            self._ogrenci_aile_listele(parent)
        
        except Exception as e:
            messagebox.showerror("Hata", f"Veli silinemedi:\n{e}")
            log_exception("_veli_sil", e)
    
    def _ogrenci_kimlik_yukle(self, parent):
        """Öğrenci kimlik bilgilerini yükle"""
        cmb_ogrenci = parent._cmb_ogrenci_kimlik
        ent_tc = parent._ent_tc
        ent_dogum = parent._ent_dogum
        ent_dogum_yeri = parent._ent_dogum_yeri
        text_notlar = parent._text_notlar
        
        ogrenci_text = cmb_ogrenci.get()
        if not ogrenci_text:
            return
        
        try:
            ogrenci_id = int(ogrenci_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            return
        
        # Form'u temizle
        ent_tc.delete(0, END)
        ent_dogum.delete(0, END)
        ent_dogum_yeri.delete(0, END)
        text_notlar.delete("1.0", END)
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                "SELECT tc_kimlik_no, dogum_tarihi, dogum_yeri, notlar FROM ogrenci_kimlik_bilgileri WHERE ogrenci_id = ?",
                (ogrenci_id,)
            )
            row = cur.fetchone()
            conn.close()
            
            if row:
                ent_tc.insert(0, row[0] or "")
                ent_dogum.insert(0, row[1] or "")
                ent_dogum_yeri.insert(0, row[2] or "")
                text_notlar.insert("1.0", row[3] or "")
        
        except Exception as e:
            log_exception("_ogrenci_kimlik_yukle", e)
    
    def _kimlik_duzenle(self, parent):
        """Kimlik bilgilerini düzenle penceresi"""
        cmb_ogrenci = parent._cmb_ogrenci_kimlik
        ogrenci_text = cmb_ogrenci.get()
        
        if not ogrenci_text:
            messagebox.showwarning("Uyarı", "Lütfen bir öğrenci seçin.")
            return
        
        try:
            ogrenci_id = int(ogrenci_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            messagebox.showerror("Hata", "Geçersiz öğrenci seçimi.")
            return
        
        # Form'u yükle
        self._ogrenci_kimlik_yukle(parent)
    
    def _kimlik_kaydet(self, parent):
        """Kimlik bilgilerini kaydet"""
        cmb_ogrenci = parent._cmb_ogrenci_kimlik
        ent_tc = parent._ent_tc
        ent_dogum = parent._ent_dogum
        ent_dogum_yeri = parent._ent_dogum_yeri
        text_notlar = parent._text_notlar
        
        ogrenci_text = cmb_ogrenci.get()
        if not ogrenci_text:
            messagebox.showwarning("Uyarı", "Lütfen bir öğrenci seçin.")
            return
        
        try:
            ogrenci_id = int(ogrenci_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            messagebox.showerror("Hata", "Geçersiz öğrenci seçimi.")
            return
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # Mevcut kaydı kontrol et
            cur.execute("SELECT id FROM ogrenci_kimlik_bilgileri WHERE ogrenci_id = ?", (ogrenci_id,))
            row = cur.fetchone()
            
            if row:
                # Güncelle
                cur.execute(
                    """
                    UPDATE ogrenci_kimlik_bilgileri
                    SET tc_kimlik_no = ?, dogum_tarihi = ?, dogum_yeri = ?, notlar = ?, guncelleme_tarihi = ?
                    WHERE ogrenci_id = ?
                    """,
                    (
                        ent_tc.get().strip(),
                        ent_dogum.get().strip(),
                        ent_dogum_yeri.get().strip(),
                        text_notlar.get("1.0", END).strip(),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        ogrenci_id
                    )
                )
            else:
                # Yeni ekle
                cur.execute(
                    """
                    INSERT INTO ogrenci_kimlik_bilgileri
                    (ogrenci_id, tc_kimlik_no, dogum_tarihi, dogum_yeri, notlar, olusturma_tarihi, guncelleme_tarihi)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ogrenci_id,
                        ent_tc.get().strip(),
                        ent_dogum.get().strip(),
                        ent_dogum_yeri.get().strip(),
                        text_notlar.get("1.0", END).strip(),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                )
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Başarılı", "Kimlik bilgileri kaydedildi!")
        
        except Exception as e:
            messagebox.showerror("Hata", f"Kimlik bilgileri kaydedilemedi:\n{e}")
            log_exception("_kimlik_kaydet", e)
    
    def _tum_danisanlari_listele(self, parent):
        """Tüm danışanların bilgilerini listele"""
        tree = parent._tree_danisanlar
        ent_ara = parent._ent_ara_danisan
        
        for iid in tree.get_children():
            tree.delete(iid)
        
        ara_metni = (ent_ara.get() or "").strip().upper()
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # Borç: records'tan canlı hesapla; danışan adı büyük/küçük harf farkına duyarsız eşleştir
            subq = "(SELECT COALESCE(SUM(r.kalan_borc), 0) FROM records r WHERE UPPER(TRIM(r.danisan_adi)) = UPPER(TRIM(d.ad_soyad)))"
            if ara_metni:
                cur.execute(
                    f"""
                    SELECT d.id, d.ad_soyad, d.dogum_tarihi, d.veli_adi, d.veli_telefon, COALESCE(d.adres,''),
                           {subq} as balance, d.aktif
                    FROM danisanlar d
                    WHERE UPPER(d.ad_soyad) LIKE ? OR UPPER(d.veli_adi) LIKE ? OR d.veli_telefon LIKE ?
                    ORDER BY d.ad_soyad
                    """,
                    (f"%{ara_metni}%", f"%{ara_metni}%", f"%{ara_metni}%")
                )
            else:
                cur.execute(
                    f"""
                    SELECT d.id, d.ad_soyad, d.dogum_tarihi, d.veli_adi, d.veli_telefon, COALESCE(d.adres,''),
                           {subq} as balance, d.aktif
                    FROM danisanlar d
                    ORDER BY d.ad_soyad
                    """
                )
            
            rows = cur.fetchall()
            conn.close()
            
            # row: id, ad_soyad, dogum_tarihi, veli_adi, veli_telefon, adres, balance, aktif
            for idx, row in enumerate(rows):
                durum = "Aktif" if row[7] else "Pasif"
                bakiye = format_money(row[6] or 0)
                dogum_tarihi = row[2] or ""
                if dogum_tarihi:
                    try:
                        dt = datetime.datetime.strptime(dogum_tarihi, "%Y-%m-%d")
                        dogum_tarihi = dt.strftime("%d.%m.%Y")
                    except Exception:
                        pass
                
                tag = "even" if idx % 2 == 0 else "odd"
                if row[6] and float(row[6] or 0) > 0:
                    tag = "borclu"
                
                veli_bilgisi = row[3] or ""
                if veli_bilgisi:
                    try:
                        conn2 = self.veritabani_baglan()
                        cur2 = conn2.cursor()
                        cur2.execute(
                            """
                            SELECT veli_yakinlik_derecesi 
                            FROM ogrenci_aile_bilgileri 
                            WHERE ogrenci_id = ? AND veli_adi = ? 
                            LIMIT 1
                            """,
                            (row[0], row[3])
                        )
                        yakinlik_row = cur2.fetchone()
                        conn2.close()
                        if yakinlik_row and yakinlik_row[0]:
                            veli_bilgisi = f"{row[3]} ({yakinlik_row[0]})"
                    except Exception:
                        pass
                
                tree.insert("", END, values=(
                    row[0], row[1], dogum_tarihi, veli_bilgisi, row[4] or "", row[5] or "", bakiye, durum
                ), tags=(tag,))
        
        except Exception as e:
            log_exception("_tum_danisanlari_listele", e)
    
    def _danisan_duzenle_from_tree(self, tree):
        """Tree'den seçilen danışanı düzenle"""
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir danışan seçin.")
            return
        # ✅ DÜZELTME: Gerçek düzenleme fonksiyonunu kullan
        # Önce parent window'u bul (danisanlar tab'ı)
        parent_win = None
        for widget in self.winfo_children():
            if isinstance(widget, ttk.Notebook):
                for tab_id in widget.tabs():
                    tab_widget = widget.nametowidget(tab_id)
                    if hasattr(tab_widget, 'danisan_tree') and tab_widget.danisan_tree == tree:
                        parent_win = tab_widget
                        break
                if parent_win:
                    break
        
        if parent_win:
            # Geçici olarak parent_win'e danisan_tree ekle
            parent_win.danisan_tree = tree
            self.danisan_duzenle(parent_win)
        else:
            # Fallback: Direkt danisan_duzenle çağır (parent olarak self kullan)
            # Geçici bir parent oluştur
            temp_parent = ttk.Frame(self)
            temp_parent.danisan_tree = tree
            self.danisan_duzenle(temp_parent)
    
    def _danisan_fiyatlandirma(self, tree):
        """Danışan için fiyatlandırma penceresi"""
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir danışan seçin.")
            return

        # 1. ADIM: Danışan adını EN BAŞTA al (Hata buradan kaynaklanıyordu)
        try:
            values = tree.item(sel[0])["values"]
            # Genellikle 0: ID, 1: Ad Soyad'dır. Listeye göre değişebilir ama standart bu.
            danisan_adi = values[1] 
        except Exception:
            danisan_adi = "Bilinmeyen Danışan"

        # Önce parent window'u bul
        parent_win = None
        for widget in self.winfo_children():
            if isinstance(widget, ttk.Notebook):
                for tab_id in widget.tabs():
                    tab_widget = widget.nametowidget(tab_id)
                    if hasattr(tab_widget, 'danisan_tree') and tab_widget.danisan_tree == tree:
                        parent_win = tab_widget
                        break
                if parent_win:
                    break
        
        if parent_win:
            # Çocuk Ücret Takibi tab'ındaki tree'yi bul ve fiyatlandırma penceresini aç
            try:
                # Ücret Takibi tab'ını bul
                ucret_tab = None
                for child in self.winfo_children():
                    if isinstance(child, ttk.Notebook):
                        for tab_id in child.tabs():
                            w = child.nametowidget(tab_id)
                            # Tab isminde veya içeriğinde 'ucret' geçiyor mu kontrol et
                            # Veya direkt class attribute kontrolü
                            if hasattr(w, 'cocuk_ucret_tree') or hasattr(w, '_tree_cocuk'):
                                ucret_tab = w
                                break
                
                # Eğer self.tab_ucret_takibi varsa ona da bakabiliriz
                if not ucret_tab and hasattr(self, 'tab_ucret_takibi'):
                     # Notebook içindeki frame'e ulaşmamız lazım
                     for child in self.tab_ucret_takibi.winfo_children():
                         if isinstance(child, ttk.Frame): # Wrapper
                             for sub in child.winfo_children():
                                 if isinstance(sub, ttk.Notebook):
                                     # İlk sayfa çocuk ücretleridir
                                     try:
                                         ucret_tab = sub.nametowidget(sub.tabs()[0])
                                     except Exception:
                                         pass

                if ucret_tab and (hasattr(ucret_tab, 'cocuk_ucret_tree') or hasattr(ucret_tab, '_tree_cocuk')):
                    target_tree = getattr(ucret_tab, 'cocuk_ucret_tree', None) or getattr(ucret_tab, '_tree_cocuk', None)
                    
                    if target_tree:
                        # Ücret takibi tab'ındaki tree'de bu danışanı bulmaya çalış
                        found = False
                        for item in target_tree.get_children():
                            v = target_tree.item(item)["values"]
                            # v[1] genelde çocuk adıdır
                            if len(v) > 1 and str(v[1]).upper() == str(danisan_adi).upper():
                                target_tree.selection_set(item)
                                self._fiyatlandirma_guncelle(ucret_tab, target_tree)
                                found = True
                                return # Başarılı çıkış

                        if not found:
                             messagebox.showinfo("Bilgi", f"'{danisan_adi}' için Ücret Takibi listesinde kayıt bulunamadı.\n\nLütfen önce 'Seans Takip' veya 'Ücret Takibi' ekranından bu öğrenciye ait bir kayıt oluşturun.")
                    else:
                         messagebox.showinfo("Bilgi", f"'{danisan_adi}' fiyatlandırması için 'Ücret Takibi' sekmesini kullanın.")
                else:
                    messagebox.showinfo("Bilgi", f"'{danisan_adi}' fiyatlandırması için 'Ücret Takibi' sekmesini kullanın.")
            
            except Exception as e:
                log_exception("_danisan_fiyatlandirma", e)
                messagebox.showinfo("Bilgi", f"'{danisan_adi}' fiyatlandırması için 'Ücret Takibi' sekmesini kullanın.")
        else:
            messagebox.showinfo("Bilgi", f"'{danisan_adi}' fiyatlandırması için 'Ücret Takibi' sekmesini kullanın.")
    
    def _danisan_detayli_bilgi(self, tree):
        """Danışan detaylı bilgi penceresi"""
        sel = tree.selection()
        if not sel:
            return
        danisan_id = tree.item(sel[0])["values"][0]
        danisan_adi = tree.item(sel[0])["values"][1]
        
        # ✅ DÜZELTME: Gerçek detaylı bilgi penceresi oluştur
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT ad_soyad, telefon, email, veli_adi, veli_telefon, dogum_tarihi, adres, notlar, olusturma_tarihi, aktif
                FROM danisanlar WHERE id=?
                """,
                (danisan_id,)
            )
            row = cur.fetchone()
            conn.close()
            
            if not row:
                messagebox.showerror("Hata", "Danışan bulunamadı.")
                return
            
            # Detaylı bilgi penceresi
            win = ttk.Toplevel(self)
            win.title(f"Danışan Detayları - {danisan_adi}")
            center_window_smart(win, 700, 600, max_ratio=0.9)
            win.transient(self)
            win.grab_set()
            self._brand_window(win)
            
            wrapper = ttk.Frame(win, padding=20)
            wrapper.pack(fill=BOTH, expand=True)
            
            ttk.Label(wrapper, text=f"Danışan Detayları: {danisan_adi}", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
            
            info_frame = ttk.Frame(wrapper)
            info_frame.pack(fill=BOTH, expand=True)
            
            def info_row(label, value):
                f = ttk.Frame(info_frame)
                f.pack(fill=X, pady=5)
                ttk.Label(f, text=f"{label}:", font=("Segoe UI", 10, "bold"), width=20).pack(side=LEFT, anchor=W)
                ttk.Label(f, text=str(value or "-"), font=("Segoe UI", 10)).pack(side=LEFT, anchor=W, padx=10)
            
            info_row("Ad Soyad", row[0])
            info_row("Telefon", row[1])
            info_row("E-posta", row[2])
            info_row("Veli Adı", row[3])
            info_row("Veli Telefon", row[4])
            info_row("Doğum Tarihi", row[5])
            info_row("Adres", row[6] or "-")
            info_row("Durum", "Aktif" if (row[9] or 1) == 1 else "Pasif")
            info_row("Kayıt Tarihi", row[8])
            
            if row[7]:
                ttk.Separator(info_frame).pack(fill=X, pady=10)
                ttk.Label(info_frame, text="Notlar:", font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(5, 0))
                txt_not = ttk.Text(info_frame, height=5, wrap=WORD)
                txt_not.insert("1.0", row[7])
                txt_not.config(state="disabled")
                txt_not.pack(fill=X, pady=5)
            
            ttk.Button(wrapper, text="Kapat", bootstyle="secondary", command=win.destroy).pack(pady=20)
            
        except Exception as e:
            messagebox.showerror("Hata", f"Detaylar yüklenemedi:\n{e}")
            log_exception("_danisan_detayli_bilgi", e)
    
    def _danisan_aktif_pasif_from_tree(self, tree):
        """Seçili danışanın aktif/pasif durumunu değiştir."""
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0]).get("values") or []
        if len(vals) < 8:
            messagebox.showerror("Hata", "Seçili satır bilgisi okunamadı.")
            return

        danisan_id = vals[0]
        danisan_adi = vals[1]
        mevcut_durum = str(vals[7] or "").strip().lower()
        yeni_aktif = 0 if mevcut_durum == "aktif" else 1

        if not messagebox.askyesno(
            "Onay",
            f"{danisan_adi} için durum {'Aktif' if yeni_aktif else 'Pasif'} yapılsın mı?",
        ):
            return

        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("UPDATE danisanlar SET aktif=? WHERE id=?", (yeni_aktif, danisan_id))
            conn.commit()
            conn.close()

            messagebox.showinfo("Başarılı", f"{danisan_adi} durumu {'Aktif' if yeni_aktif else 'Pasif'} olarak güncellendi.")
            parent = tree.master.master
            self._tum_danisanlari_listele(parent)
        except Exception as e:
            messagebox.showerror("Hata", f"Durum güncellenemedi:\n{e}")
            log_exception("_danisan_aktif_pasif_from_tree", e)

    def _danisan_sil_from_tree(self, tree):
        """Seçili danışanı veritabanından kalıcı olarak sil."""
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0]).get("values") or []
        if len(vals) < 2:
            messagebox.showerror("Hata", "Seçili satır bilgisi okunamadı.")
            return

        danisan_id = vals[0]
        danisan_adi = vals[1]

        if not messagebox.askyesno(
            "Kalıcı Silme Onayı",
            f"{danisan_adi} danışanı veritabanından KALICI olarak silinsin mi?\n\n"
            "Bu işlem geri alınamaz.",
        ):
            return

        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("DELETE FROM danisanlar WHERE id=?", (danisan_id,))
            conn.commit()
            conn.close()

            messagebox.showinfo("Başarılı", f"{danisan_adi} danışanı kalıcı olarak silindi.")
            parent = tree.master.master
            self._tum_danisanlari_listele(parent)
            self._refresh_borc_tables()
        except sqlite3.IntegrityError:
            messagebox.showerror(
                "Silme Engellendi",
                "Danışana bağlı kayıtlar bulunduğu için doğrudan silinemedi.\n"
                "Önce bağlı seans/ödeme kayıtlarını temizleyin.",
            )
        except Exception as e:
            messagebox.showerror("Hata", f"Danışan silinemedi:\n{e}")
            log_exception("_danisan_sil_from_tree", e)


    # ==================== RAPOR OLUŞTURMA FONKSİYONLARI ====================
    
    def _bep_rapor_olustur(self, cmb_cocuk, cmb_yil):
        """BEP raporu oluştur (PDF - Form tıpkısı)"""
        if not PDF_AVAILABLE:
            messagebox.showerror(
                "Hata",
                "PDF oluşturma kütüphanesi bulunamadı.\n\n"
                "PowerShell:\n"
                "  py -m pip install -r requirements.txt\n\n"
                "Sonra uygulamayı kapatıp tekrar aç.",
            )
            return
        
        cocuk_text = cmb_cocuk.get()
        yil = cmb_yil.get()
        
        if not cocuk_text:
            messagebox.showwarning("Uyarı", "Lütfen bir çocuk seçin.")
            return
        
        try:
            cocuk_id = int(cocuk_text.split("(ID: ")[1].split(")")[0])
        except Exception:
            messagebox.showerror("Hata", "Geçersiz çocuk seçimi.")
            return
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            
            # BEP programını çek
            cur.execute("SELECT id FROM bep_programlari WHERE cocuk_id=? AND program_yili=?", (cocuk_id, yil))
            bep_row = cur.fetchone()
            if not bep_row:
                messagebox.showwarning("Uyarı", f"{yil} yılı için BEP programı bulunamadı.")
                conn.close()
                return
            
            bep_id = bep_row[0]
            
            # Çocuk bilgilerini çek
            cur.execute("SELECT ad_soyad, dogum_tarihi FROM danisanlar WHERE id=?", (cocuk_id,))
            cocuk_row = cur.fetchone()
            cocuk_adi = cocuk_row[0]
            dogum_tarihi = cocuk_row[1] or ""
            
            # Hedef becerileri çek (sistemdeki formla aynı: beceri x ay -> durum)
            cur.execute("""
                SELECT ay, hedef_beceri, durum
                FROM bep_hedef_beceriler
                WHERE bep_id=?
            """, (bep_id,))
            
            hedefler = cur.fetchall()
            conn.close()
            
            # (beceri, ay) -> sembol (tablo sıkışık olmasın; açıklama raporda altta)
            durum_sembol = {"planlandi": "P", "devam_ediyor": "D", "tamamlandi": "✓"}
            grid = {}
            for h in hedefler:
                ay, beceri, durum = h[0], h[1], (h[2] or "").strip()
                if 1 <= ay <= 12 and beceri:
                    grid[(beceri, ay)] = durum_sembol.get(durum, "")
            
            # Formdaki sıra ile aynı
            HEDEF_BECERILER_PDF = [
                "Erken Okur Yazarlık Becerisi",
                "Yazı Farkındalığı",
                "Hece Bilgisi",
                "Uyak Farkındalığı",
                "Sesbilişsel Farkındalık",
                "İnce Motor Becerileri",
                "İşitsel ve Görsel Algı Dikkat",
                "Neden Sonuç İlişkisi",
                "Muhakeme Tahmin Etme"
            ]
            AYLAR_KISA = ["Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
            
            # PDF kaydet
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=f"BEP_Raporu_{cocuk_adi}_{yil}.pdf"
            )
            
            if not path:
                return
            
            # PDF oluştur (üst/alt margin: header/footer için)
            doc = SimpleDocTemplate(
                path, pagesize=A4,
                leftMargin=0.8*cm, rightMargin=0.8*cm,
                topMargin=2*cm, bottomMargin=2*cm
            )
            story = []
            styles = getSampleStyleSheet()

            # Türkçe font desteği
            font_name = TURKISH_FONT_NAME or "Helvetica"
            on_first, on_later = _pdf_page_canvas_callbacks("BEP Raporu")

            # Başlık
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=16,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            story.append(Paragraph("BİREYSEL EĞİTİM PROGRAMI (BEP)", title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Çocuk bilgileri
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=11,
                spaceAfter=12
            )
            story.append(Paragraph(f"<b>Çocuk Adı:</b> {cocuk_adi}", info_style))
            story.append(Paragraph(f"<b>Doğum Tarihi:</b> {dogum_tarihi}", info_style))
            story.append(Paragraph(f"<b>Program Yılı:</b> {yil}", info_style))
            story.append(Spacer(1, 0.5*cm))
            
            # Tablo: sistemdeki gibi grid (Hedef Beceriler | Oca | Şub | ... | Ara)
            header_row = ["Hedef Beceriler"] + AYLAR_KISA
            data = [header_row]
            for beceri in HEDEF_BECERILER_PDF:
                row = [beceri]
                for ay in range(1, 13):
                    row.append(grid.get((beceri, ay), ""))
                data.append(row)
            
            # Sütun genişlikleri: beceri geniş, aylar dar
            col_beceri = 4.2*cm
            col_ay = 1.15*cm
            col_widths = [col_beceri] + [col_ay] * 12
            
            table = Table(data, colWidths=col_widths)
            bold_font = font_name
            if PDF_AVAILABLE and pdfmetrics:
                try:
                    registered_fonts = [f[0] for f in pdfmetrics.getRegisteredFontNames()]
                    if f"{font_name}-Bold" in registered_fonts:
                        bold_font = f"{font_name}-Bold"
                except Exception:
                    pass
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), bold_font),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
            
            # Raporun altına sembol açıklaması
            story.append(Spacer(1, 0.6*cm))
            legend_style = ParagraphStyle(
                'BEP_Legend',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=9,
                spaceAfter=0,
                leftIndent=0,
                textColor=colors.HexColor('#333333'),
            )
            story.append(Paragraph(
                "<b>Açıklama:</b> &nbsp; <b>P</b> = Planlandı &nbsp;&nbsp; "
                "<b>D</b> = Devam Ediyor &nbsp;&nbsp; <b>✓</b> = Tamamlandı",
                legend_style
            ))
            
            doc.build(story, onFirstPage=on_first, onLaterPages=on_later)
            messagebox.showinfo("Başarılı", f"BEP raporu PDF olarak oluşturuldu:\n{path}")
        
        except Exception as e:
            messagebox.showerror("Hata", f"Rapor oluşturulamadı:\n{e}")
            log_exception("_bep_rapor_olustur", e)
    
    def _onam_rapor_olustur(self, tree):
        """ONAM formu raporu oluştur (PDF - Form tıpkısı)"""
        if not PDF_AVAILABLE:
            messagebox.showerror(
                "Hata",
                "PDF oluşturma kütüphanesi bulunamadı.\n\n"
                "PowerShell:\n"
                "  py -m pip install -r requirements.txt\n\n"
                "Sonra uygulamayı kapatıp tekrar aç.",
            )
            return
        
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir onam formu seçin.")
            return
        
        try:
            form_id = tree.item(sel[0])["values"][0]
            
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("""
                SELECT danisan_adi, danisan_tarih, terapist_adi, terapist_tarih, 
                       onam_verildi, olusturma_tarihi
                FROM onam_formlari
                WHERE id=?
            """, (form_id,))
            
            row = cur.fetchone()
            conn.close()
            
            if not row:
                messagebox.showwarning("Uyarı", "Onam formu bulunamadı.")
                return
            
            danisan_adi, danisan_tarih, terapist_adi, terapist_tarih, onam_verildi, olusturma_tarihi = row
            
            # PDF kaydet
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=f"ONAM_Formu_{danisan_adi}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
            )
            
            if not path:
                return
            
            # PDF oluştur (üst/alt margin: header/footer için)
            doc = SimpleDocTemplate(
                path, pagesize=A4,
                topMargin=2*cm, bottomMargin=2*cm,
                leftMargin=1.5*cm, rightMargin=1.5*cm
            )
            story = []
            styles = getSampleStyleSheet()
            
            # Türkçe font desteği
            font_name = TURKISH_FONT_NAME or "Helvetica"
            on_first, on_later = _pdf_page_canvas_callbacks("ONAM Formu")
            
            # Başlık
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=16,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            story.append(Paragraph("KİŞİSEL VERİ KORUMA ONAM FORMU", title_style))
            story.append(Spacer(1, 1*cm))
            
            # Form bilgileri
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=11,
                spaceAfter=15,
                leftIndent=1*cm
            )
            
            story.append(Paragraph(f"<b>Danışan Adı:</b> {danisan_adi or ''}", info_style))
            story.append(Paragraph(f"<b>Danışan İmza Tarihi:</b> {danisan_tarih or ''}", info_style))
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(f"<b>Terapist Adı:</b> {terapist_adi or ''}", info_style))
            story.append(Paragraph(f"<b>Terapist İmza Tarihi:</b> {terapist_tarih or ''}", info_style))
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(f"<b>Onam Durumu:</b> {'Onaylandı' if onam_verildi else 'Beklemede'}", info_style))
            story.append(Paragraph(f"<b>Oluşturma Tarihi:</b> {olusturma_tarihi or ''}", info_style))
            story.append(Spacer(1, 1*cm))
            
            # Onam metni
            normal_style = ParagraphStyle(
                'NormalTurkish',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10
            )
            onam_text = """
            Bu form, kişisel verilerin korunması ve işlenmesi konusunda danışanın bilgilendirilmesi ve onayının alınması amacıyla düzenlenmiştir.
            
            Danışan, kişisel verilerinin işlenmesi konusunda bilgilendirilmiş ve bu konuda onay vermiştir.
            """
            story.append(Paragraph(onam_text, normal_style))
            
            doc.build(story, onFirstPage=on_first, onLaterPages=on_later)
            messagebox.showinfo("Başarılı", f"ONAM formu PDF olarak oluşturuldu:\n{path}")
        
        except Exception as e:
            messagebox.showerror("Hata", f"Rapor oluşturulamadı:\n{e}")
            log_exception("_onam_rapor_olustur", e)
    
    def _cocuk_takip_rapor_olustur(self, tree):
        """Çocuk takip formu raporu oluştur (PDF — kağıt form tıpkısı: bölümler, soru/cevap, ( X ) stili)"""
        if not PDF_AVAILABLE:
            messagebox.showerror(
                "Hata",
                "PDF oluşturma kütüphanesi bulunamadı.\n\n"
                "PowerShell:\n"
                "  py -m pip install -r requirements.txt\n\n"
                "Sonra uygulamayı kapatıp tekrar aç.",
            )
            return
        
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir çocuk takip formu seçin.")
            return
        
        try:
            form_id = tree.item(sel[0])["values"][0]
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    d.ad_soyad, d.veli_adi, d.veli_telefon, d.adres,
                    ctf.form_tarihi, ctf.cinsiyet, ctf.dogum_tarihi, ctf.dogum_yeri,
                    ctf.gebelik_sekli, ctf.gebelik_sorun, ctf.dogum_sekli, ctf.dogum_hafta, ctf.dogum_kilo, ctf.dogum_boy,
                    ctf.dogum_sorun, ctf.dogum_sorun_detay, ctf.anne_sutu, ctf.anne_sutu_sure, ctf.bakim_veren,
                    ctf.yurme_yas, ctf.yurme_gec_neden, ctf.tuvalet_yas, ctf.tuvalet_gec_neden, ctf.konusma_yas, ctf.konusma_gec_neden,
                    ctf.gdb_tani, ctf.gdb_tani_detay, ctf.okul_adi, ctf.sinif, ctf.egitim_turu,
                    ctf.destek_egitim, ctf.destek_egitim_sure, ctf.sinif_ogretmen, ctf.okuloncesi, ctf.okuloncesi_yil,
                    ctf.ilkokul_baslangic_ay, ctf.egitim_sorun, ctf.okuma_baslangic,
                    ctf.okuma_sorun, ctf.okuma_sorun_detay, ctf.okuma_anlama_sorun, ctf.okuma_anlama_detay,
                    ctf.yazma_sorun, ctf.yazma_sorun_detay, ctf.aritmetik_sorun, ctf.aritmetik_sorun_detay,
                    ctf.siralama_sorun, ctf.siralama_sorun_detay, ctf.yon_ayirt_sorun, ctf.yon_ayirt_detay,
                    ctf.karneturkce, ctf.karnematematik, ctf.karnehayatbilgisi, ctf.karnesosyal, ctf.karnefen,
                    ctf.aile_sira, ctf.akrabalik, ctf.akrabalik_detay, ctf.bakim_veren_suan, ctf.aile_disinda_yasayan,
                    ctf.aile_turu, ctf.ayrilik_durum, ctf.sosyoekonomik, ctf.anne_egitim, ctf.anne_yas, ctf.anne_is,
                    ctf.baba_egitim, ctf.baba_yas, ctf.baba_is, ctf.cocuk_sayisi_detay, ctf.hasta_kardes, ctf.hasta_kardes_detay,
                    ctf.olusturma_tarihi
                FROM cocuk_takip_bilgi_formlari ctf
                LEFT JOIN danisanlar d ON ctf.danisan_id = d.id
                WHERE ctf.id=?
            """, (form_id,))
            row = cur.fetchone()
            conn.close()
            if not row:
                messagebox.showwarning("Uyarı", "Çocuk takip formu bulunamadı.")
                return
            
            def _s(v):
                return (v or "").strip() if v is not None else ""
            def _evet_hayir(n):
                if n is None: return ""
                return "Evet" if n == 1 else "Hayır"
            def _radio(secili, secenekler):
                # secenekler = [("Planlı", "Planlı"), ("Plansız", "Plansız")] -> "( X ) Planlı   ( ) Plansız"
                out = []
                for val, lbl in secenekler:
                    x = "( X )" if _s(secili) == val or _s(secili) == lbl else "( )"
                    out.append(f"{x} {lbl}")
                return "   ".join(out)
            
            row_list = [row[i] if i < len(row) else None for i in range(73)]
            (danisan_adi, veli_adi, veli_telefon, adres,
             form_tarihi, cinsiyet, dogum_tarihi, dogum_yeri,
             gebelik_sekli, gebelik_sorun, dogum_sekli, dogum_hafta, dogum_kilo, dogum_boy,
             dogum_sorun, dogum_sorun_detay, anne_sutu, anne_sutu_sure, bakim_veren,
             yurme_yas, yurme_gec_neden, tuvalet_yas, tuvalet_gec_neden, konusma_yas, konusma_gec_neden,
             gdb_tani, gdb_tani_detay, okul_adi, sinif, egitim_turu,
             destek_egitim, destek_egitim_sure, sinif_ogretmen, okuloncesi, okuloncesi_yil,
             ilkokul_baslangic_ay, egitim_sorun, okuma_baslangic,
             okuma_sorun, okuma_sorun_detay, okuma_anlama_sorun, okuma_anlama_detay,
             yazma_sorun, yazma_sorun_detay, aritmetik_sorun, aritmetik_sorun_detay,
             siralama_sorun, siralama_sorun_detay, yon_ayirt_sorun, yon_ayirt_detay,
             karneturkce, karnematematik, karnehayatbilgisi, karnesosyal, karnefen,
             aile_sira, akrabalik, akrabalik_detay, bakim_veren_suan, aile_disinda_yasayan,
             aile_turu, ayrilik_durum, sosyoekonomik, anne_egitim, anne_yas, anne_is,
             baba_egitim, baba_yas, baba_is, cocuk_sayisi_detay, hasta_kardes, hasta_kardes_detay
             ) = row_list[:72]
            olusturma_tarihi = row_list[72] if len(row_list) > 72 else None
            
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=f"Cocuk_Takip_Formu_{_s(danisan_adi)}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
            )
            if not path:
                return
            
            doc = SimpleDocTemplate(
                path, pagesize=A4,
                leftMargin=1.5*cm, rightMargin=1.5*cm,
                topMargin=2*cm, bottomMargin=2*cm
            )
            story = []
            styles = getSampleStyleSheet()
            font_name = TURKISH_FONT_NAME or "Helvetica"
            on_first, on_later = _pdf_page_canvas_callbacks("Çocuk Takip Bilgi Formu")
            
            title_style = ParagraphStyle(
                'FormTitle', parent=styles['Heading1'], fontName=font_name, fontSize=16,
                textColor=colors.HexColor('#1a1a1a'), spaceAfter=20, alignment=TA_CENTER
            )
            sect_style = ParagraphStyle(
                'Section', parent=styles['Heading2'], fontName=font_name, fontSize=12,
                textColor=colors.HexColor('#333333'), spaceBefore=18, spaceAfter=10, leftIndent=0
            )
            q_style = ParagraphStyle(
                'Question', parent=styles['Normal'], fontName=font_name, fontSize=10,
                spaceAfter=4, leftIndent=0
            )
            a_style = ParagraphStyle(
                'Answer', parent=styles['Normal'], fontName=font_name, fontSize=10,
                spaceAfter=12, leftIndent=1*cm, textColor=colors.HexColor('#222222')
            )
            
            story.append(Paragraph("ÇOCUK TAKİP BİLGİ FORMU", title_style))
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph(f"<b>Adı Soyadı:</b> {_s(danisan_adi)}", a_style))
            story.append(Paragraph(f"<b>Form Tarihi:</b> {_s(form_tarihi)}", a_style))
            story.append(Spacer(1, 0.3*cm))
            
            # —— Sayfa 1: Doğum ve Gelişim ——
            story.append(Paragraph("Doğum Süreci ve Doğum Sonrasına İlişkin Bilgiler", sect_style))
            story.append(Paragraph("Cinsiyeti:", q_style))
            story.append(Paragraph(_radio(cinsiyet, [("Kız", "Kız"), ("Erkek", "Erkek")]), a_style))
            story.append(Paragraph("Doğum Tarihi (gg/aa/yyyy):", q_style))
            story.append(Paragraph(_s(dogum_tarihi), a_style))
            story.append(Paragraph("Doğum Yeri:", q_style))
            story.append(Paragraph(_s(dogum_yeri), a_style))
            story.append(Paragraph("Gebelik Şekli:", q_style))
            story.append(Paragraph(_radio(gebelik_sekli, [("Planlı", "Planlı"), ("Plansız", "Plansız")]), a_style))
            story.append(Paragraph("Gebelik döneminde sorun olduysa belirtiniz:", q_style))
            story.append(Paragraph(_s(gebelik_sorun), a_style))
            story.append(Paragraph("Doğum Şekli:", q_style))
            story.append(Paragraph(_radio(dogum_sekli, [
                ("Normal", "Normal"), ("Sezaryen", "Sezaryen"), ("Müdahaleli", "Müdahaleli-Vakum"), ("Diğer", "Diğer")
            ]), a_style))
            story.append(Paragraph("Kaç haftalık doğdu / Doğum kilosu / Doğum boyu:", q_style))
            story.append(Paragraph(f"{dogum_hafta or ''} hafta — {dogum_kilo or ''} kg — {dogum_boy or ''} cm", a_style))
            story.append(Paragraph("Çocuğunuz doğumun esnasında ya da hemen sonrasında bir problem yaşadı mı?", q_style))
            story.append(Paragraph(_radio(_evet_hayir(dogum_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]), a_style))
            story.append(Paragraph("Yaşadıysa bunlar nelerdir? (Morarma, Havale, Enfeksiyon vb.)", q_style))
            story.append(Paragraph(_s(dogum_sorun_detay), a_style))
            story.append(Paragraph("Anne sütü alma durumu: (Almadı / Aldı — Aldıysa alma süresini yazınız)", q_style))
            story.append(Paragraph(
                _radio("Aldı" if anne_sutu == 1 else "Almadı", [("Almadı", "Almadı"), ("Aldı", "Aldı")]) + (" — " + _s(anne_sutu_sure) if _s(anne_sutu_sure) else ""),
                a_style
            ))
            story.append(Paragraph("Bebekliğinin ilk yıllarında temel bakım veren kişi veya kişiler:", q_style))
            story.append(Paragraph(_s(bakim_veren), a_style))
            
            story.append(Paragraph("Temel Gelişim Bilgileri", sect_style))
            story.append(Paragraph("Çocuğunuz kaç yaşında yürüdü:", q_style))
            story.append(Paragraph(_s(yurme_yas), a_style))
            story.append(Paragraph("Geç yürüdüyse nedenlerini belirtiniz:", q_style))
            story.append(Paragraph(_s(yurme_gec_neden), a_style))
            story.append(Paragraph("Çocuğunuz kaç yaşında tuvalet eğitimini tamamladı:", q_style))
            story.append(Paragraph(_s(tuvalet_yas), a_style))
            story.append(Paragraph("Geç tamamladıysa nedenlerini belirtiniz:", q_style))
            story.append(Paragraph(_s(tuvalet_gec_neden), a_style))
            story.append(Paragraph("Çocuğunuz kaç yaşında konuştu:", q_style))
            story.append(Paragraph(_s(konusma_yas), a_style))
            story.append(PageBreak())
            
            # —— Sayfa 2: Eğitim ——
            story.append(Paragraph("Sayfa 2: Eğitim Bilgileri", sect_style))
            story.append(Paragraph("Geç konuştuysa nedenlerini belirtiniz:", q_style))
            story.append(Paragraph(_s(konusma_gec_neden), a_style))
            story.append(Paragraph("Gelişim döneminde yaygın gelişimsel bozukluk tanısı aldı mı?", q_style))
            story.append(Paragraph(_radio(_evet_hayir(gdb_tani), [("Evet", "Evet"), ("Hayır", "Hayır")]), a_style))
            story.append(Paragraph("Yanıtınız evetse hangi tanıyı almış belirtiniz:", q_style))
            story.append(Paragraph(_s(gdb_tani_detay), a_style))
            story.append(Paragraph("Eğitim Bilgileri", sect_style))
            story.append(Paragraph("Okul Adı/İl/İlçe:", q_style))
            story.append(Paragraph(_s(okul_adi), a_style))
            story.append(Paragraph("Sınıfı:", q_style))
            story.append(Paragraph(_s(sinif), a_style))
            story.append(Paragraph("Alınan Eğitim türü: Zorunlu(örgün) temel eğitim / Özel eğitim / Her ikisi", q_style))
            story.append(Paragraph(_s(egitim_turu), a_style))
            story.append(Paragraph("Örgün eğitim haricinde aldığı eğitim desteği var mı? Evet ise süresini yazınız.", q_style))
            story.append(Paragraph(
                _radio("Evet" if destek_egitim == 1 else "Hayır", [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(destek_egitim_sure) if _s(destek_egitim_sure) else ""),
                a_style
            ))
            story.append(Paragraph("Sınıf (mentör/danışman) öğretmeninin adı soyadı:", q_style))
            story.append(Paragraph(_s(sinif_ogretmen), a_style))
            story.append(Paragraph("Okul öncesi eğitim aldı mı? Evetse kaç yıl yazınız.", q_style))
            story.append(Paragraph(
                _radio("Evet" if okuloncesi == 1 else "Hayır", [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + str(okuloncesi_yil or "") + " yıl" if okuloncesi_yil else ""),
                a_style
            ))
            story.append(Paragraph("İlkokula başlama yaşını ay olarak yazınız:", q_style))
            story.append(Paragraph(str(ilkokul_baslangic_ay) if ilkokul_baslangic_ay is not None else "", a_style))
            story.append(Paragraph("Eğitim olanaklarına ilişkin sorunlar yaşandı mı?", q_style))
            story.append(Paragraph(_s(egitim_sorun), a_style))
            story.append(Paragraph("Okumaya başlama zamanı: Okul öncesi / Birinci dönem / İkinci dönem / Daha sonraki dönemler / Okuyamıyor", q_style))
            story.append(Paragraph(_s(okuma_baslangic), a_style))
            story.append(Paragraph("Okumada sorunu var mı? (örn. Harf atlama, ters yazma vb.) Varsa belirtiniz.", q_style))
            story.append(Paragraph(_radio(_evet_hayir(okuma_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(okuma_sorun_detay) if _s(okuma_sorun_detay) else ""), a_style))
            story.append(Paragraph("Okuduğunu anlamada sorunu var mı? Varsa belirtiniz.", q_style))
            story.append(Paragraph(_radio(_evet_hayir(okuma_anlama_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(okuma_anlama_detay) if _s(okuma_anlama_detay) else ""), a_style))
            story.append(Paragraph("Yazmada sorunu var mı? Varsa belirtiniz.", q_style))
            story.append(Paragraph(_radio(_evet_hayir(yazma_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(yazma_sorun_detay) if _s(yazma_sorun_detay) else ""), a_style))
            story.append(Paragraph("Aritmetikte sorunu var mı? Varsa belirtiniz.", q_style))
            story.append(Paragraph(_radio(_evet_hayir(aritmetik_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(aritmetik_sorun_detay) if _s(aritmetik_sorun_detay) else ""), a_style))
            story.append(Paragraph("Sıralamada sorunu var mı? Varsa belirtiniz.", q_style))
            story.append(Paragraph(_radio(_evet_hayir(siralama_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(siralama_sorun_detay) if _s(siralama_sorun_detay) else ""), a_style))
            story.append(PageBreak())
            
            # —— Sayfa 3: Demografik ——
            story.append(Paragraph("Sayfa 3: Demografik Bilgiler", sect_style))
            story.append(Paragraph("Yönleri ayırt etmede sorunu var mı? (örn. Sağını solunu karıştırma) Varsa belirtiniz.", q_style))
            story.append(Paragraph(_radio(_evet_hayir(yon_ayirt_sorun), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(yon_ayirt_detay) if _s(yon_ayirt_detay) else ""), a_style))
            story.append(Paragraph("Aşağıdaki alanlara en son karne notlarını yazınız:", q_style))
            story.append(Paragraph(
                f"Türkçe: {_s(karneturkce)}   Matematik: {_s(karnematematik)}   Hayat Bilgisi: {_s(karnehayatbilgisi)}   Sosyal: {_s(karnesosyal)}   Fen: {_s(karnefen)}",
                a_style
            ))
            story.append(Paragraph("Demografik Bilgiler", sect_style))
            story.append(Paragraph("Ailenin kaçıncı çocuğu:", q_style))
            story.append(Paragraph(str(aile_sira) if aile_sira is not None else "", a_style))
            story.append(Paragraph("Anne baba arasında akrabalık var mı?", q_style))
            story.append(Paragraph(_radio("Evet" if akrabalik == 1 else "Hayır", [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(akrabalik_detay) if _s(akrabalik_detay) else ""), a_style))
            story.append(Paragraph("Çocuğa şuan temel bakım veren kişi veya kişiler:", q_style))
            story.append(Paragraph(_s(bakim_veren_suan), a_style))
            story.append(Paragraph("Ailede anne, baba ve çocuklar dışında yaşayan biri var mı? Varsa kim olduğunu yazınız:", q_style))
            story.append(Paragraph(_s(aile_disinda_yasayan), a_style))
            story.append(Paragraph("Aile Türü: Çekirdek / Geniş / Sadece anne / Sadece baba", q_style))
            story.append(Paragraph(_s(aile_turu), a_style))
            story.append(Paragraph("Yukarıdaki maddede son iki şıktan biri seçilmişse: Anne ve baba ayrı yaşıyor / Boşanmış / Anne veya babadan biri ölmüş", q_style))
            story.append(Paragraph(_s(ayrilik_durum), a_style))
            story.append(Paragraph("Ailenin sosyoekonomik düzeyi: Alt / Orta / Üst", q_style))
            story.append(Paragraph(_s(sosyoekonomik), a_style))
            story.append(Paragraph("Annenin eğitim durumu nedir?", q_style))
            story.append(Paragraph(_s(anne_egitim), a_style))
            story.append(Paragraph("Annenin yaşı nedir?", q_style))
            story.append(Paragraph(str(anne_yas) if anne_yas is not None else "", a_style))
            story.append(Paragraph("Annenin yaptığı iş nedir?", q_style))
            story.append(Paragraph(_s(anne_is), a_style))
            story.append(Paragraph("Babanın eğitim durumu nedir?", q_style))
            story.append(Paragraph(_s(baba_egitim), a_style))
            story.append(Paragraph("Babanın yaşı nedir?", q_style))
            story.append(Paragraph(str(baba_yas) if baba_yas is not None else "", a_style))
            story.append(Paragraph("Babanın yaptığı iş nedir?", q_style))
            story.append(Paragraph(_s(baba_is), a_style))
            story.append(Paragraph("Ailedeki çocuk sayısı (yaşları, cinsiyetleri ve kaçıncı çocuk olduklarını yazınız)", q_style))
            story.append(Paragraph(_s(cocuk_sayisi_detay), a_style))
            story.append(Paragraph("Hastalık tanısı almış kardeşi var mı?", q_style))
            story.append(Paragraph(_radio(_evet_hayir(hasta_kardes), [("Evet", "Evet"), ("Hayır", "Hayır")]) + (" — " + _s(hasta_kardes_detay) if _s(hasta_kardes_detay) else ""), a_style))
            story.append(Spacer(1, 0.5*cm))
            story.append(Paragraph(f"<b>Veli Adı:</b> {_s(veli_adi)}   <b>Veli Telefon:</b> {_s(veli_telefon)}", a_style))
            story.append(Paragraph(f"<b>Adres:</b> {_s(adres)}", a_style))
            story.append(Paragraph(f"<b>Oluşturma Tarihi:</b> {_s(olusturma_tarihi)}", a_style))
            
            doc.build(story, onFirstPage=on_first, onLaterPages=on_later)
            messagebox.showinfo("Başarılı", f"Çocuk takip formu PDF olarak oluşturuldu:\n{path}")
        
        except Exception as e:
            messagebox.showerror("Hata", f"Rapor oluşturulamadı:\n{e}")
            log_exception("_cocuk_takip_rapor_olustur", e)
    
    def _danisan_hoca_fiyatlandirma_kurulum(self):
        """Danışan-Hoca bazlı fiyatlandırma kurulum penceresi"""
        win = ttk.Toplevel(self)
        win.title("Danışan-Hoca Fiyatlandırma Kurulumu")
        win.geometry("800x600")
        center_window(win, 800, 600)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text="DANIŞAN-HOCA BAZLI FİYATLANDIRMA KURULUMU", 
                 font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        ttk.Label(wrapper, text="Örnek: Danışan X, Hoca A'da 3500 TL, Hoca B'de 3800 TL", 
                 font=("Segoe UI", 10), foreground="gray").pack(pady=(0, 20))
        
        # Danışan seçimi
        danisan_frame = ttk.Frame(wrapper)
        danisan_frame.pack(fill=X, pady=10)
        ttk.Label(danisan_frame, text="Danışan:", font=("Segoe UI", 10)).pack(side=LEFT, padx=5)
        cmb_danisan = ttk.Combobox(danisan_frame, state="readonly", width=30)
        cmb_danisan.pack(side=LEFT, padx=5)
        
        # Danışan listesini yükle
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT id, ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            danisan_listesi = [(row[0], row[1]) for row in cur.fetchall()]
            conn.close()
            cmb_danisan["values"] = [f"{d[1]} (ID: {d[0]})" for d in danisan_listesi]
        except Exception:
            cmb_danisan["values"] = []
        
        # Hoca ve fiyat girişi
        fiyat_frame = ttk.Labelframe(wrapper, text="Hoca ve Fiyat Bilgileri", padding=10)
        fiyat_frame.pack(fill=BOTH, expand=True, pady=10)
        
        # Hoca listesi
        ttk.Label(fiyat_frame, text="Hoca:", font=("Segoe UI", 10)).pack(anchor=W, pady=5)
        cmb_hoca = ttk.Combobox(fiyat_frame, state="readonly", width=30)
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            hoca_listesi = [row[0] for row in cur.fetchall()]
            conn.close()
            cmb_hoca["values"] = hoca_listesi
        except Exception:
            cmb_hoca["values"] = []
        cmb_hoca.pack(fill=X, pady=5)
        
        ttk.Label(fiyat_frame, text="Fiyat (₺):", font=("Segoe UI", 10)).pack(anchor=W, pady=5)
        ent_fiyat = ttk.Entry(fiyat_frame, width=30)
        ent_fiyat.pack(fill=X, pady=5)
        
        # Kaydedilen fiyatlar listesi
        list_frame = ttk.Labelframe(wrapper, text="Kaydedilen Fiyatlandırmalar", padding=10)
        list_frame.pack(fill=BOTH, expand=True, pady=10)
        
        cols = ("Danışan", "Hoca", "Fiyat (₺)")
        tree_fiyat = ttk.Treeview(list_frame, columns=cols, show="headings", style="Strong.Treeview")
        for c in cols:
            tree_fiyat.heading(c, text=c)
            tree_fiyat.column(c, width=200)
        tree_fiyat.pack(side=LEFT, fill=BOTH, expand=True)
        sb_fiyat = ttk.Scrollbar(list_frame, orient=VERTICAL, command=tree_fiyat.yview)
        tree_fiyat.configure(yscroll=sb_fiyat.set)
        sb_fiyat.pack(side=RIGHT, fill=Y)
        
        def fiyatlari_yukle():
            for iid in tree_fiyat.get_children():
                tree_fiyat.delete(iid)
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute("""
                    SELECT d.ad_soyad, pp.teacher_name, pp.price
                    FROM pricing_policy pp
                    LEFT JOIN danisanlar d ON pp.student_id = d.id
                    ORDER BY d.ad_soyad, pp.teacher_name
                """)
                for row in cur.fetchall():
                    tree_fiyat.insert("", END, values=(row[0] or "", row[1] or "", format_money(row[2] or 0)))
                conn.close()
            except Exception as e:
                log_exception("fiyatlari_yukle", e)
        
        def fiyat_kaydet():
            danisan_text = cmb_danisan.get()
            hoca = cmb_hoca.get()
            fiyat_str = ent_fiyat.get().strip()
            
            if not danisan_text or not hoca or not fiyat_str:
                messagebox.showwarning("Uyarı", "Lütfen tüm alanları doldurun.")
                return
            
            try:
                danisan_id = int(danisan_text.split("(ID: ")[1].split(")")[0])
                fiyat = parse_money(fiyat_str)
            except Exception:
                messagebox.showerror("Hata", "Geçersiz değerler.")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    "INSERT OR REPLACE INTO pricing_policy (student_id, teacher_name, price, created_at) VALUES (?, ?, ?, ?)",
                    (danisan_id, hoca, fiyat, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                )
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Başarılı", f"Fiyatlandırma kaydedildi: {danisan_text.split('(')[0].strip()} - {hoca}: {format_money(fiyat)}")
                ent_fiyat.delete(0, END)
                fiyatlari_yukle()
            except Exception as e:
                messagebox.showerror("Hata", f"Fiyatlandırma kaydedilemedi:\n{e}")
                log_exception("fiyat_kaydet", e)
        
        # Butonlar
        btn_frame = ttk.Frame(wrapper)
        btn_frame.pack(fill=X, pady=10)
        ttk.Button(btn_frame, text="💾 Fiyatlandırma Kaydet", bootstyle="success", command=fiyat_kaydet).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 Yenile", bootstyle="secondary", command=fiyatlari_yukle).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="✅ Tamam", bootstyle="primary", command=win.destroy).pack(side=RIGHT, padx=5)
        
        # İlk yükleme
        fiyatlari_yukle()
    
    # Sistem Şifreleri fonksiyonları - KALDIRILDI (Kullanıcı isteği)
    
    def _sifre_duzenle(self, parent, tree):
        """Sistem şifresi düzenle"""
        sel = tree.selection()
        if not sel:
            return
        
        sifre_id = tree.item(sel[0])["values"][0]
        
        # ✅ DÜZELTME: Gerçek şifre düzenleme penceresi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT platform_adi, kullanici_adi, sifre, aciklama, olusturma_tarihi
                FROM sistem_sifreleri WHERE id = ?
                """,
                (sifre_id,)
            )
            row = cur.fetchone()
            conn.close()
            
            if not row:
                messagebox.showerror("Hata", "Şifre kaydı bulunamadı.")
                return
            
            win = ttk.Toplevel(self)
            win.title(f"Şifre Düzenle - {row[0]}")
            center_window_smart(win, 500, 400, max_ratio=0.9)
            win.transient(self)
            win.grab_set()
            self._brand_window(win)
            
            wrapper = ttk.Frame(win, padding=20)
            wrapper.pack(fill=BOTH, expand=True)
            
            ttk.Label(wrapper, text=f"Şifre Düzenle: {row[0]}", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
            
            ttk.Label(wrapper, text="Platform Adı:").pack(anchor=W, pady=(5, 0))
            ent_platform = ttk.Entry(wrapper, width=40)
            ent_platform.insert(0, row[0] or "")
            ent_platform.pack(fill=X, pady=5)
            
            ttk.Label(wrapper, text="Kullanıcı Adı:").pack(anchor=W, pady=(10, 0))
            ent_kullanici = ttk.Entry(wrapper, width=40)
            ent_kullanici.insert(0, row[1] or "")
            ent_kullanici.pack(fill=X, pady=5)
            
            ttk.Label(wrapper, text="Şifre:").pack(anchor=W, pady=(10, 0))
            ent_sifre = ttk.Entry(wrapper, width=40, show="*")
            ent_sifre.insert(0, row[2] or "")
            ent_sifre.pack(fill=X, pady=5)
            
            ttk.Label(wrapper, text="Açıklama:").pack(anchor=W, pady=(10, 0))
            txt_aciklama = ttk.Text(wrapper, height=3, wrap=WORD)
            txt_aciklama.insert("1.0", row[3] or "")
            txt_aciklama.pack(fill=X, pady=5)
            
            def _kaydet():
                try:
                    conn = self.veritabani_baglan()
                    cur = conn.cursor()
                    cur.execute(
                        """
                        UPDATE sistem_sifreleri
                        SET platform_adi=?, kullanici_adi=?, sifre=?, aciklama=?, guncelleme_tarihi=?
                        WHERE id=?
                        """,
                        (
                            (ent_platform.get() or "").strip(),
                            (ent_kullanici.get() or "").strip(),
                            (ent_sifre.get() or "").strip(),
                            (txt_aciklama.get("1.0", END) or "").strip(),
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            sifre_id
                        )
                    )
                    conn.commit()
                    conn.close()
                    
                    messagebox.showinfo("Başarılı", "Şifre kaydı güncellendi!")
                    win.destroy()
                    if hasattr(parent, "_reload"):
                        parent._reload()
                except Exception as e:
                    messagebox.showerror("Hata", f"Şifre güncellenemedi:\n{e}")
                    log_exception("_sifre_duzenle", e)
            
            ttk.Button(wrapper, text="💾 Kaydet", bootstyle="success", command=_kaydet).pack(pady=20)
            
        except Exception as e:
            messagebox.showerror("Hata", f"Şifre bilgileri yüklenemedi:\n{e}")
            log_exception("_sifre_duzenle", e)
    
    def _sifre_sil(self, parent, tree):
        """Sistem şifresi sil"""
        sel = tree.selection()
        if not sel:
            return
        
        sifre_id = tree.item(sel[0])["values"][0]
        
        if not messagebox.askyesno(
            "Onay",
            f"{danisan_adi} danışanını listeden kaldırmak istediğinize emin misiniz?\n\n"
            "Bu işlem danışanı pasife alır (aktif=0).",
        ):
            return
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("DELETE FROM sistem_sifreleri WHERE id = ?", (sifre_id,))
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Başarılı", "Şifre kaydı silindi.")
            self._sifreler_listele(parent)
        
        except Exception as e:
            messagebox.showerror("Hata", f"Şifre silinemedi:\n{e}")
            log_exception("_sifre_sil", e)
    
    def _sifre_kopyala(self, tree):
        """Şifreyi panoya kopyala"""
        sel = tree.selection()
        if not sel:
            return
        
        sifre = tree.item(sel[0])["values"][3]
        if sifre:
            self.clipboard_clear()
            self.clipboard_append(sifre)
            messagebox.showinfo("Başarılı", "Şifre panoya kopyalandı!")
    
    def _build_settings_tab(self):
        # ✅ SMART LOGS: Kurum Müdürü için Sistem Günlüğü arayüzü
        if self.kullanici_yetki == "kurum_muduru":
            # Notebook ile iki sekme: Terapist Listesi ve Sistem Günlüğü
            settings_nb = ttk.Notebook(self.tab_settings)
            settings_nb.pack(fill=BOTH, expand=True, padx=10, pady=10)
            
            # Sekme 1: Terapist Listesi
            tab_terapist = ttk.Frame(settings_nb, padding=10)
            settings_nb.add(tab_terapist, text="Terapist Listesi")
            
            box = ttk.Labelframe(tab_terapist, text="Terapist Listesi", padding=12)
            box.pack(fill=BOTH, expand=True)
        else:
            # Eğitim Görevlisi için sadece terapist listesi
            box = ttk.Labelframe(self.tab_settings, text="Terapist Listesi", padding=12)
            box.pack(fill=BOTH, expand=True)

        self.lst = ttk.Treeview(box, columns=("Terapist", "Rol"), show="headings", height=14)
        self.lst.heading("Terapist", text="Terapist")
        self.lst.heading("Rol", text="Rol")
        self.lst.column("Terapist", width=220)
        self.lst.column("Rol", width=280)
        self.lst.pack(side=LEFT, fill=BOTH, expand=True)

        sb = ttk.Scrollbar(box, orient=VERTICAL, command=self.lst.yview)
        self.lst.configure(yscroll=sb.set)
        sb.pack(side=LEFT, fill=Y)

        right = ttk.Frame(box, padding=(12, 0))
        right.pack(side=LEFT, fill=Y)
        ttk.Label(right, text="Yeni Terapist:").pack(anchor=W, pady=(0, 4))
        self.ent_yeni = ttk.Entry(right, width=28)
        self.ent_yeni.pack(fill=X, pady=(0, 8))
        ttk.Label(right, text="Rol:").pack(anchor=W, pady=(0, 4))
        self.cmb_rol = ttk.Combobox(
            right,
            state="readonly",
            values=[
                "Kurum Müdürü / Özel Eğitim Uzmanı",
                "Özel Eğitim Uzmanı",
                "Ergoterapist",
                "Dil ve Konuşma Terapisti",
                "Klinik Psikolog",
                "Diğer",
            ],
            width=26,
        )
        self.cmb_rol.set("Diğer")
        self.cmb_rol.pack(fill=X, pady=(0, 10))
        ttk.Button(right, text="Ekle", bootstyle="success", command=self.terapist_ekle).pack(fill=X, pady=(0, 6))
        ttk.Button(right, text="Rol Güncelle", bootstyle="warning", command=self.terapist_rol_guncelle).pack(fill=X, pady=(0, 6))
        ttk.Button(right, text="Seçiliyi Sil", bootstyle="danger", command=self.terapist_sil).pack(fill=X, pady=(0, 6))
        ttk.Button(right, text="Excel'e Aktar", bootstyle="primary", command=self.excel_aktar).pack(fill=X, pady=(0, 12))
        
        # Kurum müdürü için bu bölüm sadece logo yükleme olarak sadeleştirildi
        if self.kullanici_yetki == "kurum_muduru":
            ttk.Separator(right, orient=HORIZONTAL).pack(fill=X, pady=(0, 12))
            ttk.Label(right, text="Kurum Görseli", font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(0, 6))
            ttk.Button(right, text="🖼️ Logo Yükle/Değiştir", bootstyle="info", command=self.logo_yukle_degistir, width=28).pack(fill=X, pady=(0, 6))
            ttk.Label(right, text="Ayarlar içinden sadece logo yönetimi yapılır.", font=("Segoe UI", 8), foreground="gray", wraplength=200).pack(anchor=W, pady=(0, 12))

        self.lst.bind("<<TreeviewSelect>>", self._on_terapist_select)
        
        # ✅ SMART LOGS: Sistem Günlüğü sekmesi (sadece Kurum Müdürü için)
        if self.kullanici_yetki == "kurum_muduru":
            tab_logs = ttk.Frame(settings_nb, padding=10)
            settings_nb.add(tab_logs, text="📋 Sistem Günlüğü")
            
            # Filtreleme paneli
            filter_frame = ttk.Labelframe(tab_logs, text="Filtreleme", padding=10, bootstyle="secondary")
            filter_frame.pack(fill=X, pady=(0, 10))
            
            filter_row1 = ttk.Frame(filter_frame)
            filter_row1.pack(fill=X, pady=(0, 8))
            
            ttk.Label(filter_row1, text="Başlangıç Tarihi:").pack(side=LEFT, padx=(0, 5))
            tarih_bas_var = ttk.StringVar(value=(datetime.datetime.now() - datetime.timedelta(days=7)).strftime("%Y-%m-%d"))
            ttk.Entry(filter_row1, textvariable=tarih_bas_var, width=12).pack(side=LEFT, padx=(0, 15))
            
            ttk.Label(filter_row1, text="Bitiş Tarihi:").pack(side=LEFT, padx=(0, 5))
            tarih_bit_var = ttk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
            ttk.Entry(filter_row1, textvariable=tarih_bit_var, width=12).pack(side=LEFT, padx=(0, 15))
            
            ttk.Label(filter_row1, text="Kullanıcı:").pack(side=LEFT, padx=(0, 5))
            kullanici_filter_var = ttk.StringVar(value="(Tümü)")
            kullanici_cmb = ttk.Combobox(filter_row1, textvariable=kullanici_filter_var, state="readonly", width=20)
            kullanici_cmb.pack(side=LEFT, padx=(0, 15))
            
            # Kullanıcı listesini doldur
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute("SELECT DISTINCT username FROM users WHERE is_active=1 ORDER BY username")
                kullanicilar = ["(Tümü)"] + [r[0] for r in cur.fetchall()]
                kullanici_cmb["values"] = kullanicilar
                conn.close()
            except Exception:
                kullanici_cmb["values"] = ["(Tümü)"]
            
            ttk.Label(filter_row1, text="İşlem Tipi:").pack(side=LEFT, padx=(0, 5))
            action_type_var = ttk.StringVar(value="(Tümü)")
            action_type_cmb = ttk.Combobox(filter_row1, textvariable=action_type_var, state="readonly", width=20)
            action_type_cmb["values"] = [
                "(Tümü)",
                "seans_kayit",
                "odeme_ekle",
                "kayit_sil",
                "maas_hesaplandi",
                "maas_guncellendi",
                "kasa_hareketi_sil",
                "seans_not_guncelle",
                "danisan_durum_guncelle",
                "oda_durum_guncelle"
            ]
            action_type_cmb.pack(side=LEFT, padx=(0, 15))
            
            # Log tablosu
            logs_frame = ttk.Frame(tab_logs)
            logs_frame.pack(fill=BOTH, expand=True)
            
            logs_tree = ttk.Treeview(logs_frame, columns=("Tarih", "Kullanıcı", "İşlem", "Varlık", "Detaylar"), 
                                    show="headings", height=20)
            logs_tree.heading("Tarih", text="Tarih/Saat")
            logs_tree.heading("Kullanıcı", text="Kullanıcı")
            logs_tree.heading("İşlem", text="İşlem Tipi")
            logs_tree.heading("Varlık", text="Varlık Tipi")
            logs_tree.heading("Detaylar", text="Detaylar")
            logs_tree.column("Tarih", width=150)
            logs_tree.column("Kullanıcı", width=120)
            logs_tree.column("İşlem", width=150)
            logs_tree.column("Varlık", width=120)
            logs_tree.column("Detaylar", width=400)
            
            logs_sb = ttk.Scrollbar(logs_frame, orient=VERTICAL, command=logs_tree.yview)
            logs_tree.configure(yscroll=logs_sb.set)
            logs_sb.pack(side=RIGHT, fill=Y)
            logs_tree.pack(side=LEFT, fill=BOTH, expand=True)
            
            def logs_yukle():
                """Audit trail kayıtlarını yükle"""
                for iid in logs_tree.get_children():
                    logs_tree.delete(iid)
                
                try:
                    conn = self.veritabani_baglan()
                    cur = conn.cursor()
                    
                    # Filtreleme sorgusu oluştur
                    where_clauses = []
                    params = []
                    
                    tarih_bas = tarih_bas_var.get().strip()
                    tarih_bit = tarih_bit_var.get().strip()
                    kullanici = kullanici_filter_var.get()
                    action_type = action_type_var.get()
                    
                    if tarih_bas:
                        where_clauses.append("DATE(at.olusturma_tarihi) >= ?")
                        params.append(tarih_bas)
                    
                    if tarih_bit:
                        where_clauses.append("DATE(at.olusturma_tarihi) <= ?")
                        params.append(tarih_bit)
                    
                    if kullanici and kullanici != "(Tümü)":
                        where_clauses.append("u.username = ?")
                        params.append(kullanici)
                    
                    if action_type and action_type != "(Tümü)":
                        where_clauses.append("at.action_type = ?")
                        params.append(action_type)
                    
                    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
                    
                    query = f"""
                        SELECT 
                            at.olusturma_tarihi,
                            COALESCE(u.username, 'Sistem') as kullanici,
                            at.action_type,
                            at.entity_type,
                            at.details
                        FROM audit_trail at
                        LEFT JOIN users u ON at.kullanici_id = u.id
                        WHERE {where_sql}
                        ORDER BY at.olusturma_tarihi DESC
                        LIMIT 500
                    """
                    
                    cur.execute(query, params)
                    rows = cur.fetchall()
                    conn.close()
                    
                    for row in rows:
                        tarih, kullanici_adi, action_type_val, entity_type_val, details_json = row
                        
                        # Details JSON'dan özet çıkar
                        detay_text = ""
                        try:
                            if details_json:
                                details = json.loads(details_json)
                                # Önemli alanları göster
                                if isinstance(details, dict):
                                    key_fields = ["danisan_adi", "terapist", "tutar", "personel_adi", "seans_id"]
                                    detay_list = []
                                    for key in key_fields:
                                        if key in details:
                                            detay_list.append(f"{key}: {details[key]}")
                                    detay_text = " | ".join(detay_list[:3])  # İlk 3 alan
                        except Exception:
                            detay_text = details_json[:50] if details_json else ""
                        
                        logs_tree.insert("", END, values=(
                            tarih or "",
                            kullanici_adi or "Sistem",
                            action_type_val or "",
                            entity_type_val or "",
                            detay_text
                        ))
                    
                except Exception as e:
                    messagebox.showerror("Hata", f"Günlük kayıtları yüklenemedi:\n{e}")
                    log_exception("logs_yukle", e)
            
            ttk.Button(filter_row1, text="🔍 Filtrele", bootstyle="primary", command=logs_yukle).pack(side=LEFT, padx=(0, 5))
            ttk.Button(filter_row1, text="🔄 Yenile", bootstyle="secondary", command=logs_yukle).pack(side=LEFT)
            
            # İlk yükleme
            logs_yukle()

    def _on_terapist_select(self, _evt=None):
        sel = self.lst.selection()
        if not sel:
            return
        vals = self.lst.item(sel[0]).get("values") or []
        if len(vals) >= 2:
            try:
                self.ent_yeni.delete(0, END)
                self.ent_yeni.insert(0, vals[0])
                self.cmb_rol.set(vals[1] or "Diğer")
            except Exception:
                pass

    def terapistleri_yukle(self):
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name, therapist_role FROM settings WHERE is_active=1 ORDER BY therapist_name")
            rows = cur.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Terapistler yüklenemedi:\n{e}")
            rows = []

        names = [r[0] for r in rows]
        if hasattr(self, "cmb_terapist") and self.cmb_terapist:
            self.cmb_terapist["values"] = names
            if names and self.cmb_terapist.cget("state") != "disabled":
                self.cmb_terapist.current(0)

        if hasattr(self, "lst"):
            for iid in self.lst.get_children():
                self.lst.delete(iid)
            for n, rol in rows:
                self.lst.insert("", END, values=(n, rol or ""))

    def terapist_ekle(self):
        name = (self.ent_yeni.get() or "").strip()
        if not name:
            messagebox.showwarning("Uyarı", "Lütfen terapist adını giriniz!")
            return
        if "name hoca" in name.lower():
            messagebox.showwarning("Uyarı", "Name Hoca kurumdan ayrıldı. Eklenemez.")
            return
        rol = (self.cmb_rol.get() or "").strip()
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                "INSERT OR IGNORE INTO settings (therapist_name, therapist_role, is_active, created_at) VALUES (?,?,1,?)",
                (name, rol, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            conn.commit()
            conn.close()
            self.ent_yeni.delete(0, END)
            self.terapistleri_yukle()
        except Exception as e:
            messagebox.showerror("Hata", f"Terapist ekleme hatası:\n{e}")

    def terapist_rol_guncelle(self):
        sel = self.lst.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir terapist seçiniz!")
            return
        name = (self.lst.item(sel[0]).get("values") or ["", ""])[0]
        rol = (self.cmb_rol.get() or "").strip()
        if not name:
            return
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("UPDATE settings SET therapist_role=? WHERE therapist_name=?", (rol, name))
            conn.commit()
            conn.close()
            self.terapistleri_yukle()
        except Exception as e:
            messagebox.showerror("Hata", f"Rol güncelleme hatası:\n{e}")

    def terapist_sil(self):
        sel = self.lst.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir terapist seçiniz!")
            return
        name = (self.lst.item(sel[0]).get("values") or [""])[0]
        if not name:
            return
        if not messagebox.askyesno(
            "Onay",
            f"{danisan_adi} danışanını listeden kaldırmak istediğinize emin misiniz?\n\n"
            "Bu işlem danışanı pasife alır (aktif=0).",
        ):
            return
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("DELETE FROM settings WHERE therapist_name=?", (name,))
            conn.commit()
            conn.close()
            self.terapistleri_yukle()
        except Exception as e:
            messagebox.showerror("Hata", f"Terapist silme hatası:\n{e}")

    def excel_aktar(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel dosyası", "*.xlsx"), ("Tüm dosyalar", "*.*")],
            title="Excel'e Aktar",
        )
        if not path:
            return
        try:
            conn = self.veritabani_baglan()
            if self.kullanici_yetki == "kurum_muduru" or not self.kullanici_terapist:
                df = pd.read_sql_query("SELECT * FROM records ORDER BY id DESC", conn)
            else:
                df = pd.read_sql_query(
                    "SELECT * FROM records WHERE terapist = ? ORDER BY id DESC",
                    conn,
                    params=(self.kullanici_terapist,),
                )
            conn.close()
            df.to_excel(path, index=False, engine="openpyxl")
            messagebox.showinfo("Başarılı", "Excel'e aktarıldı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel aktarma hatası:\n{e}")

    # --- leta_pro MODÜLLERİ: Menü işlemleri ---
    def yedek_klasoru_ac(self):
        try:
            os.makedirs(backups_dir(), exist_ok=True)
            open_path(backups_dir())
        except Exception as e:
            messagebox.showerror("Hata", f"Yedek klasörü açılamadı:\n{e}")

    def demo_verileri_ac(self):
        try:
            os.makedirs(demo_data_dir(), exist_ok=True)
            open_path(demo_data_dir())
        except Exception as e:
            messagebox.showerror("Hata", f"Demo verileri açılamadı:\n{e}")

    def hakkinda_goster(self):
        messagebox.showinfo(
            "Hakkında",
            "Leta Aile ve Çocuk - Yönetim Sistemi\n\n"
            "Modüller: Seans Takip, Seans Takvimi, Danışanlar, Görevler, Muhasebe, Kullanıcı Yönetimi.\n"
            "Bu sürüm tek pencere (tek root) ile stabil çalışacak şekilde hazırlanmıştır.",
        )
    def popup_eski_borc(self):
        """Eskiye dönük borç (devir bakiyesi) ekler. Kurum para takibi için kayıt sisteme işlenir."""
        win = ttk.Toplevel(self)
        win.title("Eski Borç / Devir Bakiyesi Ekle")
        center_window_smart(win, 460, 320, min_w=420, min_h=300)
        
        ttk.Label(win, text="Öğrenci Seç:", font=("Segoe UI", 10, "bold")).pack(pady=5)
        values = []
        if hasattr(self, "cmb_danisan") and hasattr(self.cmb_danisan, "cget"):
            values = list(self.cmb_danisan["values"]) if self.cmb_danisan["values"] else []
        c_danisan = ttk.Combobox(win, values=values, width=30)
        c_danisan.pack(pady=5)

        ttk.Label(win, text="Eklenecek devir borcu (TL):", font=("Segoe UI", 10, "bold")).pack(pady=5)
        e_tutar = ttk.Entry(win)
        e_tutar.pack(pady=5)

        def kaydet():
            danisan = (c_danisan.get() or "").strip()
            try:
                tutar = float((e_tutar.get() or "0").replace(",", "."))
            except Exception:
                messagebox.showerror("Hata", "Lütfen geçerli bir sayı girin.")
                return
            if not danisan:
                messagebox.showwarning("Eksik", "Öğrenci seçin ve tutar girin.")
                return
            if tutar <= 0:
                messagebox.showwarning("Eksik", "Pozitif bir tutar girin.")
                return

            conn = None
            try:
                conn = self.veritabani_baglan()
                pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
                pipeline.eski_borc_ekle(danisan, tutar)
                messagebox.showinfo("Başarılı", f"{danisan} için {format_money(tutar)} devir borcu eklendi.\nKurum kayıtlarına işlendi.")
                win.destroy()
                self.kayitlari_listele()
                self._refresh_borc_tables()
            except Exception as e:
                messagebox.showerror("Hata", str(e))
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        ttk.Button(win, text="DEVİR BORCU EKLE", bootstyle="danger", command=kaydet).pack(pady=15)

    def popup_eski_borc_sil(self):
        """Yanlış girilen devir borcu kayıtlarını seçip silme ekranı."""
        win = ttk.Toplevel(self)
        win.title("Eski Borç Kaydı Sil")
        center_window_smart(win, 780, 500, min_w=720, min_h=440)

        ttk.Label(
            win,
            text="Sadece ödeme almamış devir borcu kayıtları silinebilir.",
            bootstyle="warning",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor=W, padx=12, pady=(12, 6))

        cols = ("rid", "tarih", "danisan", "hizmet", "alinan", "kalan", "not")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=14)
        headers = [
            ("rid", "Record ID", 85),
            ("tarih", "Tarih", 110),
            ("danisan", "Danışan", 170),
            ("hizmet", "Borç", 110),
            ("alinan", "Alınan", 110),
            ("kalan", "Kalan", 110),
            ("not", "Not", 220),
        ]
        for c, h, w in headers:
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor=(CENTER if c in {"rid", "hizmet", "alinan", "kalan"} else W))

        wrap = ttk.Frame(win)
        wrap.pack(fill=BOTH, expand=True, padx=12, pady=6)
        tree.pack(in_=wrap, side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(wrap, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side=RIGHT, fill=Y)

        def _load_rows():
            for iid in tree.get_children():
                tree.delete(iid)
            conn = None
            try:
                conn = self.veritabani_baglan()
                params = []
                where_extra = ""
                sql = (
                    "SELECT id, COALESCE(tarih,''), COALESCE(danisan_adi,''), COALESCE(hizmet_bedeli,0), "
                    "COALESCE(alinan_ucret,0), COALESCE(kalan_borc,0), COALESCE(notlar,'') "
                    "FROM records WHERE COALESCE(seans_id,0)=0 "
                    "AND ("
                    "UPPER(COALESCE(notlar,'')) LIKE '%DEVIR BORC%' OR "
                    "UPPER(COALESCE(notlar,'')) LIKE '%DEVİR BORÇ%' OR "
                    "UPPER(COALESCE(notlar,'')) LIKE '%DEVIR BORÇ%' OR "
                    "UPPER(COALESCE(notlar,'')) LIKE '%DEVİR BORC%'"
                    ")"
                    + where_extra
                    + " ORDER BY tarih DESC, id DESC"
                )
                cur = conn.cursor()
                cur.execute(sql, tuple(params))
                for rid, tarih, danisan, hizmet, alinan, kalan, notlar in cur.fetchall():
                    tree.insert(
                        "",
                        END,
                        values=(rid, tarih, danisan, format_money(hizmet), format_money(alinan), format_money(kalan), notlar),
                    )
            except Exception as e:
                messagebox.showerror("Hata", f"Liste yüklenemedi:\n{e}")
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        def _secili_id():
            sel = tree.selection()
            if not sel:
                return None
            vals = tree.item(sel[0], "values")
            try:
                return int(vals[0])
            except Exception:
                return None

        def _sil():
            rid = _secili_id()
            if not rid:
                messagebox.showwarning("Uyarı", "Lütfen silinecek kaydı seçin.")
                return
            if not messagebox.askyesno("Onay", "Seçili devir borcu kaydı silinsin mi?"):
                return
            conn = None
            try:
                conn = self.veritabani_baglan()
                pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
                ok = pipeline.eski_borc_sil(rid)
                if not ok:
                    neden = pipeline.get_last_error() or "Kayıt silinemedi. Bu kayıt seans bağlantılı olabilir veya ödeme almıştır."
                    messagebox.showwarning("Uyarı", neden)
                    return
                messagebox.showinfo("Başarılı", "Devir borcu kaydı silindi.")
                _load_rows()
                self.kayitlari_listele()
                self._refresh_borc_tables()
                self._refresh_open_reports()
            except Exception as e:
                messagebox.showerror("Hata", f"Silme hatası:\n{e}")
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        btns = ttk.Frame(win)
        btns.pack(fill=X, padx=12, pady=(6, 12))
        ttk.Button(btns, text="Yenile", bootstyle="info", command=_load_rows).pack(side=LEFT)
        ttk.Button(btns, text="Seçili Kaydı Sil", bootstyle="danger", command=_sil).pack(side=LEFT, padx=8)
        ttk.Button(btns, text="Kapat", bootstyle="secondary", command=win.destroy).pack(side=RIGHT)

        _load_rows()

    def popup_toplu_odeme(self):
        """Toplu ödeme alma penceresi"""
        win = ttk.Toplevel(self)
        win.title("Toplu Ödeme Al")
        center_window_smart(win, 460, 360, min_w=420, min_h=330)
        
        ttk.Label(win, text="Öğrenci Seç:", font=("Segoe UI", 10, "bold")).pack(pady=5)
        
        values = []
        if hasattr(self, 'cmb_danisan') and hasattr(self.cmb_danisan, 'cget'):
             values = self.cmb_danisan['values']

        c_danisan = ttk.Combobox(win, values=values, width=30)
        c_danisan.pack(pady=5)

        ttk.Label(win, text="Ödenen Tutar (TL):", font=("Segoe UI", 10, "bold")).pack(pady=5)
        e_tutar = ttk.Entry(win)
        e_tutar.pack(pady=5)
        
        ttk.Label(win, text="Açıklama:", font=("Segoe UI", 10)).pack(pady=5)
        e_aciklama = ttk.Entry(win)
        e_aciklama.insert(0, "Toplu Ödeme / Peşinat")
        e_aciklama.pack(pady=5)

        def kaydet():
            danisan = c_danisan.get()
            try:
                tutar = float(e_tutar.get())
            except Exception:
                messagebox.showerror("Hata", "Geçerli tutar girin.")
                return

            if not danisan or tutar <= 0: return

            conn = None
            try:
                conn = self.veritabani_baglan()
                pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
                basarili = pipeline.toplu_odeme_al(danisan, tutar, aciklama=e_aciklama.get())
                if not basarili:
                    raise ValueError(pipeline.get_last_error() or "Toplu ödeme kaydı oluşturulamadı.")
                messagebox.showinfo("Başarılı", f"{tutar} TL tahsilat alındı.\nKasa defterine işlendi ve bakiyeden düşüldü.")
                win.destroy()
                self.kayitlari_listele()
                self._refresh_borc_tables()
            except Exception as e:
                messagebox.showerror("Hata", str(e))
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        ttk.Button(win, text="TAHSİLAT YAP", bootstyle="success", command=kaydet).pack(pady=15)    

    def popup_toplu_odeme_yonet(self):
        """Toplu ödeme kayıtlarını eşleşme kontrolü ile listeler ve gerekirse siler."""
        win = ttk.Toplevel(self)
        win.title("Toplu Ödeme Yönetimi")
        center_window_smart(win, 980, 520, min_w=880, min_h=460)

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="Toplu ödeme kayıtları (eşleşme kontrolü)", font=("Segoe UI", 11, "bold")).pack(side=LEFT)

        cols = ("kasa_id", "tarih", "danisan", "record_id", "record_danisan", "tutar", "durum", "aciklama")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=16)
        headers = [
            ("kasa_id", "Kasa ID", 80),
            ("tarih", "Tarih", 100),
            ("danisan", "Açıklamadaki Danışan", 170),
            ("record_id", "Record ID", 85),
            ("record_danisan", "Record Danışan", 170),
            ("tutar", "Tutar", 110),
            ("durum", "Eşleşme", 100),
            ("aciklama", "Açıklama", 260),
        ]
        for c, h, w in headers:
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor=(CENTER if c in {"kasa_id", "record_id", "tutar", "durum"} else W))

        body = ttk.Frame(win)
        body.pack(fill=BOTH, expand=True, padx=10, pady=(0, 10))
        tree.pack(in_=body, side=LEFT, fill=BOTH, expand=True)
        sb = ttk.Scrollbar(body, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.tag_configure("ok", background="#d4edda")
        tree.tag_configure("bad", background="#f8d7da")

        def _load():
            for iid in tree.get_children():
                tree.delete(iid)
            conn = None
            try:
                conn = self.veritabani_baglan()
                pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
                rows = pipeline.toplu_odeme_kayitlari_getir()
                for r in rows:
                    durum = "Eşleşti" if r.get("matched") else "Eşleşmedi"
                    tag = "ok" if r.get("matched") else "bad"
                    tree.insert(
                        "",
                        END,
                        values=(
                            r.get("kasa_id"),
                            r.get("tarih", ""),
                            r.get("danisan", ""),
                            r.get("record_id", 0),
                            r.get("record_danisan", ""),
                            format_money(r.get("tutar", 0)),
                            durum,
                            r.get("aciklama", ""),
                        ),
                        tags=(tag,),
                    )
            except Exception as e:
                messagebox.showerror("Hata", f"Toplu ödeme listesi yüklenemedi:\n{e}")
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        def _secili_kasa_id():
            sel = tree.selection()
            if not sel:
                return None
            vals = tree.item(sel[0], "values")
            try:
                return int(vals[0])
            except Exception:
                return None

        def _sil():
            kasa_id = _secili_kasa_id()
            if not kasa_id:
                messagebox.showwarning("Uyarı", "Lütfen silinecek toplu ödeme kaydını seçin.")
                return
            if not messagebox.askyesno("Onay", "Seçili toplu ödeme kaydı silinsin mi?\n(Borç/ödeme dengesi geri alınacaktır.)"):
                return
            conn = None
            try:
                conn = self.veritabani_baglan()
                pipeline = DataPipeline(conn, self.kullanici[0] if self.kullanici else None)
                ok = pipeline.toplu_odeme_hareket_sil(kasa_id)
                if not ok:
                    messagebox.showwarning("Uyarı", pipeline.get_last_error() or "Silme işlemi başarısız.")
                    return
                messagebox.showinfo("Başarılı", "Toplu ödeme kaydı silindi.")
                _load()
                self.kayitlari_listele()
                self._refresh_borc_tables()
                self._refresh_open_reports()
            except Exception as e:
                messagebox.showerror("Hata", f"Silme hatası:\n{e}")
            finally:
                try:
                    if conn is not None:
                        conn.close()
                except Exception:
                    pass

        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill=X)
        ttk.Button(btns, text="Yenile", bootstyle="info", command=_load).pack(side=LEFT)
        ttk.Button(btns, text="Seçili Toplu Ödemeyi Sil", bootstyle="danger", command=_sil).pack(side=LEFT, padx=8)
        ttk.Button(btns, text="Kapat", bootstyle="secondary", command=win.destroy).pack(side=RIGHT)

        _load()

    def eski_veri_migration(self):
        """
        Eski Veri İçe Aktarma Wizard'ı (Dosya okuma yerine Excel template + manuel import)
        Kullanıcı Excel template'i indirir, doldurur ve sisteme yükler.
        
        ✅ YETKİ KONTROLÜ: Sadece Kurum Müdürü erişebilir.
        """
        # ✅ YETKİ KONTROLÜ: Sadece Kurum Müdürü erişebilir
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Yetki", "Bu işlem sadece Kurum Müdürü yetkisi ile yapılabilir.")
            return
        win = ttk.Toplevel(self)
        win.title("Veri İçe Aktarma Wizard'ı")
        win.transient(self)
        center_window_smart(win, 900, 700, min_w=850, min_h=650)
        maximize_window(win)
        self._brand_window(win)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text="📥 Veri İçe Aktarma", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        # Açıklama
        desc_frame = ttk.Labelframe(wrapper, text="Nasıl Çalışır?", padding=15, bootstyle="info")
        desc_frame.pack(fill=X, pady=(0, 20))
        
        desc_text = """
1. Excel Template'i İndir butonuna tıklayarak şablon dosyasını indirin
2. Template dosyasını Excel'de açın ve verilerinizi doldurun
3. Doldurduğunuz dosyayı "Excel Dosyası Yükle" butonu ile sisteme yükleyin
4. Sistem verilerinizi kontrol edip önizleme gösterecek
5. Onayladığınızda veriler veritabanına aktarılacak
        """
        ttk.Label(desc_frame, text=desc_text.strip(), font=("Segoe UI", 10), justify="left").pack(anchor=W)
        
        # Template indirme - 3 ayrı buton
        template_frame = ttk.Labelframe(wrapper, text="1. Template İndir (3 Farklı Template)", padding=15, bootstyle="secondary")
        template_frame.pack(fill=X, pady=(0, 15))
        
        template_buttons_frame = ttk.Frame(template_frame)
        template_buttons_frame.pack(fill=X)
        
        def template_indir_danisanlar():
            """Danışanlar template'i oluştur ve indir (Seans Ücret Takip ile aynı stil)"""
            try:
                template_name = "Danışanlar"
                template_data = {
                    "Danışanlar": [
                        ["AD SOYAD", "TELEFON", "EMAIL", "DOĞUM TARİHİ (GG.AA.YYYY)", "VELİ ADI", "VELİ TELEFON", "ADRES"],
                        ["AHMET YILMAZ", "05551234567", "ahmet@example.com", "15.01.2015", "VELİ ADI", "05559876543", "Adres bilgisi"],
                        ["MEHMET DEMİR", "05559876543", "mehmet@example.com", "20.03.2016", "VELİ ADI 2", "05551111111", "Adres bilgisi 2"]
                    ]
                }
                _create_template_file(template_name, template_data)
            except Exception as e:
                messagebox.showerror("Hata", f"Template oluşturulamadı:\n{e}")
                log_exception("template_indir_danisanlar", e)
        
        def template_indir_haftalik():
            """Haftalık Seans Takvimi template'i oluştur ve indir (Seans Ücret Takip ile aynı stil)"""
            try:
                template_name = "Haftalık_Seans_Takvimi"
                template_data = {
                    "Haftalık Seans Takvimi": [
                        ["HAFTA BAŞLANGIÇ TARİHİ (GG.AA.YYYY)", "GÜN", "SAAT (HH:MM)", "PERSONEL ADI", "ÖĞRENCİ ADI", "NOTLAR"],
                        ["27.01.2026", "Pazartesi", "09:00", "Pervin Hoca", "AHMET YILMAZ", "Haftalık program"],
                        ["27.01.2026", "Pazartesi", "10:00", "Arif Hoca", "MEHMET DEMİR", ""],
                        ["27.01.2026", "Salı", "14:00", "Pervin Hoca", "AHMET YILMAZ", ""]
                    ]
                }
                _create_template_file(template_name, template_data)
            except Exception as e:
                messagebox.showerror("Hata", f"Template oluşturulamadı:\n{e}")
                log_exception("template_indir_haftalik", e)
        
        def template_indir_seans_ucret():
            """Seans Ücret Takip template'i oluştur ve indir (6 sütun: Tarih, Danışan Adı, Terapist, Alınacak Ücret, Alınan Ücret, Kalan Borç)"""
            try:
                template_name = "Seans_Ücret_Takip"
                template_data = {
                    "Seans Ücret Takip": [
                        ["Tarih", "Danışan Adı", "Terapist", "Alınacak Ücret", "Alınan Ücret", "Kalan Borç"],
                        ["17.03.2025", "Defne Yılmaz", "Name Hoca", "1800", "", ""],
                        ["18.03.2025", "Ali Eymen Cive", "Pervin Hoca", "2800", "2800", "0"],
                        ["20.03.2025", "Aslı Berra Taşar", "Pervin Hoca", "2700", "", "2700"]
                    ]
                }
                _create_template_file(template_name, template_data)
            except Exception as e:
                messagebox.showerror("Hata", f"Template oluşturulamadı:\n{e}")
                log_exception("template_indir_seans_ucret", e)
        
        def _create_template_file(template_name, template_data):
            """Ortak template oluşturma fonksiyonu"""
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"Leta_{template_name}_Template_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if not path:
                return
            
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                for sheet_name, data in template_data.items():
                    df = pd.DataFrame(data[1:], columns=data[0])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Formatlama
                    worksheet = writer.sheets[sheet_name]
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Sütun genişlikleri
                    for idx, col in enumerate(df.columns, 1):
                        worksheet.column_dimensions[chr(64 + idx)].width = max(len(str(col)) + 2, 15)
            
            messagebox.showinfo(
                "Başarılı", 
                f"{template_name} template dosyası oluşturuldu:\n{path}\n\n"
                f"Lütfen:\n"
                f"1) Dosyayı Excel'de açın\n"
                f"2) Örnek satırları silin ve kendi verilerinizi girin\n"
                f"3) Dosyayı kaydedin\n"
                f"4) Buraya geri dönüp 'Excel Dosyası Yükle' ile içe aktarın"
            )
        
        ttk.Button(template_buttons_frame, text="📥 1. Danışanlar Template", bootstyle="success", 
                  command=template_indir_danisanlar, width=25).pack(side=LEFT, padx=5)
        ttk.Button(template_buttons_frame, text="📅 2. Haftalık Seans Takvimi Template", bootstyle="primary", 
                  command=template_indir_haftalik, width=30).pack(side=LEFT, padx=5)
        ttk.Button(template_buttons_frame, text="💰 3. Seans Ücret Takip Template", bootstyle="warning", 
                  command=template_indir_seans_ucret, width=30).pack(side=LEFT, padx=5)
        
        # Dosya yükleme
        upload_frame = ttk.Labelframe(wrapper, text="2. Doldurulmuş Excel Dosyasını Yükle", padding=15, bootstyle="secondary")
        upload_frame.pack(fill=X, pady=(0, 15))
        
        file_path_var = ttk.StringVar()
        ttk.Label(upload_frame, text="Dosya:", font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(0, 5))
        
        file_frame = ttk.Frame(upload_frame)
        file_frame.pack(fill=X, pady=(0, 10))
        
        ent_file = ttk.Entry(file_frame, textvariable=file_path_var, width=50, state="readonly")
        ent_file.pack(side=LEFT, fill=X, expand=True, padx=(0, 5))
        
        def dosya_sec():
            path = filedialog.askopenfilename(
                filetypes=[("Excel", "*.xlsx *.xls")],
                title="Excel Dosyası Seç"
            )
            if path:
                file_path_var.set(path)
        
        ttk.Button(file_frame, text="📂 Dosya Seç", bootstyle="primary", command=dosya_sec).pack(side=LEFT)
        
        # Önizleme alanı
        preview_frame = ttk.Labelframe(wrapper, text="3. Önizleme ve Onay", padding=15, bootstyle="secondary")
        preview_frame.pack(fill=BOTH, expand=True, pady=(0, 15))
        
        preview_tree_frame = ttk.Frame(preview_frame)
        preview_tree_frame.pack(fill=BOTH, expand=True)
        
        preview_tree = ttk.Treeview(preview_tree_frame, columns=("Veri Tipi", "Kayıt Sayısı", "Durum"), show="headings", height=8)
        preview_tree.heading("Veri Tipi", text="Veri Tipi")
        preview_tree.heading("Kayıt Sayısı", text="Kayıt Sayısı")
        preview_tree.heading("Durum", text="Durum")
        preview_tree.column("Veri Tipi", width=200)
        preview_tree.column("Kayıt Sayısı", width=150)
        preview_tree.column("Durum", width=200)
        preview_tree.pack(side=LEFT, fill=BOTH, expand=True)
        
        preview_sb = ttk.Scrollbar(preview_tree_frame, orient=VERTICAL, command=preview_tree.yview)
        preview_tree.configure(yscroll=preview_sb.set)
        preview_sb.pack(side=RIGHT, fill=Y)
        
        def dosya_yukle_ve_onizle():
            """Excel dosyasını yükle ve önizleme göster"""
            file_path = file_path_var.get()
            if not file_path or not os.path.exists(file_path):
                messagebox.showwarning("Uyarı", "Lütfen bir Excel dosyası seçin!")
                return
            
            # ✅ DOSYA DOĞRULAMA: Dosya uzantısını kontrol et
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls']:
                messagebox.showerror("Hata", 
                    f"Geçersiz dosya formatı!\n\n"
                    f"Seçilen dosya: {os.path.basename(file_path)}\n"
                    f"Dosya uzantısı: {file_ext}\n\n"
                    f"Lütfen .xlsx veya .xls formatında bir Excel dosyası seçin.")
                return
            
            # ✅ DOSYA BOYUTU KONTROLÜ: Dosya boş mu kontrol et
            try:
                file_size = os.path.getsize(file_path)
                if file_size == 0:
                    messagebox.showerror("Hata", "Seçilen dosya boş!\n\nLütfen geçerli bir Excel dosyası seçin.")
                    return
            except Exception:
                pass
            
            # Önizleme treeview'ı temizle
            for item in preview_tree.get_children():
                preview_tree.delete(item)
            
            try:
                # Excel dosyasını oku (engine belirtilmeli)
                # .xls dosyaları için xlrd engine kullan, .xlsx için openpyxl
                engine = 'xlrd' if file_ext == '.xls' else 'openpyxl'
                excel_data = pd.read_excel(file_path, sheet_name=None, engine=engine)
                
                toplam_kayit = 0
                # 3 farklı template desteği (+ Haftalık Seans Programı = Takvimi ile aynı)
                template_sheets = ["Danışanlar", "Haftalık Seans Takvimi", "Haftalık Seans Programı", "Seans Ücret Takip"]
                for sheet_name in template_sheets:
                    if sheet_name in excel_data:
                        df = excel_data[sheet_name]
                        kayit_sayisi = len(df) - 1 if len(df) > 1 else 0  # Başlık satırını çıkar
                        if kayit_sayisi > 0:
                            preview_tree.insert("", END, values=(sheet_name, kayit_sayisi, "✅ Hazır"), tags=("ok",))
                            toplam_kayit += kayit_sayisi
                        else:
                            preview_tree.insert("", END, values=(sheet_name, 0, "⚠️ Boş"), tags=("bos",))
                
                preview_tree.tag_configure("ok", foreground="green")
                preview_tree.tag_configure("bos", foreground="orange")
                
                if toplam_kayit > 0:
                    messagebox.showinfo("Önizleme", f"Toplam {toplam_kayit} kayıt bulundu.\n\n'Verileri İçe Aktar' butonuna tıklayarak devam edebilirsiniz.")
                else:
                    messagebox.showwarning("Uyarı", "Excel dosyasında veri bulunamadı!\n\nLütfen template dosyasını doğru şekilde doldurduğunuzdan emin olun.")
                    
            except zipfile.BadZipFile as e:
                messagebox.showerror("Hata", 
                    f"Excel dosyası bozuk veya geçersiz format!\n\n"
                    f"Dosya: {os.path.basename(file_path)}\n\n"
                    f"Lütfen:\n"
                    f"1) Dosyanın gerçekten bir Excel dosyası olduğundan emin olun\n"
                    f"2) Dosyayı Excel'de açıp kaydedin\n"
                    f"3) Template dosyasını kullanarak yeni bir dosya oluşturun")
                log_exception("dosya_yukle_ve_onizle", e)
            except Exception as e:
                error_msg = str(e)
                if "not a zip file" in error_msg.lower() or "badzipfile" in error_msg.lower():
                    messagebox.showerror("Hata", 
                        f"Excel dosyası bozuk veya geçersiz format!\n\n"
                        f"Dosya: {os.path.basename(file_path)}\n\n"
                        f"Lütfen:\n"
                        f"1) Dosyanın gerçekten bir Excel dosyası olduğundan emin olun\n"
                        f"2) Dosyayı Excel'de açıp kaydedin\n"
                        f"3) Template dosyasını kullanarak yeni bir dosya oluşturun")
                else:
                    messagebox.showerror("Hata", f"Excel dosyası okunamadı:\n\n{error_msg}\n\nLütfen geçerli bir Excel dosyası seçin.")
                log_exception("dosya_yukle_ve_onizle", e)
        
        def verileri_ice_aktar():
            """Excel'deki verileri veritabanına aktar"""
            file_path = file_path_var.get()
            if not file_path or not os.path.exists(file_path):
                messagebox.showwarning("Uyarı", "Lütfen bir Excel dosyası seçin!")
                return
            
            # ✅ DOSYA DOĞRULAMA: Dosya uzantısını kontrol et
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls']:
                messagebox.showerror("Hata", 
                    f"Geçersiz dosya formatı!\n\n"
                    f"Seçilen dosya: {os.path.basename(file_path)}\n"
                    f"Dosya uzantısı: {file_ext}\n\n"
                    f"Lütfen .xlsx veya .xls formatında bir Excel dosyası seçin.")
                return
            
            # ✅ DOSYA BOYUTU KONTROLÜ
            try:
                file_size = os.path.getsize(file_path)
                if file_size == 0:
                    messagebox.showerror("Hata", "Seçilen dosya boş!\n\nLütfen geçerli bir Excel dosyası seçin.")
                    return
            except Exception:
                pass
            
            try:
                # Excel dosyasını oku (engine belirtilmeli)
                # .xls dosyaları için xlrd engine kullan, .xlsx için openpyxl
                engine = 'xlrd' if file_ext == '.xls' else 'openpyxl'
                excel_data = pd.read_excel(file_path, sheet_name=None, engine=engine)
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                
                aktarilan = []
                
                # 1) Danışanlar Template'i (sütunlar: AD SOYAD, TELEFON, EMAIL, DOĞUM TARİHİ, VELİ ADI, VELİ TELEFON, ADRES)
                if "Danışanlar" in excel_data:
                    df_danisan = excel_data["Danışanlar"]
                    for idx, row in df_danisan.iterrows():
                        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        if first_cell and any(x in first_cell.upper() for x in ("AD SOYAD", "TELEFON", "EMAIL", "DOĞUM", "Tarih", "Date")):
                            continue
                        try:
                            ad_soyad = str(row.iloc[0]).strip().upper() if pd.notna(row.iloc[0]) else ""
                            if not ad_soyad or ad_soyad == "NAN":
                                continue
                            
                            cur.execute("SELECT id FROM danisanlar WHERE UPPER(ad_soyad) = UPPER(?)", (ad_soyad,))
                            if cur.fetchone():
                                continue
                            
                            telefon = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                            email = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
                            dogum_raw = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ""
                            dogum_tarihi = self._normalize_hafta_tarihi(dogum_raw) if dogum_raw else ""
                            veli_adi = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ""
                            veli_telefon = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else ""
                            adres = str(row.iloc[6]).strip() if len(row) > 6 and pd.notna(row.iloc[6]) else ""
                            
                            cur.execute(
                                """
                                INSERT INTO danisanlar 
                                (ad_soyad, telefon, email, dogum_tarihi, veli_adi, veli_telefon, adres, aktif, olusturma_tarihi)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
                                """,
                                (ad_soyad, telefon, email, dogum_tarihi, veli_adi, veli_telefon, adres, 
                                 datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                            )
                            aktarilan.append(f"Danışan: {ad_soyad}")
                        except Exception as e:
                            log_exception("danisan_aktar", e)
                            continue
                
                # 2) Haftalık Seans Takvimi / Haftalık Seans Programı Template'i
                if "Haftalık Seans Takvimi" in excel_data:
                    df_haftalik = excel_data["Haftalık Seans Takvimi"]
                elif "Haftalık Seans Programı" in excel_data:
                    df_haftalik = excel_data["Haftalık Seans Programı"]
                else:
                    df_haftalik = None
                if df_haftalik is not None:
                    haftalik_sayisi = 0
                    for idx, row in df_haftalik.iterrows():
                        # Başlık satırını atla (HAFTA BAŞLANGIÇ TARİHİ / Tarih / Hafta vb.)
                        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        fc_upper = first_cell.upper().replace("İ", "I")
                        if first_cell and any(x in fc_upper for x in ("TARIH", "HAFTA", "DATE")):
                            continue
                        try:
                            hafta_raw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                            hafta_baslangic = self._normalize_hafta_tarihi(hafta_raw) if hafta_raw else ""
                            gun = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                            saat_raw = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
                            # Saat HH:MM formatına getir (10, 10.00, 10:00 -> 10:00; 16:30 korunur)
                            saat = self._normalize_saat(saat_raw) if saat_raw else ""
                            personel_adi = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ""
                            ogrenci_adi = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ""
                            # 7 sütunlu tabloda: 5=Oda (kullanma), 6=Notlar; 6 sütunda: 5=Notlar
                            notlar = (str(row.iloc[6]).strip() if len(row) > 6 and pd.notna(row.iloc[6]) else str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else "")
                            if not hafta_baslangic or not gun or not saat or not personel_adi:
                                continue
                            cur.execute(
                                """
                                INSERT OR REPLACE INTO haftalik_seans_programi
                                (personel_adi, hafta_baslangic_tarihi, gun, saat, ogrenci_adi, oda_adi, notlar, olusturma_tarihi, olusturan_kullanici_id)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    personel_adi, hafta_baslangic, gun, saat, ogrenci_adi or "", "", notlar or "",
                                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), kullanici_id
                                )
                            )
                            haftalik_sayisi += 1
                        except Exception as e:
                            log_exception("haftalik_aktar", e)
                            continue
                    if haftalik_sayisi > 0:
                        aktarilan.append(f"Haftalık Seans Takvimi: {haftalik_sayisi} adet")
                
                # 3) Seans Ücret Takip Template'i (6 sütun: Tarih, Danışan Adı, Terapist, Alınacak Ücret, Alınan Ücret, Kalan Borç)
                if "Seans Ücret Takip" in excel_data:
                    df_seans = excel_data["Seans Ücret Takip"]
                    seans_sayisi = 0
                    seans_tekrar_atlandi = 0
                    # Sütun 3 = Alınacak Ücret (hizmet_bedeli), Sütun 4 = Alınan Ücret (alinan_ucret)
                    def _safe_float(v):
                        if v is None or (isinstance(v, float) and pd.isna(v)):
                            return 0.0
                        s = str(v).strip().replace(",", ".").replace("₺", "").replace(" ", "")
                        try:
                            return float(s) if s else 0.0
                        except (ValueError, TypeError):
                            return 0.0
                    for idx, row in df_seans.iterrows():
                        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        if first_cell and any(x in first_cell for x in ("TARİH", "Tarih", "Danışan", "Date", "tarih", "danışan")):
                            continue
                        try:
                            tarih_raw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                            tarih = self._normalize_hafta_tarihi(tarih_raw) if tarih_raw else ""  # GG.AA.YYYY → YYYY-MM-DD
                            danisan_adi = str(row.iloc[1]).strip().upper() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                            terapist = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""
                            bedel = _safe_float(row.iloc[3]) if len(row) > 3 else 0.0   # Alınacak Ücret = hizmet bedeli
                            # Yeni şablon: Alınan Ücret = sütun 4; eski 7 sütunlu şablon: GÜNCEL = sütun 6
                            if len(row) > 6:
                                alinan = _safe_float(row.iloc[6])  # Eski: GÜNCEL
                            else:
                                alinan = _safe_float(row.iloc[4]) if len(row) > 4 else 0.0  # Yeni: Alınan Ücret
                            notlar = ""
                            if not tarih or not danisan_adi or not terapist:
                                continue
                            # ✅ Tekrar önleme: aynı (tarih, danışan, terapist) zaten varsa atla
                            danisan_norm = str(danisan_adi).strip().upper()
                            terapist_norm = str(terapist).strip()
                            cur.execute(
                                "SELECT id FROM records WHERE tarih = ? AND UPPER(TRIM(COALESCE(danisan_adi,''))) = ? AND TRIM(COALESCE(terapist,'')) = ? LIMIT 1",
                                (tarih, danisan_norm, terapist_norm)
                            )
                            if cur.fetchone():
                                seans_tekrar_atlandi += 1
                                continue
                            pipeline = DataPipeline(conn, kullanici_id)
                            seans_id = pipeline.seans_kayit(
                                tarih=tarih,
                                saat="09:00",
                                danisan_adi=danisan_norm,
                                terapist=terapist_norm,
                                hizmet_bedeli=bedel,
                                alinan_ucret=alinan,
                                notlar=notlar,
                                oda=oda,
                                check_oda_cakisma=False,
                                skip_pricing_update=True,  # Tarihsel veri; güncel fiyatlar ayrı verilecek
                                ensure_danisan=False,  # Eski veri importunda danışanlar listesi şişmesin
                            )
                            if seans_id:
                                seans_sayisi += 1
                        except Exception as e:
                            log_exception("seans_aktar", e)
                            continue
                    
                    if seans_sayisi > 0 or seans_tekrar_atlandi > 0:
                        msg = f"Seans Ücret Takip: {seans_sayisi} adet"
                        if seans_tekrar_atlandi > 0:
                            msg += f" ({seans_tekrar_atlandi} tekrar atlandı)"
                        aktarilan.append(msg)
                
                conn.commit()
                conn.close()
                
                # ✅ Import sonrası listeleri yenile: danışan combobox ve Öğrenci Bilgileri tablosu
                try:
                    conn_refresh = self.veritabani_baglan()
                    c = conn_refresh.cursor()
                    c.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
                    danisan_listesi = [row[0] for row in c.fetchall()]
                    conn_refresh.close()
                    if hasattr(self, "cmb_danisan") and self.cmb_danisan.winfo_exists():
                        self.cmb_danisan["values"] = danisan_listesi
                    # Öğrenci Bilgileri sekmesindeki danışan tablosunu yenile
                    for child in self.tab_ogrenci_bilgileri.winfo_children():
                        if hasattr(child, "_tree_danisanlar"):
                            self._tum_danisanlari_listele(child)
                            break
                except Exception:
                    pass
                
                messagebox.showinfo("Başarılı", 
                    f"Veriler başarıyla içe aktarıldı!\n\nAktarılan:\n" + "\n".join(aktarilan) +
                    "\n\n💡 Danışanlar: SEANS TAKİP veya ÖĞRENCİ BİLGİLERİ sekmesinde görünür. "
                    "Haftalık program: ÖĞRENCİ BİLGİLERİ → Haftalık Seans Programı sekmesinde personel ve hafta seçip 'Yükle' ile görüntüleyin.")
                win.destroy()
                
            except zipfile.BadZipFile as e:
                messagebox.showerror("Hata", 
                    f"Excel dosyası bozuk veya geçersiz format!\n\n"
                    f"Dosya: {os.path.basename(file_path)}\n\n"
                    f"Lütfen:\n"
                    f"1) Dosyanın gerçekten bir Excel dosyası olduğundan emin olun\n"
                    f"2) Dosyayı Excel'de açıp kaydedin\n"
                    f"3) Template dosyasını kullanarak yeni bir dosya oluşturun")
                log_exception("verileri_ice_aktar", e)
            except Exception as e:
                error_msg = str(e)
                if "not a zip file" in error_msg.lower() or "badzipfile" in error_msg.lower():
                    messagebox.showerror("Hata", 
                        f"Excel dosyası bozuk veya geçersiz format!\n\n"
                        f"Dosya: {os.path.basename(file_path)}\n\n"
                        f"Lütfen:\n"
                        f"1) Dosyanın gerçekten bir Excel dosyası olduğundan emin olun\n"
                        f"2) Dosyayı Excel'de açıp kaydedin\n"
                        f"3) Template dosyasını kullanarak yeni bir dosya oluşturun")
                else:
                    messagebox.showerror("Hata", f"Veriler içe aktarılamadı:\n\n{error_msg}\n\nLütfen geçerli bir Excel dosyası seçin.")
                log_exception("verileri_ice_aktar", e)
        
        ttk.Button(upload_frame, text="👁️ Önizleme Göster", bootstyle="info", 
                  command=dosya_yukle_ve_onizle, width=25).pack(side=LEFT, padx=5)
        ttk.Button(upload_frame, text="✅ Verileri İçe Aktar", bootstyle="success", 
                  command=verileri_ice_aktar, width=25).pack(side=LEFT, padx=5)
        
        # Kapat butonu
        ttk.Button(wrapper, text="Kapat", bootstyle="secondary", command=win.destroy).pack(pady=(10, 0))
    
    def eski_veri_migration_legacy(self):
        """Eski veritabanından yeni sisteme veri aktarımı"""
        win = ttk.Toplevel(self)
        win.title("Eski Veri Migration")
        win.geometry("600x400")
        center_window(win, 600, 400)
        win.transient(self)
        self._brand_window(win)
        
        wrapper = ttk.Frame(win, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        ttk.Label(wrapper, text="Eski Veri Migration", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        ttk.Label(wrapper, text="Bu araç eski veritabanından yeni sisteme veri aktarımı yapar.", 
                 font=("Segoe UI", 10), wraplength=550).pack(pady=(0, 20))
        
        # Eski DB seçimi
        frm_db = ttk.Frame(wrapper)
        frm_db.pack(fill=X, pady=10)
        
        ttk.Label(frm_db, text="Eski Veritabanı:", width=20).pack(side=LEFT, padx=5)
        ent_eski_db = ttk.Entry(frm_db, width=40)
        ent_eski_db.pack(side=LEFT, padx=5, fill=X, expand=True)
        
        def db_sec():
            dosya = filedialog.askopenfilename(
                title="Eski Veritabanı Seç",
                filetypes=[("Veritabanı Dosyaları", "*.db"), ("Tüm Dosyalar", "*.*")]
            )
            if dosya:
                ent_eski_db.delete(0, END)
                ent_eski_db.insert(0, dosya)
        
        ttk.Button(frm_db, text="📁 Seç", bootstyle="primary", command=db_sec).pack(side=LEFT, padx=5)
        
        # Sonuç alanı
        frm_sonuc = ttk.Labelframe(wrapper, text="Migration Sonuçları", padding=10, bootstyle="secondary")
        frm_sonuc.pack(fill=BOTH, expand=True, pady=10)
        
        text_sonuc = tk.Text(frm_sonuc, height=10, wrap=tk.WORD)
        text_sonuc.pack(fill=BOTH, expand=True)
        sb_sonuc = ttk.Scrollbar(frm_sonuc, orient=VERTICAL, command=text_sonuc.yview)
        text_sonuc.configure(yscrollcommand=sb_sonuc.set)
        sb_sonuc.pack(side=RIGHT, fill=Y)
        
        def migration_baslat():
            eski_db_yolu = ent_eski_db.get().strip()
            if not eski_db_yolu:
                messagebox.showerror("Hata", "Lütfen eski veritabanı dosyasını seçin.")
                return
            
            if not os.path.exists(eski_db_yolu):
                messagebox.showerror("Hata", f"Veritabanı dosyası bulunamadı:\n{eski_db_yolu}")
                return
            
            text_sonuc.delete("1.0", END)
            text_sonuc.insert(END, "🔄 Migration başlatılıyor...\n\n")
            win.update()
            
            try:
                # Migration script'ini import et ve çalıştır
                import migration_eski_veriler
                results = migration_eski_veriler.migrate_eski_veriler(eski_db_yolu)
                
                if results["success"]:
                    text_sonuc.insert(END, "✅ Migration tamamlandı!\n\n")
                    text_sonuc.insert(END, "Aktarılan Tablolar:\n")
                    for table, count in results["migrated_tables"].items():
                        text_sonuc.insert(END, f"  • {table}: {count} kayıt\n")
                    
                    if results["errors"]:
                        text_sonuc.insert(END, "\n⚠️ Hatalar:\n")
                        for error in results["errors"]:
                            text_sonuc.insert(END, f"  • {error}\n")
                    
                    messagebox.showinfo("Başarılı", "Migration tamamlandı! Sonuçları kontrol edin.")
                else:
                    text_sonuc.insert(END, "❌ Migration başarısız!\n\n")
                    text_sonuc.insert(END, "Hatalar:\n")
                    for error in results["errors"]:
                        text_sonuc.insert(END, f"  • {error}\n")
                    
                    messagebox.showerror("Hata", "Migration başarısız! Detaylar için sonuç alanına bakın.")
            
            except Exception as e:
                text_sonuc.insert(END, f"❌ Hata: {e}\n")
                messagebox.showerror("Hata", f"Migration hatası:\n{e}")
                log_exception("eski_veri_migration", e)
        
        ttk.Button(wrapper, text="🚀 Migration Başlat", bootstyle="success", 
                  command=migration_baslat).pack(pady=10)

    def kullanim_kilavuzu_ac(self):
        try:
            ensure_user_guide_present()
            candidates = [
                str(data_dir() / "KULLANIM_KILAVUZU.txt"),
                str(app_dir() / "KULLANIM_KILAVUZU.txt"),
                str(app_dir() / "script" / "assets" / "KULLANIM_KILAVUZU.txt"),
                str(assets_dir() / "KULLANIM_KILAVUZU.txt"),
            ]
            found = next((c for c in candidates if os.path.exists(c)), None)
            if not found:
                raise FileNotFoundError(f"Kılavuz bulunamadı. Beklenen yer: {data_dir()}")
            open_path(found)
        except Exception as e:
            messagebox.showerror("Hata", f"Kılavuz açılamadı:\n{e}")

    def ilk_5_adim_goster(self):
        messagebox.showinfo(
            "İlk 5 Adım",
            "1) İlk gün: 'İLK KURULUM' ile Kurum Müdürü oluştur.\n"
            "2) Çalışanlar: 'KAYIT OL' ile hesap açar.\n"
            "3) SEANS TAKİP: Seansı yaz → KAYDET.\n"
            "4) Tahsilat: Seansı seç → '💰 Ödeme Ekle' butonuna bas.\n"
            "5) Haftalık durum: Muhasebe → Haftalık Ders/Ücret Takip.\n",
        )

    def ucret_takibi_goster(self):
        messagebox.showinfo("Bilgi", "Ücret takibi ana ekranda (SEANS TAKİP) listelenmektedir.")

    def odeme_islemleri(self):
        messagebox.showinfo("Bilgi", "Ödeme eklemek için listeden kaydı seçip '💰 Ödeme Ekle' butonunu kullanabilirsiniz.")

    def kendi_seanslarim(self):
        # Eğitim görevlisi için filtre zaten aktif; sadece ilgili sekmeye geç
        try:
            self.nb.select(self.tab_records)
        except Exception:
            pass
        self.kayitlari_listele()

    def kendi_ucretlerim(self):
        self.kendi_seanslarim()

    def randevu_yonetimi(self):
        # leta_pro ile uyumlu: randevu = seans takvimi
        self.seans_takvimi_goster()

    def gelir_gider_raporu(self):
        win = ttk.Toplevel(self)
        win.title("Gelir-Gider Raporu")
        center_window_smart(win, 900, 620)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)

        ttk.Label(win, text="GELİR-GİDER RAPORU", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill=X)

        ttk.Label(frm, text="Başlangıç:").grid(row=0, column=0, padx=6, pady=6, sticky=W)
        ent_bas = ttk.Entry(frm, width=14)
        ent_bas.insert(0, datetime.datetime.now().strftime("%Y-%m-01"))
        ent_bas.grid(row=0, column=1, padx=6, pady=6, sticky=W)

        ttk.Label(frm, text="Bitiş:").grid(row=0, column=2, padx=6, pady=6, sticky=W)
        ent_bit = ttk.Entry(frm, width=14)
        ent_bit.insert(0, datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_bit.grid(row=0, column=3, padx=6, pady=6, sticky=W)

        out = ttk.Text(win, height=18)
        out.pack(fill=BOTH, expand=True, padx=12, pady=10)

        def _rapor():
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                bas = ent_bas.get().strip()
                bit = ent_bit.get().strip()

                # Kurum müdürü: tüm kasa + seanslar
                if self.kullanici_yetki == "kurum_muduru":
                    cur.execute(
                        """
                        SELECT
                            COALESCE(SUM(alinan_ucret),0),
                            COALESCE(SUM(hizmet_bedeli - alinan_ucret),0),
                            COUNT(*)
                        FROM records
                        WHERE tarih >= ? AND tarih <= ?
                        """,
                        (bas, bit),
                    )
                    gelir_kayit, alacak, toplam = cur.fetchone() or (0, 0, 0)

                    cur.execute(
                        """
                        SELECT
                            COALESCE(SUM(CASE WHEN tip='giren' THEN tutar ELSE 0 END),0),
                            COALESCE(SUM(CASE WHEN tip='cikan' THEN tutar ELSE 0 END),0)
                        FROM kasa_hareketleri
                        WHERE tarih >= ? AND tarih <= ?
                        """,
                        (bas, bit),
                    )
                    kasa_giren, kasa_cikan = cur.fetchone() or (0, 0)
                else:
                    # Eğitim görevlisi: kendi seansları + kendi tahsilatları (record_id üzerinden)
                    ter = self.kullanici_terapist or ""
                    cur.execute(
                        """
                        SELECT
                            COALESCE(SUM(alinan_ucret),0),
                            COALESCE(SUM(hizmet_bedeli - alinan_ucret),0),
                            COUNT(*)
                        FROM records
                        WHERE tarih >= ? AND tarih <= ? AND terapist = ?
                        """,
                        (bas, bit, ter),
                    )
                    gelir_kayit, alacak, toplam = cur.fetchone() or (0, 0, 0)

                    cur.execute(
                        """
                        SELECT
                            COALESCE(SUM(CASE WHEN kh.tip='giren' THEN kh.tutar ELSE 0 END),0),
                            COALESCE(SUM(CASE WHEN kh.tip='cikan' THEN kh.tutar ELSE 0 END),0)
                        FROM kasa_hareketleri kh
                        LEFT JOIN records r ON r.id = kh.record_id
                        WHERE kh.tarih >= ? AND kh.tarih <= ? AND r.terapist = ?
                        """,
                        (bas, bit, ter),
                    )
                    kasa_giren, kasa_cikan = cur.fetchone() or (0, 0)

                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Rapor oluşturulamadı:\n{e}")
                return

            net = float(kasa_giren or 0) - float(kasa_cikan or 0)
            txt = (
                "GELİR-GİDER RAPORU\n"
                + "=" * 45
                + "\n"
                + f"Tarih Aralığı: {ent_bas.get().strip()} - {ent_bit.get().strip()}\n\n"
                + f"Toplam Seans: {toplam}\n"
                + f"Toplam Tahsilat (Kasa): {float(kasa_giren or 0):,.2f} ₺\n"
                + f"Toplam Gider (Kasa): {float(kasa_cikan or 0):,.2f} ₺\n"
                + f"Net (Tahsilat - Gider): {net:,.2f} ₺\n\n"
                + f"Seans Geliri (Kayıt): {float(gelir_kayit or 0):,.2f} ₺\n"
                + f"Toplam Alacak: {alacak:,.2f} ₺\n"
            )
            out.delete("1.0", END)
            out.insert("1.0", txt)

        ttk.Button(frm, text="Rapor Oluştur", bootstyle="primary", command=_rapor).grid(row=0, column=4, padx=10, pady=6)

    def kasa_defteri_goster(self):
        # Kurum müdürü tam yetki; eğitim görevlisi kendi tahsilatlarını görebilir (record_id -> records.terapist)
        win = ttk.Toplevel(self)
        win.title("Kasa Defteri (Günlük)")
        center_window_smart(win, 1200, 780)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        ttk.Label(win, text="KASA DEFTERİ (GÜNLÜK)", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)

        ttk.Label(top, text="Tarih (YYYY-AA-GG):", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=6)
        tarih_var = ttk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_t = ttk.Entry(top, textvariable=tarih_var, width=14)
        ent_t.pack(side=LEFT, padx=6)

        info_lbl = ttk.Label(top, text="", font=("Segoe UI", 10))
        info_lbl.pack(side=LEFT, padx=10)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)

        # ✅ ANLAŞILIR SÜTUN İSİMLERİ
        cols = ("ID", "Hareket Tipi", "Açıklama", "Tutar", "Ödeme Şekli", "İlgili Kayıt", "Oluşturma Tarihi")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140)
        tree.column("ID", width=70)
        tree.column("Tip", width=80)
        tree.column("Açıklama", width=360)
        tree.column("Tutar", width=120)
        tree.column("KayıtID", width=90)
        tree.column("Oluşturma", width=160)

        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _load():
            for iid in tree.get_children():
                tree.delete(iid)
            tarih = (tarih_var.get() or "").strip()
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                if self.kullanici_yetki == "kurum_muduru":
                    cur.execute(
                        """
                        SELECT id, tip, aciklama, tutar, COALESCE(odeme_sekli,''), COALESCE(record_id,''), COALESCE(olusturma_tarihi,'')
                        FROM kasa_hareketleri
                        WHERE tarih=?
                        ORDER BY id ASC
                        """,
                        (tarih,),
                    )
                    rows = cur.fetchall()

                    cur.execute(
                        """
                        SELECT COALESCE(SUM(CASE WHEN tip IN ('giren') THEN tutar ELSE 0 END),0),
                               COALESCE(SUM(CASE WHEN tip IN ('cikan','çıkan') THEN tutar ELSE 0 END),0)
                        FROM kasa_hareketleri WHERE tarih=?
                        """,
                        (tarih,),
                    )
                    giren, cikan = cur.fetchone() or (0, 0)

                    cur.execute(
                        """
                        SELECT COALESCE(SUM(CASE WHEN tip='giren' THEN tutar ELSE -tutar END),0)
                        FROM kasa_hareketleri
                        WHERE tarih < ?
                        """,
                        (tarih,),
                    )
                    devreden = float((cur.fetchone() or [0])[0] or 0)
                else:
                    ter = self.kullanici_terapist or ""
                    cur.execute(
                        """
                        SELECT kh.id, kh.tip, kh.aciklama, kh.tutar, COALESCE(kh.odeme_sekli,''), COALESCE(kh.record_id,''), COALESCE(kh.olusturma_tarihi,'')
                        FROM kasa_hareketleri kh
                        LEFT JOIN records r ON r.id = kh.record_id
                        WHERE kh.tarih=? AND (r.terapist = ? OR kh.record_id IS NULL)
                        ORDER BY kh.id ASC
                        """,
                        (tarih, ter),
                    )
                    rows = cur.fetchall()

                    cur.execute(
                        """
                        SELECT COALESCE(SUM(CASE WHEN kh.tip='giren' THEN kh.tutar ELSE 0 END),0),
                               COALESCE(SUM(CASE WHEN kh.tip IN ('cikan','çıkan') THEN kh.tutar ELSE 0 END),0)
                        FROM kasa_hareketleri kh
                        LEFT JOIN records r ON r.id = kh.record_id
                        WHERE kh.tarih=? AND (r.terapist = ? OR kh.record_id IS NULL)
                        """,
                        (tarih, ter),
                    )
                    giren, cikan = cur.fetchone() or (0, 0)

                    cur.execute(
                        """
                        SELECT COALESCE(SUM(CASE WHEN kh.tip='giren' THEN kh.tutar ELSE -kh.tutar END),0)
                        FROM kasa_hareketleri kh
                        LEFT JOIN records r ON r.id = kh.record_id
                        WHERE kh.tarih < ? AND r.terapist=?
                        """,
                        (tarih, ter),
                    )
                    devreden = float((cur.fetchone() or [0])[0] or 0)

                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Kasa yüklenemedi:\n{e}")
                return

            for idx, r in enumerate(rows):
                tip_display = (r[1] or "").strip()
                if tip_display.lower() in ("cikan", "çıkan"):
                    tip_display = "Çıkan"
                elif tip_display.lower() == "giren":
                    tip_display = "Giren"
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=(r[0], tip_display, r[2], format_money(r[3]), r[4], r[5], r[6]), tags=(tag,))

            bugun_net = float(giren or 0) - float(cikan or 0)
            bugun_kasa = devreden + bugun_net
            info_lbl.config(
                text=f"Devreden: {devreden:,.2f} ₺ | Giren: {float(giren or 0):,.2f} ₺ | Çıkan: {float(cikan or 0):,.2f} ₺ | Bugünkü Kasa: {bugun_kasa:,.2f} ₺"
            )

        def _add(is_gider: bool):
            if self.kullanici_yetki != "kurum_muduru":
                messagebox.showwarning("Yetki", "Bu işlem sadece Kurum Müdürü tarafından yapılabilir.")
                return
            d = ttk.Toplevel(win)
            d.title("Gider Ekle" if is_gider else "Gelir Ekle")
            d.resizable(False, False)
            center_window(d, 520, 280)
            d.transient(win)
            d.grab_set()

            f = ttk.Frame(d, padding=14)
            f.pack(fill=BOTH, expand=True)

            ttk.Label(f, text="Tarih (YYYY-AA-GG):").grid(row=0, column=0, sticky=W, padx=6, pady=6)
            e_t = ttk.Entry(f, width=16)
            e_t.insert(0, (tarih_var.get() or "").strip() or datetime.datetime.now().strftime("%Y-%m-%d"))
            e_t.grid(row=0, column=1, sticky=W, padx=6, pady=6)

            ttk.Label(f, text="Açıklama:").grid(row=1, column=0, sticky=W, padx=6, pady=6)
            e_a = ttk.Entry(f, width=42)
            e_a.grid(row=1, column=1, sticky=W, padx=6, pady=6)
            e_a.focus_set()

            ttk.Label(f, text="Tutar (₺):").grid(row=2, column=0, sticky=W, padx=6, pady=6)
            e_u = ttk.Entry(f, validate="key", validatecommand=self._vcmd_money, width=18)
            e_u.grid(row=2, column=1, sticky=W, padx=6, pady=6)

            ttk.Label(f, text="Ödeme Şekli:").grid(row=3, column=0, sticky=W, padx=6, pady=6)
            cb = ttk.Combobox(f, state="readonly", values=["Nakit", "Havale/EFT", "Kart", "Diğer"], width=16)
            cb.current(0)
            cb.grid(row=3, column=1, sticky=W, padx=6, pady=6)
            
            # ✅ FINANSAL ZEKA: Gider kategorisi (sadece gider için)
            cb_kategori = None
            if is_gider:
                ttk.Label(f, text="Gider Kategorisi:").grid(row=4, column=0, sticky=W, padx=6, pady=6)
                cb_kategori = ttk.Combobox(f, state="readonly", values=["Maaş", "Kira", "Kırtasiye", "Vergi/Fatura", "Diğer"], width=16)
                cb_kategori.current(4)  # Varsayılan: "Diğer"
                cb_kategori.grid(row=4, column=1, sticky=W, padx=6, pady=6)

            def _save():
                try:
                    tutar = parse_money(e_u.get())
                except Exception:
                    messagebox.showerror("Hata", "Lütfen geçerli bir tutar giriniz!")
                    return
                if tutar <= 0:
                    messagebox.showwarning("Uyarı", "Tutar 0'dan büyük olmalıdır!")
                    return
                ac = (e_a.get() or "").strip()
                if not ac:
                    messagebox.showwarning("Uyarı", "Açıklama zorunludur!")
                    return
                try:
                    conn = self.veritabani_baglan()
                    cur = conn.cursor()
                    # ✅ FINANSAL ZEKA: gider_kategorisi ekle
                    gider_kategori = (cb_kategori.get() or "").strip() if is_gider and cb_kategori else ""
                    cur.execute(
                        """
                        INSERT INTO kasa_hareketleri (tarih, tip, aciklama, tutar, odeme_sekli, gider_kategorisi, record_id, seans_id, olusturan_kullanici_id, olusturma_tarihi)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            (e_t.get() or "").strip(),
                            ("cikan" if is_gider else "giren"),
                            ac,
                            tutar,
                            (cb.get() or "").strip(),
                            gider_kategori,  # ✅ FINANSAL ZEKA: Gider kategorisi
                            None,
                            None,
                            (self.kullanici[0] if self.kullanici else None),
                            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        ),
                    )
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Hata", f"Kayıt eklenemedi:\n{e}")
                    return
                d.destroy()
                _load()

            ttk.Button(f, text="KAYDET", bootstyle="success", command=_save).grid(row=5 if is_gider else 4, column=0, columnspan=2, sticky=EW, padx=6, pady=(12, 0))

        def _delete():
            if self.kullanici_yetki != "kurum_muduru":
                messagebox.showwarning("Yetki", "Silme işlemi sadece Kurum Müdürü tarafından yapılabilir.")
                return
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values") or ()
            if not vals:
                return
            kid = vals[0]
            if not messagebox.askyesno("Onay", f"Kasa hareketi silinsin mi?\nID: {kid}"):
                return
            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                pipeline = DataPipeline(conn, kullanici_id)
                basarili = pipeline.kasa_hareketi_sil(kid)
                conn.close()

                if not basarili:
                    messagebox.showerror("Hata", "Kasa kaydı silinemedi!")
                    return

                messagebox.showinfo("Başarılı", "Kasa kaydı silindi!\n\nİlgili tablolar otomatik güncellendi.")
            except Exception as e:
                messagebox.showerror("Hata", f"Silinemedi:\n{e}")
                log_exception("kasa_hareketi_sil_ui", e)
                return
            _load()


        ttk.Button(top, text="Raporu Göster", bootstyle="primary", command=_load).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Yenile", bootstyle="secondary", command=_load).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Gelir Ekle", bootstyle="success", command=lambda: _add(False)).pack(side=RIGHT, padx=6)
        ttk.Button(top, text="Gider Ekle", bootstyle="danger", command=lambda: _add(True)).pack(side=RIGHT, padx=6)
        ttk.Button(top, text="Sil", bootstyle="warning", command=_delete).pack(side=RIGHT, padx=6)

        _load()

    def haftalik_ders_ucret_takip(self):
        win = ttk.Toplevel(self)
        win.title("Haftalık Ders/Ücret Takip")
        center_window_smart(win, 1400, 820)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        ttk.Label(win, text="HAFTALIK DERS / ÜCRET TAKİP", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)

        bugun = datetime.datetime.now()
        hafta_bas = bugun - datetime.timedelta(days=bugun.weekday())
        hafta_var = ttk.StringVar(value=hafta_bas.strftime("%Y-%m-%d"))
        ttk.Label(top, text="Hafta Başlangıcı:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=6)
        ttk.Entry(top, textvariable=hafta_var, width=14).pack(side=LEFT, padx=6)

        ttk.Label(top, text="Terapist:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(16, 6))
        cb = ttk.Combobox(top, state="readonly", width=26)
        cb.pack(side=LEFT, padx=6)

        # terapist listesi
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            tnames = [r[0] for r in cur.fetchall()]
            conn.close()
        except Exception:
            tnames = DEFAULT_THERAPISTS[:]

        if self.kullanici_yetki == "kurum_muduru":
            cb["values"] = ["(Tümü)"] + tnames
            cb.current(0)
        else:
            cb["values"] = [self.kullanici_terapist] if self.kullanici_terapist else tnames[:1]
            cb.current(0)
            cb.configure(state="disabled")

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)

        cols = ("ID", "Tarih", "Saat", "Danışan", "Terapist", "Seans", "Ücret", "Tutar", "Ödeme", "Not")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140)
        tree.column("ID", width=70)
        tree.column("Tarih", width=110)
        tree.column("Saat", width=80)
        tree.column("Danışan", width=220)
        tree.column("Terapist", width=170)
        tree.column("Seans", width=70)
        tree.column("Ücret", width=70)
        tree.column("Tutar", width=110)
        tree.column("Ödeme", width=110)
        tree.column("Not", width=260)

        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        summary = ttk.Label(win, text="", font=("Segoe UI", 10))
        summary.pack(fill=X, padx=12, pady=(0, 10))

        def _load():
            for iid in tree.get_children():
                tree.delete(iid)
            try:
                bas = datetime.datetime.strptime((hafta_var.get() or "").strip(), "%Y-%m-%d")
            except Exception:
                bas = hafta_bas
            bit = (bas + datetime.timedelta(days=6)).strftime("%Y-%m-%d")
            bas_s = bas.strftime("%Y-%m-%d")
            ter_filter = cb.get()

            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                if self.kullanici_yetki == "kurum_muduru" and ter_filter and ter_filter != "(Tümü)":
                    cur.execute(
                        """
                        SELECT id, tarih, saat, danisan_adi, terapist,
                               COALESCE(seans_alindi,0), COALESCE(ucret_alindi,0), COALESCE(ucret_tutar,0), COALESCE(odeme_sekli,''), COALESCE(notlar,'')
                        FROM seans_takvimi
                        WHERE tarih >= ? AND tarih <= ? AND terapist = ?
                        ORDER BY tarih, saat
                        """,
                        (bas_s, bit, ter_filter),
                    )
                elif self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                    cur.execute(
                        """
                        SELECT id, tarih, saat, danisan_adi, terapist,
                               COALESCE(seans_alindi,0), COALESCE(ucret_alindi,0), COALESCE(ucret_tutar,0), COALESCE(odeme_sekli,''), COALESCE(notlar,'')
                        FROM seans_takvimi
                        WHERE tarih >= ? AND tarih <= ? AND terapist = ?
                        ORDER BY tarih, saat
                        """,
                        (bas_s, bit, self.kullanici_terapist),
                    )
                else:
                    cur.execute(
                        """
                        SELECT id, tarih, saat, danisan_adi, terapist,
                               COALESCE(seans_alindi,0), COALESCE(ucret_alindi,0), COALESCE(ucret_tutar,0), COALESCE(odeme_sekli,''), COALESCE(notlar,'')
                        FROM seans_takvimi
                        WHERE tarih >= ? AND tarih <= ?
                        ORDER BY tarih, saat
                        """,
                        (bas_s, bit),
                    )
                rows = cur.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Haftalık kayıtlar yüklenemedi:\n{e}")
                return

            seans_top = len(rows)
            seans_alindi = sum(1 for r in rows if int(r[5] or 0) == 1)
            seans_alinmadi = seans_top - seans_alindi
            ucret_alindi = sum(1 for r in rows if int(r[6] or 0) == 1)
            toplam_tutar = sum(float(r[7] or 0) for r in rows if int(r[6] or 0) == 1)

            for idx, r in enumerate(rows):
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert(
                    "",
                    END,
                    values=(
                        r[0],
                        r[1],
                        r[2],
                        r[3],
                        r[4],
                        ("✓" if int(r[5] or 0) == 1 else ""),
                        ("✓" if int(r[6] or 0) == 1 else ""),
                        format_money(r[7]),
                        r[8],
                        r[9],
                    ),
                    tags=(tag,),
                )

            summary.config(
                text=f"Aralık: {bas_s} - {bit} | Toplam: {seans_top} | Alındı: {seans_alindi} | Alınmadı: {seans_alinmadi} | Ücret Alındı: {ucret_alindi} | Tahsilat: {toplam_tutar:,.2f} ₺"
            )

        def _selected_sid():
            sel = tree.selection()
            if not sel:
                return None
            vals = tree.item(sel[0], "values") or ()
            return vals[0] if vals else None

        def _toggle(col: str):
            """
            PIPELINE ENTEGRASYONU: Seans durum toggle (ATOMIC TRANSACTION)
            → DataPipeline.seans_durum_guncelle() ile tek transaction içinde güncelleme
            """
            sid = _selected_sid()
            if not sid:
                return
            try:
                conn = self.veritabani_baglan()
                kullanici_id = self.kullanici[0] if self.kullanici else None
                
                # ✅ PIPELINE KULLAN (Tek transaction ile atomik güncelleme)
                pipeline = DataPipeline(conn, kullanici_id)
                
                if col == "seans":
                    # Seans alındı/alınmadı toggle
                    basarili = pipeline.seans_durum_guncelle(seans_id=sid, seans_alindi=None)
                else:
                    # Ücret alındı/alınmadı toggle
                    basarili = pipeline.seans_durum_guncelle(seans_id=sid, ucret_alindi=None)
                
                if not basarili:
                    messagebox.showerror("Hata", "Durum güncellenemedi!")
                    return
                
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
                log_exception("_toggle_pipeline", e)
                return
            _load()

        def _edit_amount():
            sid = _selected_sid()
            if not sid:
                return
            d = ttk.Toplevel(win)
            d.title("Ücret / Ödeme Güncelle")
            d.resizable(False, False)
            center_window(d, 520, 320)
            d.transient(win)
            d.grab_set()

            f = ttk.Frame(d, padding=14)
            f.pack(fill=BOTH, expand=True)

            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT tarih, danisan_adi, terapist,
                           COALESCE(hizmet_bedeli,0), COALESCE(odeme_sekli,''), COALESCE(notlar,''),
                           COALESCE(seans_alindi,0), COALESCE(ucret_alindi,0)
                    FROM seans_takvimi WHERE id=?
                    """,
                    (sid,),
                )
                row = cur.fetchone() or ("", "", "", 0, "", "", 0, 0)
                conn.close()
            except Exception:
                row = ("", "", "", 0, "", "", 0, 0)

            seans_tarih = row[0]
            seans_dan = row[1]
            seans_ter = row[2]
            var_seans = ttk.IntVar(value=1 if int(row[6] or 0) == 1 else 0)
            var_ucret = ttk.IntVar(value=1 if int(row[7] or 0) == 1 else 0)

            ttk.Checkbutton(f, text="Seans Alındı", variable=var_seans, bootstyle="success").pack(anchor=W, pady=(0, 6))
            ttk.Checkbutton(f, text="Ücret Alındı", variable=var_ucret, bootstyle="success").pack(anchor=W, pady=(0, 10))

            # ✅ DÜZELTME: "Ücret Tutarı" yerine "Seans Ücreti" (hizmet_bedeli güncellenmeli, alınan ücret değil)
            ttk.Label(f, text="Seans Ücreti (₺):").pack(anchor=W)
            e_u = ttk.Entry(f, validate="key", validatecommand=self._vcmd_money)
            e_u.insert(0, str(row[3] or 0))  # hizmet_bedeli
            e_u.pack(fill=X, pady=6)

            ttk.Label(f, text="Ödeme Şekli:").pack(anchor=W)
            cb2 = ttk.Combobox(f, state="readonly", values=["", "Nakit", "Havale/EFT", "Kart", "Diğer"])
            cb2.set(row[4] or "")
            cb2.pack(fill=X, pady=6)

            ttk.Label(f, text="Not:").pack(anchor=W)
            e_n = ttk.Entry(f)
            e_n.insert(0, row[5] or "")
            e_n.pack(fill=X, pady=6)

            def _save():
                """
                PIPELINE ENTEGRASYONU: Seans durum ve ücret güncelleme (ATOMIC TRANSACTION)
                → DataPipeline.seans_durum_guncelle() ile tek transaction içinde güncelleme
                """
                try:
                    seans_ucreti = parse_money(e_u.get())  # Seans ücreti (hizmet_bedeli)
                except Exception:
                    messagebox.showerror("Hata", "Lütfen geçerli bir tutar giriniz!")
                    return
                try:
                    conn = self.veritabani_baglan()
                    kullanici_id = self.kullanici[0] if self.kullanici else None
                    
                    # ✅ PIPELINE KULLAN (Tek transaction ile atomik güncelleme)
                    # ✅ DÜZELTME: ucret_tutar parametresi hizmet_bedeli güncellemesi için kullanılıyor
                    pipeline = DataPipeline(conn, kullanici_id)
                    basarili = pipeline.seans_durum_guncelle(
                        seans_id=sid,
                        seans_alindi=bool(var_seans.get()),
                        ucret_alindi=bool(var_ucret.get()),
                        ucret_tutar=seans_ucreti,  # Bu aslında hizmet_bedeli güncellemesi için
                        odeme_sekli=(cb2.get() or "").strip(),
                    )
                    
                    if basarili:
                        # ✅ ENTERPRISE: Notları Pipeline üzerinden güncelle (audit trail)
                        try:
                            notlar_text = (e_n.get() or "").strip()
                            if notlar_text:
                                pipeline_not = DataPipeline(conn, kullanici_id)
                                pipeline_not.seans_not_guncelle(sid, notlar_text)
                        except Exception as e:
                            log_exception("seans_not_guncelle_ui", e)
                        
                        messagebox.showinfo("Başarılı", "Seans durumu güncellendi!\n\nTüm tablolar otomatik senkronize edildi!")
                    else:
                        messagebox.showerror("Hata", "Güncellenemedi!")
                        return
                    
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
                    log_exception("_edit_amount_pipeline", e)
                    return
                d.destroy()
                # ✅ OTOMATIK YENİLENME: Tabloları otomatik yenile (Windows/macOS)
                _load()
                if hasattr(self, 'kayitlari_listele'):
                    self.kayitlari_listele()
                self._refresh_borc_tables()

            ttk.Button(f, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(10, 0))

        def _export():
            try:
                bas = datetime.datetime.strptime((hafta_var.get() or "").strip(), "%Y-%m-%d")
            except Exception:
                bas = hafta_bas
            bit = (bas + datetime.timedelta(days=6)).strftime("%Y-%m-%d")
            bas_s = bas.strftime("%Y-%m-%d")
            ter_filter = cb.get()

            try:
                conn = self.veritabani_baglan()
                if self.kullanici_yetki == "kurum_muduru" and ter_filter and ter_filter != "(Tümü)":
                    df = pd.read_sql_query(
                        """
                        SELECT tarih, saat, danisan_adi, terapist,
                               COALESCE(seans_alindi,0) AS seans_alindi,
                               COALESCE(ucret_alindi,0) AS ucret_alindi,
                               COALESCE(ucret_tutar,0) AS ucret_tutar,
                               COALESCE(odeme_sekli,'') AS odeme_sekli,
                               COALESCE(notlar,'') AS notlar
                        FROM seans_takvimi
                        WHERE tarih >= ? AND tarih <= ? AND terapist = ?
                        ORDER BY tarih, saat
                        """,
                        conn,
                        params=(bas_s, bit, ter_filter),
                    )
                elif self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                    df = pd.read_sql_query(
                        """
                        SELECT tarih, saat, danisan_adi, terapist,
                               COALESCE(seans_alindi,0) AS seans_alindi,
                               COALESCE(ucret_alindi,0) AS ucret_alindi,
                               COALESCE(ucret_tutar,0) AS ucret_tutar,
                               COALESCE(odeme_sekli,'') AS odeme_sekli,
                               COALESCE(notlar,'') AS notlar
                        FROM seans_takvimi
                        WHERE tarih >= ? AND tarih <= ? AND terapist = ?
                        ORDER BY tarih, saat
                        """,
                        conn,
                        params=(bas_s, bit, self.kullanici_terapist),
                    )
                else:
                    df = pd.read_sql_query(
                        """
                        SELECT tarih, saat, danisan_adi, terapist,
                               COALESCE(seans_alindi,0) AS seans_alindi,
                               COALESCE(ucret_alindi,0) AS ucret_alindi,
                               COALESCE(ucret_tutar,0) AS ucret_tutar,
                               COALESCE(odeme_sekli,'') AS odeme_sekli,
                               COALESCE(notlar,'') AS notlar
                        FROM seans_takvimi
                        WHERE tarih >= ? AND tarih <= ?
                        ORDER BY tarih, saat
                        """,
                        conn,
                        params=(bas_s, bit),
                    )
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Excel verisi oluşturulamadı:\n{e}")
                return

            path = filedialog.asksaveasfilename(
                title="Excel Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"haftalik_ders_ucret_{bas_s}_{bit}.xlsx",
            )
            if not path:
                return
            try:
                df.to_excel(path, index=False, engine="openpyxl")
                messagebox.showinfo("Başarılı", "Excel'e aktarıldı.")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel aktarma hatası:\n{e}")

        ttk.Button(top, text="Göster", bootstyle="primary", command=_load).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Seans Alındı/Alınmadı", bootstyle="success", command=lambda: _toggle("seans")).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Ücret Alındı/Alınmadı", bootstyle="success", command=lambda: _toggle("ucret")).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Ücret/Not Düzenle", bootstyle="warning", command=_edit_amount).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Excel'e Aktar", bootstyle="secondary", command=_export).pack(side=RIGHT, padx=6)

        # Çift tık: direkt "Ücret/Not Düzenle" (yeni kullanıcılar için kolay)
        tree.bind("<Double-1>", lambda e: _edit_amount())

        _load()

    # --- Seans Takvimi (Günlük / Haftalık) ---
    def seans_takvimi_goster(self):
        win = ttk.Toplevel(self)
        win.title("Günlük Seans Takvimi")
        center_window_smart(win, 1400, 720)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="Tarih (YYYY-AA-GG):", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=6)
        tarih_var = ttk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        ent = ttk.Entry(top, textvariable=tarih_var, width=14)
        ent.pack(side=LEFT, padx=6)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)

        # Terapistler settings tablosundan
        terapistler = []
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            terapistler = [r[0] for r in cur.fetchall()]
            conn.close()
        except Exception:
            terapistler = DEFAULT_THERAPISTS[:]

        # Eğitim görevlisi ise sadece kendisi
        if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
            terapistler = [self.kullanici_terapist]

        saatler = [f"{h:02d}:00" for h in range(9, 20)]
        cols = ["Saat"] + terapistler
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", height=20, style="Strong.Treeview")
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=150)
        tree.column("Saat", width=80)

        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yenile():
            for i in tree.get_children():
                tree.delete(i)
            tarih = (tarih_var.get() or "").strip()
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    "SELECT saat, danisan_adi, terapist, COALESCE(oda,'') FROM seans_takvimi WHERE tarih=? ORDER BY saat",
                    (tarih,),
                )
                seanslar = cur.fetchall()
                conn.close()
            except Exception:
                seanslar = []

            takvim = {s: {t: "" for t in terapistler} for s in saatler}
            for saat, danisan, terapist, oda in seanslar:
                if terapist in takvim.get(saat or "", {}):
                    takvim[saat][terapist] = danisan

            for saat in saatler:
                row = [saat] + [takvim[saat].get(t, "") for t in terapistler]
                idx = saatler.index(saat)
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=row, tags=(tag,))

        ttk.Button(top, text="Göster", bootstyle="primary", command=_yenile).pack(side=LEFT, padx=6)
        _yenile()

    def haftalik_takvim_goster(self):
        win = ttk.Toplevel(self)
        win.title("Haftalık Seans Takvimi")
        center_window_smart(win, 1500, 820)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="Hafta Başlangıcı (YYYY-AA-GG):", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=6)
        bugun = datetime.datetime.now()
        hafta_bas = bugun - datetime.timedelta(days=bugun.weekday())
        hafta_var = ttk.StringVar(value=hafta_bas.strftime("%Y-%m-%d"))
        ent = ttk.Entry(top, textvariable=hafta_var, width=14)
        ent.pack(side=LEFT, padx=6)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)

        gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        saatler = [f"{h:02d}:00" for h in range(9, 20)]
        cols = ["Saat"] + gunler
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", height=20, style="Strong.Treeview")
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=190)
        tree.column("Saat", width=80)

        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yenile():
            for i in tree.get_children():
                tree.delete(i)
            try:
                bas = datetime.datetime.strptime((hafta_var.get() or "").strip(), "%Y-%m-%d")
            except Exception:
                bas = hafta_bas

            takvim = {s: {g: "" for g in gunler} for s in saatler}
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                for gun_idx, gun in enumerate(gunler):
                    tarih = (bas + datetime.timedelta(days=gun_idx)).strftime("%Y-%m-%d")
                    # seans_takvimi varsa oradan, yoksa records'tan
                    try:
                        cur.execute(
                            "SELECT danisan_adi, terapist FROM seans_takvimi WHERE tarih=? ORDER BY saat",
                            (tarih,),
                        )
                        seanslar = cur.fetchall()
                    except Exception:
                        cur.execute("SELECT danisan_adi, terapist FROM records WHERE tarih=? ORDER BY id", (tarih,))
                        seanslar = cur.fetchall()
                    for danisan, terapist in seanslar:
                        if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                            if terapist != self.kullanici_terapist:
                                continue
                        for saat in saatler:
                            if not takvim[saat][gun]:
                                takvim[saat][gun] = f"{danisan} ({terapist})"
                                break
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Haftalık takvim yüklenemedi:\n{e}")

            for saat in saatler:
                row = [saat] + [takvim[saat].get(g, "") for g in gunler]
                idx = saatler.index(saat)
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=row, tags=(tag,))

        ttk.Button(top, text="Göster", bootstyle="primary", command=_yenile).pack(side=LEFT, padx=6)
        _yenile()

    def yeni_seans_ekle(self):
        win = ttk.Toplevel(self)
        win.title("Yeni Seans Ekle")
        center_window_smart(win, 620, 620, max_ratio=0.85)
        win.transient(self)
        win.grab_set()
        self._brand_window(win)

        ttk.Label(win, text="YENİ SEANS EKLE", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text="Tarih (YYYY-AA-GG):").pack(anchor=W)
        tarih_var = ttk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(frm, textvariable=tarih_var).pack(fill=X, pady=6)

        ttk.Label(frm, text="Saat:").pack(anchor=W)
        saat_cb = ttk.Combobox(frm, values=[f"{h:02d}:00" for h in range(9, 20)], state="readonly")
        saat_cb.current(0)
        saat_cb.pack(fill=X, pady=6)

        ttk.Label(frm, text="Danışan:").pack(anchor=W)
        dan_cb = ttk.Combobox(frm)
        dan_cb.pack(fill=X, pady=6)

        ttk.Label(frm, text="Terapist:").pack(anchor=W)
        ter_cb = ttk.Combobox(frm, state="readonly")
        ter_cb.pack(fill=X, pady=6)

        ttk.Label(frm, text="Oda:").pack(anchor=W)
        oda_cb = ttk.Combobox(frm, state="readonly")
        oda_cb.pack(fill=X, pady=6)

        ttk.Label(frm, text="Not:").pack(anchor=W)
        ent_not = ttk.Entry(frm)
        ent_not.pack(fill=X, pady=6)

        # load lists
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            dan_cb["values"] = [r[0] for r in cur.fetchall()]
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            ter_names = [r[0] for r in cur.fetchall()]
            if self.kullanici_yetki != "kurum_muduru" and self.kullanici_terapist:
                ter_names = [self.kullanici_terapist]
            ter_cb["values"] = ter_names
            if ter_names:
                ter_cb.current(0)
            cur.execute("SELECT oda_adi FROM odalar WHERE aktif=1 ORDER BY oda_adi")
            oda_names = [r[0] for r in cur.fetchall()]
            oda_cb["values"] = oda_names
            if oda_names:
                oda_cb.current(0)
            conn.close()
        except Exception:
            pass

        def _kaydet():
            if not (dan_cb.get() or "").strip():
                messagebox.showwarning("Uyarı", "Lütfen danışan seçiniz!")
                return
            if not (ter_cb.get() or "").strip():
                messagebox.showwarning("Uyarı", "Lütfen terapist seçiniz!")
                return
            if "name hoca" in (ter_cb.get() or "").lower():
                messagebox.showwarning("Uyarı", "Name Hoca kurumdan ayrıldı.")
                return
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                t = (tarih_var.get() or "").strip()
                s = (saat_cb.get() or "").strip() or self._default_saat()
                d = (dan_cb.get() or "").strip()
                ter = (ter_cb.get() or "").strip()
                nt = (ent_not.get() or "").strip()
                cur.execute(
                    """
                    INSERT INTO seans_takvimi (tarih, saat, danisan_adi, terapist, oda, durum, notlar, olusturma_tarihi, olusturan_kullanici_id, record_id)
                    VALUES (?,?,?,?,?,?,?,?,?,NULL)
                    """,
                    (
                        t,
                        s,
                        d,
                        ter,
                        (oda_cb.get() or "").strip(),
                        "planlandi",
                        nt,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        self.kullanici[0] if self.kullanici else None,
                    ),
                )
                sid = int(cur.lastrowid or 0)
                try:
                    self._sync_from_seans_to_record(cur, sid, t, s, d, ter, nt)
                except Exception:
                    pass
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Seans eklenemedi:\n{e}")
                return
            win.destroy()
            messagebox.showinfo("Başarılı", "Seans eklendi.")

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_kaydet).pack(fill=X, pady=(12, 0))

    # --- Danışan Yönetimi ---
    def danisan_yonetimi(self):
        win = ttk.Toplevel(self)
        win.title("Danışan Yönetimi")
        center_window_smart(win, 1300, 760)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="DANIŞAN YÖNETİMİ", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(side=LEFT)
        ttk.Button(top, text="Yeni Danışan Ekle", bootstyle="success", command=lambda: self.yeni_danisan_ekle(win)).pack(side=RIGHT)
        ttk.Button(top, text="Düzenle", bootstyle="warning", command=lambda: self.danisan_duzenle(win)).pack(side=RIGHT, padx=6)
        ttk.Button(top, text="Aktif/Pasif", bootstyle="secondary", command=lambda: self.danisan_aktif_pasif(win)).pack(side=RIGHT, padx=6)

        ar = ttk.Frame(win, padding=(10, 0))
        ar.pack(fill=X)
        ttk.Label(ar, text="Ara:").pack(side=LEFT, padx=6)
        ent_ara = ttk.Entry(ar)
        ent_ara.pack(side=LEFT, fill=X, expand=True, padx=6)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("ID", "Ad Soyad", "Telefon", "Veli Adı", "Veli Telefon", "E-posta", "Durum")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.column("ID", width=60)
        tree.column("Ad Soyad", width=240)
        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yukle():
            for i in tree.get_children():
                tree.delete(i)
            kelime = (ent_ara.get() or "").strip()
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                sql = "SELECT id, ad_soyad, telefon, veli_adi, veli_telefon, email, aktif FROM danisanlar"
                params = []
                if kelime:
                    sql += " WHERE ad_soyad LIKE ?"
                    params.append(f"%{kelime}%")
                sql += " ORDER BY ad_soyad"
                cur.execute(sql, tuple(params))
                rows = cur.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Danışanlar yüklenemedi:\n{e}")
                return
            for idx, r in enumerate(rows):
                durum = "Aktif" if (r[6] or 0) == 1 else "Pasif"
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=(r[0], r[1], r[2] or "", r[3] or "", r[4] or "", r[5] or "", durum), tags=(tag,))

        ent_ara.bind("<KeyRelease>", lambda e: _yukle())
        _yukle()
        win.danisan_tree = tree
        win._reload = _yukle

    def _yeni_danisan_ekle_ve_guncelle(self, combobox, parent):
        """Yeni danışan ekle ve combobox'ı otomatik güncelle (terapist mantığı gibi)"""
        win = ttk.Toplevel(parent)
        win.title("Yeni Danışan Ekle")
        center_window_smart(win, 620, 760, max_ratio=0.88)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)

        ttk.Label(win, text="YENİ DANIŞAN EKLE", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)

        def field(label):
            ttk.Label(frm, text=label).pack(anchor=W, pady=(8, 0))
            e = ttk.Entry(frm)
            e.pack(fill=X, pady=4)
            return e

        ent_ad = field("Ad Soyad *:")
        ent_tel = field("Telefon:")
        ent_email = field("E-posta:")
        ent_veli = field("Veli Adı:")
        ent_veli_tel = field("Veli Telefon:")
        ent_dogum = field("Doğum Tarihi:")
        ttk.Label(frm, text="Adres:").pack(anchor=W, pady=(8, 0))
        txt_adres = ttk.Text(frm, height=3)
        txt_adres.pack(fill=X, pady=4)
        ttk.Label(frm, text="Notlar:").pack(anchor=W, pady=(8, 0))
        txt_not = ttk.Text(frm, height=3)
        txt_not.pack(fill=X, pady=4)

        def _kaydet():
            if not (ent_ad.get() or "").strip():
                messagebox.showwarning("Uyarı", "Ad Soyad zorunludur!")
                return
            try:
                danisan_adi = (ent_ad.get() or "").strip()
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    INSERT INTO danisanlar (ad_soyad, telefon, email, veli_adi, veli_telefon, dogum_tarihi, adres, notlar, olusturma_tarihi, aktif)
                    VALUES (?,?,?,?,?,?,?,?,?,1)
                    """,
                    (
                        danisan_adi,
                        (ent_tel.get() or "").strip(),
                        (ent_email.get() or "").strip(),
                        (ent_veli.get() or "").strip(),
                        (ent_veli_tel.get() or "").strip(),
                        (ent_dogum.get() or "").strip(),
                        (txt_adres.get("1.0", END) or "").strip(),
                        (txt_not.get("1.0", END) or "").strip(),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                conn.commit()
                conn.close()
                
                # Tüm danışan combobox'larını güncelle
                try:
                    conn2 = self.veritabani_baglan()
                    cur2 = conn2.cursor()
                    cur2.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
                    danisan_listesi = [row[0] for row in cur2.fetchall()]
                    conn2.close()
                    
                    # Ana combobox'ı güncelle
                    if hasattr(self, 'cmb_danisan'):
                        self.cmb_danisan["values"] = danisan_listesi
                        self.cmb_danisan.set(danisan_adi)  # Yeni eklenen danışanı seç
                    
                    # Parametre olarak gelen combobox'ı güncelle
                    combobox["values"] = danisan_listesi
                    combobox.set(danisan_adi)  # Yeni eklenen danışanı seç
                    
                except Exception:
                    pass
                
            except Exception as e:
                messagebox.showerror("Hata", f"Danışan eklenemedi:\n{e}")
                return
            messagebox.showinfo("Başarılı", f"Danışan eklendi: {danisan_adi}")
            win.destroy()

        btns = ttk.Frame(frm)
        btns.pack(fill=X, pady=(16, 0))
        ttk.Button(btns, text="KAYDET", bootstyle="success", command=_kaydet).pack(side=LEFT, fill=X, expand=True, padx=6)
        ttk.Button(btns, text="İptal", bootstyle="secondary", command=win.destroy).pack(side=LEFT, fill=X, expand=True, padx=6)

    def yeni_danisan_ekle(self, parent):
        """Eski fonksiyon - Geriye uyumluluk için"""
        # Ana combobox yoksa normal pencere aç
        if hasattr(self, 'cmb_danisan'):
            self._yeni_danisan_ekle_ve_guncelle(self.cmb_danisan, parent)
            return
        
        # Eski yöntem - sadece pencere aç
        win = ttk.Toplevel(parent)
        win.title("Yeni Danışan Ekle")
        center_window_smart(win, 620, 760, max_ratio=0.88)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)

        ttk.Label(win, text="YENİ DANIŞAN EKLE", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)

        def field(label):
            ttk.Label(frm, text=label).pack(anchor=W, pady=(8, 0))
            e = ttk.Entry(frm)
            e.pack(fill=X, pady=4)
            return e

        ent_ad = field("Ad Soyad *:")
        ent_tel = field("Telefon:")
        ent_email = field("E-posta:")
        ent_veli = field("Veli Adı:")
        ent_veli_tel = field("Veli Telefon:")
        ent_dogum = field("Doğum Tarihi:")
        ttk.Label(frm, text="Adres:").pack(anchor=W, pady=(8, 0))
        txt_adres = ttk.Text(frm, height=3)
        txt_adres.pack(fill=X, pady=4)
        ttk.Label(frm, text="Notlar:").pack(anchor=W, pady=(8, 0))
        txt_not = ttk.Text(frm, height=3)
        txt_not.pack(fill=X, pady=4)

        def _kaydet():
            if not (ent_ad.get() or "").strip():
                messagebox.showwarning("Uyarı", "Ad Soyad zorunludur!")
                return
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    INSERT INTO danisanlar (ad_soyad, telefon, email, veli_adi, veli_telefon, dogum_tarihi, adres, notlar, olusturma_tarihi, aktif)
                    VALUES (?,?,?,?,?,?,?,?,?,1)
                    """,
                    (
                        (ent_ad.get() or "").strip(),
                        (ent_tel.get() or "").strip(),
                        (ent_email.get() or "").strip(),
                        (ent_veli.get() or "").strip(),
                        (ent_veli_tel.get() or "").strip(),
                        (ent_dogum.get() or "").strip(),
                        (txt_adres.get("1.0", END) or "").strip(),
                        (txt_not.get("1.0", END) or "").strip(),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                conn.commit()
                conn.close()
                
                # Danışan eklendikten sonra combobox'ları yenile
                if hasattr(self, 'cmb_danisan'):
                        try:
                            conn2 = self.veritabani_baglan()
                            cur2 = conn2.cursor()
                            cur2.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
                            danisan_listesi = [row[0] for row in cur2.fetchall()]
                            conn2.close()
                            self.cmb_danisan["values"] = danisan_listesi
                        except Exception:
                            pass
                    
            except Exception as e:
                messagebox.showerror("Hata", f"Danışan eklenemedi:\n{e}")
                return
            messagebox.showinfo("Başarılı", "Danışan eklendi.")
            win.destroy()

        btns = ttk.Frame(win, padding=(16, 0))
        btns.pack(fill=X, pady=10)
        ttk.Button(btns, text="KAYDET", bootstyle="success", command=_kaydet).pack(side=LEFT, fill=X, expand=True, padx=6)
        ttk.Button(btns, text="İptal", bootstyle="secondary", command=win.destroy).pack(side=LEFT, fill=X, expand=True, padx=6)

    # --- Görev Takibi ---
    def gorev_takibi(self):
        win = ttk.Toplevel(self)
        win.title("Görev Takibi")
        center_window_smart(win, 1300, 760)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="GÖREV TAKİBİ", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(side=LEFT)
        ttk.Button(top, text="Yeni Görev Ekle", bootstyle="success", command=lambda: self.yeni_gorev_ekle(win)).pack(side=RIGHT)
        ttk.Button(top, text="Durum Güncelle", bootstyle="warning", command=lambda: self.gorev_durum_guncelle(win)).pack(side=RIGHT, padx=6)
        ttk.Button(top, text="Tamamla", bootstyle="success-outline", command=lambda: self.gorev_tamamla(win)).pack(side=RIGHT, padx=6)
        ttk.Button(top, text="Sil", bootstyle="danger-outline", command=lambda: self.gorev_sil(win)).pack(side=RIGHT, padx=6)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("ID", "Başlık", "Açıklama", "Atanan", "Durum", "Öncelik", "Başlangıç", "Bitiş")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140)
        tree.column("ID", width=60)
        tree.column("Başlık", width=220)
        tree.column("Açıklama", width=320)
        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yukle():
            for i in tree.get_children():
                tree.delete(i)
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT g.id, g.baslik, g.aciklama, u.full_name, g.durum, g.oncelik, g.baslangic_tarihi, g.bitis_tarihi
                    FROM gorevler g
                    LEFT JOIN users u ON g.atanan_kullanici_id = u.id
                    ORDER BY g.id DESC
                    """
                )
                rows = cur.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Görevler yüklenemedi:\n{e}")
                return
            for idx, r in enumerate(rows):
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=r, tags=(tag,))

        _yukle()
        win.gorev_tree = tree
        win._reload = _yukle

    def yeni_gorev_ekle(self, parent):
        win = ttk.Toplevel(parent)
        win.title("Yeni Görev Ekle")
        center_window_smart(win, 700, 720, max_ratio=0.9)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)

        ttk.Label(win, text="YENİ GÖREV EKLE", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text="Başlık *:").pack(anchor=W)
        ent_bas = ttk.Entry(frm)
        ent_bas.pack(fill=X, pady=6)

        ttk.Label(frm, text="Açıklama:").pack(anchor=W)
        txt = ttk.Text(frm, height=5)
        txt.pack(fill=X, pady=6)

        ttk.Label(frm, text="Atanan:").pack(anchor=W)
        cb = ttk.Combobox(frm, state="readonly")
        cb.pack(fill=X, pady=6)

        ttk.Label(frm, text="Öncelik:").pack(anchor=W)
        cb_on = ttk.Combobox(frm, values=["Düşük", "Normal", "Yüksek", "Acil"], state="readonly")
        cb_on.current(1)
        cb_on.pack(fill=X, pady=6)

        ttk.Label(frm, text="Bitiş Tarihi (YYYY-AA-GG):").pack(anchor=W)
        ent_bit = ttk.Entry(frm)
        ent_bit.pack(fill=X, pady=6)

        # load users
        users_map = {}
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT id, COALESCE(full_name, username) FROM users WHERE is_active=1 ORDER BY full_name, username")
            vals = []
            for uid, nm in cur.fetchall():
                users_map[f"{nm} ({uid})"] = uid
                vals.append(f"{nm} ({uid})")
            cb["values"] = vals
            if vals:
                cb.current(0)
            conn.close()
        except Exception:
            pass

        def _kaydet():
            if not (ent_bas.get() or "").strip():
                messagebox.showwarning("Uyarı", "Başlık zorunludur!")
                return
            atanan_id = users_map.get(cb.get()) if cb.get() else None
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    INSERT INTO gorevler (baslik, aciklama, atanan_kullanici_id, olusturan_kullanici_id, durum, oncelik, baslangic_tarihi, bitis_tarihi, olusturma_tarihi)
                    VALUES (?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        (ent_bas.get() or "").strip(),
                        (txt.get("1.0", END) or "").strip(),
                        atanan_id,
                        self.kullanici[0] if self.kullanici else None,
                        "beklemede",
                        (cb_on.get() or "normal").lower(),
                        datetime.datetime.now().strftime("%Y-%m-%d"),
                        (ent_bit.get() or "").strip() or None,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Görev eklenemedi:\n{e}")
                return
            messagebox.showinfo("Başarılı", "Görev eklendi.")
            win.destroy()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_kaydet).pack(fill=X, pady=(12, 0))

    # --- Kullanıcı Yönetimi ---
    def kullanicilari_listele(self):
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Uyarı", "Bu işlem sadece Kurum Müdürü yetkisi ile yapılabilir.")
            return
        win = ttk.Toplevel(self)
        win.title("Kullanıcılar")
        center_window_smart(win, 1300, 760)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        ttk.Label(win, text="KULLANICI LİSTESİ", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        btns = ttk.Frame(win, padding=(10, 0, 10, 10))
        btns.pack(fill=X)
        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("ID", "Kullanıcı Adı", "Ad Soyad", "Rol (Unvan)", "Terapist", "Yetki", "Durum", "Oluşturma", "Son Giriş")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.column("ID", width=60)
        tree.column("Rol (Unvan)", width=190)
        tree.column("Terapist", width=170)
        tree.column("Yetki", width=140)
        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yukle():
            for i in tree.get_children():
                tree.delete(i)
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT
                        u.id,
                        u.username,
                        COALESCE(u.full_name,''),
                        COALESCE(NULLIF(u.title_role,''), s.therapist_role, ''),
                        COALESCE(u.therapist_name,''),
                        COALESCE(u.access_role, u.role, ''),
                        COALESCE(u.is_active,1),
                        COALESCE(u.created_at,''),
                        COALESCE(u.last_login,'')
                    FROM users u
                    LEFT JOIN settings s ON s.therapist_name = u.therapist_name
                    ORDER BY u.id
                    """
                )
                rows = cur.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Kullanıcılar yüklenemedi:\n{e}")
                return
            for idx, r in enumerate(rows):
                durum = "Aktif" if int(r[6] or 0) == 1 else "Pasif"
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=(r[0], r[1], r[2], r[3], r[4], r[5], durum, r[7], r[8]), tags=(tag,))

        def _selected_user_id():
            sel = tree.selection()
            if not sel:
                return None
            vals = tree.item(sel[0]).get("values") or []
            if not vals:
                return None
            try:
                return int(vals[0])
            except Exception:
                return None

        def _rol_duzenle():
            uid = _selected_user_id()
            if not uid:
                messagebox.showwarning("Uyarı", "Lütfen bir kullanıcı seçiniz!")
                return
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT
                        u.username,
                        COALESCE(u.full_name,''),
                        COALESCE(u.access_role, u.role, 'egitim_gorevlisi'),
                        COALESCE(u.therapist_name,''),
                        COALESCE(NULLIF(u.title_role,''), s.therapist_role, ''),
                        COALESCE(u.is_active,1)
                    FROM users u
                    LEFT JOIN settings s ON s.therapist_name = u.therapist_name
                    WHERE u.id=?
                    """,
                    (uid,),
                )
                row = cur.fetchone()
                # terapistler + unvanlar
                cur.execute("SELECT therapist_name, COALESCE(therapist_role,'') FROM settings WHERE is_active=1 ORDER BY therapist_name")
                terapist_rows = cur.fetchall()
                terapistler = [r[0] for r in terapist_rows]
                role_map = {r[0]: (r[1] or "") for r in terapist_rows}
                cur.execute("SELECT DISTINCT COALESCE(therapist_role,'') FROM settings WHERE is_active=1 AND COALESCE(therapist_role,'')<>'' ORDER BY therapist_role")
                unvanlar = [r[0] for r in cur.fetchall()]
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Kullanıcı bilgisi alınamadı:\n{e}")
                return
            if not row:
                return
            uname, full_name, access_role, therapist_name, title_role, is_active = row

            dlg = ttk.Toplevel(win)
            dlg.title("Rol / Terapist Düzenle")
            dlg.geometry("420x340")
            dlg.transient(win)
            dlg.grab_set()
            frm = ttk.Frame(dlg, padding=16)
            frm.pack(fill=BOTH, expand=True)

            ttk.Label(frm, text=f"Kullanıcı: {uname}", font=("Segoe UI", 11, "bold")).pack(anchor=W, pady=(0, 10))

            ttk.Label(frm, text="Terapist (opsiyonel):").pack(anchor=W)
            cb_ter = ttk.Combobox(frm, state="readonly", values=[""] + terapistler)
            cb_ter.pack(fill=X, pady=6)
            cb_ter.set(therapist_name or "")

            ttk.Label(frm, text="Rol (Unvan):").pack(anchor=W)
            cb_unvan = ttk.Combobox(frm, state="readonly", values=unvanlar + (["Diğer"] if "Diğer" not in unvanlar else []))
            cb_unvan.pack(fill=X, pady=6)
            cb_unvan.set(title_role or (role_map.get(therapist_name or "", "") or ""))

            ttk.Label(frm, text="Yetki:").pack(anchor=W)
            cb_yetki = ttk.Combobox(frm, state="readonly", values=["kurum_muduru", "egitim_gorevlisi"])
            cb_yetki.pack(fill=X, pady=6)
            cb_yetki.set(access_role or "egitim_gorevlisi")

            ttk.Label(
                frm,
                text="Not: Yetki 'egitim_gorevlisi' ise terapist seçilmesi önerilir (kendi kayıtları filtrelenir).",
                foreground="gray",
                wraplength=380,
            ).pack(anchor=W, pady=(8, 0))

            def _sync_unvan(_evt=None):
                tname = (cb_ter.get() or "").strip()
                if not tname:
                    return
                auto = role_map.get(tname, "")
                if auto:
                    cb_unvan.set(auto)

            cb_ter.bind("<<ComboboxSelected>>", _sync_unvan)

            def _save():
                new_access = (cb_yetki.get() or "egitim_gorevlisi").strip()
                new_ter = (cb_ter.get() or "").strip() or None
                new_title = (cb_unvan.get() or "").strip()
                try:
                    conn2 = self.veritabani_baglan()
                    cur2 = conn2.cursor()
                    # Son kurum müdürü demote edilemesin
                    if (access_role or "") == "kurum_muduru" and new_access != "kurum_muduru" and int(is_active or 0) == 1:
                        cur2.execute(
                            "SELECT COUNT(*) FROM users WHERE is_active=1 AND COALESCE(access_role, role)='kurum_muduru' AND id<>?",
                            (uid,),
                        )
                        if int((cur2.fetchone() or [0])[0] or 0) <= 0:
                            conn2.close()
                            messagebox.showwarning("Uyarı", "Son aktif kurum müdürü rolü değiştirilemez.")
                            return
                    cur2.execute(
                        "UPDATE users SET access_role=?, role=?, title_role=?, therapist_name=? WHERE id=?",
                        (new_access, new_access, new_title, new_ter, uid),
                    )
                    conn2.commit()
                    conn2.close()
                except Exception as e:
                    messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
                    return
                dlg.destroy()
                _yukle()

            ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(14, 0))

        ttk.Button(btns, text="Rol/Terapist Düzenle", bootstyle="warning", command=_rol_duzenle).pack(side=LEFT, padx=4)
        ttk.Button(btns, text="Yenile", bootstyle="secondary", command=_yukle).pack(side=LEFT, padx=4)

        _yukle()

    def kullanici_sil(self):
        if self.kullanici_yetki != "kurum_muduru":
            messagebox.showwarning("Uyarı", "Bu işlem sadece Kurum Müdürü yetkisi ile yapılabilir.")
            return
        win = ttk.Toplevel(self)
        win.title("Kullanıcı Sil / Pasif Et")
        center_window_smart(win, 900, 620)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        ttk.Label(win, text="KULLANICI SİL / PASİF ET", font=("Segoe UI", 14, "bold"), bootstyle="danger").pack(pady=10)
        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("ID", "Kullanıcı", "Ad Soyad", "Yetki", "Terapist", "Durum")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140)
        tree.column("ID", width=70)
        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yukle():
            for i in tree.get_children():
                tree.delete(i)
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                "SELECT id, username, COALESCE(full_name,''), COALESCE(access_role, role, ''), COALESCE(therapist_name,''), COALESCE(is_active,1) FROM users ORDER BY id"
            )
            rows = cur.fetchall()
            conn.close()
            for idx, r in enumerate(rows):
                # r: (id, username, full_name, access_role, therapist_name, is_active)
                durum = "Aktif" if int(r[5] or 0) == 1 else "Pasif"
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=(r[0], r[1], r[2], r[3], r[4], durum), tags=(tag,))

        def _pasif_et():
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0]).get("values") or []
            uid = int(vals[0])
            uname = str(vals[1])
            rol = str(vals[3])
            try:
                conn_chk = self.veritabani_baglan()
                cur_chk = conn_chk.cursor()
                if rol == "kurum_muduru":
                    cur_chk.execute(
                        "SELECT COUNT(*) FROM users WHERE is_active=1 AND COALESCE(access_role, role)='kurum_muduru' AND id<>?",
                        (uid,),
                    )
                    if int((cur_chk.fetchone() or [0])[0] or 0) <= 0:
                        conn_chk.close()
                        messagebox.showwarning("Uyarı", "Son aktif kurum müdürü pasif edilemez.")
                        return
                conn_chk.close()
            except Exception:
                pass
            if not messagebox.askyesno("Onay", f"'{uname}' kullanıcısı pasif yapılsın mı?"):
                return
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("UPDATE users SET is_active=0 WHERE id=?", (uid,))
            conn.commit()
            conn.close()
            _yukle()

        ttk.Button(win, text="Seçiliyi Pasif Et", bootstyle="danger", command=_pasif_et).pack(pady=10)
        _yukle()

    # --- EK MODÜL: ODA YÖNETİMİ (leta_pro) ---
    def odalar_yonetimi(self):
        win = ttk.Toplevel(self)
        win.title("Oda Yönetimi")
        center_window_smart(win, 1100, 720)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="ODA YÖNETİMİ", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(side=LEFT)
        ttk.Button(top, text="Yeni Oda Ekle", bootstyle="success", command=lambda: self.oda_ekle(win)).pack(side=RIGHT)
        ttk.Button(top, text="Düzenle", bootstyle="warning", command=lambda: self.oda_duzenle(win)).pack(side=RIGHT, padx=6)
        ttk.Button(top, text="Aktif/Pasif", bootstyle="secondary", command=lambda: self.oda_aktif_pasif(win)).pack(side=RIGHT, padx=6)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("ID", "Oda", "Tip", "Kapasite", "Açıklama", "Durum")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.column("ID", width=60)
        tree.column("Oda", width=220)
        tree.column("Açıklama", width=260)

        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        def _yukle():
            for i in tree.get_children():
                tree.delete(i)
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute("SELECT id, oda_adi, COALESCE(oda_tipi,''), COALESCE(kapasite,''), COALESCE(aciklama,''), COALESCE(aktif,1) FROM odalar ORDER BY oda_adi")
                rows = cur.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Odalar yüklenemedi:\n{e}")
                return
            for idx, r in enumerate(rows):
                durum = "Aktif" if int(r[5] or 0) == 1 else "Pasif"
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert("", END, values=(r[0], r[1], r[2], r[3], r[4], durum), tags=(tag,))

        _yukle()
        win.oda_tree = tree
        win._reload = _yukle

    def oda_ekle(self, parent):
        win = ttk.Toplevel(parent)
        win.title("Yeni Oda Ekle")
        center_window_smart(win, 600, 640, max_ratio=0.9)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)

        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)
        ttk.Label(frm, text="ODA ADI *:").pack(anchor=W, pady=(4, 0))
        ent_ad = ttk.Entry(frm)
        ent_ad.pack(fill=X, pady=6)

        ttk.Label(frm, text="Oda Tipi:").pack(anchor=W, pady=(4, 0))
        ent_tip = ttk.Entry(frm)
        ent_tip.pack(fill=X, pady=6)

        ttk.Label(frm, text="Kapasite:").pack(anchor=W, pady=(4, 0))
        ent_kap = ttk.Entry(frm)
        ent_kap.pack(fill=X, pady=6)

        ttk.Label(frm, text="Açıklama:").pack(anchor=W, pady=(4, 0))
        txt = ttk.Text(frm, height=4)
        txt.pack(fill=X, pady=6)

        def _save():
            oda = (ent_ad.get() or "").strip()
            if not oda:
                messagebox.showwarning("Uyarı", "Oda adı zorunludur!")
                return
            try:
                kapasite = int((ent_kap.get() or "0").strip() or "0")
            except Exception:
                kapasite = 0
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    "INSERT OR IGNORE INTO odalar (oda_adi, oda_tipi, kapasite, aciklama, aktif) VALUES (?,?,?,?,1)",
                    (oda, (ent_tip.get() or "").strip(), kapasite, (txt.get('1.0', END) or '').strip()),
                )
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Oda eklenemedi:\n{e}")
                return
            win.destroy()
            if hasattr(parent, "_reload"):
                parent._reload()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(10, 0))

    def oda_duzenle(self, parent):
        if not hasattr(parent, "oda_tree"):
            return
        sel = parent.oda_tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir oda seçiniz!")
            return
        vals = parent.oda_tree.item(sel[0]).get("values") or []
        oid = int(vals[0])

        win = ttk.Toplevel(parent)
        win.title("Oda Düzenle")
        center_window_smart(win, 600, 640, max_ratio=0.9)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)

        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)
        ttk.Label(frm, text="ODA ADI *:").pack(anchor=W, pady=(4, 0))
        ent_ad = ttk.Entry(frm)
        ent_ad.insert(0, vals[1] or "")
        ent_ad.pack(fill=X, pady=6)

        ttk.Label(frm, text="Oda Tipi:").pack(anchor=W, pady=(4, 0))
        ent_tip = ttk.Entry(frm)
        ent_tip.insert(0, vals[2] or "")
        ent_tip.pack(fill=X, pady=6)

        ttk.Label(frm, text="Kapasite:").pack(anchor=W, pady=(4, 0))
        ent_kap = ttk.Entry(frm)
        ent_kap.insert(0, str(vals[3] or ""))
        ent_kap.pack(fill=X, pady=6)

        ttk.Label(frm, text="Açıklama:").pack(anchor=W, pady=(4, 0))
        txt = ttk.Text(frm, height=4)
        txt.insert("1.0", vals[4] or "")
        txt.pack(fill=X, pady=6)

        def _save():
            oda = (ent_ad.get() or "").strip()
            if not oda:
                messagebox.showwarning("Uyarı", "Oda adı zorunludur!")
                return
            try:
                kapasite = int((ent_kap.get() or "0").strip() or "0")
            except Exception:
                kapasite = 0
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    "UPDATE odalar SET oda_adi=?, oda_tipi=?, kapasite=?, aciklama=? WHERE id=?",
                    (oda, (ent_tip.get() or "").strip(), kapasite, (txt.get('1.0', END) or '').strip(), oid),
                )
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Oda güncellenemedi:\n{e}")
                return
            win.destroy()
            if hasattr(parent, "_reload"):
                parent._reload()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(10, 0))

    def oda_aktif_pasif(self, parent):
        if not hasattr(parent, "oda_tree"):
            return
        sel = parent.oda_tree.selection()
        if not sel:
            return
        vals = parent.oda_tree.item(sel[0]).get("values") or []
        oid = int(vals[0])
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT COALESCE(aktif,1) FROM odalar WHERE id=?", (oid,))
            aktif_mevcut = int((cur.fetchone() or [1])[0] or 0)
            yeni_aktif = (aktif_mevcut == 0)  # Toggle
            
            # ✅ ENTERPRISE: Pipeline üzerinden güncelle (audit trail)
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            basarili = pipeline.oda_durum_guncelle(oid, yeni_aktif)
            conn.close()
            
            if not basarili:
                messagebox.showerror("Hata", "Oda durumu güncellenemedi!")
                return
                
            messagebox.showinfo("Başarılı", f"Oda durumu {'aktif' if yeni_aktif else 'pasif'} olarak güncellendi!")
        except Exception as e:
            messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
            log_exception("oda_aktif_pasif", e)
            return
        if hasattr(parent, "_reload"):
            parent._reload()

    # --- EK MODÜL: Danışan Düzenle / Aktif-Pasif ---
    def danisan_duzenle(self, parent):
        if not hasattr(parent, "danisan_tree"):
            return
        sel = parent.danisan_tree.selection()
        if not sel:
            messagebox.showwarning("Uyarı", "Lütfen bir danışan seçiniz!")
            return
        vals = parent.danisan_tree.item(sel[0]).get("values") or []
        did = int(vals[0])
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad, telefon, email, veli_adi, veli_telefon, dogum_tarihi, adres, notlar FROM danisanlar WHERE id=?", (did,))
            row = cur.fetchone()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Danışan okunamadı:\n{e}")
            return
        if not row:
            return

        win = ttk.Toplevel(parent)
        win.title("Danışan Düzenle")
        center_window_smart(win, 720, 820, max_ratio=0.9)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)

        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)

        def field(label, init=""):
            ttk.Label(frm, text=label).pack(anchor=W, pady=(8, 0))
            e = ttk.Entry(frm)
            e.insert(0, init or "")
            e.pack(fill=X, pady=4)
            return e

        # Tabloya göre: AD SOYAD, DOĞUM TARİHİ, VELİ ADI, VELİ TELEFON, ADRES, NOTLAR (danışan telefon/email yok; iletişim veli bilgilerinden)
        ent_ad = field("AD SOYAD *:", row[0])
        ent_dogum = field("DOĞUM TARİHİ (GG.AA.YYYY):", row[5] or "")
        ent_veli = field("VELİ ADI:", row[3] or "")
        ent_veli_tel = field("VELİ TELEFON:", row[4] or "")
        ttk.Label(frm, text="ADRES:").pack(anchor=W, pady=(8, 0))
        txt_adres = ttk.Text(frm, height=3)
        txt_adres.insert("1.0", row[6] or "")
        txt_adres.pack(fill=X, pady=4)
        ttk.Label(frm, text="NOTLAR:").pack(anchor=W, pady=(8, 0))
        txt_not = ttk.Text(frm, height=3)
        txt_not.insert("1.0", row[7] or "")
        txt_not.pack(fill=X, pady=4)

        def _save():
            if not (ent_ad.get() or "").strip():
                messagebox.showwarning("Uyarı", "AD SOYAD zorunludur!")
                return
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    UPDATE danisanlar
                    SET ad_soyad=?, veli_adi=?, veli_telefon=?, dogum_tarihi=?, adres=?, notlar=?
                    WHERE id=?
                    """,
                    (
                        (ent_ad.get() or "").strip(),
                        (ent_veli.get() or "").strip(),
                        (ent_veli_tel.get() or "").strip(),
                        (ent_dogum.get() or "").strip(),
                        (txt_adres.get("1.0", END) or "").strip(),
                        (txt_not.get("1.0", END) or "").strip(),
                        did,
                    ),
                )
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
                return
            win.destroy()
            if hasattr(parent, "_reload"):
                parent._reload()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(12, 0))

    def danisan_aktif_pasif(self, parent):
        if not hasattr(parent, "danisan_tree"):
            return
        sel = parent.danisan_tree.selection()
        if not sel:
            return
        vals = parent.danisan_tree.item(sel[0]).get("values") or []
        did = int(vals[0])
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT COALESCE(aktif,1) FROM danisanlar WHERE id=?", (did,))
            aktif_mevcut = int((cur.fetchone() or [1])[0] or 0)
            yeni_aktif = (aktif_mevcut == 0)  # Toggle
            
            # ✅ ENTERPRISE: Pipeline üzerinden güncelle (audit trail)
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            basarili = pipeline.danisan_durum_guncelle(did, yeni_aktif)
            conn.close()
            
            if not basarili:
                messagebox.showerror("Hata", "Danışan durumu güncellenemedi!")
                return
                
            messagebox.showinfo("Başarılı", f"Danışan durumu {'aktif' if yeni_aktif else 'pasif'} olarak güncellendi!")
        except Exception as e:
            messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
            log_exception("danisan_aktif_pasif", e)
            return
        if hasattr(parent, "_reload"):
            parent._reload()

    # --- EK MODÜL: Görev Durumu Güncelle/Tamamla/Sil ---
    def _selected_tree_id(self, tree):
        sel = tree.selection()
        if not sel:
            return None
        vals = tree.item(sel[0]).get("values") or []
        if not vals:
            return None
        try:
            return int(vals[0])
        except Exception:
            return None

    def gorev_durum_guncelle(self, parent):
        if not hasattr(parent, "gorev_tree"):
            return
        gid = self._selected_tree_id(parent.gorev_tree)
        if not gid:
            messagebox.showwarning("Uyarı", "Lütfen bir görev seçiniz!")
            return
        win = ttk.Toplevel(parent)
        win.title("Durum Güncelle")
        center_window_smart(win, 520, 420, max_ratio=0.85)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)
        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)
        ttk.Label(frm, text="Durum:").pack(anchor=W)
        cb = ttk.Combobox(frm, state="readonly", values=["beklemede", "devam", "tamamlandi", "iptal"])
        cb.current(0)
        cb.pack(fill=X, pady=8)
        ttk.Label(frm, text="Öncelik:").pack(anchor=W)
        cb2 = ttk.Combobox(frm, state="readonly", values=["dusuk", "normal", "yuksek", "acil"])
        cb2.current(1)
        cb2.pack(fill=X, pady=8)

        def _save():
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute("UPDATE gorevler SET durum=?, oncelik=? WHERE id=?", (cb.get(), cb2.get(), gid))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
                return
            win.destroy()
            if hasattr(parent, "_reload"):
                parent._reload()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(12, 0))

    def gorev_tamamla(self, parent):
        if not hasattr(parent, "gorev_tree"):
            return
        gid = self._selected_tree_id(parent.gorev_tree)
        if not gid:
            return
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                "UPDATE gorevler SET durum='tamamlandi', tamamlanma_tarihi=? WHERE id=?",
                (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), gid),
            )
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
            return
        if hasattr(parent, "_reload"):
            parent._reload()

    def gorev_sil(self, parent):
        if not hasattr(parent, "gorev_tree"):
            return
        gid = self._selected_tree_id(parent.gorev_tree)
        if not gid:
            return
        if not messagebox.askyesno(
            "Onay",
            f"{danisan_adi} danışanını listeden kaldırmak istediğinize emin misiniz?\n\n"
            "Bu işlem danışanı pasife alır (aktif=0).",
        ):
            return
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("DELETE FROM gorevler WHERE id=?", (gid,))
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Silinemedi:\n{e}")
            return
        if hasattr(parent, "_reload"):
            parent._reload()

    # --- EK MODÜL: Seans Listesi (Düzenle/Sil) ---
    def seans_listesi(self):
        win = ttk.Toplevel(self)
        win.title("Seans Listesi (Düzenle/Sil)")
        center_window_smart(win, 1400, 820)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        self._style_table_strong()

        top = ttk.Frame(win, padding=10)
        top.pack(fill=X)
        ttk.Label(top, text="Tarih (YYYY-AA-GG):").pack(side=LEFT, padx=6)
        tarih_var = ttk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d"))
        ent_t = ttk.Entry(top, textvariable=tarih_var, width=14)
        ent_t.pack(side=LEFT, padx=6)
        ttk.Label(top, text="Terapist:").pack(side=LEFT, padx=6)
        cb = ttk.Combobox(top, state="readonly", width=24)
        cb.pack(side=LEFT, padx=6)
        ttk.Button(top, text="Yenile", bootstyle="primary", command=lambda: _yukle()).pack(side=LEFT, padx=6)
        ttk.Button(top, text="Düzenle", bootstyle="warning", command=lambda: self.seans_duzenle(win)).pack(side=RIGHT)
        ttk.Button(top, text="Sil", bootstyle="danger", command=lambda: self.seans_sil(win)).pack(side=RIGHT, padx=6)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=BOTH, expand=True)
        cols = ("ID", "Tarih", "Saat", "Danışan", "Terapist", "Oda", "Durum", "Seans", "Ücret", "Tutar", "Ödeme", "Not")
        tree = ttk.Treeview(frame, columns=cols, show="headings", bootstyle="info", style="Strong.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140)
        tree.column("ID", width=70)
        tree.column("Seans", width=70)
        tree.column("Ücret", width=70)
        tree.column("Tutar", width=110)
        tree.column("Ödeme", width=110)
        tree.column("Not", width=260)
        sb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        self._apply_stripes(tree)

        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            names = [r[0] for r in cur.fetchall()]
            conn.close()
        except Exception:
            names = DEFAULT_THERAPISTS[:]
        cb["values"] = ["(Tümü)"] + names
        cb.current(0)

        def _yukle():
            for i in tree.get_children():
                tree.delete(i)
            tarih = (tarih_var.get() or "").strip()
            terapist = cb.get()
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                if terapist and terapist != "(Tümü)":
                    cur.execute(
                        """
                        SELECT id, tarih, saat, danisan_adi, terapist, COALESCE(oda,''), COALESCE(durum,''),
                               COALESCE(seans_alindi,0), COALESCE(ucret_alindi,0), COALESCE(ucret_tutar,0), COALESCE(odeme_sekli,''),
                               COALESCE(notlar,'')
                        FROM seans_takvimi
                        WHERE tarih=? AND terapist=?
                        ORDER BY saat
                        """,
                        (tarih, terapist),
                    )
                else:
                    cur.execute(
                        """
                        SELECT id, tarih, saat, danisan_adi, terapist, COALESCE(oda,''), COALESCE(durum,''),
                               COALESCE(seans_alindi,0), COALESCE(ucret_alindi,0), COALESCE(ucret_tutar,0), COALESCE(odeme_sekli,''),
                               COALESCE(notlar,'')
                        FROM seans_takvimi
                        WHERE tarih=?
                        ORDER BY saat
                        """,
                        (tarih,),
                    )
                rows = cur.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Seanslar yüklenemedi:\n{e}")
                return
            for idx, r in enumerate(rows):
                tag = "even" if idx % 2 == 0 else "odd"
                tree.insert(
                    "",
                    END,
                    values=(
                        r[0],
                        r[1],
                        r[2],
                        r[3],
                        r[4],
                        r[5],
                        r[6],
                        ("✓" if int(r[7] or 0) == 1 else ""),
                        ("✓" if int(r[8] or 0) == 1 else ""),
                        format_money(r[9]),
                        r[10],
                        r[11],
                    ),
                    tags=(tag,),
                )

        _yukle()
        win.seans_tree = tree
        win._reload = _yukle
        win._tarih_var = tarih_var

    def seans_duzenle(self, parent):
        if not hasattr(parent, "seans_tree"):
            return
        sid = self._selected_tree_id(parent.seans_tree)
        if not sid:
            messagebox.showwarning("Uyarı", "Lütfen bir seans seçiniz!")
            return
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT tarih, saat, danisan_adi, terapist, COALESCE(oda,''), COALESCE(durum,''), COALESCE(seans_alindi,0),
                       COALESCE(ucret_alindi,0), COALESCE(ucret_tutar,0), COALESCE(odeme_sekli,''), COALESCE(notlar,'')
                FROM seans_takvimi WHERE id=?
                """,
                (sid,),
            )
            row = cur.fetchone()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Seans okunamadı:\n{e}")
            return
        if not row:
            return

        win = ttk.Toplevel(parent)
        win.title("Seans Düzenle")
        center_window_smart(win, 700, 720, max_ratio=0.9)
        win.transient(parent)
        win.grab_set()
        self._brand_window(win)
        frm = ttk.Frame(win, padding=16)
        frm.pack(fill=BOTH, expand=True)

        def field(label, init=""):
            ttk.Label(frm, text=label).pack(anchor=W, pady=(8, 0))
            e = ttk.Entry(frm)
            e.insert(0, init or "")
            e.pack(fill=X, pady=4)
            return e

        ent_tarih = field("Tarih (YYYY-AA-GG):", row[0])
        ent_saat = field("Saat (HH:MM):", row[1])
        ent_dan = field("Danışan:", row[2])
        ent_ter = field("Terapist:", row[3])
        ent_oda = field("Oda:", row[4])
        ent_durum = field("Durum:", row[5] or "planlandi")

        var_seans = ttk.IntVar(value=1 if int(row[6] or 0) == 1 else 0)
        var_ucret = ttk.IntVar(value=1 if int(row[7] or 0) == 1 else 0)
        ttk.Checkbutton(frm, text="Seans Alındı", variable=var_seans, bootstyle="success").pack(anchor=W, pady=(10, 0))
        ttk.Checkbutton(frm, text="Ücret Alındı", variable=var_ucret, bootstyle="success").pack(anchor=W, pady=(6, 0))
        ent_tutar = field("Ücret Tutarı (₺):", str(row[8] or 0))
        ent_odeme = field("Ödeme Şekli:", row[9] or "")
        ent_not = field("Not:", row[10] or "")

        def _save():
            if "name hoca" in (ent_ter.get() or "").lower():
                messagebox.showwarning("Uyarı", "Name Hoca kurumdan ayrıldı.")
                return
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                cur.execute(
                    """
                    UPDATE seans_takvimi
                    SET tarih=?, saat=?, danisan_adi=?, terapist=?, oda=?, durum=?, seans_alindi=?, ucret_alindi=?, ucret_tutar=?, odeme_sekli=?, notlar=?
                    WHERE id=?
                    """,
                    (
                        (ent_tarih.get() or "").strip(),
                        (ent_saat.get() or "").strip(),
                        (ent_dan.get() or "").strip().upper(),
                        (ent_ter.get() or "").strip(),
                        (ent_oda.get() or "").strip(),
                        (ent_durum.get() or "").strip(),
                        int(var_seans.get() or 0),
                        int(var_ucret.get() or 0),
                        parse_money(ent_tutar.get()),
                        (ent_odeme.get() or "").strip(),
                        (ent_not.get() or "").strip(),
                        sid,
                    ),
                )
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Güncellenemedi:\n{e}")
                return
            win.destroy()
            if hasattr(parent, "_reload"):
                parent._reload()

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(12, 0))

    def seans_sil(self, parent):
        if not hasattr(parent, "seans_tree"):
            return
        sid = self._selected_tree_id(parent.seans_tree)
        if not sid:
            return
        if not messagebox.askyesno(
            "Onay",
            f"{danisan_adi} danışanını listeden kaldırmak istediğinize emin misiniz?\n\n"
            "Bu işlem danışanı pasife alır (aktif=0).",
        ):
            return
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            # bağlı records kaydı varsa onu da sil
            try:
                cur.execute("SELECT COALESCE(record_id,NULL) FROM seans_takvimi WHERE id=?", (sid,))
                rid = (cur.fetchone() or [None])[0]
                if rid:
                    cur.execute("DELETE FROM records WHERE id=?", (rid,))
            except Exception:
                pass
            # ✅ ENTERPRISE: Pipeline üzerinden sil (cascade + audit trail)
            kullanici_id = self.kullanici[0] if self.kullanici else None
            pipeline = DataPipeline(conn, kullanici_id)
            basarili = pipeline.kayit_sil(seans_id=sid)
            conn.close()
            
            if not basarili:
                messagebox.showerror("Hata", "Seans kaydı silinemedi!")
                return
                
            messagebox.showinfo("Başarılı", "Seans kaydı silindi!\n\nTüm bağlı kayıtlar otomatik temizlendi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Silinemedi:\n{e}")
            log_exception("seans_sil_ui", e)
            return
        if hasattr(parent, "_reload"):
            parent._reload()

    # ✅ ONAM FORMU TAB
    def _build_onam_formu_tab(self):
        """Bu modül kaldırıldı."""
        messagebox.showinfo("Bilgi", "Bu modül bu sürümde kaldırılmıştır.")

    def _onam_formlari_listele(self, tree):
        """Onam formlarını listele"""
        for iid in tree.get_children():
            tree.delete(iid)
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("""
                SELECT id, danisan_adi, danisan_tarih, terapist_adi, terapist_tarih, 
                       CASE WHEN onam_verildi=1 THEN 'Onaylandı' ELSE 'Beklemede' END
                FROM onam_formlari
                ORDER BY olusturma_tarihi DESC
            """)
            for row in cur.fetchall():
                tree.insert("", END, values=row)
            conn.close()
        except Exception as e:
            log_exception("_onam_formlari_listele", e)
    
    def _onam_formu_ekle(self):
        """Yeni Onam Formu ekle/düzenle"""
        win = ttk.Toplevel(self)
        win.title("ONAM FORMU")
        win.geometry("700x600")
        center_window(win, 700, 600)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        
        # Scrollable frame
        canvas = tk.Canvas(win)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        wrapper = ttk.Frame(scrollable_frame, padding=20)
        wrapper.pack(fill=BOTH, expand=True)
        
        # Logo üstte (resimdeki gibi)
        if getattr(self, "_logo_small", None):
            ttk.Label(wrapper, image=self._logo_small).pack(anchor=E, pady=(0, 10))
        
        # Başlık: ONAM FORMU (ortada, resimdeki gibi)
        ttk.Label(wrapper, text="ONAM FORMU", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(pady=(0, 20))
        
        # Onam metni - resimdeki metin birebir
        onam_text = (
            "6698 sayılı Kişisel Verilerin Korunması Kanunu kapsamında Leta Çocuk ve Aile Danışmanlık "
            "kurumunda çocuğunuza ve size verilmiş olan hizmetler, danışmanlık kapsamında video ve fotoğraf "
            "olarak kullanılmaktadır. Konu ile ilgili bilgilendirme ve kapsam şu şekildedir; kurumumuz tarafından "
            "sizlerle paylaşılan video ve fotoğraflar herhangi bir kurum, sosyal medya, resmi internet siteleri, "
            "uzman ve kişilerle paylaşılmayacaktır. Video veya fotoğraflar danışmanlık hizmetimizin sonunda yok edilecektir."
        )
        ttk.Label(wrapper, text=onam_text, wraplength=620, justify="left", font=("Segoe UI", 10)).pack(anchor=W, pady=(0, 16))
        
        ttk.Label(wrapper, text="Yukarıdaki yazıları anladım ve kabul ediyorum.", font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(0, 24))
        
        # İmza blokları - resimdeki düzen: Danışanın Adı Soyadı (sol) | Tarih (sağ)
        form = ttk.Frame(wrapper)
        form.pack(fill=X, pady=10)
        
        # Satır 1: Danışanın Adı Soyadı (sol) — Tarih (sağ)
        row1 = ttk.Frame(form)
        row1.pack(fill=X, pady=(0, 20))
        ttk.Label(row1, text="Danışanın Adı Soyadı:", font=("Segoe UI", 10)).pack(side=LEFT, padx=(0, 8))
        danisan_var = tk.StringVar()
        cb_danisan = ttk.Combobox(row1, textvariable=danisan_var, width=36, state="normal")
        cb_danisan.pack(side=LEFT, fill=X, expand=True, padx=4)
        ttk.Label(row1, text="Tarih:", font=("Segoe UI", 10)).pack(side=LEFT, padx=(20, 8))
        danisan_tarih_var = tk.StringVar(value=datetime.datetime.now().strftime("%d.%m.%Y"))
        ent_danisan_tarih = ttk.Entry(row1, textvariable=danisan_tarih_var, width=14)
        ent_danisan_tarih.pack(side=LEFT, padx=4)
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            cb_danisan["values"] = [row[0] for row in cur.fetchall()]
            conn.close()
        except Exception:
            cb_danisan["values"] = []
        
        # Satır 2: Terapistin Adı Soyadı (sol) — Tarih (sağ)
        row2 = ttk.Frame(form)
        row2.pack(fill=X, pady=(0, 10))
        ttk.Label(row2, text="Terapistin Adı Soyadı:", font=("Segoe UI", 10)).pack(side=LEFT, padx=(0, 8))
        terapist_var = tk.StringVar()
        cb_terapist = ttk.Combobox(row2, textvariable=terapist_var, width=36, state="readonly")
        cb_terapist.pack(side=LEFT, fill=X, expand=True, padx=4)
        ttk.Label(row2, text="Tarih:", font=("Segoe UI", 10)).pack(side=LEFT, padx=(20, 8))
        terapist_tarih_var = tk.StringVar(value=datetime.datetime.now().strftime("%d.%m.%Y"))
        ent_terapist_tarih = ttk.Entry(row2, textvariable=terapist_tarih_var, width=14)
        ent_terapist_tarih.pack(side=LEFT, padx=4)
        
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT therapist_name FROM settings WHERE is_active=1 ORDER BY therapist_name")
            terapistler = [row[0] for row in cur.fetchall()]
            cb_terapist["values"] = terapistler
            if terapistler:
                cb_terapist.current(0)
            conn.close()
        except Exception:
            cb_terapist["values"] = []
        
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        def _save():
            danisan = danisan_var.get().strip().upper()
            terapist = terapist_var.get().strip()
            if not danisan or not terapist:
                messagebox.showwarning("Uyarı", "Lütfen danışan ve terapist seçiniz!")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                # Danışan ID'sini bul
                cur.execute("SELECT id FROM danisanlar WHERE ad_soyad=?", (danisan,))
                d_row = cur.fetchone()
                if not d_row:
                    messagebox.showerror("Hata", "Danışan bulunamadı!")
                    conn.close()
                    return
                danisan_id = d_row[0]
                
                # Onam formunu kaydet
                cur.execute("""
                    INSERT INTO onam_formlari 
                    (danisan_id, danisan_adi, danisan_tarih, terapist_adi, terapist_tarih, onam_verildi, olusturma_tarihi)
                    VALUES (?, ?, ?, ?, ?, 1, ?)
                """, (
                    danisan_id, danisan, danisan_tarih_var.get(), 
                    terapist, terapist_tarih_var.get(),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Başarılı", "Onam formu kaydedildi!")
                win.destroy()
                # Liste yenile
                if hasattr(self, "tab_onam"):
                    for child in self.tab_onam.winfo_children():
                        if isinstance(child, ttk.Frame):
                            for widget in child.winfo_children():
                                if isinstance(widget, ttk.Treeview):
                                    self._onam_formlari_listele(widget)
            except Exception as e:
                messagebox.showerror("Hata", f"Kayıt hatası:\n{e}")
                log_exception("_onam_formu_ekle", e)
        
        btns = ttk.Frame(win)
        btns.pack(fill=X, padx=20, pady=10)
        ttk.Button(btns, text="💾 Kaydet", bootstyle="success", command=_save).pack(side=RIGHT, padx=5)
        ttk.Button(btns, text="İptal", bootstyle="secondary", command=win.destroy).pack(side=RIGHT, padx=5)

    # ✅ ÇOCUK TAKİP BİLGİ FORMU TAB (3 Sayfa)
    def _build_cocuk_takip_formu_tab(self):
        """Bu modül kaldırıldı."""
        messagebox.showinfo("Bilgi", "Bu modül bu sürümde kaldırılmıştır.")

    def _cocuk_takip_formlari_listele(self, tree):
        """Çocuk takip formlarını listele"""
        for iid in tree.get_children():
            tree.delete(iid)
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("""
                SELECT id, danisan_id, form_tarihi, cinsiyet, dogum_tarihi, okul_adi
                FROM cocuk_takip_bilgi_formlari
                ORDER BY form_tarihi DESC
            """)
            for row in cur.fetchall():
                form_id, danisan_id, form_tarih, cinsiyet, dogum_tarih, okul = row
                # Danışan adını bul
                cur.execute("SELECT ad_soyad FROM danisanlar WHERE id=?", (danisan_id,))
                d_row = cur.fetchone()
                danisan_adi = d_row[0] if d_row else "Bilinmiyor"
                tree.insert("", END, values=(form_id, danisan_adi, form_tarih or "", cinsiyet or "", dogum_tarih or "", okul or ""))
            conn.close()
        except Exception as e:
            log_exception("_cocuk_takip_formlari_listele", e)
    
    def _cocuk_takip_formu_ekle(self):
        """Çocuk Takip Bilgi Formu ekle/düzenle (3 Sayfa)"""
        win = ttk.Toplevel(self)
        win.title("ÇOCUK TAKİP BİLGİ FORMU")
        win.geometry("900x750")
        center_window(win, 900, 750)
        maximize_window(win)
        win.transient(self)
        self._brand_window(win)
        
        # Notebook - 3 sayfa
        nb = ttk.Notebook(win)
        nb.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # SAYFA 1: Doğum ve Gelişim Bilgileri
        page1 = ttk.Frame(nb, padding=15)
        nb.add(page1, text="Sayfa 1: Doğum & Gelişim")
        
        # SAYFA 2: Eğitim Bilgileri
        page2 = ttk.Frame(nb, padding=15)
        nb.add(page2, text="Sayfa 2: Eğitim Bilgileri")
        
        # SAYFA 3: Demografik Bilgiler
        page3 = ttk.Frame(nb, padding=15)
        nb.add(page3, text="Sayfa 3: Demografik Bilgiler")
        
        # Form değişkenleri (tüm sayfalarda paylaşılacak) - closure ile paylaş
        form_data = {"danisan_var": None, "sayfa1_vars": {}, "sayfa2_vars": {}, "sayfa3_vars": {}}
        
        # Sayfa 1'den değişkenleri almak için callback
        def _set_sayfa1_vars(vars_dict):
            form_data["sayfa1_vars"] = vars_dict
        
        # Sayfa 2'den değişkenleri almak için callback
        def _set_sayfa2_vars(vars_dict):
            form_data["sayfa2_vars"] = vars_dict
        
        # Sayfa 3'ten değişkenleri almak için callback
        def _set_sayfa3_vars(vars_dict):
            form_data["sayfa3_vars"] = vars_dict
        
        # Sayfaları oluştur ve değişkenleri paylaş
        self._build_cocuk_takip_sayfa1(page1, form_data, _set_sayfa1_vars)
        self._build_cocuk_takip_sayfa2(page2, form_data, _set_sayfa2_vars)
        self._build_cocuk_takip_sayfa3(page3, form_data, _set_sayfa3_vars)
        
        # Butonlar
        btns = ttk.Frame(win)
        btns.pack(fill=X, padx=20, pady=10)
        
        def _save_all():
            danisan_var_obj = form_data.get("danisan_var")
            if not danisan_var_obj:
                messagebox.showwarning("Uyarı", "Lütfen danışan seçiniz!")
                return
            danisan = danisan_var_obj.get().strip().upper()
            if not danisan:
                messagebox.showwarning("Uyarı", "Lütfen danışan seçiniz!")
                return
            
            try:
                conn = self.veritabani_baglan()
                cur = conn.cursor()
                # Danışan ID'sini bul
                cur.execute("SELECT id FROM danisanlar WHERE ad_soyad=?", (danisan,))
                d_row = cur.fetchone()
                if not d_row:
                    messagebox.showerror("Hata", "Danışan bulunamadı!")
                    conn.close()
                    return
                danisan_id = d_row[0]
                
                # Tüm form verilerini topla
                s1 = form_data.get("sayfa1_vars", {})
                s2 = form_data.get("sayfa2_vars", {})
                s3 = form_data.get("sayfa3_vars", {})
                
                def _evet_hayir(v):
                    if v in ("Evet", "evet", "1"): return 1
                    if v in ("Hayır", "hayır", "0"): return 0
                    return None
                def _almadi_aldi(v):
                    if v in ("Aldı", "aldı"): return 1
                    if v in ("Almadı", "almadı"): return 0
                    return None
                
                # Tam 70 sütun — değerleri sırayla listeye alıp tuple yapıyoruz (74/77 değer hatası kalmasın)
                def _int(x):
                    try: return int(x) if x and str(x).strip() else None
                    except (TypeError, ValueError): return None
                def _float(x):
                    try: return float(x) if x and str(x).strip() else None
                    except (TypeError, ValueError): return None
                vals = [
                    danisan_id,
                    datetime.datetime.now().strftime("%Y-%m-%d"),
                    s1.get("cinsiyet"),
                    s1.get("dogum_tarih"),
                    s1.get("dogum_yeri"),
                    s1.get("gebelik_sekli"),
                    s1.get("gebelik_sorun"),
                    s1.get("dogum_sekli"),
                    _int(s1.get("dogum_hafta")),
                    _float(s1.get("dogum_kilo")),
                    _float(s1.get("dogum_boy")),
                    _evet_hayir(s1.get("dogum_sorun")),
                    s1.get("dogum_sorun_detay"),
                    _almadi_aldi(s1.get("anne_sutu")),
                    s1.get("anne_sutu_sure"),
                    s1.get("bakim_veren"),
                    s1.get("yurme_yas"),
                    s1.get("yurme_gec_neden"),
                    s1.get("tuvalet_yas"),
                    s1.get("tuvalet_gec_neden"),
                    s1.get("konusma_yas"),
                    s2.get("konusma_gec_neden"),
                    _evet_hayir(s2.get("gdb_tani")),
                    s2.get("gdb_tani_detay"),
                    s2.get("okul_adi"),
                    s2.get("sinif"),
                    s2.get("egitim_turu"),
                    1 if s2.get("destek_egitim") == "Evet" else 0,
                    s2.get("destek_egitim_sure"),
                    s2.get("sinif_ogretmen"),
                    1 if s2.get("okuloncesi") == "Evet" else 0,
                    _int(s2.get("okuloncesi_yil")),
                    _int(s2.get("ilkokul_baslangic")),
                    s2.get("egitim_sorun"),
                    s2.get("okuma_baslangic"),
                    _evet_hayir(s2.get("okuma_sorun")),
                    s2.get("okuma_sorun_detay"),
                    _evet_hayir(s2.get("okuma_anlama")),
                    s2.get("okuma_anlama_detay"),
                    _evet_hayir(s2.get("yazma_sorun")),
                    s2.get("yazma_sorun_detay"),
                    _evet_hayir(s2.get("aritmetik_sorun")),
                    s2.get("aritmetik_sorun_detay"),
                    _evet_hayir(s2.get("siralama_sorun")),
                    s2.get("siralama_sorun_detay"),
                    _evet_hayir(s3.get("yon_ayirt")),
                    s3.get("yon_ayirt_detay"),
                    s3.get("karneturkce"),
                    s3.get("karnematematik"),
                    s3.get("karnehayatbilgisi"),
                    s3.get("karnesosyal"),
                    s3.get("karnefen"),
                    _int(s3.get("aile_sira")),
                    1 if s3.get("akrabalik") == "Evet" else 0,
                    s3.get("akrabalik_detay"),
                    s3.get("bakim_veren_suan"),
                    s3.get("aile_disinda"),
                    s3.get("aile_turu"),
                    s3.get("ayrilik_durum"),
                    s3.get("sosyoekonomik"),
                    s3.get("anne_egitim"),
                    _int(s3.get("anne_yas")),
                    s3.get("anne_is"),
                    s3.get("baba_egitim"),
                    _int(s3.get("baba_yas")),
                    s3.get("baba_is"),
                    s3.get("cocuk_sayisi_detay"),
                    1 if s3.get("hasta_kardes") == "Evet" else 0,
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    self.kullanici[0] if self.kullanici else None,
                ]
                # Tam 70 değer: fazlaysa kes, azsa None ile doldur (tablo 70 sütun)
                vlist = list(vals)[:70]
                while len(vlist) < 70:
                    vlist.append(None)
                vals = tuple(vlist)
                cur.execute("""
                    INSERT INTO cocuk_takip_bilgi_formlari (
                        danisan_id, form_tarihi,
                        cinsiyet, dogum_tarihi, dogum_yeri, gebelik_sekli, gebelik_sorun,
                        dogum_sekli, dogum_hafta, dogum_kilo, dogum_boy,
                        dogum_sorun, dogum_sorun_detay, anne_sutu, anne_sutu_sure, bakim_veren,
                        yurme_yas, yurme_gec_neden, tuvalet_yas, tuvalet_gec_neden, konusma_yas, konusma_gec_neden,
                        gdb_tani, gdb_tani_detay, okul_adi, sinif, egitim_turu,
                        destek_egitim, destek_egitim_sure, sinif_ogretmen, okuloncesi, okuloncesi_yil, ilkokul_baslangic_ay, egitim_sorun,
                        okuma_baslangic, okuma_sorun, okuma_sorun_detay, okuma_anlama_sorun, okuma_anlama_detay,
                        yazma_sorun, yazma_sorun_detay, aritmetik_sorun, aritmetik_sorun_detay, siralama_sorun, siralama_sorun_detay,
                        yon_ayirt_sorun, yon_ayirt_detay, karneturkce, karnematematik, karnehayatbilgisi, karnesosyal, karnefen,
                        aile_sira, akrabalik, akrabalik_detay, bakim_veren_suan, aile_disinda_yasayan, aile_turu, ayrilik_durum, sosyoekonomik,
                        anne_egitim, anne_yas, anne_is, baba_egitim, baba_yas, baba_is, cocuk_sayisi_detay, hasta_kardes,
                        olusturma_tarihi, olusturan_kullanici_id
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, vals)
                conn.commit()
                conn.close()
                
                messagebox.showinfo("✅ Başarılı", "Çocuk takip bilgi formu kaydedildi!")
                win.destroy()
                # Liste yenile
                if hasattr(self, "tab_cocuk_takip"):
                    for child in self.tab_cocuk_takip.winfo_children():
                        if isinstance(child, ttk.Frame):
                            for widget in child.winfo_children():
                                if isinstance(widget, ttk.Labelframe):
                                    for w2 in widget.winfo_children():
                                        if isinstance(w2, ttk.Treeview):
                                            self._cocuk_takip_formlari_listele(w2)
            except Exception as e:
                messagebox.showerror("Hata", f"Form kaydetme hatası:\n{e}")
                log_exception("_cocuk_takip_formu_kaydet", e)
        
        ttk.Button(btns, text="💾 Kaydet", bootstyle="success", command=_save_all).pack(side=RIGHT, padx=5)
        ttk.Button(btns, text="İptal", bootstyle="secondary", command=win.destroy).pack(side=RIGHT, padx=5)
    
    def _paper_radio_group(self, parent, var, options, padx_between=12):
        """Kağıt formu gibi ( ) seçilince ( X ) gösteren radyo grubu. options = [(value, display_text), ...]"""
        fr = ttk.Frame(parent)
        brackets = []
        def _update():
            cur = var.get()
            for i, (val, _) in enumerate(options):
                brackets[i].config(text="( X )" if cur == val else "( )")
        for value, text in options:
            sub = ttk.Frame(fr)
            lbl_b = ttk.Label(sub, text="( )", width=4, anchor="w")
            lbl_t = ttk.Label(sub, text=text)
            brackets.append(lbl_b)
            def _on_click(e, v=value):
                var.set(v)
            for w in (sub, lbl_b, lbl_t):
                w.bind("<Button-1>", _on_click)
                try: w.configure(cursor="hand2")
                except Exception: pass
            lbl_b.pack(side=LEFT)
            lbl_t.pack(side=LEFT, padx=(2, padx_between))
            sub.pack(side=LEFT)
        try:
            var.trace_add("write", lambda *a: _update())
        except Exception:
            var.trace("w", lambda *a: _update())
        _update()
        return fr

    def _paper_radio_group_wrapped(self, parent, var, options, padx_between=8, columns=4):
        """Çok seçenekli radyo grubu; seçenekleri grid ile birkaç sütunda sarar."""
        fr = ttk.Frame(parent)
        brackets = []
        def _update():
            cur = var.get()
            for i, (val, _) in enumerate(options):
                brackets[i].config(text="( X )" if cur == val else "( )")
        for idx, (value, text) in enumerate(options):
            sub = ttk.Frame(fr)
            lbl_b = ttk.Label(sub, text="( )", width=4, anchor="w")
            lbl_t = ttk.Label(sub, text=text)
            brackets.append(lbl_b)
            def _on_click(e, v=value):
                var.set(v)
            for w in (sub, lbl_b, lbl_t):
                w.bind("<Button-1>", _on_click)
                try: w.configure(cursor="hand2")
                except Exception: pass
            lbl_b.pack(side=LEFT)
            lbl_t.pack(side=LEFT, padx=(2, padx_between))
            sub.grid(row=idx // columns, column=idx % columns, sticky=W, padx=(0, 12), pady=2)
        try:
            var.trace_add("write", lambda *a: _update())
        except Exception:
            var.trace("w", lambda *a: _update())
        _update()
        return fr
    
    def _paper_check_one(self, parent, var, value_yes, text_yes, text_no, padx_between=12):
        """Evet/Hayır gibi tek seçim: ( ) Evet  ( ) Hayır — seçilene ( X )"""
        return self._paper_radio_group(parent, var, [(value_yes, text_yes), ("Hayır", text_no)], padx_between)
    
    def _build_cocuk_takip_sayfa1(self, parent, form_data, set_vars_callback):
        """Sayfa 1: Resimdeki düzen - ÇOCUK GELİŞİM BİLGİ FORMU, Doğum Süreci, Temel Gelişim Bilgileri"""
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        ttk.Label(scrollable, text="ÇOCUK GELİŞİM BİLGİ FORMU", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 16))
        
        form = ttk.Frame(scrollable, padding=(0, 4))
        form.pack(fill=X, pady=8)
        lbl_w, wrap = 36, 380  # etiket genişliği ve uzun metin satır uzunluğu
        r = 0
        
        # --- Temel bilgiler ---
        ttk.Label(form, text="Adı Soyadı:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        danisan_var = tk.StringVar()
        cb_danisan = ttk.Combobox(form, textvariable=danisan_var, width=40, state="normal")
        cb_danisan.grid(row=r, column=1, sticky=W, padx=5, pady=6)
        form_data["danisan_var"] = danisan_var
        try:
            conn = self.veritabani_baglan()
            cur = conn.cursor()
            cur.execute("SELECT ad_soyad FROM danisanlar WHERE aktif=1 ORDER BY ad_soyad")
            cb_danisan["values"] = [row[0] for row in cur.fetchall()]
            conn.close()
        except Exception:
            cb_danisan["values"] = []
        r += 1
        
        ttk.Label(form, text="Cinsiyeti:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        cinsiyet_var = tk.StringVar()
        fr_cins = self._paper_radio_group(form, cinsiyet_var, [("Kız", "Kız"), ("Erkek", "Erkek")], padx_between=14)
        fr_cins.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Doğum Tarihi (gg/aa/yyyy):", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        dogum_tarih_var = tk.StringVar()
        ttk.Entry(form, textvariable=dogum_tarih_var, width=24).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Doğum Yeri:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        dogum_yeri_var = tk.StringVar()
        ttk.Entry(form, textvariable=dogum_yeri_var, width=40).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        # --- Bölüm: Doğum Süreci ve Doğum Sonrasına İlişkin Bilgiler ---
        ttk.Label(scrollable, text="Doğum Süreci ve Doğum Sonrasına İlişkin Bilgiler", font=("Segoe UI", 11, "bold")).pack(anchor=W, padx=5, pady=(20, 10))
        
        ttk.Label(form, text="Gebelik Şekli:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        gebelik_var = tk.StringVar()
        fr_geb = self._paper_radio_group(form, gebelik_var, [("Planlı", "Planlı"), ("Plansız", "Plansız")], padx_between=14)
        fr_geb.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Gebelik döneminde sorun olduysa belirtiniz:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        gebelik_sorun_var = tk.StringVar()
        ttk.Entry(form, textvariable=gebelik_sorun_var, width=52).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Doğum Şekli:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        dogum_sekli_var = tk.StringVar()
        fr_dog = self._paper_radio_group(form, dogum_sekli_var, [
            ("Normal", "Normal"), ("Sezaryen", "Sezaryen"), ("Müdahaleli", "Müdahaleli-Vakum"), ("Diğer", "Diğer")
        ], padx_between=10)
        fr_dog.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Kaç haftalık doğdu:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        dogum_hafta_var = tk.StringVar()
        ttk.Entry(form, textvariable=dogum_hafta_var, width=10).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Doğum kilosu:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        dogum_kilo_var = tk.StringVar()
        ttk.Entry(form, textvariable=dogum_kilo_var, width=10).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Doğum boyu:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        dogum_boy_var = tk.StringVar()
        ttk.Entry(form, textvariable=dogum_boy_var, width=10).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Çocuğunuz doğumun esnasında ya da hemen sonrasında bir problem yaşadı mı?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        dogum_sorun_var = tk.StringVar()
        fr_sorun = self._paper_radio_group(form, dogum_sorun_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=14)
        fr_sorun.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Yaşadıysa bunlar nelerdir? (Morarma, Havale, Enfeksiyon, Hemen Ağlamama, Diğer)", width=lbl_w, anchor="w", wraplength=wrap).grid(row=r, column=0, sticky=NW, padx=5, pady=6)
        dogum_sorun_detay_var = tk.StringVar()
        ttk.Entry(form, textvariable=dogum_sorun_detay_var, width=52).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Başka bir problem varsa yazınız:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        diger_sorun_var = tk.StringVar()
        ttk.Entry(form, textvariable=diger_sorun_var, width=52).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Anne sütü alma durumu: (Almadı / Aldı — Aldıysa alma süresini yazınız)", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        anne_sutu_var = tk.StringVar()
        anne_sutu_sure_var = tk.StringVar()
        fr_anne = ttk.Frame(form)
        fr_anne.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_anne, anne_sutu_var, [("Almadı", "Almadı"), ("Aldı", "Aldı")], padx_between=12).pack(side=LEFT)
        ttk.Label(fr_anne, text="Alma Süresi:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_anne, textvariable=anne_sutu_sure_var, width=14).pack(side=LEFT)
        r += 2
        
        ttk.Label(form, text="Bebekliğinin ilk yıllarında temel bakım veren kişi veya kişiler:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        bakim_veren_var = tk.StringVar()
        bakim_opts = [
            ("Anne ve baba", "Anne ve baba"), ("Sadece anne", "Sadece anne"), ("Sadece baba", "Sadece baba"),
            ("Bakıcı", "Bakıcı"), ("Büyük anne", "Büyük anne"), ("Anneanne", "Anneanne"), ("Akrabalar", "Akrabalar"),
            ("Edinilmiş aile", "Edinilmiş aile"), ("Koruyucu aile", "Koruyucu aile"), ("Kurum", "Kurum"), ("Diğer", "Diğer")
        ]
        fr_bakim = self._paper_radio_group_wrapped(form, bakim_veren_var, bakim_opts, padx_between=8, columns=4)
        fr_bakim.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        # --- Bölüm: Temel Gelişim Bilgileri ---
        ttk.Label(scrollable, text="Temel Gelişim Bilgileri", font=("Segoe UI", 11, "bold")).pack(anchor=W, padx=5, pady=(20, 10))
        
        ttk.Label(form, text="Çocuğunuz kaç yaşında yürüdü:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        yurme_yas_var = tk.StringVar()
        ttk.Entry(form, textvariable=yurme_yas_var, width=12).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Geç yürüdüyse nedenlerini belirtiniz:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        yurme_gec_neden_var = tk.StringVar()
        ttk.Entry(form, textvariable=yurme_gec_neden_var, width=52).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Çocuğunuz kaç yaşında tuvalet eğitimini tamamladı:", width=lbl_w, anchor="w", wraplength=wrap).grid(row=r, column=0, sticky=W, padx=5, pady=6)
        tuvalet_yas_var = tk.StringVar()
        ttk.Entry(form, textvariable=tuvalet_yas_var, width=12).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Geç tamamladıysa nedenlerini belirtiniz:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        tuvalet_gec_neden_var = tk.StringVar()
        ttk.Entry(form, textvariable=tuvalet_gec_neden_var, width=52).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        ttk.Label(form, text="Çocuğunuz kaç yaşında konuştu:", width=lbl_w, anchor="w").grid(row=r, column=0, sticky=W, padx=5, pady=6)
        konusma_yas_var = tk.StringVar()
        ttk.Entry(form, textvariable=konusma_yas_var, width=12).grid(row=r, column=1, sticky=W, padx=5, pady=6)
        r += 1
        
        def _update_vars():
            set_vars_callback({
                "cinsiyet": cinsiyet_var.get(),
                "dogum_tarih": dogum_tarih_var.get(),
                "dogum_yeri": dogum_yeri_var.get(),
                "gebelik_sekli": gebelik_var.get(),
                "gebelik_sorun": gebelik_sorun_var.get(),
                "dogum_sekli": dogum_sekli_var.get(),
                "dogum_hafta": dogum_hafta_var.get(),
                "dogum_kilo": dogum_kilo_var.get(),
                "dogum_boy": dogum_boy_var.get(),
                "dogum_sorun": dogum_sorun_var.get(),
                "dogum_sorun_detay": dogum_sorun_detay_var.get(),
                "anne_sutu": anne_sutu_var.get(),
                "anne_sutu_sure": anne_sutu_sure_var.get(),
                "bakim_veren": bakim_veren_var.get(),
                "yurme_yas": yurme_yas_var.get(),
                "yurme_gec_neden": yurme_gec_neden_var.get(),
                "tuvalet_yas": tuvalet_yas_var.get(),
                "tuvalet_gec_neden": tuvalet_gec_neden_var.get(),
                "konusma_yas": konusma_yas_var.get(),
            })
        
        for v in [cinsiyet_var, dogum_tarih_var, dogum_yeri_var, gebelik_var, gebelik_sorun_var, dogum_sekli_var,
                  dogum_hafta_var, dogum_kilo_var, dogum_boy_var, dogum_sorun_var, dogum_sorun_detay_var,
                  anne_sutu_var, anne_sutu_sure_var, bakim_veren_var, yurme_yas_var, yurme_gec_neden_var,
                  tuvalet_yas_var, tuvalet_gec_neden_var, konusma_yas_var]:
            v.trace_add("write", lambda *a: _update_vars())
        _update_vars()
        
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
    
    def _build_cocuk_takip_sayfa2(self, parent, form_data, set_vars_callback):
        """Sayfa 2: Eğitim Bilgileri — soru üstte, seçenekler altta"""
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        form = ttk.Frame(scrollable, padding=(0, 4))
        form.pack(fill=X, pady=8)
        r, wrap = 0, 560

        ttk.Label(form, text="Geç konuştuysa nedenlerini belirtiniz:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        konusma_gec_neden_var = tk.StringVar()
        ttk.Entry(form, textvariable=konusma_gec_neden_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Gelişim döneminde yaygın gelişimsel bozukluk tanısı aldı mı?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        gdb_tani_var = tk.StringVar()
        self._paper_radio_group(form, gdb_tani_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=12).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Yanıtınız evetse hangi tanıyı almış belirtiniz:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        gdb_tani_detay_var = tk.StringVar()
        ttk.Entry(form, textvariable=gdb_tani_detay_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Eğitim Bilgileri", font=("Segoe UI", 11, "bold")).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(16, 8))
        r += 1
        ttk.Label(form, text="Okul Adı/İl/İlçe:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        okul_adi_var = tk.StringVar()
        ttk.Entry(form, textvariable=okul_adi_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Sınıfı:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        sinif_var = tk.StringVar()
        ttk.Entry(form, textvariable=sinif_var, width=24).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Alınan Eğitim türü: Zorunlu(örgün) temel eğitim / Özel eğitim / Her ikisi", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        egitim_turu_var = tk.StringVar()
        self._paper_radio_group(form, egitim_turu_var, [
            ("Zorunlu", "Zorunlu(örgün) temel eğitim"), ("Özel eğitim", "Özel eğitim"), ("Her ikisi", "Her ikisi")
        ], padx_between=10).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Örgün eğitim haricinde aldığı eğitim desteği (özel eğitim, özel ders, etüt vb.) var mı? Evet ise süresini yazınız.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        destek_egitim_var = tk.StringVar()
        destek_egitim_sure_var = tk.StringVar()
        fr_destek = ttk.Frame(form)
        fr_destek.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_destek, destek_egitim_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_destek, text="Süresini yazınız:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_destek, textvariable=destek_egitim_sure_var, width=22).pack(side=LEFT)
        r += 2
        ttk.Label(form, text="Sınıf (mentör/danışman) öğretmeninin adı soyadı:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        sinif_ogretmen_var = tk.StringVar()
        ttk.Entry(form, textvariable=sinif_ogretmen_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Okul öncesi eğitim aldı mı? Evetse kaç yıl yazınız.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        okuloncesi_var = tk.StringVar()
        okuloncesi_yil_var = tk.StringVar()
        fr_ok = ttk.Frame(form)
        fr_ok.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_ok, okuloncesi_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_ok, text="Evetse kaç yıl:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_ok, textvariable=okuloncesi_yil_var, width=10).pack(side=LEFT)
        r += 2
        ttk.Label(form, text="İlkokula başlama yaşını ay olarak yazınız:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        ilkokul_baslangic_var = tk.StringVar()
        ttk.Entry(form, textvariable=ilkokul_baslangic_var, width=14).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Eğitim olanaklarına ilişkin sorunlar yaşandı mı? (örn. Sağlık problemleri, hava koşulları nedeniyle okula gitmeme, sık öğretmen ve okul değişikliği, ailede yaşanan travmatik bir olaydan dolayı okula düzenli gidememe gibi)", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        egitim_sorun_var = tk.StringVar()
        ttk.Entry(form, textvariable=egitim_sorun_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Okumaya başlama zamanı: Okul öncesi / Birinci dönem / İkinci dönem / Daha sonraki dönemler / Okuyamıyor", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        okuma_baslangic_var = tk.StringVar()
        okb_opts = [("Okul öncesi", "Okul öncesi"), ("Birinci dönem", "Birinci dönem"), ("İkinci dönem", "İkinci dönem"), ("Daha sonraki dönemler", "Daha sonraki dönemler"), ("Okuyamıyor", "Okuyamıyor")]
        self._paper_radio_group(form, okuma_baslangic_var, okb_opts, padx_between=8).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Okumada sorunu var mı? (örn. Harf atlama, ters yazma, harf karıştırma gibi) Varsa belirtiniz.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        okuma_sorun_var = tk.StringVar()
        okuma_sorun_detay_var = tk.StringVar()
        fr_okum = ttk.Frame(form)
        fr_okum.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_okum, okuma_sorun_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_okum, text="Varsa belirtiniz:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_okum, textvariable=okuma_sorun_detay_var, width=28).pack(side=LEFT)
        r += 2
        ttk.Label(form, text="Okuduğunu anlamada sorunu var mı? Varsa belirtiniz.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        okuma_anlama_var = tk.StringVar()
        okuma_anlama_detay_var = tk.StringVar()
        fr_okan = ttk.Frame(form)
        fr_okan.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_okan, okuma_anlama_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_okan, text="Varsa belirtiniz:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_okan, textvariable=okuma_anlama_detay_var, width=28).pack(side=LEFT)
        r += 2
        ttk.Label(form, text="Yazmada sorunu var mı? (örn. Harf atlama, ters yazma, harf karıştırma gibi) Varsa belirtiniz.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        yazma_sorun_var = tk.StringVar()
        yazma_sorun_detay_var = tk.StringVar()
        fr_yaz = ttk.Frame(form)
        fr_yaz.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_yaz, yazma_sorun_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_yaz, text="Varsa belirtiniz:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_yaz, textvariable=yazma_sorun_detay_var, width=28).pack(side=LEFT)
        r += 2
        ttk.Label(form, text="Aritmetikte sorunu var mı? (örn. Sayıları tanıyamama, ters yazma, zihinden problem çözememe gibi) Varsa belirtiniz.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        aritmetik_sorun_var = tk.StringVar()
        aritmetik_sorun_detay_var = tk.StringVar()
        fr_ar = ttk.Frame(form)
        fr_ar.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_ar, aritmetik_sorun_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_ar, text="Varsa belirtiniz:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_ar, textvariable=aritmetik_sorun_detay_var, width=28).pack(side=LEFT)
        r += 2
        ttk.Label(form, text="Sıralamada sorunu var mı? (örn. Haftanın günlerini veya ayları sıralayamama, ileri geri ritmik sayma gibi) Varsa belirtiniz.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        siralama_sorun_var = tk.StringVar()
        siralama_sorun_detay_var = tk.StringVar()
        fr_sir = ttk.Frame(form)
        fr_sir.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_sir, siralama_sorun_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_sir, text="Varsa belirtiniz:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_sir, textvariable=siralama_sorun_detay_var, width=28).pack(side=LEFT)
        r += 2
        
        def _update_vars():
            set_vars_callback({
                "konusma_gec_neden": konusma_gec_neden_var.get(),
                "gdb_tani": gdb_tani_var.get(),
                "gdb_tani_detay": gdb_tani_detay_var.get(),
                "okul_adi": okul_adi_var.get(),
                "sinif": sinif_var.get(),
                "egitim_turu": egitim_turu_var.get(),
                "destek_egitim": destek_egitim_var.get(),
                "destek_egitim_sure": destek_egitim_sure_var.get(),
                "sinif_ogretmen": sinif_ogretmen_var.get(),
                "okuloncesi": okuloncesi_var.get(),
                "okuloncesi_yil": okuloncesi_yil_var.get(),
                "ilkokul_baslangic": ilkokul_baslangic_var.get(),
                "egitim_sorun": egitim_sorun_var.get(),
                "okuma_baslangic": okuma_baslangic_var.get(),
                "okuma_sorun": okuma_sorun_var.get(),
                "okuma_sorun_detay": okuma_sorun_detay_var.get(),
                "okuma_anlama": okuma_anlama_var.get(),
                "okuma_anlama_detay": okuma_anlama_detay_var.get(),
                "yazma_sorun": yazma_sorun_var.get(),
                "yazma_sorun_detay": yazma_sorun_detay_var.get(),
                "aritmetik_sorun": aritmetik_sorun_var.get(),
                "aritmetik_sorun_detay": aritmetik_sorun_detay_var.get(),
                "siralama_sorun": siralama_sorun_var.get(),
                "siralama_sorun_detay": siralama_sorun_detay_var.get(),
            })
        
        all_vars_s2 = [konusma_gec_neden_var, gdb_tani_var, gdb_tani_detay_var, okul_adi_var, sinif_var, egitim_turu_var,
                       destek_egitim_var, destek_egitim_sure_var, sinif_ogretmen_var, okuloncesi_var, okuloncesi_yil_var,
                       ilkokul_baslangic_var, egitim_sorun_var, okuma_baslangic_var, okuma_sorun_var, okuma_sorun_detay_var,
                       okuma_anlama_var, okuma_anlama_detay_var, yazma_sorun_var, yazma_sorun_detay_var,
                       aritmetik_sorun_var, aritmetik_sorun_detay_var, siralama_sorun_var, siralama_sorun_detay_var]
        for v in all_vars_s2:
            if hasattr(v, "trace_add"):
                v.trace_add("write", lambda *a: _update_vars())
        _update_vars()
        
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
    
    def _build_cocuk_takip_sayfa3(self, parent, form_data, set_vars_callback):
        """Sayfa 3: Demografik Bilgiler — soru üstte, seçenekler altta"""
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        form = ttk.Frame(scrollable, padding=(0, 4))
        form.pack(fill=X, pady=8)
        r, wrap = 0, 560

        ttk.Label(form, text="Yönleri ayırt etmede sorunu var mı? (örn. Sağını solunu karıştırma veya ayırt edememe gibi) Varsa belirtiniz.", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        yon_ayirt_var = tk.StringVar()
        yon_ayirt_detay_var = tk.StringVar()
        fr_yon = ttk.Frame(form)
        fr_yon.grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        self._paper_radio_group(fr_yon, yon_ayirt_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=8).pack(side=LEFT)
        ttk.Label(fr_yon, text="Varsa belirtiniz:").pack(side=LEFT, padx=(12, 4))
        ttk.Entry(fr_yon, textvariable=yon_ayirt_detay_var, width=28).pack(side=LEFT)
        r += 2
        
        ttk.Label(form, text="Aşağıdaki alanlara en son karne notlarını yazınız:", font=("Segoe UI", 10, "bold"), wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(12, 6))
        r += 1
        ttk.Label(form, text="Türkçe:", wraplength=wrap).grid(row=r, column=0, sticky=W, padx=5, pady=(4, 2))
        karneturkce_var = tk.StringVar()
        ttk.Entry(form, textvariable=karneturkce_var, width=14).grid(row=r, column=1, sticky=W, padx=5, pady=(4, 2))
        r += 1
        ttk.Label(form, text="Matematik:", wraplength=wrap).grid(row=r, column=0, sticky=W, padx=5, pady=(4, 2))
        karnematematik_var = tk.StringVar()
        ttk.Entry(form, textvariable=karnematematik_var, width=14).grid(row=r, column=1, sticky=W, padx=5, pady=(4, 2))
        r += 1
        ttk.Label(form, text="Hayat Bilgisi:", wraplength=wrap).grid(row=r, column=0, sticky=W, padx=5, pady=(4, 2))
        karnehayatbilgisi_var = tk.StringVar()
        ttk.Entry(form, textvariable=karnehayatbilgisi_var, width=14).grid(row=r, column=1, sticky=W, padx=5, pady=(4, 2))
        r += 1
        ttk.Label(form, text="Sosyal Bilgiler:", wraplength=wrap).grid(row=r, column=0, sticky=W, padx=5, pady=(4, 2))
        karnesosyal_var = tk.StringVar()
        ttk.Entry(form, textvariable=karnesosyal_var, width=14).grid(row=r, column=1, sticky=W, padx=5, pady=(4, 2))
        r += 1
        ttk.Label(form, text="Fen bilgisi:", wraplength=wrap).grid(row=r, column=0, sticky=W, padx=5, pady=(4, 2))
        karnefen_var = tk.StringVar()
        ttk.Entry(form, textvariable=karnefen_var, width=14).grid(row=r, column=1, sticky=W, padx=5, pady=(4, 2))
        r += 1
        
        ttk.Label(form, text="Demografik Bilgiler", font=("Segoe UI", 11, "bold")).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(20, 10))
        r += 1
        ttk.Label(form, text="Ailenin kaçıncı çocuğu:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        aile_sira_var = tk.StringVar()
        ttk.Entry(form, textvariable=aile_sira_var, width=14).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Anne baba arasında akrabalık var mı?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        akrabalik_var = tk.StringVar()
        akrabalik_detay_var = tk.StringVar()
        self._paper_radio_group(form, akrabalik_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=12).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Varsa nasıl bir akrabalık olduğunu yazınız: (Örneğin; amca, hala, dayı, teyze çocuğu ya da anne baba veya babanın amca, hala, dayı, teyze çocukları gibi)", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        ttk.Entry(form, textvariable=akrabalik_detay_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Çocuğa şuan temel bakım veren kişi veya kişiler:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        bakim_veren_suan_var = tk.StringVar()
        bakim_suan_opts = [
            ("Anne ve baba", "Anne ve baba"), ("Sadece anne", "Sadece anne"), ("Sadece baba", "Sadece baba"),
            ("Bakıcı", "Bakıcı"), ("Büyük anne", "Büyük anne"), ("Anneanne", "Anneanne"), ("Akrabalar", "Akrabalar"),
            ("Edinilmiş aile", "Edinilmiş aile"), ("Koruyucu aile", "Koruyucu aile"), ("Kurum", "Kurum"), ("Diğer", "Diğer")
        ]
        self._paper_radio_group_wrapped(form, bakim_veren_suan_var, bakim_suan_opts, padx_between=6, columns=4).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Ailede anne, baba ve çocuklar dışında yaşayan biri var mı? Varsa kim olduğunu yazınız:", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        aile_disinda_var = tk.StringVar()
        ttk.Entry(form, textvariable=aile_disinda_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Aile Türü: Çekirdek / Geniş / Sadece anne / Sadece baba", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        aile_turu_var = tk.StringVar()
        self._paper_radio_group(form, aile_turu_var, [
            ("Çekirdek", "Çekirdek"), ("Geniş", "Geniş"), ("Sadece anne", "Sadece anne"), ("Sadece baba", "Sadece baba")
        ], padx_between=8).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Yukarıdaki maddede son iki şıktan biri seçilmişse, aşağıdakileri yanıtlayınız: Anne ve baba ayrı yaşıyor / Anne ve boşanmış / Anne veya babadan biri ölmüş", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        ayrilik_durum_var = tk.StringVar()
        self._paper_radio_group(form, ayrilik_durum_var, [
            ("Ayrı yaşıyor", "Anne ve baba ayrı yaşıyor"), ("Boşanmış", "Anne ve boşanmış"), ("Ölmüş", "Anne veya babadan biri ölmüş")
        ], padx_between=8).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        ttk.Label(form, text="Ailenin sosyoekonomik düzeyi: Alt / Orta / Üst", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        sosyoekonomik_var = tk.StringVar()
        self._paper_radio_group(form, sosyoekonomik_var, [("Alt", "Alt"), ("Orta", "Orta"), ("Üst", "Üst")], padx_between=8).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Annenin eğitim durumu nedir?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        anne_egitim_var = tk.StringVar()
        ttk.Entry(form, textvariable=anne_egitim_var, width=44).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Annenin yaşı nedir?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        anne_yas_var = tk.StringVar()
        ttk.Entry(form, textvariable=anne_yas_var, width=14).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Annenin yaptığı iş nedir?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        anne_is_var = tk.StringVar()
        ttk.Entry(form, textvariable=anne_is_var, width=44).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Babanın eğitim durumu nedir?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        baba_egitim_var = tk.StringVar()
        ttk.Entry(form, textvariable=baba_egitim_var, width=44).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Babanın yaşı nedir?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        baba_yas_var = tk.StringVar()
        ttk.Entry(form, textvariable=baba_yas_var, width=14).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Babanın yaptığı iş nedir?", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        baba_is_var = tk.StringVar()
        ttk.Entry(form, textvariable=baba_is_var, width=44).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Ailedeki çocuk sayısı (yaşları, cinsiyetleri ve kaçıncı çocuk olduklarını yazınız)", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        cocuk_sayisi_detay_var = tk.StringVar()
        ttk.Entry(form, textvariable=cocuk_sayisi_detay_var, width=52).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        ttk.Label(form, text="Hastalık tanısı almış kardeşi var mı? (dahili, nörolojik, psikiyatrik vb.)", wraplength=wrap).grid(row=r, column=0, columnspan=2, sticky=W, padx=5, pady=(6, 2))
        hasta_kardes_var = tk.StringVar()
        self._paper_radio_group(form, hasta_kardes_var, [("Evet", "Evet"), ("Hayır", "Hayır")], padx_between=12).grid(row=r+1, column=0, columnspan=2, sticky=W, padx=(24, 5), pady=(0, 6))
        r += 2
        
        def _update_vars():
            set_vars_callback({
                "yon_ayirt": yon_ayirt_var.get(),
                "yon_ayirt_detay": yon_ayirt_detay_var.get(),
                "karneturkce": karneturkce_var.get(),
                "karnematematik": karnematematik_var.get(),
                "karnehayatbilgisi": karnehayatbilgisi_var.get(),
                "karnesosyal": karnesosyal_var.get(),
                "karnefen": karnefen_var.get(),
                "aile_sira": aile_sira_var.get(),
                "akrabalik": akrabalik_var.get(),
                "akrabalik_detay": akrabalik_detay_var.get(),
                "bakim_veren_suan": bakim_veren_suan_var.get(),
                "aile_disinda": aile_disinda_var.get(),
                "aile_turu": aile_turu_var.get(),
                "ayrilik_durum": ayrilik_durum_var.get(),
                "sosyoekonomik": sosyoekonomik_var.get(),
                "anne_egitim": anne_egitim_var.get(),
                "anne_yas": anne_yas_var.get(),
                "anne_is": anne_is_var.get(),
                "baba_egitim": baba_egitim_var.get(),
                "baba_yas": baba_yas_var.get(),
                "baba_is": baba_is_var.get(),
                "cocuk_sayisi_detay": cocuk_sayisi_detay_var.get(),
                "hasta_kardes": hasta_kardes_var.get(),
            })
        
        all_vars_s3 = [yon_ayirt_var, yon_ayirt_detay_var, karneturkce_var, karnematematik_var, karnehayatbilgisi_var,
                       karnesosyal_var, karnefen_var, aile_sira_var, akrabalik_var, akrabalik_detay_var,
                       bakim_veren_suan_var, aile_disinda_var, aile_turu_var, ayrilik_durum_var, sosyoekonomik_var,
                       anne_egitim_var, anne_yas_var, anne_is_var, baba_egitim_var, baba_yas_var, baba_is_var,
                       cocuk_sayisi_detay_var, hasta_kardes_var]
        for v in all_vars_s3:
            if hasattr(v, "trace_add"):
                v.trace_add("write", lambda *a: _update_vars())
        _update_vars()
        
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
    
    def safe_call(self, fn, label="UI"):
        def wrapper(*args, **kwargs):
            try:
                if not callable(fn):
                    raise TypeError(f"safe_call: fn callable değil -> {type(fn)} / {fn!r}")
                return fn(*args, **kwargs)
            except Exception as e:
                try:
                    log_exception(label, e)
                except Exception:
                    pass
                try:
                    messagebox.showerror("Hata", f"{label}\n\n{e}")
                except Exception:
                    pass
                return None
        return wrapper
    
    def _tk_report_callback_exception(self, exc, val, tb):
        try:
            from core.logging_utils import log_exception
            import traceback

            err = val if isinstance(val, BaseException) else Exception(str(val))
            log_exception("TK_CALLBACK", err)

            # ayrıca dosyaya detay bas
            try:
                with open("leta_error.log", "a", encoding="utf-8") as f:
                    f.write("\n--- TK CALLBACK TRACEBACK ---\n")
                    traceback.print_exception(exc, val, tb, file=f)
            except Exception:
                pass
        except Exception:
            pass


class RegisterDialog(ttk.Toplevel):
    def __init__(self, parent: App, first_setup: bool = False):
        super().__init__(parent)
        self.parent = parent
        self.first_setup = first_setup
        self.title("Yeni Kullanıcı Kaydı")
        self.geometry("420x420")
        self.transient(parent)
        self.grab_set()

        header = "İLK KURULUM - KURUM MÜDÜRÜ" if self.first_setup else "YENİ KULLANICI KAYDI"
        ttk.Label(self, text=header, font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        frm = ttk.Frame(self, padding=16)
        frm.pack(fill=BOTH, expand=True)

        def field(label, show=None):
            ttk.Label(frm, text=label).pack(anchor=W, pady=(8, 0))
            e = ttk.Entry(frm, show=show or "")
            e.pack(fill=X, pady=4)
            return e

        self.ent_user = field("Kullanıcı Adı *:")
        self.ent_name = field("Ad Soyad:")
        self.ent_email = field("E-posta:")
        self.ent_pw = field("Şifre *:", show="*")
        self.ent_pw2 = field("Şifre Tekrar *:", show="*")

        note = "Not: Bu sürümde tüm kullanıcılar Kurum Müdürü yetkisiyle açılır."
        ttk.Label(frm, text=note, foreground="gray").pack(pady=(10, 0))

        def _save():
            u = (self.ent_user.get() or "").strip()
            pw = self.ent_pw.get() or ""
            pw2 = self.ent_pw2.get() or ""
            if not u or not pw:
                messagebox.showwarning("Uyarı", "Kullanıcı adı ve şifre zorunludur!")
                return
            if pw != pw2:
                messagebox.showwarning("Uyarı", "Şifreler eşleşmiyor!")
                return
            if u.lower() == "name":
                messagebox.showwarning("Uyarı", "Name Hoca kurumdan ayrıldı. Bu kullanıcı adı kullanılamaz.")
                return
            try:
                conn = parent.veritabani_baglan()
                cur = conn.cursor()
                # güvenlik: ilk kurulum sadece DB boşken
                if self.first_setup:
                    cur.execute("SELECT COUNT(*) FROM users WHERE is_active=1")
                    if int((cur.fetchone() or [0])[0] or 0) > 0:
                        conn.close()
                        messagebox.showwarning("Uyarı", "İlk kurulum zaten tamamlanmış.")
                        self.destroy()
                        parent._refresh_first_run_state()
                        return
                cur.execute("SELECT COUNT(*) FROM users WHERE username=?", (u,))
                if (cur.fetchone() or [0])[0] > 0:
                    conn.close()
                    messagebox.showerror("Hata", "Bu kullanıcı adı zaten var!")
                    return
                access_role = "kurum_muduru"
                cur.execute(
                    """
                    INSERT INTO users (username, password_hash, role, access_role, title_role, full_name, email, created_at, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
                    """,
                    (
                        u,
                        hash_pass(pw),
                        access_role,        # legacy role (back-compat)
                        access_role,        # yetki
                        "",                 # unvan (sonradan kullanıcı yönetiminden set edilir)
                        (self.ent_name.get() or "").strip(),
                        (self.ent_email.get() or "").strip(),
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Hata", f"Kayıt oluşturulamadı:\n{e}")
                return
            messagebox.showinfo("Başarılı", "Kayıt oluşturuldu. Artık giriş yapabilirsiniz.")
            self.destroy()
            try:
                parent._refresh_first_run_state()
            except Exception:
                pass

        ttk.Button(frm, text="KAYDET", bootstyle="success", command=_save).pack(fill=X, pady=(14, 0))


    def _relaunch_cmd(extra_args: list[str] | None = None) -> list[str]: 
     extra_args = extra_args or []
     if getattr(sys, "frozen", False):
        return [sys.executable, *extra_args]
     return [sys.executable, os.path.abspath(__file__), *extra_args]


    def _run_reset_worker() -> None:
   
     deadline = time.time() + 60.0
     last = ""
     while time.time() < deadline:
        ok, msg = safe_delete_db_files()
        last = msg
        db = db_path()
        if (not os.path.exists(db)) and (not os.path.exists(db + "-wal")) and (not os.path.exists(db + "-shm")):
            break
        # Silme başarısız olsa bile, kilit kalkana kadar tekrar dene
        time.sleep(0.25)
        if ok and (msg or "").startswith("RENAMED:"):
            # rename ile DB fiilen devre dışı kaldı; yeni DB açılabilir
            break
     try:
        with open(error_log_path(), "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | RESET_WORKER | {last}\n")
     except Exception:
        pass
     spawn_detached(_relaunch_cmd())

    def get_personel_cuzdan(self, personel_adi: str) -> dict:
        """
        app_ui'nin beklediği:
        {"beklemede_toplam": float, "odendi_toplam": float, "toplam_hak_edis": float}
        """
        def table_exists(name: str) -> bool:
            try:
                self.cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (name,))
                return self.cur.fetchone() is not None
            except Exception:
                return False

        out = {"beklemede_toplam": 0.0, "odendi_toplam": 0.0, "toplam_hak_edis": 0.0}
        if not table_exists("personel_ucret_takibi"):
            return out

        try:
            p = (personel_adi or "").strip()
            self.cur.execute(
                """
                SELECT
                    COALESCE(SUM(CASE WHEN COALESCE(odeme_durumu,'')='beklemede' THEN personel_ucreti ELSE 0 END),0) AS beklemede,
                    COALESCE(SUM(CASE WHEN COALESCE(odeme_durumu,'')='odendi' THEN personel_ucreti ELSE 0 END),0) AS odendi,
                    COALESCE(SUM(personel_ucreti),0) AS toplam
                FROM personel_ucret_takibi
                WHERE TRIM(COALESCE(personel_adi,'')) = ?
                """,
                (p,),
            )
            row = self.cur.fetchone()
            if row:
                out["beklemede_toplam"] = float(row[0] or 0.0)
                out["odendi_toplam"] = float(row[1] or 0.0)
                out["toplam_hak_edis"] = float(row[2] or 0.0)
        except Exception as e:
            try:
                log_exception("get_personel_cuzdan", e)
            except Exception:
                pass
        return out

    

    
